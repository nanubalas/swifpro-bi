import re
from decimal import Decimal, InvalidOperation
from django.db import transaction, IntegrityError
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.http import Http404, HttpResponse
from django.core.mail import EmailMessage
from django.template.loader import render_to_string

from core.models import (
    Tenant, Location, PurchaseOrder, PurchaseOrderLine, PurchaseOrderAmendment, Shipment, ShipmentLine, Container, ShipmentEvent,
    PurchaseRequisition, PurchaseRequisitionLine,
    InventoryBalance, InventoryMovement, ChannelSnapshot, SalesChannel, Product,
    Supplier, ChannelConnection, SalesOrder, SalesOrderLine,
    ProductBarcode, UnitOfMeasure, UOMConversion, BillOfMaterials, BillOfMaterialsLine,
    InventoryTransfer, InventoryTransferLine,
    GoodsReceipt, GoodsReceiptLine, LandedCostCharge,
    SupplierInvoice, SupplierInvoiceLine,
    ReturnAuthorization, ReturnLine,
    TaxCode, Customer, CustomerInvoice, CustomerInvoiceLine,
    GLAccount, JournalEntry, Payment, PaymentAllocation, VatReturn, Expense,
    CreditNote, CreditNoteLine, BankTransaction,
    SalesQuote, SalesQuoteLine, CustomerOrder, CustomerOrderLine,
    RecurringInvoice, RecurringInvoiceLine,
    OrgMembership, AuditLog, AccessRequest, UserProfile, UserPermissionOverride
)
from django.core.exceptions import PermissionDenied
from core import roles as roles_mod
from core.access import (
    get_active_role, get_memberships, default_landing_url, SESSION_TENANT_KEY,
    SESSION_LOCATION_KEY, SESSION_SITE_KEY, get_active_location, get_active_site,
    selectable_locations, selectable_sites, active_site_id,
    can_access_company, can_access_site, active_location_ids,
)
from core.audit import log_audit
from core.forms import (
    PurchaseOrderForm, PurchaseOrderLineFormSet,
    PurchaseRequisitionForm, PurchaseRequisitionLineFormSet,
    ShipmentUpdateForm, ProductForm, SupplierForm,
    LocationForm, ChannelConnectionForm,
    SalesOrderForm, SalesOrderLineFormSet, TenantSettingsForm,
    UnitOfMeasureForm, UOMConversionForm, BillOfMaterialsForm, BOMLineFormSet,
    InventoryTransferForm, InventoryTransferLineFormSet,
    GoodsReceiptForm, GoodsReceiptLineFormSet, LandedCostChargeForm,
    SupplierInvoiceForm, SupplierInvoiceLineFormSet,
    ReturnAuthorizationForm, ReturnLineFormSet,
    TaxCodeForm, CustomerForm,
    CustomerInvoiceForm, CustomerInvoiceLineFormSet,
    GLAccountForm, ReceiptForm, SupplierPaymentForm, RefundForm, AccessRequestForm,
    NewOrganisationForm, InviteUserForm, ExpenseForm,
    CreditNoteForm, CreditNoteLineFormSet, BankTransactionForm,
    SalesQuoteForm, SalesQuoteLineFormSet, CustomerOrderForm, CustomerOrderLineFormSet,
    RecurringInvoiceForm, RecurringInvoiceLineFormSet
)
from core.services.inventory import apply_movement, reserve_stock, release_reservations, consume_reservations
from core.services.bom import explode_product
from core.services.gl import post_customer_invoice, post_supplier_invoice, post_payment, post_inventory_receipt, post_cogs, post_expense, post_credit_note
from core.services import reports as reports_service
from core.services import vat as vat_service
from core.services import importer as importer_service
from django.db.utils import OperationalError
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from core.auth import role_required, permission_required, effective_groups, ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_SALES, ROLE_FINANCE, ROLE_READONLY
from core import permissions as permissions_mod




def _get_default_tenant(request=None):
    # Resolve the active tenant for the request (session org -> membership ->
    # profile -> first tenant). See core.access.
    from core.access import get_active_tenant
    return get_active_tenant(request)


def _scope_location_fields(form, request, tenant, *field_names):
    """Narrow a form's location dropdown(s) to the locations the user may use
    across the whole company (any site). Used for cross-site destinations."""
    from core.access import accessible_location_ids
    allowed = accessible_location_ids(getattr(request, "user", None), tenant)
    if allowed is None:
        return
    for name in field_names:
        field = form.fields.get(name)
        if field is not None and getattr(field, "queryset", None) is not None:
            field.queryset = field.queryset.filter(id__in=allowed)


def _scope_location_fields_by_site(form, request, *field_names):
    """Narrow a form's location dropdown(s) to the inventory locations under the
    SELECTED site (the working context). Inventory workflows pick a location here."""
    allowed = active_location_ids(request)  # accessible locations under the active site
    if allowed is None:
        return
    for name in field_names:
        field = form.fields.get(name)
        if field is not None and getattr(field, "queryset", None) is not None:
            field.queryset = field.queryset.filter(id__in=allowed)


def _can_access_location(request, tenant, *location_ids):
    """True if the user may use every given location (None ids are ignored).

    Mirrors accessible_location_ids: Admins/superusers/users with no grants are
    unrestricted; otherwise the location must be in their granted set."""
    from core.access import accessible_location_ids
    allowed = accessible_location_ids(getattr(request, "user", None), tenant)
    if allowed is None:
        return True
    return all(lid in allowed for lid in location_ids if lid is not None)


def _generate_po_number():
    return "PO-" + timezone.now().strftime("%Y%m%d-%H%M%S-%f")


def _default_vat_rate(tenant):
    """Standard VAT rate for the tenant (from the 'STD' tax code), default 20%."""
    code = TaxCode.objects.filter(tenant=tenant, code="STD", is_active=True).first()
    return code.rate if code else Decimal("0.20")


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_FINANCE, ROLE_READONLY])

def po_list(request):
    tenant = _get_default_tenant(request)
    pos = PurchaseOrder.objects.filter(tenant=tenant).order_by("-created_at")
    site = active_site_id(request)  # scope to the selected site
    if site:
        pos = pos.filter(site_id=site)
    return render(request, "po_list.html", {"tenant": tenant, "pos": pos})

@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_FINANCE, ROLE_READONLY])

def po_print(request, po_id):
    tenant = _get_default_tenant(request)
    po = get_object_or_404(PurchaseOrder, id=po_id, tenant=tenant)
    
    subtotal = sum((line.line_total for line in po.lines.all()), Decimal("0.00"))
    vat_rate = _default_vat_rate(tenant)
    vat_amount = subtotal * vat_rate
    total = subtotal + vat_amount

    return render(request, "po_print.html", {
        "tenant": tenant,
        "po": po,
        "subtotal": subtotal,
        "vat_amount": vat_amount,
        "total": total,
        "vat_rate_percent": vat_rate * 100,
    })


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_FINANCE, ROLE_READONLY])
def po_pdf(request, po_id):
    tenant = _get_default_tenant(request)
    po = get_object_or_404(PurchaseOrder, id=po_id, tenant=tenant)
    from core.services.pdf import pdf_response
    return pdf_response(f"purchase-order-{po.po_number}.pdf", "documents/po_pdf.html",
                        {"tenant": tenant, "po": po, "doc_title": "PURCHASE ORDER",
                         "number": po.po_number, "notes": po.notes}, download=False)


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_FINANCE, ROLE_READONLY])
def supplier_prices_json(request, supplier_id):
    """Last known unit cost per product for a supplier (for PO price prefill)."""
    from django.http import JsonResponse
    from core.services.purchasing import last_prices_for_supplier
    tenant = _get_default_tenant(request)
    supplier = get_object_or_404(Supplier, id=supplier_id, tenant=tenant)
    prices = {str(pid): str(cost) for pid, cost in last_prices_for_supplier(tenant, supplier).items()}
    return JsonResponse({"prices": prices})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_READONLY])
def po_backorders(request):
    """Outstanding PO lines: ordered but not yet fully received (backorders)."""
    tenant = _get_default_tenant(request)
    rows = []
    lines = (PurchaseOrderLine.objects
             .filter(po__tenant=tenant, po__is_current=True)
             .exclude(po__status__in=[PurchaseOrder.Status.CANCELLED, PurchaseOrder.Status.CLOSED, PurchaseOrder.Status.DRAFT])
             .select_related("po", "po__supplier", "product"))
    for l in lines:
        if l.open_qty and l.open_qty > 0:
            rows.append({"line": l, "po": l.po, "open_qty": l.open_qty})
    rows.sort(key=lambda r: (r["po"].expected_date or timezone.localdate(), r["po"].po_number))
    return render(request, "po/backorders.html", {"tenant": tenant, "rows": rows})



@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
def po_send(request, po_id):
    tenant = _get_default_tenant(request)
    po = get_object_or_404(PurchaseOrder, id=po_id, tenant=tenant)

    if request.method != "POST":
        return redirect("po_detail", po_id=po.id)

    if po.status in [PurchaseOrder.Status.CANCELLED, PurchaseOrder.Status.CLOSED]:
        messages.error(request, "Cannot send a cancelled/closed PO.")
        return redirect("po_detail", po_id=po.id)

    supplier_email = getattr(po.supplier, "email", None)
    if not supplier_email:
        messages.error(request, "Supplier has no email address. Add one and try again.")
        return redirect("po_detail", po_id=po.id)

    subject = f"Purchase Order {po.po_number}"
    html_body = render_to_string("po_email.html", {"tenant": tenant, "po": po})
    msg = EmailMessage(subject=subject, body=html_body, to=[supplier_email])
    msg.content_subtype = "html"
    # Attach the PO as a PDF.
    from core.services.pdf import render_to_pdf
    pdf = render_to_pdf("documents/po_pdf.html", {
        "tenant": tenant, "po": po, "doc_title": "PURCHASE ORDER",
        "number": po.po_number, "notes": po.notes})
    if pdf:
        msg.attach(f"purchase-order-{po.po_number}.pdf", pdf, "application/pdf")
    try:
        msg.send(fail_silently=False)
    except Exception as e:
        messages.error(request, f"Email failed: {e}")
        return redirect("po_detail", po_id=po.id)

    po.sent_to = supplier_email
    po.sent_at = timezone.now()
    po.sent_subject = subject
    po.status = PurchaseOrder.Status.SENT if po.status in [PurchaseOrder.Status.SUBMITTED, PurchaseOrder.Status.APPROVED, PurchaseOrder.Status.APPROVAL_PENDING] else po.status
    po.save(update_fields=["sent_to", "sent_at", "sent_subject", "status"])

    messages.success(request, f"PO emailed to {supplier_email}.")
    return redirect("po_detail", po_id=po.id)



@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
def po_amend(request, po_id):
    tenant = _get_default_tenant(request)
    po = get_object_or_404(PurchaseOrder, id=po_id, tenant=tenant)

    if request.method != "POST":
        return redirect("po_detail", po_id=po.id)

    if po.status in [PurchaseOrder.Status.DRAFT]:
        messages.info(request, "Draft POs can be edited directly.")
        return redirect("po_detail", po_id=po.id)

    if po.status in [PurchaseOrder.Status.CANCELLED, PurchaseOrder.Status.CLOSED]:
        messages.error(request, "Cannot amend a cancelled/closed PO.")
        return redirect("po_detail", po_id=po.id)

    reason = request.POST.get("reason", "").strip()
    if not reason:
        messages.error(request, "Amendment reason is required.")
        return redirect("po_detail", po_id=po.id)

    # New version needs a unique po_number (Meta.unique_together = tenant, po_number).
    # Derive a stable base by stripping any prior "-vN" suffix, then re-version.
    new_version = po.version + 1
    base_number = re.sub(r"-v\d+$", "", po.po_number)
    new_number = f"{base_number}-v{new_version}"

    with transaction.atomic():
        # Create new version
        new_po = PurchaseOrder.objects.create(
            tenant=tenant,
            po_number=new_number,
            supplier=po.supplier,
            currency_code=po.currency_code,
            version=new_version,
            supersedes=po,
            is_current=True,
            status=PurchaseOrder.Status.DRAFT,
            expected_date=po.expected_date,
            notes=po.notes,
        )
        from core.services.purchasing import sync_po_line_received
        for line in po.lines.all():
            new_line = PurchaseOrderLine.objects.create(
                po=new_po,
                product=line.product,
                ordered_qty=line.ordered_qty,
                unit_cost=line.unit_cost,
                # Carry the stable identity forward so receipts against any
                # version feed this line's open/received quantities (M17).
                root_line_id=(line.root_line_id or line.id),
            )
            # received_qty is a cache: rebuild it from actual receipts rather
            # than statically copying the old version's counter.
            sync_po_line_received(new_line)

        po.is_current = False
        po.save(update_fields=["is_current"])

        PurchaseOrderAmendment.objects.create(
            tenant=tenant,
            from_po=po,
            to_po=new_po,
            reason=reason,
            created_by=request.user,
        )

    messages.success(request, f"Created PO amendment v{new_po.version}. Update it and submit again.")
    return redirect("po_detail", po_id=new_po.id)



@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
def po_cancel(request, po_id):
    tenant = _get_default_tenant(request)
    po = get_object_or_404(PurchaseOrder, id=po_id, tenant=tenant)

    if request.method != "POST":
        return redirect("po_detail", po_id=po.id)

    if po.status in [PurchaseOrder.Status.CANCELLED, PurchaseOrder.Status.CLOSED]:
        messages.info(request, "PO is already cancelled/closed.")
        return redirect("po_detail", po_id=po.id)

    reason = request.POST.get("reason", "").strip() or "Cancelled"
    any_received = any((l.received_qty > Decimal("0.00") for l in po.lines.all()))
    if any_received:
        # Cancel remaining qty only (close open qty by setting ordered=received for remaining)
        for l in po.lines.all():
            if l.open_qty > Decimal("0.00"):
                l.ordered_qty = l.received_qty
                l.save(update_fields=["ordered_qty"])
        po.status = PurchaseOrder.Status.CLOSED if all((l.open_qty == Decimal("0.00") for l in po.lines.all())) else po.status
        po.cancelled_reason = reason
        po.save(update_fields=["status", "cancelled_reason"])
        messages.warning(request, "Partial receipts exist. Cancelled remaining quantities (PO effectively closed for open lines).")
        return redirect("po_detail", po_id=po.id)

    po.status = PurchaseOrder.Status.CANCELLED
    po.cancelled_reason = reason
    po.save(update_fields=["status", "cancelled_reason"])
    messages.success(request, "PO cancelled.")
    return redirect("po_detail", po_id=po.id)



def _safe_default_tenant(request=None):
    try:
        return _get_default_tenant(request)
    except OperationalError:
        return None

@login_required
def landing(request):
    """Post-login dispatcher: route the user to their role dashboard (or the
    org picker when they belong to more than one organisation)."""
    tenant = _safe_default_tenant(request)
    if not tenant:
        return render(request, "landing.html", {"tenant": None, "needs_setup": True})

    memberships = get_memberships(request.user)
    if len(memberships) > 1 and not request.session.get(SESSION_TENANT_KEY):
        return redirect("select_org")

    # Best-effort daily housekeeping: expire stale quotes, generate due
    # recurring invoices (throttled to once per day per tenant).
    from core.services import housekeeping
    housekeeping.opportunistic(request)

    role = get_active_role(request)
    return redirect(default_landing_url(tenant, role))


def _select_company(request, tid, *, switching):
    """Set the active company, clearing any selected site (the user must pick a
    site for the new company). Returns True on success; 403s are handled by the
    caller. Audits the (un)authorised access."""
    if not can_access_company(request.user, tid):
        log_audit(action="UNAUTHORISED_COMPANY_ACCESS", request=request, user=request.user,
                  detail=f"tenant_id={tid}")
        return False
    request.session[SESSION_TENANT_KEY] = int(tid)
    # Switching company clears the selected site (and any stale workflow location).
    request.session.pop(SESSION_SITE_KEY, None)
    request.session.pop(SESSION_LOCATION_KEY, None)
    tenant = OrgMembership.objects.filter(user=request.user, tenant_id=tid).first()
    log_audit(action="COMPANY_SWITCHED" if switching else "COMPANY_SELECTED",
              request=request, user=request.user, tenant=getattr(tenant, "tenant", None),
              detail=f"tenant_id={tid}")
    return True


@login_required
def select_org(request):
    """Company chooser for users belonging to multiple organisations."""
    memberships = get_memberships(request.user)
    if request.method == "POST":
        tid = request.POST.get("tenant")
        if tid and _select_company(request, tid, switching=False):
            return redirect("landing")  # landing then resolves the site
        raise PermissionDenied("You do not have access to that company.")
    return render(request, "select_org.html", {"memberships": memberships})


@login_required
def select_site(request):
    """Mandatory Site chooser - the global context is Company + Site. There is no
    'all sites' option, and inventory locations are NOT chosen here."""
    from core.access import get_active_tenant
    tenant = get_active_tenant(request)
    if tenant is None:
        return redirect("landing")
    sites = list(selectable_sites(request.user, tenant))
    next_url = request.POST.get("next") or request.GET.get("next") or ""
    if request.method == "POST":
        sid = request.POST.get("site")
        if sid and can_access_site(request.user, tenant, sid):
            request.session[SESSION_SITE_KEY] = int(sid)
            request.session.pop(SESSION_LOCATION_KEY, None)  # drop stale workflow location
            site = next((s for s in sites if str(s.id) == str(sid)), None)
            log_audit(action="SITE_SELECTED", request=request, user=request.user, tenant=tenant,
                      detail=getattr(site, "name", sid))
            return redirect(_safe_next(next_url) or default_landing_url(tenant, get_active_role(request)))
        log_audit(action="UNAUTHORISED_SITE_ACCESS", request=request, user=request.user, tenant=tenant,
                  detail=f"site_id={sid}")
        raise PermissionDenied("You do not have access to that site.")
    if not sites:
        return redirect("no_site")
    if len(sites) == 1:  # nothing to choose - take it and move on
        request.session[SESSION_SITE_KEY] = sites[0].id
        return redirect("landing")
    return render(request, "select_site.html", {
        "tenant": tenant, "sites": sites, "next": next_url})


@login_required
def switch_company(request):
    """Header company switcher: set company, clear site, force a site choice."""
    if request.method == "POST":
        tid = request.POST.get("tenant")
        if tid and _select_company(request, tid, switching=True):
            return redirect("landing")
        raise PermissionDenied("You do not have access to that company.")
    return redirect("landing")


@login_required
def switch_site(request):
    """Header site switcher: change the working site, staying on the same page
    when possible."""
    from core.access import get_active_tenant
    tenant = get_active_tenant(request)
    if request.method == "POST" and tenant is not None:
        sid = request.POST.get("site")
        next_url = _safe_next(request.POST.get("next") or "")
        if sid and can_access_site(request.user, tenant, sid):
            request.session[SESSION_SITE_KEY] = int(sid)
            request.session.pop(SESSION_LOCATION_KEY, None)  # clear stale workflow location
            site = selectable_sites(request.user, tenant).filter(id=sid).first()
            log_audit(action="SITE_SWITCHED", request=request, user=request.user, tenant=tenant,
                      detail=getattr(site, "name", sid))
            return redirect(next_url or default_landing_url(tenant, get_active_role(request)))
        log_audit(action="UNAUTHORISED_SITE_ACCESS", request=request, user=request.user, tenant=tenant,
                  detail=f"site_id={sid}")
        raise PermissionDenied("You do not have access to that site.")
    return redirect("landing")


@login_required
def switch_workspace(request):
    """Combined company + site switcher used by the workspace picker.

    Validates access to BOTH the chosen company and a site within it, then
    applies them atomically (so the session is never left with a company but no
    matching site). Stays on the current page when possible."""
    from core.access import get_active_tenant
    if request.method != "POST":
        return redirect("landing")
    tid = request.POST.get("tenant")
    sid = request.POST.get("site")
    next_url = _safe_next(request.POST.get("next") or "")

    if not tid or not can_access_company(request.user, tid):
        log_audit(action="UNAUTHORISED_COMPANY_ACCESS", request=request, user=request.user,
                  detail=f"tenant_id={tid}")
        raise PermissionDenied("You do not have access to that company.")
    tenant = Tenant.objects.filter(id=tid).first()
    if not sid or not can_access_site(request.user, tenant, sid):
        log_audit(action="UNAUTHORISED_SITE_ACCESS", request=request, user=request.user, tenant=tenant,
                  detail=f"site_id={sid}")
        raise PermissionDenied("You do not have access to that site.")

    # Both valid - apply together.
    company_changed = str(request.session.get(SESSION_TENANT_KEY)) != str(tid)
    request.session[SESSION_TENANT_KEY] = int(tid)
    request.session[SESSION_SITE_KEY] = int(sid)
    request.session.pop(SESSION_LOCATION_KEY, None)  # drop stale workflow location
    site = selectable_sites(request.user, tenant).filter(id=sid).first()
    if company_changed:
        log_audit(action="COMPANY_SWITCHED", request=request, user=request.user, tenant=tenant,
                  detail=f"tenant_id={tid}")
    log_audit(action="SITE_SWITCHED", request=request, user=request.user, tenant=tenant,
              detail=getattr(site, "name", sid))
    return redirect(next_url or default_landing_url(tenant, get_active_role(request)))


@login_required
def no_site(request):
    """Shown when a user has no site they can work in (no accessible active
    location in the selected company)."""
    from core.access import get_active_tenant
    tenant = get_active_tenant(request)
    return render(request, "no_site.html", {"tenant": tenant})


def _safe_next(url):
    """Only allow same-site relative redirects (no open redirects)."""
    if url and url.startswith("/") and not url.startswith("//"):
        return url
    return ""


def _dashboard_kpis(tenant, role_code, site_id=None, location_ids=None):
    """Role-relevant headline metrics for the dashboard, scoped to the selected
    site where the data has a site dimension. Company-level figures (bank
    balance) stay company-wide. Each card is
    {label, value, is_money, icon, url, tone, sub}."""
    from django.db.models import Sum
    from core.models import (JournalLine, CustomerInvoice, PurchaseOrder, PurchaseRequisition,
                             SalesQuote, CustomerOrder, StockAdjustment, InventoryBalance)
    from core.services import reports as rpt
    today = timezone.localdate()
    site_ids = [site_id] if site_id else None  # None = company-wide

    def gl_balance(code, normal="debit"):
        agg = JournalLine.objects.filter(entry__tenant=tenant, account__code=code).aggregate(d=Sum("debit"), c=Sum("credit"))
        d = agg["d"] or Decimal("0.00"); c = agg["c"] or Decimal("0.00")
        return (d - c) if normal == "debit" else (c - d)

    def sales_mtd():
        first = today.replace(day=1)
        qs = JournalLine.objects.filter(entry__tenant=tenant, account__code="4000", entry__entry_date__gte=first)
        if site_ids is not None:
            qs = qs.filter(entry__site_id__in=site_ids)
        agg = qs.aggregate(d=Sum("debit"), c=Sum("credit"))
        return (agg["c"] or Decimal("0.00")) - (agg["d"] or Decimal("0.00"))

    def overdue_invoices():
        qs = CustomerInvoice.objects.filter(tenant=tenant, status__in=CustomerInvoice.OPEN_STATES, due_date__lt=today)
        if site_ids is not None:
            qs = qs.filter(site_id__in=site_ids)
        return qs.count()

    def low_stock_count():
        # On-hand within the selected site's locations vs the product reorder level.
        n = 0
        for p in Product.objects.filter(tenant=tenant, reorder_level__gt=0, is_active=True):
            bals = InventoryBalance.objects.filter(tenant=tenant, product=p)
            if location_ids is not None:
                bals = bals.filter(location_id__in=location_ids)
            on_hand = bals.aggregate(s=Sum("on_hand"))["s"] or Decimal("0.00")
            if on_hand < p.reorder_level:
                n += 1
        return n

    def _po_scope(qs):
        return qs.filter(site_id__in=site_ids) if site_ids is not None else qs

    def open_pos():
        return _po_scope(PurchaseOrder.objects.filter(
            tenant=tenant, is_current=True,
            status__in=[PurchaseOrder.Status.SUBMITTED, PurchaseOrder.Status.APPROVAL_PENDING,
                        PurchaseOrder.Status.APPROVED, PurchaseOrder.Status.SENT,
                        PurchaseOrder.Status.IN_TRANSIT, PurchaseOrder.Status.PARTIALLY_RECEIVED])).count()

    # --- card builders (computed lazily, only what each role needs) ---
    def c_sales():   return {"label": "Sales (this month)", "value": sales_mtd(), "is_money": True, "icon": "graph-up-arrow", "url": "/sales/reports/", "tone": "success"}
    def c_ar():      return {"label": "Receivables outstanding", "value": rpt.aged_receivables(tenant, site_ids=site_ids)["total"], "is_money": True, "icon": "cash-coin", "url": "/reports/aged-receivables/", "tone": "primary"}
    def c_ap():      return {"label": "Payables outstanding", "value": rpt.aged_payables(tenant, site_ids=site_ids)["total"], "is_money": True, "icon": "credit-card", "url": "/reports/aged-payables/", "tone": "primary"}
    def c_bank():    return {"label": "Bank balance", "value": gl_balance("1050"), "is_money": True, "icon": "bank2", "url": "/reports/balance-sheet/", "tone": "success", "sub": "Company"}
    def c_overdue(): n = overdue_invoices(); return {"label": "Overdue invoices", "value": n, "is_money": False, "icon": "exclamation-octagon", "url": "/ar/invoices/", "tone": "danger" if n else "muted"}
    def c_lowstock():n = low_stock_count(); return {"label": "Low-stock items", "value": n, "is_money": False, "icon": "exclamation-triangle", "url": "/inventory/low-stock/", "tone": "warning" if n else "muted"}
    def c_stockval():return {"label": "Stock value", "value": rpt.stock_valuation(tenant, location_ids=location_ids)["total"], "is_money": True, "icon": "box-seam", "url": "/reports/stock-valuation/", "tone": "primary"}
    def c_openpo():  return {"label": "Open purchase orders", "value": open_pos(), "is_money": False, "icon": "file-earmark-text", "url": "/po/", "tone": "primary"}
    def c_req():     n = PurchaseRequisition.objects.filter(tenant=tenant, status=PurchaseRequisition.Status.SUBMITTED).count(); return {"label": "Requisitions to approve", "value": n, "is_money": False, "icon": "card-checklist", "url": "/requisitions/", "tone": "warning" if n else "muted"}
    def c_backorder():
        lines = PurchaseOrderLine.objects.filter(po__tenant=tenant, po__is_current=True).exclude(
            po__status__in=[PurchaseOrder.Status.CANCELLED, PurchaseOrder.Status.CLOSED, PurchaseOrder.Status.DRAFT])
        if site_ids is not None:
            lines = lines.filter(po__site_id__in=site_ids)
        n = sum(1 for l in lines if l.open_qty and l.open_qty > 0)
        return {"label": "Backorder lines", "value": n, "is_money": False, "icon": "hourglass-split", "url": "/po/backorders/", "tone": "warning" if n else "muted"}
    def c_pendadj():
        qs = StockAdjustment.objects.filter(tenant=tenant, status=StockAdjustment.Status.PENDING)
        if location_ids is not None:
            qs = qs.filter(location_id__in=location_ids)
        n = qs.count(); return {"label": "Adjustments to approve", "value": n, "is_money": False, "icon": "sliders", "url": "/inventory/adjustments/", "tone": "warning" if n else "muted"}
    def c_quotes():  n = SalesQuote.objects.filter(tenant=tenant, status__in=[SalesQuote.Status.DRAFT, SalesQuote.Status.SENT]).count(); return {"label": "Open quotes", "value": n, "is_money": False, "icon": "file-text", "url": "/quotes/", "tone": "primary"}
    def c_orders():
        qs = CustomerOrder.objects.filter(tenant=tenant, status=CustomerOrder.Status.CONFIRMED)
        if site_ids is not None:
            qs = qs.filter(site_id__in=site_ids)
        n = qs.count(); return {"label": "Orders to invoice", "value": n, "is_money": False, "icon": "bag-check", "url": "/customer-orders/", "tone": "primary"}
    def c_torecv():
        n = _po_scope(PurchaseOrder.objects.filter(tenant=tenant, is_current=True,
            status__in=[PurchaseOrder.Status.APPROVED, PurchaseOrder.Status.SENT,
                        PurchaseOrder.Status.IN_TRANSIT, PurchaseOrder.Status.PARTIALLY_RECEIVED])).count()
        return {"label": "POs to receive", "value": n, "is_money": False, "icon": "truck", "url": "/po/", "tone": "primary"}

    layout = {
        roles_mod.ACCOUNTANT: [c_ar, c_ap, c_overdue, c_bank],
        roles_mod.FINANCE:    [c_ar, c_ap, c_overdue, c_bank],
        roles_mod.SALES:      [c_sales, c_quotes, c_orders, c_overdue],
        roles_mod.WAREHOUSE:  [c_lowstock, c_pendadj, c_torecv, c_stockval],
        roles_mod.PURCHASING: [c_openpo, c_req, c_backorder, c_lowstock],
        roles_mod.READONLY:   [c_sales, c_ar, c_stockval],
        roles_mod.MANAGER:    [c_sales, c_openpo, c_lowstock, c_ar],
        roles_mod.ADMIN:      [c_sales, c_ar, c_lowstock, c_openpo],
    }
    builders = layout.get(role_code, [])
    cards = []
    for b in builders:
        try:
            cards.append(b())
        except Exception:
            continue  # never let a KPI break the dashboard
    return cards


def _render_dashboard(request, role_code):
    """Role-based home: headline KPIs for the role plus a navigation launcher of
    the modules/reports the role can access."""
    tenant = _get_default_tenant(request)
    if not tenant:
        return render(request, "landing.html", {"tenant": None, "needs_setup": True})
    # Cards come straight from the role's accessible navigation (single source
    # of truth for per-role access); skip the self-referential Dashboard entry.
    sections = [(t, items) for (t, items) in roles_mod.sidebar_for_role(role_code) if t != "Dashboard"]
    site = get_active_site(request)
    now = timezone.localtime()
    hour = now.hour
    if hour < 12:
        greeting = "Good morning"
    elif hour < 18:
        greeting = "Good afternoon"
    else:
        greeting = "Good evening"
    return render(request, "dashboards/home.html", {
        "tenant": tenant,
        "role_code": role_code,
        "role_label": roles_mod.ROLE_LABELS.get(role_code, role_code),
        "title": roles_mod.DASHBOARD_TITLE.get(role_code, "Dashboard"),
        "sections": sections,
        "kpis": _dashboard_kpis(tenant, role_code, site_id=getattr(site, "id", None),
                                location_ids=active_location_ids(request)),
        "active_site": site,
        "greeting": greeting,
        "now": now,
        "onboarding_complete": tenant.onboarding_complete,
    })


def _make_dashboard(role_code):
    @login_required
    def _view(request):
        active = get_active_role(request)
        # Owner/Admin may view any dashboard; others only their own.
        if active != role_code and active != roles_mod.ADMIN:
            raise PermissionDenied("This dashboard is not available for your role.")
        return _render_dashboard(request, role_code)
    return _view


dashboard_admin = _make_dashboard(roles_mod.ADMIN)
dashboard_accountant = _make_dashboard(roles_mod.ACCOUNTANT)
dashboard_manager = _make_dashboard(roles_mod.MANAGER)
dashboard_sales = _make_dashboard(roles_mod.SALES)
dashboard_warehouse = _make_dashboard(roles_mod.WAREHOUSE)
dashboard_purchasing = _make_dashboard(roles_mod.PURCHASING)
dashboard_finance = _make_dashboard(roles_mod.FINANCE)
dashboard_readonly = _make_dashboard(roles_mod.READONLY)


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def audit_log_list(request):
    tenant = _get_default_tenant(request)
    qs = AuditLog.objects.filter(tenant=tenant).select_related("user", "site")
    action = (request.GET.get("action") or "").strip()
    q = (request.GET.get("q") or "").strip()
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    site_raw = (request.GET.get("site") or "").strip()
    if action:
        qs = qs.filter(action=action)
    if q:
        qs = qs.filter(Q(username__icontains=q) | Q(detail__icontains=q) |
                       Q(entity_type__icontains=q) | Q(entity_id__icontains=q))
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    selected_site = None
    if site_raw.isdigit():
        selected_site = int(site_raw)
        qs = qs.filter(site_id=selected_site)
    actions = list(AuditLog.objects.filter(tenant=tenant).values_list("action", flat=True).distinct().order_by("action"))
    sites = list(selectable_sites(request.user, tenant))
    return render(request, "audit_log.html", {
        "tenant": tenant, "logs": qs[:500], "actions": actions, "sites": sites,
        "f_action": action, "q": q, "date_from": date_from, "date_to": date_to,
        "selected_site": selected_site,
    })


def _csv_response(filename, columns, rows):
    import csv as _csv
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = _csv.writer(resp)
    writer.writerow(columns)
    for r in rows:
        writer.writerow(r)
    return resp


def _xlsx_response(filename, columns, rows, sheet_title="Export"):
    """Return an .xlsx workbook with a bold, frozen header row and tidy widths."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Export")[:31]  # Excel sheet-name limit
    ws.append(list(columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append(list(r))
    ws.freeze_panes = "A2"
    for i, col in enumerate(columns, start=1):
        width = max([len(str(col))] + [len(str(r[i - 1])) for r in rows]) if rows else len(str(col))
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(width + 2, 10), 60)

    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _export_response(request, filename, columns, rows, sheet_title="Export"):
    """Pick CSV or Excel from ?format=xlsx; CSV is the default."""
    if (request.GET.get("format") or "").lower() in ("xlsx", "excel"):
        stem = filename.rsplit(".", 1)[0]
        return _xlsx_response(f"{stem}.xlsx", columns, rows, sheet_title=sheet_title)
    return _csv_response(filename, columns, rows)


@permission_required(permissions_mod.EXPORT_DATA)
def data_export(request, kind):
    """Download a tenant's products/customers/suppliers as CSV (gated by export_data)."""
    from core.services import importer
    if kind not in importer.CONFIG:
        raise Http404("Unknown export type.")
    tenant = _get_default_tenant(request)
    columns, rows = importer.export_rows(tenant, kind)
    log_audit(action="DATA_EXPORTED", request=request, user=request.user, tenant=tenant,
              detail=f"{kind} ({len(rows)} rows)")
    return _export_response(request, f"{kind}.csv", columns, rows, sheet_title=kind)


def _finance_export_data(tenant, kind, date_from, date_to, as_of):
    """Return (filename, columns, rows) for an accountant CSV export."""
    from core.models import JournalLine
    money = lambda v: f"{(v or Decimal('0.00')):.2f}"

    if kind == "trial-balance":
        data = reports_service.trial_balance(tenant, date_to=as_of)
        cols = ["Account code", "Account", "Debit", "Credit"]
        rows = [[r["account"].code, r["account"].name, money(r["debit"]), money(r["credit"])] for r in data["rows"]]
        rows.append(["", "TOTAL", money(data["total_debit"]), money(data["total_credit"])])
        return f"trial-balance-{as_of}.csv", cols, rows

    if kind == "profit-and-loss":
        data = reports_service.profit_and_loss(tenant, date_from=date_from, date_to=date_to)
        cols = ["Section", "Account", "Amount"]
        rows = []
        for r in data["income"]:
            rows.append(["Income", f"{r['account'].code} {r['account'].name}", money(r["amount"])])
        rows.append(["Income", "Total income", money(data["income_total"])])
        for r in data["cogs"]:
            rows.append(["Cost of goods sold", f"{r['account'].code} {r['account'].name}", money(r["amount"])])
        rows.append(["Cost of goods sold", "Total COGS", money(data["cogs_total"])])
        rows.append(["", "Gross profit", money(data["gross_profit"])])
        for r in data["expense"]:
            rows.append(["Operating expenses", f"{r['account'].code} {r['account'].name}", money(r["amount"])])
        rows.append(["Operating expenses", "Total expenses", money(data["expense_total"])])
        rows.append(["", "Net profit", money(data["net_profit"])])
        return f"profit-and-loss-{date_from}-to-{date_to}.csv", cols, rows

    if kind == "balance-sheet":
        data = reports_service.balance_sheet(tenant, as_of=as_of)
        cols = ["Section", "Account", "Amount"]
        rows = []
        for key, label in [("assets", "Assets"), ("liabilities", "Liabilities"), ("equity", "Equity")]:
            for r in data[key]:
                rows.append([label, f"{r['account'].code} {r['account'].name}", money(r["amount"])])
        rows.append(["Equity", "Retained earnings", money(data["retained_earnings"])])
        rows.append(["", "Total assets", money(data["asset_total"])])
        rows.append(["", "Total liabilities + equity", money(data["liabilities_equity_total"])])
        return f"balance-sheet-{as_of}.csv", cols, rows

    if kind == "cash-flow":
        data = reports_service.cash_flow_summary(tenant, date_from=date_from, date_to=date_to)
        cols = ["Source / use of cash", "Amount"]
        rows = [["Opening balance", money(data["opening"])]]
        rows += [[r["account"].name, money(r["amount"])] for r in data["rows"]]
        rows.append(["Net movement", money(data["net"])])
        rows.append(["Closing balance", money(data["closing"])])
        return f"cash-flow-{date_from}-to-{date_to}.csv", cols, rows

    if kind in ("aged-receivables", "aged-payables"):
        data = (reports_service.aged_receivables if kind == "aged-receivables" else reports_service.aged_payables)(tenant, as_of=as_of)
        cols = ["Party", "Reference", "Date", "Due", "Days overdue", "Bucket", "Amount"]
        rows = [[r["party"], r["ref"], r["date"], r["due"], r["days"], r["bucket"], money(r["amount"])] for r in data["rows"]]
        rows.append(["TOTAL", "", "", "", "", "", money(data["total"])])
        return f"{kind}-{as_of}.csv", cols, rows

    if kind == "journal":
        cols = ["Date", "JE", "Ref type", "Ref", "Account code", "Account", "Description", "Debit", "Credit"]
        lines = (JournalLine.objects.filter(entry__tenant=tenant)
                 .select_related("entry", "account").order_by("entry__entry_date", "entry_id", "id"))
        if date_from:
            lines = lines.filter(entry__entry_date__gte=date_from)
        if date_to:
            lines = lines.filter(entry__entry_date__lte=date_to)
        rows = [[l.entry.entry_date, l.entry_id, l.entry.ref_type or "", l.entry.ref_id or "",
                 l.account.code, l.account.name, l.description or "", money(l.debit), money(l.credit)] for l in lines]
        return "general-ledger.csv", cols, rows

    if kind == "expenses":
        cols = ["Date", "Payee", "Supplier", "Category", "Description", "Method", "Net", "VAT", "Total",
                "Paid", "Reimbursable", "Status"]
        rows = [[e.expense_date, e.payee, (e.supplier.name if e.supplier_id else ""), e.category.name,
                 e.description or "", e.get_method_display(), money(e.net_amount), money(e.tax_amount),
                 money(e.total), "Yes" if e.paid else "No", "Yes" if e.reimbursable else "No", e.status]
                for e in Expense.objects.filter(tenant=tenant).select_related("category", "supplier", "tax_code").order_by("-expense_date")]
        return "expenses.csv", cols, rows

    if kind == "payments":
        cols = ["Date", "Direction", "Party", "Method", "Reference", "Amount", "Allocated", "Reconciled"]
        rows = [[p.payment_date, p.get_direction_display(), p.party_name, p.get_method_display(),
                 p.reference or "", money(p.amount), money(p.allocated), "Yes" if p.is_reconciled else "No"]
                for p in Payment.objects.filter(tenant=tenant).select_related("customer", "supplier").order_by("-payment_date")]
        return "payments.csv", cols, rows

    if kind == "invoices":
        cols = ["Number", "Date", "Customer", "Subtotal", "Tax", "Total", "Paid", "Outstanding", "Status"]
        rows = [[i.invoice_number, i.invoice_date, i.customer.name, money(i.subtotal), money(i.tax_total),
                 money(i.total), money(i.amount_paid), money(i.outstanding), i.status]
                for i in CustomerInvoice.objects.filter(tenant=tenant).select_related("customer").prefetch_related("lines", "lines__tax_code", "payment_allocations", "credit_notes").order_by("-invoice_date")]
        return "customer-invoices.csv", cols, rows

    if kind == "bills":
        cols = ["Number", "Date", "Supplier", "Subtotal", "Tax", "Total", "Outstanding", "Status"]
        rows = [[b.invoice_number, b.invoice_date, b.supplier.name, money(b.subtotal), money(b.tax_total),
                 money(b.total), money(b.outstanding), b.status]
                for b in SupplierInvoice.objects.filter(tenant=tenant).select_related("supplier").prefetch_related("lines", "lines__tax_code", "payment_allocations", "credit_notes").order_by("-invoice_date")]
        return "supplier-bills.csv", cols, rows

    if kind == "credit-notes":
        cols = ["Number", "Date", "Type", "Party", "Net", "VAT", "Total", "Applied to", "Status"]
        rows = [[c.credit_note_number, c.credit_note_date, c.get_kind_display(), c.party_name,
                 money(c.subtotal), money(c.tax_total), money(c.total),
                 (c.customer_invoice.invoice_number if c.customer_invoice_id else c.supplier_invoice.invoice_number if c.supplier_invoice_id else ""), c.status]
                for c in CreditNote.objects.filter(tenant=tenant).select_related("customer", "supplier", "customer_invoice", "supplier_invoice").prefetch_related("lines", "lines__tax_code").order_by("-credit_note_date")]
        return "credit-notes.csv", cols, rows

    if kind == "bank-transactions":
        cols = ["Date", "Description", "Reference", "Amount", "Matched to", "Reconciled"]
        rows = [[t.txn_date, t.description, t.reference or "", money(t.amount), t.matched_label, "Yes" if t.is_reconciled else "No"]
                for t in BankTransaction.objects.filter(tenant=tenant).select_related("matched_payment", "matched_expense")]
        return "bank-transactions.csv", cols, rows

    if kind == "vat-return":
        from core.services import vat as vat_svc
        b = vat_svc.compute_vat_return(tenant, date_from, date_to)
        cols = ["Box", "Description", "Amount"]
        labels = [
            ("1", "VAT due on sales", b["box1_vat_due_sales"]),
            ("2", "VAT due on acquisitions", b["box2_vat_due_acquisitions"]),
            ("3", "Total VAT due", b["box3_total_vat_due"]),
            ("4", "VAT reclaimed on purchases", b["box4_vat_reclaimed"]),
            ("5", "Net VAT to pay / reclaim", b["box5_net_vat"]),
            ("6", "Total sales ex VAT", b["box6_total_sales_ex_vat"]),
            ("7", "Total purchases ex VAT", b["box7_total_purchases_ex_vat"]),
            ("8", "EU supplies", b["box8_eu_supplies"]),
            ("9", "EU acquisitions", b["box9_eu_acquisitions"]),
        ]
        rows = [[bx, desc, money(amt)] for bx, desc, amt in labels]
        return f"vat-return-{date_from}-to-{date_to}.csv", cols, rows

    if kind == "vat-transactions":
        from core.services import vat as vat_svc
        cols = ["Date", "Document", "Direction", "Reference", "Party", "Description",
                "Treatment", "Rate", "Net", "VAT", "In VAT boxes"]
        rows = [[r["date"], r["doc_type"], r["direction"], r["ref"], r["party"], r["description"],
                 r["treatment"], f"{r['rate']:.4f}", money(r["net"]), money(r["vat"]),
                 "Yes" if r["in_boxes"] else "No"]
                for r in vat_svc.vat_transactions(tenant, date_from, date_to)]
        return f"vat-records-{date_from}-to-{date_to}.csv", cols, rows

    if kind in ("sales-history", "sales-by-product", "sales-by-customer", "sales-by-channel"):
        from core.services import sales_reports
        if kind == "sales-history":
            d = sales_reports.sales_history(tenant, date_from, date_to)
            cols = ["Invoice", "Date", "Customer", "Status", "Net", "VAT", "Total"]
            rows = [[r["invoice"].invoice_number, r["invoice"].invoice_date, r["invoice"].customer.name,
                     r["invoice"].display_status, money(r["net"]), money(r["vat"]), money(r["total"])]
                    for r in d["rows"]]
            rows.append(["", "", "", "TOTAL", money(d["net_total"]), money(d["vat_total"]), money(d["grand_total"])])
            return f"sales-history-{date_from}-to-{date_to}.csv", cols, rows
        if kind == "sales-by-product":
            d = sales_reports.sales_by_product(tenant, date_from, date_to)
            cols = ["Product", "Name", "Qty", "Net", "Total"]
            rows = [[r["key"], r["name"], money(r["qty"]), money(r["net"]), money(r["total"])] for r in d["rows"]]
            rows.append(["", "TOTAL", "", money(d["net_total"]), money(d["grand_total"])])
            return f"sales-by-product-{date_from}-to-{date_to}.csv", cols, rows
        if kind == "sales-by-customer":
            d = sales_reports.sales_by_customer(tenant, date_from, date_to)
            cols = ["Customer", "Invoices", "Net", "Total"]
            rows = [[r["name"], r["count"], money(r["net"]), money(r["total"])] for r in d["rows"]]
            rows.append(["TOTAL", "", money(d["net_total"]), money(d["grand_total"])])
            return f"sales-by-customer-{date_from}-to-{date_to}.csv", cols, rows
        d = sales_reports.sales_by_channel(tenant, date_from, date_to)
        cols = ["Channel", "Orders", "Total"]
        rows = [[r["channel"], r["count"], money(r["total"])] for r in d["rows"]]
        rows.append(["TOTAL", "", money(d["grand_total"])])
        return f"sales-by-channel-{date_from}-to-{date_to}.csv", cols, rows

    return None


@permission_required(permissions_mod.EXPORT_DATA)
def finance_export(request, kind):
    """Accountant CSV/Excel export of finance reports and ledgers (export_data)."""
    tenant = _get_default_tenant(request)
    as_of = _parse_date(request.GET.get("as_of")) or _parse_date(request.GET.get("to")) or timezone.localdate()
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    if kind in ("profit-and-loss", "cash-flow", "vat-return", "vat-transactions",
                "sales-history", "sales-by-product", "sales-by-customer", "sales-by-channel") and not date_from and not date_to:
        date_from, date_to = reports_service.current_financial_year(tenant)
    result = _finance_export_data(tenant, kind, date_from, date_to, as_of)
    if result is None:
        raise Http404("Unknown export type.")
    filename, columns, rows = result
    log_audit(action="DATA_EXPORTED", request=request, user=request.user, tenant=tenant,
              detail=f"{kind} ({len(rows)} rows)")
    return _export_response(request, filename, columns, rows, sheet_title=kind)


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def audit_log_export(request):
    """Download the audit log for the active organisation as CSV/Excel (admin)."""
    tenant = _get_default_tenant(request)
    logs = AuditLog.objects.filter(tenant=tenant)[:5000]
    columns = ["timestamp", "action", "user", "entity_type", "entity_id", "old_value",
               "new_value", "detail", "path", "ip", "user_agent"]
    rows = [[l.created_at.strftime("%Y-%m-%d %H:%M:%S"), l.action, l.username or "",
             l.entity_type or "", l.entity_id or "", l.old_value or "", l.new_value or "",
             l.detail or "", l.path or "", l.ip or "", l.user_agent or ""] for l in logs]
    log_audit(action="DATA_EXPORTED", request=request, user=request.user, tenant=tenant,
              detail=f"audit log ({len(rows)} rows)")
    return _export_response(request, "audit-log.csv", columns, rows, sheet_title="Audit log")


@login_required
def change_password(request):
    """Self-service password change; audited as PASSWORD_CHANGED."""
    from django.contrib.auth.forms import PasswordChangeForm
    from django.contrib.auth import update_session_auth_hash
    form = PasswordChangeForm(request.user, request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.save()
        update_session_auth_hash(request, user)
        log_audit(action="PASSWORD_CHANGED", request=request, user=request.user,
                  tenant=_get_default_tenant(request))
        messages.success(request, "Your password has been changed.")
        return redirect("change_password")
    return render(request, "auth/change_password.html", {"form": form})


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def settings_role_landing(request):
    tenant = _get_default_tenant(request)
    if request.method == "POST":
        mapping = {}
        valid = {c for c, _ in roles_mod.LANDING_CHOICES}
        for code, _label in roles_mod.ROLE_CHOICES:
            chosen = request.POST.get(f"landing_{code}")
            if chosen and chosen in valid:
                mapping[code] = chosen
        tenant.role_landing = mapping
        tenant.save(update_fields=["role_landing"])
        messages.success(request, "Default landing pages updated.")
        return redirect("settings_role_landing")
    current = tenant.role_landing or {}
    rows = [{"code": code, "label": label, "current": current.get(code, "")}
            for code, label in roles_mod.ROLE_CHOICES]
    return render(request, "settings/role_landing.html", {
        "tenant": tenant,
        "rows": rows,
        "landing_choices": roles_mod.LANDING_CHOICES,
    })


def permission_denied_view(request, exception=None):
    log_audit(action="ACCESS_DENIED", request=request, user=getattr(request, "user", None),
              detail=(str(exception)[:200] if exception else None))
    return render(request, "403.html", status=403)


# ============================
# Access requests (public sign-up -> admin provisions the account)
# ============================

def request_access(request):
    """Public form: a prospective user asks the admin for an account."""
    if request.user.is_authenticated:
        return redirect("landing")
    form = AccessRequestForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        req = form.save()
        log_audit(action="ACCESS_REQUEST", request=request,
                  detail=f"{req.name} <{req.email}> team={req.team or '-'}", username=req.email)
        from core import notify
        notify.notify_admins_new_request(req, request)
        messages.success(request, "Thanks! Your request has been sent to the administrator.")
        return redirect("login")
    return render(request, "auth/request_access.html", {"form": form})


def _unique_username(base):
    from django.contrib.auth.models import User
    base = "".join(ch for ch in (base or "user").lower() if ch.isalnum() or ch in "._-") or "user"
    candidate = base[:140]
    i = 1
    while User.objects.filter(username=candidate).exists():
        i += 1
        candidate = f"{base[:135]}{i}"
    return candidate


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def access_request_list(request):
    from core.models import Department
    tenant = _get_default_tenant(request)
    requests = AccessRequest.objects.all()
    departments = Department.objects.filter(tenant=tenant, is_active=True).order_by("name")
    return render(request, "access_requests.html", {
        "tenant": tenant, "requests": requests, "roles": roles_mod.ROLE_CHOICES,
        "departments": departments,
    })


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
@transaction.atomic
def access_request_action(request, req_id):
    from django.contrib.auth.models import User
    tenant = _get_default_tenant(request)
    req = get_object_or_404(AccessRequest, id=req_id)
    if request.method != "POST":
        return redirect("access_request_list")

    action = request.POST.get("action")
    if req.status != AccessRequest.Status.PENDING:
        messages.info(request, "This request has already been handled.")
        return redirect("access_request_list")

    if action == "reject":
        req.status = AccessRequest.Status.REJECTED
        req.reviewed_by = request.user
        req.reviewed_at = timezone.now()
        req.save()
        log_audit(action="ACCESS_REQUEST_REJECTED", request=request, user=request.user, tenant=tenant, detail=req.email)
        from core import notify
        notify.notify_applicant_rejected(req, request)
        messages.warning(request, f"Request from {req.name} rejected.")
        return redirect("access_request_list")

    if action == "approve":
        role = request.POST.get("role")
        if role not in dict(roles_mod.ROLE_CHOICES):
            messages.error(request, "Please choose a valid role.")
            return redirect("access_request_list")

        username = _unique_username((req.email.split("@")[0] if req.email else req.employee_id) or req.name)
        temp_password = get_random_string(10)
        parts = (req.name or "").split(" ", 1)
        user = User.objects.create_user(
            username=username, email=req.email,
            first_name=parts[0], last_name=(parts[1] if len(parts) > 1 else ""),
            password=temp_password,
        )
        # Optionally assign the new member to a structured department.
        from core.models import Department
        dept = None
        dept_id = (request.POST.get("department") or "").strip()
        if dept_id.isdigit():
            dept = Department.objects.filter(id=int(dept_id), tenant=tenant).first()

        UserProfile.objects.update_or_create(user=user, defaults={"tenant": tenant})
        OrgMembership.objects.get_or_create(
            user=user, tenant=tenant,
            defaults={"role": role, "is_default": True, "department": dept})

        req.status = AccessRequest.Status.APPROVED
        req.reviewed_by = request.user
        req.reviewed_at = timezone.now()
        req.created_user = user
        req.tenant = tenant
        req.save()
        log_audit(action="ACCESS_REQUEST_APPROVED", request=request, user=request.user, tenant=tenant,
                  detail=f"{req.email} -> {username} ({role})")
        from core import notify
        notify.notify_applicant_approved(req, username, temp_password, request)
        messages.success(
            request,
            f"Account created for {req.name}: username '{username}', temporary password '{temp_password}' "
            f"(role {dict(roles_mod.ROLE_CHOICES)[role]}). Share these and ask them to change the password.",
        )
        return redirect("access_request_list")

    return redirect("access_request_list")


# ============================
# Onboarding (guided setup)
# ============================

def _onboarding_steps(tenant):
    """Compute the guided-setup checklist with completion status for a tenant."""
    has_details = bool(tenant.company_number or tenant.legal_name or tenant.address_line1)
    vat_done = bool(tenant.vat_number) if tenant.vat_registered else TaxCode.objects.filter(tenant=tenant).exists()
    return [
        {"key": "details", "label": "Business details", "icon": "building",
         "desc": "Legal name, registration number and address.", "url": "/settings/tenant/",
         "done": has_details},
        {"key": "tax", "label": "VAT & tax", "icon": "percent",
         "desc": "VAT registration and tax codes.", "url": "/tax-codes/",
         "done": vat_done},
        {"key": "location", "label": "First location", "icon": "geo-alt",
         "desc": "A warehouse, store or office.", "url": "/locations/new/",
         "done": Location.objects.filter(tenant=tenant).exists()},
        {"key": "team", "label": "Invite your team", "icon": "people",
         "desc": "Add users and assign roles.", "url": "/team/invite/",
         "done": OrgMembership.objects.filter(tenant=tenant).count() > 1},
        {"key": "products", "label": "Add products", "icon": "box-seam",
         "desc": "Create or import your catalogue.", "url": "/products/",
         "done": Product.objects.filter(tenant=tenant).exists()},
        {"key": "customers", "label": "Add customers", "icon": "people-fill",
         "desc": "Create or import customers.", "url": "/customers/",
         "done": Customer.objects.filter(tenant=tenant).exists()},
        {"key": "suppliers", "label": "Add suppliers", "icon": "shop",
         "desc": "Create or import suppliers.", "url": "/suppliers/",
         "done": Supplier.objects.filter(tenant=tenant).exists()},
    ]


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def onboarding(request):
    tenant = _get_default_tenant(request)
    if not tenant:
        return redirect("new_organisation")
    steps = _onboarding_steps(tenant)
    done = sum(1 for s in steps if s["done"])
    return render(request, "onboarding/onboarding.html", {
        "tenant": tenant, "steps": steps, "done": done, "total": len(steps),
        "percent": int(done * 100 / len(steps)) if steps else 0,
    })


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def onboarding_finish(request):
    tenant = _get_default_tenant(request)
    if request.method == "POST" and tenant:
        tenant.onboarding_complete = True
        tenant.save(update_fields=["onboarding_complete"])
        messages.success(request, "Setup complete - welcome aboard!")
    return redirect("landing")


@login_required
@transaction.atomic
def new_organisation(request):
    """Create a brand-new organisation; the creator becomes its Owner/Admin."""
    form = NewOrganisationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        tenant = form.save()  # signals seed tax codes + GL accounts
        OrgMembership.objects.create(user=request.user, tenant=tenant, role=roles_mod.ADMIN, is_default=False)
        UserProfile.objects.update_or_create(user=request.user, defaults={"tenant": tenant})
        request.session[SESSION_TENANT_KEY] = tenant.id  # switch to the new org
        log_audit(action="ORG_CREATED", request=request, user=request.user, tenant=tenant, detail=tenant.name)
        messages.success(request, f"Organisation '{tenant.name}' created. Let's finish setting it up.")
        return redirect("onboarding")
    return render(request, "onboarding/new_organisation.html", {"form": form})


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def roles_permissions(request):
    """Matrix of roles x permissions, plus the org's access policy toggle."""
    from core import permissions as perms_mod
    tenant = _get_default_tenant(request)
    if request.method == "POST":
        new_val = request.POST.get("keep_permissions_on_role_change") == "on"
        if new_val != tenant.keep_permissions_on_role_change:
            tenant.keep_permissions_on_role_change = new_val
            tenant.save(update_fields=["keep_permissions_on_role_change"])
            log_audit(action="SETTINGS_CHANGED", request=request, user=request.user, tenant=tenant,
                      detail=f"keep_permissions_on_role_change={new_val}")
        messages.success(request, "Access policy updated.")
        return redirect("roles_permissions")
    matrix = []
    for code, label, category in perms_mod.PERMISSIONS:
        matrix.append({
            "label": label, "category": category,
            "cells": [perms_mod.role_has_permission(rc, code) for rc, _ in roles_mod.ROLE_CHOICES],
        })
    return render(request, "team/permissions_matrix.html", {
        "tenant": tenant,
        "role_labels": [lbl for _, lbl in roles_mod.ROLE_CHOICES],
        "matrix": matrix,
        "keep_permissions_on_role_change": tenant.keep_permissions_on_role_change,
    })


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
@transaction.atomic
def invite_user(request):
    """Admin invites a teammate directly: creates the account + role membership
    in the active org and emails them a temporary password."""
    from django.contrib.auth.models import User
    tenant = _get_default_tenant(request)
    form = InviteUserForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        email = form.cleaned_data["email"]
        name = form.cleaned_data["name"]
        role = form.cleaned_data["role"]
        username = _unique_username(email.split("@")[0] if email else name)
        temp_password = get_random_string(10)
        parts = name.split(" ", 1)
        user = User.objects.create_user(
            username=username, email=email,
            first_name=parts[0], last_name=(parts[1] if len(parts) > 1 else ""),
            password=temp_password,
        )
        UserProfile.objects.update_or_create(user=user, defaults={"tenant": tenant})
        OrgMembership.objects.get_or_create(user=user, tenant=tenant, defaults={"role": role, "is_default": True})
        from core import notify
        notify.notify_credentials(email, name, username, temp_password, request)
        log_audit(action="USER_INVITED", request=request, user=request.user, tenant=tenant,
                  detail=f"{email} -> {username} ({role})")
        messages.success(
            request,
            f"Invited {name}: username '{username}', temporary password '{temp_password}' "
            f"({dict(roles_mod.ROLE_CHOICES)[role]}). We've emailed them these details.",
        )
        return redirect("invite_user")
    return render(request, "team/invite.html", {"tenant": tenant, "form": form})


# ============================
# Users & Roles management (admin)
# ============================

def _active_admin_count(tenant, exclude_user_id=None):
    qs = OrgMembership.objects.filter(tenant=tenant, role=roles_mod.ADMIN, user__is_active=True)
    if exclude_user_id:
        qs = qs.exclude(user_id=exclude_user_id)
    return qs.count()


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def members_list(request):
    tenant = _get_default_tenant(request)
    members = OrgMembership.objects.filter(tenant=tenant).select_related("user").order_by("user__username")
    return render(request, "team/members.html", {
        "tenant": tenant, "members": members, "roles": roles_mod.ROLE_CHOICES,
    })


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
@transaction.atomic
def member_change_role(request, membership_id):
    tenant = _get_default_tenant(request)
    m = get_object_or_404(OrgMembership, id=membership_id, tenant=tenant)
    if request.method == "POST":
        new_role = request.POST.get("role")
        if new_role not in dict(roles_mod.ROLE_CHOICES):
            messages.error(request, "Invalid role.")
        elif m.role == roles_mod.ADMIN and new_role != roles_mod.ADMIN and _active_admin_count(tenant, exclude_user_id=m.user_id) == 0:
            messages.error(request, "You can't change the last Owner/Admin's role.")
        elif new_role != m.role:
            old = m.role
            m.role = new_role
            m.save()  # signal re-syncs Django groups
            # Custom per-user permissions are deltas on the role baseline. By
            # default we reset them on a role change so access stays predictable;
            # if the org opts to keep them, we prune the ones the new role makes
            # redundant so the remaining overrides still mean something.
            existing = list(UserPermissionOverride.objects.filter(tenant=tenant, user=m.user))
            note = ""
            if existing:
                if getattr(tenant, "keep_permissions_on_role_change", False):
                    new_base = permissions_mod.role_permissions(new_role)
                    pruned = 0
                    for o in existing:
                        redundant = ((o.effect == UserPermissionOverride.GRANT and o.permission in new_base) or
                                     (o.effect == UserPermissionOverride.REVOKE and o.permission not in new_base))
                        if redundant:
                            o.delete()
                            pruned += 1
                    kept = len(existing) - pruned
                    note = f" {kept} custom permission(s) kept." if kept else " Custom permissions were already covered by the new role."
                else:
                    UserPermissionOverride.objects.filter(tenant=tenant, user=m.user).delete()
                    note = " Custom permissions were reset to the role default."
            log_audit(action="ROLE_CHANGED", request=request, user=request.user, tenant=tenant,
                      detail=f"{m.user.username}: {old} -> {new_role}")
            messages.success(request, f"{m.user.username}'s role changed to {dict(roles_mod.ROLE_CHOICES)[new_role]}.{note}")
    return redirect("members_list")


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
@transaction.atomic
def member_toggle_active(request, membership_id):
    tenant = _get_default_tenant(request)
    m = get_object_or_404(OrgMembership, id=membership_id, tenant=tenant)
    if request.method == "POST":
        if m.user_id == request.user.id:
            messages.error(request, "You can't deactivate your own account.")
        elif m.user.is_active and m.role == roles_mod.ADMIN and _active_admin_count(tenant, exclude_user_id=m.user_id) == 0:
            messages.error(request, "You can't deactivate the last Owner/Admin.")
        else:
            m.user.is_active = not m.user.is_active
            m.user.save(update_fields=["is_active"])
            action = "USER_REACTIVATED" if m.user.is_active else "USER_DEACTIVATED"
            log_audit(action=action, request=request, user=request.user, tenant=tenant, detail=m.user.username)
            messages.success(request, f"{m.user.username} {'reactivated' if m.user.is_active else 'deactivated'}.")
    return redirect("members_list")


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
@transaction.atomic
def member_remove(request, membership_id):
    tenant = _get_default_tenant(request)
    m = get_object_or_404(OrgMembership, id=membership_id, tenant=tenant)
    if request.method == "POST":
        if m.user_id == request.user.id:
            messages.error(request, "You can't remove yourself from the organisation.")
        elif m.role == roles_mod.ADMIN and _active_admin_count(tenant, exclude_user_id=m.user_id) == 0:
            messages.error(request, "You can't remove the last Owner/Admin.")
        else:
            uname = m.user.username
            m.delete()
            log_audit(action="USER_REMOVED", request=request, user=request.user, tenant=tenant, detail=uname)
            messages.success(request, f"{uname} removed from {tenant.name}.")
    return redirect("members_list")


@role_required([ROLE_ADMIN], [ROLE_ADMIN])
@transaction.atomic
def member_permissions(request, membership_id):
    """Admin editor for a single member's effective permissions: the role
    baseline plus per-user grants/revokes. Owners/Admins always have everything,
    so their permissions are shown read-only."""
    tenant = _get_default_tenant(request)
    m = get_object_or_404(OrgMembership, id=membership_id, tenant=tenant)
    base = permissions_mod.role_permissions(m.role)
    is_admin_role = (m.role == roles_mod.ADMIN)
    role_label = dict(roles_mod.ROLE_CHOICES)[m.role]

    if request.method == "POST" and not is_admin_role:
        if request.POST.get("reset"):
            removed = UserPermissionOverride.objects.filter(tenant=tenant, user=m.user).delete()[0]
            if removed:
                log_audit(action="PERMISSION_CHANGED", request=request, user=request.user, tenant=tenant,
                          detail=f"{m.user.username}: reset to {role_label} default")
                messages.success(request, f"Reset {m.user.username}'s permissions to the {role_label} default.")
            else:
                messages.info(request, "No custom permissions to reset.")
            return redirect("member_permissions", membership_id=m.id)

        existing = {o.permission: o for o in UserPermissionOverride.objects.filter(tenant=tenant, user=m.user)}
        changes = []
        for code, _label, _cat in permissions_mod.PERMISSIONS:
            desired = request.POST.get(f"perm_{code}") == "on"
            in_base = code in base
            o = existing.get(code)
            if desired == in_base:
                if o:
                    o.delete()
                    changes.append(f"={code}")
            else:
                effect = UserPermissionOverride.GRANT if desired else UserPermissionOverride.REVOKE
                if not o or o.effect != effect:
                    UserPermissionOverride.objects.update_or_create(
                        tenant=tenant, user=m.user, permission=code, defaults={"effect": effect})
                    changes.append(f"{'+' if desired else '-'}{code}")
        if changes:
            log_audit(action="PERMISSION_CHANGED", request=request, user=request.user, tenant=tenant,
                      detail=f"{m.user.username}: {', '.join(changes)}")
            messages.success(request, f"Permissions updated for {m.user.username}.")
        else:
            messages.info(request, "No permission changes.")
        return redirect("member_permissions", membership_id=m.id)

    overrides = dict(UserPermissionOverride.objects.filter(tenant=tenant, user=m.user).values_list("permission", "effect"))
    rows = []
    for code, label, category in permissions_mod.PERMISSIONS:
        in_base = code in base
        effect = overrides.get(code)
        effective = True if is_admin_role else in_base
        if effect == UserPermissionOverride.GRANT:
            effective = True
        elif effect == UserPermissionOverride.REVOKE:
            effective = False
        rows.append({
            "code": code, "label": label, "category": category,
            "in_base": in_base, "effective": effective,
            "override": ("granted" if effect == UserPermissionOverride.GRANT
                         else "revoked" if effect == UserPermissionOverride.REVOKE else ""),
        })
    return render(request, "team/member_permissions.html", {
        "tenant": tenant, "member": m, "rows": rows,
        "is_admin_role": is_admin_role, "role_label": role_label,
        "override_count": len(overrides),
    })


# ============================
# CSV import (products / customers / suppliers)
# ============================

def _run_import(request, kind):
    cfg = importer_service.CONFIG[kind]
    tenant = _get_default_tenant(request)
    summary = None
    if request.method == "POST":
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Please choose a CSV file.")
        elif not f.name.lower().endswith(".csv"):
            messages.error(request, "Please upload a .csv file.")
        else:
            try:
                _, rows = importer_service.read_rows(f)
                summary = cfg["fn"](tenant, rows)
                messages.success(
                    request,
                    f"Imported {cfg['label']}: {summary['created']} created, "
                    f"{summary['updated']} updated, {len(summary['errors'])} skipped.",
                )
            except Exception as e:
                messages.error(request, f"Could not read the file: {e}")
    return render(request, "imports/import.html", {
        "tenant": tenant, "kind": kind, "cfg": cfg, "summary": summary,
    })


@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
def import_products(request):
    return _run_import(request, "products")


@role_required([ROLE_ADMIN, ROLE_SALES, ROLE_FINANCE], [ROLE_ADMIN, ROLE_SALES, ROLE_FINANCE])
def import_customers(request):
    return _run_import(request, "customers")


@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
def import_suppliers(request):
    return _run_import(request, "suppliers")


@login_required
def import_template(request, kind):
    cfg = importer_service.CONFIG.get(kind)
    if not cfg:
        raise Http404
    import csv as _csv
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = f'attachment; filename="{kind}_template.csv"'
    writer = _csv.writer(resp)
    writer.writerow(cfg["columns"])
    writer.writerow(cfg["sample"])
    return resp


@transaction.atomic
@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
def po_create(request):
    tenant = _get_default_tenant(request)
    if not tenant:
        return render(request, "base.html", {"content": "Create a Tenant in admin first."})

    po = PurchaseOrder(tenant=tenant, po_number=_generate_po_number())

    if request.method == "POST":
        form = PurchaseOrderForm(request.POST, instance=po)
        formset = PurchaseOrderLineFormSet(request.POST, instance=po)

        if form.is_valid() and formset.is_valid():
            po = form.save(commit=False)
            po.tenant = tenant
            po.site = get_active_site(request)  # PO belongs to the working site
            if not po.receiving_location_id:
                po.receiving_location = _active_site_default_location(request, tenant)  # receive into the site

            action = form.cleaned_data.get("action") or "save"
            po.status = PurchaseOrder.Status.SUBMITTED if action == "submit" else PurchaseOrder.Status.DRAFT

            # Save first so we have an ID for lines
            po.save()
            formset.save()

            # Set currency from supplier (per supplier currency)
            try:
                po.currency_code = po.supplier.currency_code or tenant.currency_code
            except Exception:
                po.currency_code = tenant.currency_code
            po.save()

            # Threshold-based approvals
            total = sum((l.line_total for l in po.lines.all()), Decimal("0.00"))
            threshold = getattr(tenant, "po_approval_threshold", Decimal("0.00")) or Decimal("0.00")
            if po.status == PurchaseOrder.Status.SUBMITTED and threshold > 0 and total > threshold:
                po.approval_required = True
                po.status = PurchaseOrder.Status.APPROVAL_PENDING
                po.save()
                messages.warning(
                    request,
                    f"PO submitted but requires approval (total {total} > threshold {threshold})."
                )

            # Create shipment record on submit only if approval not required
            if po.status == PurchaseOrder.Status.SUBMITTED and not getattr(po, "approval_required", False):
                dest = _po_destination(po)
                if dest:
                    Shipment.objects.get_or_create(
                        tenant=tenant,
                        po=po,
                        defaults={
                            "from_supplier": po.supplier,
                            "destination": dest,
                            "status": Shipment.Status.CREATED,
                        },
                    )

            return redirect("po_detail", po_id=po.id)
    else:
        form = PurchaseOrderForm(instance=po)
        formset = PurchaseOrderLineFormSet(instance=po)

    return render(request, "po_create.html", {"tenant": tenant, "form": form, "formset": formset, "po": po})


@login_required
@role_required([ROLE_ADMIN, ROLE_FINANCE], [ROLE_ADMIN, ROLE_FINANCE])
@transaction.atomic
def po_approve(request, po_id):
    tenant = _get_default_tenant(request)
    po = get_object_or_404(PurchaseOrder, id=po_id, tenant=tenant)

    if request.method != "POST":
        return redirect("po_detail", po_id=po.id)

    if po.status not in [PurchaseOrder.Status.SUBMITTED, PurchaseOrder.Status.APPROVAL_PENDING]:
        messages.info(request, "PO is not awaiting approval.")
        return redirect("po_detail", po_id=po.id)

    po.approval_required = False
    po.approved_by = request.user
    po.approved_at = timezone.now()
    po.status = PurchaseOrder.Status.APPROVED
    po.save()

    # Ensure there is at least one shipment + shipment lines
    dest = _po_destination(po)
    if dest:
        shipment, _ = Shipment.objects.get_or_create(
            tenant=tenant,
            po=po,
            defaults={"from_supplier": po.supplier, "destination": dest, "status": Shipment.Status.CREATED},
        )
        for pol in po.lines.all():
            ShipmentLine.objects.get_or_create(
                shipment=shipment,
                po_line=pol,
                defaults={"expected_qty": pol.open_qty},
            )

    messages.success(request, f"PO {po.po_number} approved.")
    return redirect("po_detail", po_id=po.id)

@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_FINANCE, ROLE_READONLY])
def po_detail(request, po_id):
    tenant = _get_default_tenant(request)
    po = get_object_or_404(PurchaseOrder, id=po_id, tenant=tenant)

    subtotal = sum((line.line_total for line in po.lines.all()), Decimal("0.00"))
    vat_rate = _default_vat_rate(tenant)
    vat_amount = subtotal * vat_rate
    total = subtotal + vat_amount

    return render(request, "po_detail.html", {
        "tenant": tenant,
        "po": po,
        "subtotal": subtotal,
        "vat_amount": vat_amount,
        "total": total,
        "vat_rate_percent": vat_rate * 100,
    })


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
@transaction.atomic
def po_submit(request, po_id):
    tenant = _get_default_tenant(request)
    po = get_object_or_404(PurchaseOrder, id=po_id, tenant=tenant)

    if request.method != "POST":
        return redirect("po_detail", po_id=po.id)

    if po.status != PurchaseOrder.Status.DRAFT:
        messages.info(request, "Only Draft POs can be submitted.")
        return redirect("po_detail", po_id=po.id)

    # Set currency from supplier (per-supplier currency), fallback to tenant
    po.currency_code = getattr(po.supplier, "currency_code", None) or tenant.currency_code

    total = sum((l.line_total for l in po.lines.all()), Decimal("0.00"))
    threshold = getattr(tenant, "po_approval_threshold", Decimal("0.00")) or Decimal("0.00")

    po.approval_required = bool(threshold and threshold > 0 and total > threshold)
    po.status = PurchaseOrder.Status.APPROVAL_PENDING if po.approval_required else PurchaseOrder.Status.SUBMITTED
    po.save()

    if po.approval_required:
        from core import notify
        from django.urls import reverse
        notify.notify_roles(
            tenant, [roles_mod.ADMIN, roles_mod.MANAGER, roles_mod.PURCHASING],
            exclude_user=request.user, category="APPROVAL_REQUEST", actor=request.user,
            url=reverse("po_detail", kwargs={"po_id": po.id}),
            title=f"PO awaiting approval: {po.po_number}",
            message=f"{request.user.username} submitted PO {po.po_number} ({po.currency_code} {total:.2f}) to {po.supplier.name} for approval.",
            request=request,
        )

    # Record the agreed supplier prices for this PO's lines.
    from core.services.purchasing import record_po_prices
    record_po_prices(po)

    # Always create at least 1 shipment on submit (planned), with shipment lines (expected qty = open qty)
    dest = _po_destination(po)
    if not dest:
        messages.error(request, "Create at least one Location before submitting POs.")
        return redirect("po_detail", po_id=po.id)

    shipment, _ = Shipment.objects.get_or_create(
        tenant=tenant,
        po=po,
        defaults={
            "from_supplier": po.supplier,
            "destination": dest,
            "status": Shipment.Status.CREATED,
        },
    )

    # Create/refresh shipment lines from PO open qty
    for pol in po.lines.all():
        exp = pol.open_qty
        sl, _ = ShipmentLine.objects.get_or_create(
            shipment=shipment,
            po_line=pol,
            defaults={"expected_qty": exp},
        )
        # If draft submission and no receipts yet, keep expected in sync with ordered
        if sl.received_qty == Decimal("0.00"):
            sl.expected_qty = exp
            sl.save(update_fields=["expected_qty"])

    if po.approval_required:
        messages.warning(request, f"PO submitted. Approval required (total {total} > threshold {threshold}).")
    else:
        messages.success(request, f"PO {po.po_number} submitted.")
    return redirect("po_detail", po_id=po.id)

# ---------------------------------------------------------------------------
# Purchase Requisitions (internal purchase request -> PO)
# ---------------------------------------------------------------------------

def _generate_req_number():
    return "PR-" + timezone.now().strftime("%Y%m%d-%H%M%S-%f")


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_FINANCE, ROLE_READONLY])
def requisition_list(request):
    tenant = _get_default_tenant(request)
    reqs = (PurchaseRequisition.objects
            .filter(tenant=tenant)
            .select_related("preferred_supplier", "requested_by", "converted_po")
            .order_by("-created_at"))
    return render(request, "requisitions/requisition_list.html", {"tenant": tenant, "reqs": reqs})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE])
def requisition_create(request):
    tenant = _get_default_tenant(request)
    if not tenant:
        return render(request, "base.html", {"content": "Create a Tenant in admin first."})

    req = PurchaseRequisition(tenant=tenant, req_number=_generate_req_number())

    if request.method == "POST":
        form = PurchaseRequisitionForm(request.POST, instance=req)
        formset = PurchaseRequisitionLineFormSet(request.POST, instance=req)
        if form.is_valid() and formset.is_valid():
            req = form.save(commit=False)
            req.tenant = tenant
            req.requested_by = request.user
            action = form.cleaned_data.get("action") or "save"
            req.status = (PurchaseRequisition.Status.SUBMITTED if action == "submit"
                          else PurchaseRequisition.Status.DRAFT)
            req.save()
            formset.save()
            log_audit(action="requisition_create", request=request, user=request.user, tenant=tenant,
                      detail=f"{req.req_number} ({req.status})")
            messages.success(request, f"Requisition {req.req_number} created.")
            return redirect("requisition_detail", req_id=req.id)
    else:
        form = PurchaseRequisitionForm(instance=req)
        formset = PurchaseRequisitionLineFormSet(instance=req)

    return render(request, "requisitions/requisition_create.html",
                  {"tenant": tenant, "form": form, "formset": formset, "req": req})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_FINANCE, ROLE_READONLY])
def requisition_detail(request, req_id):
    tenant = _get_default_tenant(request)
    req = get_object_or_404(PurchaseRequisition, id=req_id, tenant=tenant)
    return render(request, "requisitions/requisition_detail.html", {"tenant": tenant, "req": req})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE])
def requisition_submit(request, req_id):
    tenant = _get_default_tenant(request)
    req = get_object_or_404(PurchaseRequisition, id=req_id, tenant=tenant)
    if request.method != "POST":
        return redirect("requisition_detail", req_id=req.id)
    if req.status != PurchaseRequisition.Status.DRAFT:
        messages.info(request, "Only Draft requisitions can be submitted.")
        return redirect("requisition_detail", req_id=req.id)
    if not req.lines.exists():
        messages.error(request, "Add at least one line before submitting.")
        return redirect("requisition_detail", req_id=req.id)
    req.status = PurchaseRequisition.Status.SUBMITTED
    req.save(update_fields=["status"])
    log_audit(action="requisition_submit", request=request, user=request.user, tenant=tenant, detail=req.req_number)
    messages.success(request, f"Requisition {req.req_number} submitted for approval.")
    return redirect("requisition_detail", req_id=req.id)


@login_required
@role_required([ROLE_ADMIN, ROLE_FINANCE], [ROLE_ADMIN, ROLE_FINANCE])
def requisition_approve(request, req_id):
    tenant = _get_default_tenant(request)
    req = get_object_or_404(PurchaseRequisition, id=req_id, tenant=tenant)
    if request.method != "POST":
        return redirect("requisition_detail", req_id=req.id)
    if req.status != PurchaseRequisition.Status.SUBMITTED:
        messages.info(request, "Requisition is not awaiting approval.")
        return redirect("requisition_detail", req_id=req.id)
    req.status = PurchaseRequisition.Status.APPROVED
    req.approved_by = request.user
    req.approved_at = timezone.now()
    req.save(update_fields=["status", "approved_by", "approved_at"])
    log_audit(action="requisition_approve", request=request, user=request.user, tenant=tenant, detail=req.req_number)
    messages.success(request, f"Requisition {req.req_number} approved.")
    return redirect("requisition_detail", req_id=req.id)


@login_required
@role_required([ROLE_ADMIN, ROLE_FINANCE], [ROLE_ADMIN, ROLE_FINANCE])
def requisition_reject(request, req_id):
    tenant = _get_default_tenant(request)
    req = get_object_or_404(PurchaseRequisition, id=req_id, tenant=tenant)
    if request.method != "POST":
        return redirect("requisition_detail", req_id=req.id)
    if req.status != PurchaseRequisition.Status.SUBMITTED:
        messages.info(request, "Requisition is not awaiting approval.")
        return redirect("requisition_detail", req_id=req.id)
    req.status = PurchaseRequisition.Status.REJECTED
    req.rejected_reason = (request.POST.get("reason") or "").strip() or None
    req.save(update_fields=["status", "rejected_reason"])
    log_audit(action="requisition_reject", request=request, user=request.user, tenant=tenant,
              detail=f"{req.req_number}: {req.rejected_reason or ''}")
    messages.success(request, f"Requisition {req.req_number} rejected.")
    return redirect("requisition_detail", req_id=req.id)


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE])
def requisition_cancel(request, req_id):
    tenant = _get_default_tenant(request)
    req = get_object_or_404(PurchaseRequisition, id=req_id, tenant=tenant)
    if request.method != "POST":
        return redirect("requisition_detail", req_id=req.id)
    if req.status in (PurchaseRequisition.Status.CONVERTED, PurchaseRequisition.Status.CANCELLED):
        messages.info(request, "Requisition cannot be cancelled.")
        return redirect("requisition_detail", req_id=req.id)
    req.status = PurchaseRequisition.Status.CANCELLED
    req.save(update_fields=["status"])
    log_audit(action="requisition_cancel", request=request, user=request.user, tenant=tenant, detail=req.req_number)
    messages.success(request, f"Requisition {req.req_number} cancelled.")
    return redirect("requisition_detail", req_id=req.id)


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
@transaction.atomic
def requisition_convert(request, req_id):
    """Create a Draft Purchase Order from an approved requisition."""
    tenant = _get_default_tenant(request)
    req = get_object_or_404(PurchaseRequisition, id=req_id, tenant=tenant)
    if request.method != "POST":
        return redirect("requisition_detail", req_id=req.id)

    if req.status != PurchaseRequisition.Status.APPROVED:
        messages.error(request, "Only approved requisitions can be converted to a PO.")
        return redirect("requisition_detail", req_id=req.id)
    if req.converted_po_id:
        messages.info(request, "This requisition has already been converted.")
        return redirect("po_detail", po_id=req.converted_po_id)

    supplier = req.preferred_supplier
    if supplier is None:
        # Fall back to a product's preferred supplier, else first supplier on file.
        for l in req.lines.select_related("product"):
            sup = getattr(l.product, "preferred_supplier", None)
            if sup is not None:
                supplier = sup
                break
    if supplier is None:
        supplier = Supplier.objects.filter(tenant=tenant).order_by("id").first()
    if supplier is None:
        messages.error(request, "No supplier available. Set a preferred supplier on the requisition first.")
        return redirect("requisition_detail", req_id=req.id)

    po = PurchaseOrder.objects.create(
        tenant=tenant,
        po_number=_generate_po_number(),
        supplier=supplier,
        expected_date=req.needed_by,
        notes=f"From requisition {req.req_number}." + (f" {req.justification}" if req.justification else ""),
        status=PurchaseOrder.Status.DRAFT,
        currency_code=getattr(supplier, "currency_code", None) or tenant.currency_code,
    )
    std_tx = TaxCode.objects.filter(tenant=tenant, code="STD", is_active=True).first()
    for l in req.lines.select_related("product"):
        unit_cost = l.estimated_unit_cost
        if unit_cost is None:
            unit_cost = getattr(l.product, "standard_cost", None) or Decimal("0.00")
        PurchaseOrderLine.objects.create(
            po=po,
            product=l.product,
            ordered_qty=l.quantity,
            unit_cost=unit_cost,
            tax_code=getattr(l.product, "tax_code", None) or std_tx,
        )

    req.status = PurchaseRequisition.Status.CONVERTED
    req.converted_po = po
    req.save(update_fields=["status", "converted_po"])
    log_audit(action="requisition_convert", request=request, user=request.user, tenant=tenant,
              detail=f"{req.req_number} -> {po.po_number}")
    messages.success(request, f"Requisition {req.req_number} converted to draft PO {po.po_number}.")
    return redirect("po_detail", po_id=po.id)


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_READONLY])
def shipment_list(request):
    tenant = _get_default_tenant(request)
    shipments = Shipment.objects.filter(tenant=tenant).select_related("po", "from_supplier", "destination").order_by("-created_at")
    return render(request, "shipments/shipment_list.html", {"tenant": tenant, "shipments": shipments})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_READONLY], [ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE])
def shipment_detail(request, shipment_id):
    tenant = _get_default_tenant(request)
    shipment = get_object_or_404(Shipment, id=shipment_id, tenant=tenant)
    po = shipment.po

    form = ShipmentUpdateForm(request.POST or None, instance=shipment)

    if request.method == "POST":
        action = request.POST.get("action", "")

        if action == "update_shipment" and form.is_valid():
            shipment = form.save()
            # Keep PO in transit if any shipment is in transit-ish
            if shipment.status in [Shipment.Status.IN_TRANSIT, Shipment.Status.PICKED_UP]:
                po.status = PurchaseOrder.Status.IN_TRANSIT
                po.save(update_fields=["status"])
            messages.success(request, "Shipment updated.")
            return redirect("shipment_detail", shipment_id=shipment.id)

        if action == "allocate":
            # Validate first, then persist atomically - so an invalid value
            # gives a friendly message instead of a 500.
            try:
                with transaction.atomic():
                    for sl in shipment.lines.select_related("po_line", "po_line__product"):
                        raw = (request.POST.get(f"exp_{sl.id}") or "").strip()
                        if raw == "":
                            continue
                        try:
                            exp = Decimal(raw)
                        except InvalidOperation:
                            raise ValueError(f"Invalid quantity '{raw}' for {sl.po_line.product.sku}.")
                        if exp < sl.received_qty:
                            raise ValueError(
                                f"Expected qty cannot be below received for {sl.po_line.product.sku}."
                            )
                        sl.expected_qty = exp
                        sl.save(update_fields=["expected_qty"])
            except ValueError as e:
                messages.error(request, str(e))
                return redirect("shipment_detail", shipment_id=shipment.id)
            messages.success(request, "Allocation updated.")
            return redirect("shipment_detail", shipment_id=shipment.id)

        if action == "add_container":
            cn = (request.POST.get("container_number") or "").strip()
            if not cn:
                messages.error(request, "Container number is required.")
                return redirect("shipment_detail", shipment_id=shipment.id)
            Container.objects.get_or_create(
                shipment=shipment,
                container_number=cn,
                defaults={
                    "seal_number": (request.POST.get("seal_number") or "").strip() or None,
                    "mode": (request.POST.get("mode") or "").strip() or None,
                }
            )
            messages.success(request, "Container added.")
            return redirect("shipment_detail", shipment_id=shipment.id)

        if action == "add_event":
            event_type = (request.POST.get("event_type") or "").strip()
            if not event_type:
                messages.error(request, "Event type is required.")
                return redirect("shipment_detail", shipment_id=shipment.id)
            cont_id = (request.POST.get("container_id") or "").strip()
            cont = None
            if cont_id:
                cont = get_object_or_404(Container, id=cont_id, shipment=shipment)
            ShipmentEvent.objects.create(
                shipment=shipment,
                container=cont,
                event_type=event_type,
                status=(request.POST.get("status") or "").strip() or None,
                notes=(request.POST.get("notes") or "").strip() or None,
            )
            messages.success(request, "Event added.")
            return redirect("shipment_detail", shipment_id=shipment.id)

    # Derived
    lines = shipment.lines.select_related("po_line", "po_line__product").all()
    containers = shipment.containers.all()
    events = shipment.events.select_related("container").order_by("-occurred_at")
    return render(request, "shipments/shipment_detail.html", {
        "tenant": tenant,
        "po": po,
        "shipment": shipment,
        "form": form,
        "lines": lines,
        "containers": containers,
        "events": events,
    })

@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_READONLY], [ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE])
def shipment_update(request, shipment_id):
    return shipment_detail(request, shipment_id)

@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_WAREHOUSE])
def receive_po(request, po_id):
    tenant = _get_default_tenant(request)
    po = get_object_or_404(PurchaseOrder, id=po_id, tenant=tenant)

    if po.approval_required or po.status == PurchaseOrder.Status.APPROVAL_PENDING:
        messages.error(request, "PO requires approval before receiving.")
        return redirect("po_detail", po_id=po.id)

    shipment_id = request.GET.get("shipment_id") or request.POST.get("shipment_id")
    if shipment_id:
        shipment = get_object_or_404(Shipment, id=shipment_id, tenant=tenant, po=po)
    else:
        shipment = po.shipments.order_by("-created_at").first()

    if not shipment:
        messages.error(request, "No shipment found for this PO. Submit the PO first.")
        return redirect("po_detail", po_id=po.id)

    if not shipment.lines.exists():
        messages.error(request, "This shipment has no planned lines. Allocate quantities to the shipment first.")
        return redirect("shipment_detail", shipment_id=shipment.id)

    dest_location = shipment.destination
    if not _can_access_location(request, tenant, getattr(dest_location, "id", None)):
        messages.error(request, "You do not have access to this shipment's receiving location.")
        return redirect("po_detail", po_id=po.id)
    default_grn = "GRN-" + timezone.now().strftime("%Y%m%d-%H%M%S-%f")

    if request.method == "POST":
        grn_number = (request.POST.get("grn_number") or default_grn).strip()
        received_at = timezone.now()

        try:
            with transaction.atomic():
                receipt = GoodsReceipt.objects.create(
                    tenant=tenant,
                    po=po,
                    shipment=shipment,
                    grn_number=grn_number,
                    received_at=received_at,
                    received_to=dest_location,
                    attachment=request.FILES.get("attachment"),
                    status=GoodsReceipt.Status.DRAFT,
                )

                # Optional single landed cost (MVP). A malformed amount must abort
                # the whole receipt (caught below) rather than be silently dropped:
                # otherwise stock capitalizes without its landed cost and no error
                # surfaces (M13).
                lc_name = (request.POST.get("landed_cost_name") or "").strip()
                lc_amount_raw = (request.POST.get("landed_cost_amount") or "").strip()
                if lc_name and lc_amount_raw:
                    try:
                        lc_amount = Decimal(lc_amount_raw)
                    except InvalidOperation:
                        raise ValueError(f"Invalid landed cost amount '{lc_amount_raw}'.")
                    if lc_amount < Decimal("0.00"):
                        raise ValueError("Landed cost amount cannot be negative.")
                    LandedCostCharge.objects.create(
                        tenant=tenant,
                        receipt=receipt,
                        name=lc_name,
                        amount=lc_amount,
                        currency_code=po.currency_code,
                    )

                # Pass 1: validate + collect received lines and the goods total.
                received_lines = []
                goods_total = Decimal("0.00")
                for sl in shipment.lines.select_related("po_line", "po_line__product"):
                    qty_raw = (request.POST.get(f"recv_{sl.id}") or "").strip()
                    if not qty_raw:
                        continue
                    try:
                        qty = Decimal(qty_raw)
                    except InvalidOperation:
                        raise ValueError(f"Invalid quantity '{qty_raw}' for {sl.po_line.product.sku}.")
                    if qty <= 0:
                        continue
                    if qty > sl.open_qty:
                        raise ValueError(
                            f"Cannot receive more than open qty for {sl.po_line.product.sku}."
                        )
                    expiry_raw = (request.POST.get(f"expiry_{sl.id}") or "").strip()
                    expiry = None
                    if expiry_raw:
                        try:
                            expiry = timezone.datetime.fromisoformat(expiry_raw).date()
                        except Exception:
                            expiry = None
                    received_lines.append({
                        "sl": sl, "qty": qty,
                        "lot_code": (request.POST.get(f"lot_{sl.id}") or "").strip() or None,
                        "serial": (request.POST.get(f"serial_{sl.id}") or "").strip() or None,
                        "expiry": expiry,
                    })
                    goods_total += qty * sl.po_line.unit_cost

                if not received_lines:
                    # Nothing actually received: abort the whole transaction.
                    raise ValueError("Nothing received.")

                # Landed costs apportion across received value, raising unit cost.
                landed_total = sum((lc.amount for lc in receipt.landed_costs.all()), Decimal("0.00"))
                ratio = (landed_total / goods_total) if goods_total > 0 else Decimal("0.00")

                # Pass 2: create GRN lines, apply costed movements.
                from core.services.purchasing import sync_po_line_received
                inventory_value = Decimal("0.00")
                touched_lines = []
                for item in received_lines:
                    sl = item["sl"]
                    qty = item["qty"]
                    base_cost = sl.po_line.unit_cost
                    landed_unit_cost = (base_cost * (Decimal("1") + ratio)).quantize(Decimal("0.0001"))

                    GoodsReceiptLine.objects.create(
                        receipt=receipt, po_line=sl.po_line, product=sl.po_line.product,
                        # Link to the stable PO-line identity so this receipt
                        # counts toward the current version's open/received (M17).
                        root_line_id=(sl.po_line.root_line_id or sl.po_line_id),
                        qty_received=qty, unit_cost=base_cost,
                        lot_code=item["lot_code"], serial_number=item["serial"], expiry_date=item["expiry"],
                    )
                    movement = apply_movement(
                        tenant=tenant,
                        product=sl.po_line.product,
                        location=dest_location,
                        movement_type=InventoryMovement.MovementType.RECEIVE,
                        qty_delta=qty,
                        ref_type="GRN",
                        ref_id=receipt.grn_number,
                        notes=f"Receipt against PO {po.po_number}", user=request.user,
                        lot_code=item["lot_code"], serial_number=item["serial"], expiry_date=item["expiry"],
                        unit_cost=landed_unit_cost,
                    )
                    # Actual capitalized value (standard products differ from cost).
                    inventory_value += movement.value or Decimal("0.00")
                    # Shipment line tracks its own (version-specific) progress.
                    sl.received_qty += qty
                    sl.save(update_fields=["received_qty"])
                    touched_lines.append(sl.po_line)

                # Mark the receipt POSTED *and persist it* before syncing, so the
                # received_qty rebuild (which sums POSTED receipts from the DB)
                # includes this one.
                receipt.status = GoodsReceipt.Status.POSTED
                receipt.posted_at = received_at
                receipt.save(update_fields=["status", "posted_at"])

                # Rebuild received_qty from the actual receipts for every stable
                # identity we touched (writes to all versions of each line).
                for pol in touched_lines:
                    sync_po_line_received(pol)

                # Capitalize: DR Inventory (at cost basis) / CR GRNI (goods) / CR
                # Accruals (landed) / +/- Purchase Price Variance (standard costing).
                post_inventory_receipt(tenant, goods_total, receipt.grn_number, user=request.user,
                                       entry_date=received_at.date(), landed_value=landed_total,
                                       inventory_value=inventory_value,
                                       site_id=getattr(dest_location, "site_id", None))

                # Update PO status (refresh lines so open_qty reflects the sync).
                po.refresh_from_db()
                if all((l.open_qty == Decimal("0.00") for l in po.lines.all())):
                    po.status = PurchaseOrder.Status.RECEIVED
                else:
                    po.status = PurchaseOrder.Status.PARTIALLY_RECEIVED
                po.save(update_fields=["status"])
        except ValueError as e:
            # Validation failure (over-receipt / nothing received): roll back and re-prompt.
            messages.error(request, str(e))
            return redirect("receive_po", po_id=po.id)

        messages.success(request, "Receipt posted.")
        return redirect("po_detail", po_id=po.id)

    return render(request, "receive_po.html", {
        "tenant": tenant,
        "po": po,
        "shipment": shipment,
        "default_grn": default_grn,
    })

@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE, ROLE_PROCUREMENT, ROLE_FINANCE, ROLE_READONLY])
def inventory_list(request):
    from core.access import accessible_location_ids
    tenant = _get_default_tenant(request)
    balances = (
        InventoryBalance.objects
        .filter(tenant=tenant)
        .select_related("product", "location")
        .order_by("product__sku", "location__name")
    )
    # Restrict to the locations this user may see.
    allowed = active_location_ids(request)  # scope to the selected site
    if allowed is not None:
        balances = balances.filter(location_id__in=allowed)
    # Optional location filter.
    loc_id = request.GET.get("location") or ""
    if loc_id:
        balances = balances.filter(location_id=loc_id)
    locations = Location.objects.filter(tenant=tenant)
    if allowed is not None:
        locations = locations.filter(id__in=allowed)
    return render(request, "inventory_list.html", {
        "tenant": tenant, "balances": balances,
        "locations": locations.order_by("name"), "location": loc_id})


# ============================
# Stock adjustments (damage / loss / write-off / return-to-supplier) + approval
# ============================

def _post_stock_adjustment(adj, user):
    """Write the inventory movement for an adjustment, post its GL impact, and
    mark it posted. The GL value follows the costed movement value so damage /
    write-off / shrinkage hit the books (DR Inventory Adjustments / CR Inventory),
    keeping the inventory control account in step with the stock ledger."""
    from core.models import StockAdjustment
    from core.services.gl import post_stock_adjustment
    movement = apply_movement(
        tenant=adj.tenant, product=adj.product, location=adj.location,
        movement_type=adj.movement_type, qty_delta=adj.qty_delta,
        ref_type="STOCK_ADJ", ref_id=str(adj.id),
        notes=(adj.get_reason_display() + (f": {adj.notes}" if adj.notes else "")), user=user,
        bin=adj.bin,
    )
    value = getattr(movement, "value", None) or Decimal("0.00")
    fields = ["status", "posted_at", "approved_by"]
    if adj.reason == StockAdjustment.Reason.RETURN_SUPPLIER and adj.supplier_id:
        # Return to supplier: raise a purchase credit note (DR AP / CR Inventory)
        # rather than booking the value as shrinkage.
        from core.services.purchasing import create_return_credit_note
        cn = create_return_credit_note(adj, value, user=user)
        if cn is not None:
            adj.credit_note = cn
            fields.append("credit_note")
    else:
        # Book the GL impact using the costed movement value (signed like qty_delta).
        post_stock_adjustment(adj, value, user=user)
    adj.status = StockAdjustment.Status.POSTED
    adj.posted_at = timezone.now()
    adj.approved_by = user
    adj.save(update_fields=fields)


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_READONLY], [ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE])
def adjustment_list(request):
    from core.models import StockAdjustment
    from core.access import accessible_location_ids
    tenant = _get_default_tenant(request)
    adjustments = StockAdjustment.objects.filter(tenant=tenant).select_related("product", "location", "requested_by")
    allowed = active_location_ids(request)  # scope to the selected site
    if allowed is not None:
        adjustments = adjustments.filter(location_id__in=allowed)
    return render(request, "inventory/adjustment_list.html", {
        "tenant": tenant, "adjustments": adjustments,
        "threshold": tenant.stock_adjustment_approval_threshold,
        "can_approve": bool({ROLE_ADMIN, ROLE_PROCUREMENT} & effective_groups(request)) or request.user.is_superuser,
    })


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE])
@transaction.atomic
def adjustment_create(request):
    from core.models import StockAdjustment
    from core.forms import StockAdjustmentForm
    tenant = _get_default_tenant(request)
    form = StockAdjustmentForm(request.POST or None)
    _scope_location_fields_by_site(form, request, "location")
    if request.method == "POST" and form.is_valid():
        adj = form.save(commit=False)
        adj.tenant = tenant
        adj.requested_by = request.user
        adj.estimated_value = (abs(adj.qty_delta) * (adj.product.cost_price or Decimal("0.00"))).quantize(Decimal("0.01"))
        threshold = tenant.stock_adjustment_approval_threshold or Decimal("0.00")
        needs_approval = threshold > 0 and adj.estimated_value >= threshold
        adj.status = StockAdjustment.Status.PENDING
        adj.save()
        if needs_approval:
            log_audit(action="STOCK_ADJ_REQUESTED", request=request, user=request.user, tenant=tenant,
                      detail=f"{adj.product.sku} {adj.qty_delta} ({adj.get_reason_display()}) - awaiting approval")
            messages.warning(request, f"Adjustment for {adj.product.sku} needs approval (value {adj.estimated_value}). It's pending.")
        else:
            _post_stock_adjustment(adj, request.user)
            log_audit(action="STOCK_ADJUSTED", request=request, user=request.user, tenant=tenant,
                      detail=f"{adj.product.sku} {adj.qty_delta} ({adj.get_reason_display()})")
            messages.success(request, f"Stock adjusted: {adj.product.sku} {adj.qty_delta}.")
        return redirect("adjustment_list")
    return render(request, "inventory/adjustment_form.html", {"tenant": tenant, "form": form})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
@transaction.atomic
def adjustment_approve(request, adj_id):
    from core.models import StockAdjustment
    tenant = _get_default_tenant(request)
    adj = get_object_or_404(StockAdjustment, id=adj_id, tenant=tenant)
    if request.method == "POST" and adj.status == StockAdjustment.Status.PENDING:
        _post_stock_adjustment(adj, request.user)
        log_audit(action="STOCK_ADJ_APPROVED", request=request, user=request.user, tenant=tenant,
                  detail=f"{adj.product.sku} {adj.qty_delta} ({adj.get_reason_display()})")
        messages.success(request, f"Adjustment approved and posted: {adj.product.sku} {adj.qty_delta}.")
    return redirect("adjustment_list")


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
@transaction.atomic
def adjustment_reject(request, adj_id):
    from core.models import StockAdjustment
    tenant = _get_default_tenant(request)
    adj = get_object_or_404(StockAdjustment, id=adj_id, tenant=tenant)
    if request.method == "POST" and adj.status == StockAdjustment.Status.PENDING:
        adj.status = StockAdjustment.Status.REJECTED
        adj.approved_by = request.user
        adj.save(update_fields=["status", "approved_by"])
        log_audit(action="STOCK_ADJ_REJECTED", request=request, user=request.user, tenant=tenant,
                  detail=f"{adj.product.sku} {adj.qty_delta}")
        messages.info(request, f"Adjustment rejected: {adj.product.sku} {adj.qty_delta}.")
    return redirect("adjustment_list")


# ============================
# Low-stock alerts + stock movement history
# ============================

@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_READONLY])
def low_stock(request):
    """Products whose total on-hand has fallen below their reorder level."""
    tenant = _get_default_tenant(request)
    rows = []
    for p in (Product.objects.filter(tenant=tenant, reorder_level__gt=0, is_active=True)
              .select_related("preferred_supplier")):
        on_hand = p.on_hand_total
        if on_hand < p.reorder_level:
            rows.append({"product": p, "on_hand": on_hand, "reorder": p.reorder_level,
                         "shortfall": p.reorder_level - on_hand})
    rows.sort(key=lambda r: r["shortfall"], reverse=True)
    return render(request, "inventory/low_stock.html", {"tenant": tenant, "rows": rows})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE])
@transaction.atomic
def low_stock_reorder(request):
    """Create Purchase Requisitions from selected low-stock lines, grouped by
    preferred supplier. Items without a preferred supplier go into one
    'unassigned' requisition the buyer can complete later."""
    tenant = _get_default_tenant(request)
    if request.method != "POST":
        return redirect("low_stock")

    selected_ids = request.POST.getlist("select")
    if not selected_ids:
        messages.info(request, "Select at least one product to reorder.")
        return redirect("low_stock")

    products = {str(p.id): p for p in Product.objects.filter(tenant=tenant, id__in=selected_ids)
                .select_related("preferred_supplier", "tax_code")}

    # Build {supplier_or_None: [(product, qty), ...]}
    groups = {}
    for pid in selected_ids:
        p = products.get(pid)
        if p is None:
            continue
        raw = (request.POST.get(f"qty_{pid}") or "").strip()
        try:
            qty = Decimal(raw)
        except (InvalidOperation, ValueError):
            qty = max(p.reorder_level - p.on_hand_total, Decimal("0"))
        if qty <= 0:
            continue
        groups.setdefault(p.preferred_supplier, []).append((p, qty))

    if not groups:
        messages.info(request, "Nothing to order (quantities were zero).")
        return redirect("low_stock")

    created = []
    for supplier, items in groups.items():
        req = PurchaseRequisition.objects.create(
            tenant=tenant,
            req_number=_generate_req_number(),
            preferred_supplier=supplier,
            justification="Raised automatically from low-stock alerts.",
            status=PurchaseRequisition.Status.DRAFT,
            requested_by=request.user,
        )
        for p, qty in items:
            PurchaseRequisitionLine.objects.create(
                requisition=req,
                product=p,
                quantity=qty,
                estimated_unit_cost=(p.standard_cost or None),
                notes="Below reorder level",
            )
        created.append(req)
        log_audit(action="requisition_create", request=request, user=request.user, tenant=tenant,
                  detail=f"{req.req_number} auto-reorder ({len(items)} lines)")

    if len(created) == 1:
        messages.success(request, f"Created requisition {created[0].req_number} from low-stock.")
        return redirect("requisition_detail", req_id=created[0].id)
    messages.success(request, f"Created {len(created)} requisitions from low-stock (grouped by supplier).")
    return redirect("requisition_list")


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_READONLY])
def stock_movements(request):
    """The stock ledger: every movement, filterable by product/location/type/date."""
    tenant = _get_default_tenant(request)
    product_id = request.GET.get("product") or ""
    location_id = request.GET.get("location") or ""
    mtype = request.GET.get("type") or ""
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))

    from core.access import accessible_location_ids
    allowed = active_location_ids(request)  # scope to the selected site

    qs = InventoryMovement.objects.filter(tenant=tenant).select_related("product", "location", "user")
    if allowed is not None:
        qs = qs.filter(location_id__in=allowed)
    if product_id:
        qs = qs.filter(product_id=product_id)
    if location_id:
        qs = qs.filter(location_id=location_id)
    if mtype:
        qs = qs.filter(movement_type=mtype)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    movements = qs.order_by("-created_at", "-id")[:300]

    loc_choices = Location.objects.filter(tenant=tenant)
    if allowed is not None:
        loc_choices = loc_choices.filter(id__in=allowed)

    return render(request, "inventory/stock_movements.html", {
        "tenant": tenant, "movements": movements,
        "product": product_id, "location": location_id, "type": mtype,
        "date_from": date_from, "date_to": date_to,
        "products": Product.objects.filter(tenant=tenant).order_by("sku"),
        "locations": loc_choices.order_by("name"),
        "type_choices": InventoryMovement.MovementType.choices,
        "filtered": bool(product_id or location_id or mtype or date_from or date_to),
    })


@login_required
@role_required([ROLE_ADMIN, ROLE_FINANCE, ROLE_READONLY])

def reconcile(request):
    tenant = _get_default_tenant(request)

    # Latest snapshot per SKU for Shopify (MVP)
    latest = {}
    for s in ChannelSnapshot.objects.filter(tenant=tenant, channel=SalesChannel.SHOPIFY).order_by("sku", "-as_of"):
        if s.sku not in latest:
            latest[s.sku] = s

    # SwifPro BI totals per SKU (sum across locations)
    swifpro_totals = {}
    for b in InventoryBalance.objects.filter(tenant=tenant).select_related("product"):
        swifpro_totals.setdefault(b.product.sku, Decimal("0.00"))
        swifpro_totals[b.product.sku] += (b.on_hand or Decimal("0.00"))

    rows = []
    all_skus = sorted(set(list(swifpro_totals.keys()) + list(latest.keys())))
    for sku in all_skus:
        sk_qty = swifpro_totals.get(sku, Decimal("0.00"))
        ch_qty = latest.get(sku).quantity if sku in latest else Decimal("0.00")
        drift = ch_qty - sk_qty
        rows.append({"sku": sku, "swifpro_qty": sk_qty, "channel_qty": ch_qty, "drift": drift})

    return render(request, "reconcile.html", {"tenant": tenant, "rows": rows})

@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_SALES, ROLE_READONLY])

def product_list(request):
    from core.models import ProductCategory
    tenant = _get_default_tenant(request)
    q = (request.GET.get("q") or "").strip()
    ptype = request.GET.get("type") or ""
    category = request.GET.get("category") or ""
    status = request.GET.get("status") or ""
    qs = Product.objects.filter(tenant=tenant).select_related("category")
    if q:
        qs = qs.filter(Q(sku__icontains=q) | Q(name__icontains=q) | Q(brand__icontains=q)
                       | Q(barcodes__code__icontains=q)).distinct()
    if ptype:
        qs = qs.filter(product_type=ptype)
    if category:
        qs = qs.filter(category_id=category)
    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)
    qs = qs.order_by("sku")
    return render(request, "products/product_list.html", {
        "tenant": tenant, "products": qs,
        "q": q, "type": ptype, "category": category, "status": status,
        "type_choices": Product.Type.choices,
        "categories": ProductCategory.objects.filter(tenant=tenant).order_by("name"),
        "filtered": bool(q or ptype or category or status),
    })

@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])

def product_create(request):
    tenant = _get_default_tenant(request)
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tenant = tenant
            obj.save()
            barcode = form.cleaned_data.get("barcode")
            if barcode:
                ProductBarcode.objects.get_or_create(tenant=tenant, code=barcode, defaults={"product": obj})
            # Optional opening stock -> a one-off receipt movement (sets cost).
            opening = form.cleaned_data.get("opening_stock")
            loc = form.cleaned_data.get("opening_location")
            if opening and opening > 0 and loc:
                apply_movement(tenant=tenant, product=obj, location=loc,
                               movement_type=InventoryMovement.MovementType.RECEIVE,
                               qty_delta=opening, ref_type="OPENING", ref_id=obj.sku,
                               notes="Opening stock", unit_cost=(obj.standard_cost or None), user=request.user)
            messages.success(request, "Product created.")
            return redirect("product_detail", product_id=obj.id)
    else:
        form = ProductForm()

    return render(request, "products/product_form.html", {
        "tenant": tenant, "form": form, "mode": "create"
    })

@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])

def product_edit(request, product_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Product, id=product_id, tenant=tenant)

    if request.method == "POST":
        old_cost = obj.standard_cost
        form = ProductForm(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            obj = form.save()
            # Audit a change to the product's standard (cost) price.
            new_cost = obj.standard_cost
            if old_cost != new_cost:
                log_audit(action="PRODUCT_COST_CHANGED", request=request, user=request.user, tenant=tenant,
                          entity_type="Product", entity_id=obj.sku,
                          old_value=old_cost, new_value=new_cost,
                          detail=f"{obj.sku} standard cost {old_cost} -> {new_cost}")
            barcode = form.cleaned_data.get("barcode")
            if barcode:
                ProductBarcode.objects.update_or_create(
                    tenant=tenant, code=barcode, defaults={"product": obj}
                )
            messages.success(request, "Product updated.")
            return redirect("product_detail", product_id=obj.id)
    else:
        initial = {}
        bc = obj.barcodes.order_by('id').first() if hasattr(obj, 'barcodes') else None
        if bc:
            initial['barcode'] = bc.code
        form = ProductForm(instance=obj, initial=initial)

    return render(request, "products/product_form.html", {
        "tenant": tenant, "form": form, "mode": "edit", "product": obj
    })


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_SALES, ROLE_READONLY])
def product_detail(request, product_id):
    """Product profile: info, stock by location, sales & purchase history,
    supplier info, margin, stock movements and price history."""
    from core.models import (InventoryBalance, InventoryMovement, CustomerInvoiceLine,
                             PurchaseOrderLine, ProductBarcode)
    tenant = _get_default_tenant(request)
    p = get_object_or_404(Product, id=product_id, tenant=tenant)

    balances = list(InventoryBalance.objects.filter(tenant=tenant, product=p).select_related("location").order_by("location__name"))
    movements = list(InventoryMovement.objects.filter(tenant=tenant, product=p).select_related("location").order_by("-created_at", "-id")[:30])
    barcodes = list(ProductBarcode.objects.filter(tenant=tenant, product=p).values_list("code", flat=True))

    sales_lines = list(CustomerInvoiceLine.objects.filter(invoice__tenant=tenant, product=p)
                       .select_related("invoice", "invoice__customer")
                       .order_by("-invoice__invoice_date")[:20])
    qty_sold = sum((l.qty or Decimal("0.00") for l in sales_lines), Decimal("0.00"))
    revenue = sum((l.line_total for l in sales_lines), Decimal("0.00"))

    po_lines = list(PurchaseOrderLine.objects.filter(po__tenant=tenant, product=p)
                    .select_related("po", "po__supplier").order_by("-po__created_at")[:20])
    qty_purchased = sum((l.ordered_qty or Decimal("0.00") for l in po_lines), Decimal("0.00"))
    # Recorded supplier price history (agreed on PO + actual on bill).
    from core.models import SupplierPriceHistory
    price_history = [{"date": r.recorded_at, "supplier": r.supplier.name, "ref": r.reference,
                      "unit_cost": r.unit_cost, "source": r.get_source_display()}
                     for r in (SupplierPriceHistory.objects.filter(tenant=tenant, product=p)
                               .select_related("supplier")[:30])]

    supplier_names, seen = [], set()
    if p.preferred_supplier_id:
        supplier_names.append((p.preferred_supplier.name, True)); seen.add(p.preferred_supplier_id)
    for l in po_lines:
        if l.po.supplier_id not in seen:
            seen.add(l.po.supplier_id); supplier_names.append((l.po.supplier.name, False))

    return render(request, "products/product_detail.html", {
        "tenant": tenant, "p": p, "balances": balances, "movements": movements, "barcodes": barcodes,
        "sales_lines": sales_lines, "qty_sold": qty_sold, "revenue": revenue,
        "po_lines": po_lines, "qty_purchased": qty_purchased, "price_history": price_history,
        "supplier_names": supplier_names,
    })


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_READONLY], [ROLE_ADMIN, ROLE_PROCUREMENT])
def product_category_list(request):
    from core.models import ProductCategory
    from core.forms import ProductCategoryForm
    tenant = _get_default_tenant(request)
    if request.method == "POST":
        form = ProductCategoryForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tenant = tenant
            try:
                obj.save()
                messages.success(request, "Category saved.")
            except IntegrityError:
                messages.error(request, "That category already exists.")
            return redirect("product_category_list")
    else:
        form = ProductCategoryForm()
    cats = ProductCategory.objects.filter(tenant=tenant, parent__isnull=True).prefetch_related("subcategories", "products").order_by("name")
    return render(request, "products/category_list.html", {"tenant": tenant, "form": form, "categories": cats})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
def product_category_delete(request, category_id):
    from core.models import ProductCategory
    tenant = _get_default_tenant(request)
    cat = get_object_or_404(ProductCategory, id=category_id, tenant=tenant)
    if request.method == "POST":
        name = str(cat)
        cat.delete()  # products' category FK is SET_NULL; subcategories' parent SET_NULL
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant, detail=f"Category {name}")
        messages.success(request, f"Category '{name}' deleted.")
    return redirect("product_category_list")


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])

def product_delete(request, product_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Product, id=product_id, tenant=tenant)

    if request.method == "POST":
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant,
                  detail=f"Product {obj.sku} - {obj.name}")
        obj.delete()
        messages.success(request, "Product deleted.")
        return redirect("product_list")

    return render(request, "products/product_delete.html", {
        "tenant": tenant, "product": obj
    })

@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_FINANCE, ROLE_READONLY])

def supplier_list(request):
    tenant = _get_default_tenant(request)
    q = (request.GET.get("q") or "").strip()
    status = request.GET.get("status") or ""
    category = (request.GET.get("category") or "").strip()
    suppliers = Supplier.objects.filter(tenant=tenant)
    if q:
        suppliers = suppliers.filter(
            Q(name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q)
            | Q(vat_number__icontains=q) | Q(company_number__icontains=q) | Q(contact_person__icontains=q))
    if status:
        suppliers = suppliers.filter(status=status)
    if category:
        suppliers = suppliers.filter(categories__icontains=category)
    suppliers = suppliers.order_by("name")
    return render(request, "suppliers/supplier_list.html", {
        "tenant": tenant, "suppliers": suppliers, "q": q, "status": status, "category": category,
        "status_choices": Supplier.Status.choices, "filtered": bool(q or status or category)})


def _find_supplier_duplicates(tenant, obj, exclude_id=None):
    """Possible duplicate suppliers matching email/phone/VAT/company number/name."""
    checks = [("email", obj.email, "email"), ("phone", obj.phone, "phone"),
              ("vat_number", obj.vat_number, "VAT number"),
              ("company_number", obj.company_number, "company number"),
              ("name", obj.name, "name")]
    seen, out = set(), []
    for field, value, label in checks:
        value = (value or "").strip()
        if not value:
            continue
        qs = Supplier.objects.filter(tenant=tenant, **{f"{field}__iexact": value})
        if exclude_id:
            qs = qs.exclude(id=exclude_id)
        for match in qs:
            if match.id not in seen:
                seen.add(match.id)
                out.append({"id": match.id, "name": match.name, "match": label})
    return out

@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])

def supplier_create(request):
    tenant = _get_default_tenant(request)
    form = SupplierForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        dups = _find_supplier_duplicates(tenant, obj)
        if dups and request.POST.get("confirm_duplicate") != "1":
            return render(request, "suppliers/supplier_form.html",
                          {"tenant": tenant, "form": form, "mode": "create", "duplicates": dups})
        try:
            obj.save()
            messages.success(request, "Supplier created.")
            return redirect("supplier_detail", supplier_id=obj.id)
        except IntegrityError:
            form.add_error("name", "Supplier name already exists.")
    return render(request, "suppliers/supplier_form.html", {"tenant": tenant, "form": form, "mode": "create"})

@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])

def supplier_edit(request, supplier_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Supplier, id=supplier_id, tenant=tenant)
    form = SupplierForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        edited = form.save(commit=False)
        dups = _find_supplier_duplicates(tenant, edited, exclude_id=obj.id)
        if dups and request.POST.get("confirm_duplicate") != "1":
            return render(request, "suppliers/supplier_form.html",
                          {"tenant": tenant, "form": form, "mode": "edit", "duplicates": dups})
        try:
            form.save()
            messages.success(request, "Supplier updated.")
            return redirect("supplier_detail", supplier_id=obj.id)
        except IntegrityError:
            form.add_error("name", "Supplier name already exists.")
    return render(request, "suppliers/supplier_form.html", {"tenant": tenant, "form": form, "mode": "edit"})

@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])

def supplier_delete(request, supplier_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Supplier, id=supplier_id, tenant=tenant)
    if request.method == "POST":
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant,
                  detail=f"Supplier {obj.name}")
        obj.delete()
        messages.success(request, "Supplier deleted.")
        return redirect("supplier_list")
    return render(request, "suppliers/supplier_delete.html", {"tenant": tenant, "supplier": obj})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_FINANCE, ROLE_READONLY])
def supplier_detail(request, supplier_id):
    """Supplier profile: details, purchase orders, bills, payments, outstanding
    payables, products supplied, price history, notes and an activity timeline."""
    tenant = _get_default_tenant(request)
    s = get_object_or_404(Supplier, id=supplier_id, tenant=tenant)

    pos = list(PurchaseOrder.objects.filter(tenant=tenant, supplier=s)
               .prefetch_related("lines", "lines__product").order_by("-created_at", "-id"))
    bills = list(SupplierInvoice.objects.filter(tenant=tenant, supplier=s)
                 .prefetch_related("lines", "lines__tax_code", "payment_allocations", "credit_notes")
                 .order_by("-invoice_date", "-id"))
    payments = list(Payment.objects.filter(tenant=tenant, supplier=s).order_by("-payment_date", "-id"))
    expenses = list(Expense.objects.filter(tenant=tenant, supplier=s).select_related("category").order_by("-expense_date", "-id"))
    credit_notes = list(CreditNote.objects.filter(tenant=tenant, supplier=s, kind=CreditNote.Kind.PURCHASE)
                        .order_by("-credit_note_date", "-id"))

    # Products supplied + price history, derived from this supplier's PO lines.
    products = {}
    price_history = []
    for po in pos:
        po_date = po.created_at.date()
        for line in po.lines.all():
            p = line.product
            info = products.setdefault(p.id, {"product": p, "times": 0, "last_cost": None, "last_date": None})
            info["times"] += 1
            if info["last_date"] is None or po_date >= info["last_date"]:
                info["last_date"] = po_date
                info["last_cost"] = line.unit_cost
            price_history.append({"date": po_date, "product": p, "unit_cost": line.unit_cost,
                                  "ref": po.po_number})
    products_supplied = sorted(products.values(), key=lambda x: x["product"].sku)
    # Products that name this supplier as their preferred supplier also count.
    for p in Product.objects.filter(tenant=tenant, preferred_supplier=s):
        if p.id not in products:
            products_supplied.append({"product": p, "times": 0, "last_cost": p.standard_cost, "last_date": None})
    price_history.sort(key=lambda r: r["date"], reverse=True)

    timeline = []
    for po in pos:
        timeline.append({"date": po.created_at.date(), "icon": "file-earmark-text", "kind": "PO",
                         "text": f"PO {po.po_number} ({po.get_status_display()})", "url": f"/po/{po.id}/"})
    for b in bills:
        timeline.append({"date": b.invoice_date, "icon": "receipt-cutoff", "kind": "Bill",
                         "text": f"Bill {b.invoice_number} ({b.get_status_display()}) - {b.total}", "url": f"/invoices/{b.id}/"})
    for p in payments:
        timeline.append({"date": p.payment_date, "icon": "cash-stack", "kind": "Payment",
                         "text": f"Payment {p.amount} ({p.get_method_display()})", "url": f"/payments/{p.id}/"})
    for cn in credit_notes:
        timeline.append({"date": cn.credit_note_date, "icon": "arrow-return-left", "kind": "Credit note",
                         "text": f"Credit note {cn.credit_note_number} - {cn.total}", "url": f"/credit-notes/{cn.id}/"})
    for ex in expenses:
        timeline.append({"date": ex.expense_date, "icon": "receipt", "kind": "Expense",
                         "text": f"Expense {ex.category.name} - {ex.total} ({ex.get_status_display()})",
                         "url": f"/expenses/{ex.id}/"})
    timeline.sort(key=lambda e: e["date"], reverse=True)

    purchases_total = sum((b.total for b in bills if b.status == "POSTED"), Decimal("0.00"))

    return render(request, "suppliers/supplier_detail.html", {
        "tenant": tenant, "s": s,
        "pos": pos[:10], "bills": bills[:10], "payments": payments[:10],
        "expenses": expenses[:10],
        "products_supplied": products_supplied, "price_history": price_history[:30],
        "timeline": timeline[:40], "purchases_total": purchases_total,
        "po_count": len(pos),
    })


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE, ROLE_READONLY])

def location_list(request):
    tenant = _get_default_tenant(request)
    locations = Location.objects.filter(tenant=tenant).order_by("name")
    return render(request, "locations/location_list.html", {"tenant": tenant, "locations": locations})

@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_WAREHOUSE])

def location_create(request):
    tenant = _get_default_tenant(request)
    form = LocationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        try:
            obj.save()
            messages.success(request, "Location created.")
            return redirect("location_list")
        except IntegrityError:
            form.add_error("name", "Location name already exists.")
    return render(request, "locations/location_form.html", {"tenant": tenant, "form": form, "mode": "create"})

@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_WAREHOUSE])

def location_edit(request, location_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Location, id=location_id, tenant=tenant)
    form = LocationForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        try:
            form.save()
            messages.success(request, "Location updated.")
            return redirect("location_list")
        except IntegrityError:
            form.add_error("name", "Location name already exists.")
    return render(request, "locations/location_form.html", {"tenant": tenant, "form": form, "mode": "edit"})

@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_WAREHOUSE])

def location_delete(request, location_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Location, id=location_id, tenant=tenant)
    if request.method == "POST":
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant,
                  detail=f"Location {obj.name}")
        obj.delete()
        messages.success(request, "Location deleted.")
        return redirect("location_list")
    return render(request, "locations/location_delete.html", {"tenant": tenant, "location": obj})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
@transaction.atomic
def location_access(request):
    """Admin matrix to grant users access to specific locations. A user with no
    ticks sees all locations (open by default); ticking any narrows them."""
    from core.models import OrgMembership, UserLocationAccess
    tenant = _get_default_tenant(request)
    members = list(OrgMembership.objects.filter(tenant=tenant).select_related("user").order_by("user__username"))
    locations = list(Location.objects.filter(tenant=tenant).order_by("name"))

    if request.method == "POST":
        UserLocationAccess.objects.filter(tenant=tenant).delete()
        bulk = []
        for m in members:
            for loc in locations:
                if request.POST.get(f"grant_{m.user_id}_{loc.id}"):
                    bulk.append(UserLocationAccess(tenant=tenant, user=m.user, location=loc))
        if bulk:
            UserLocationAccess.objects.bulk_create(bulk)
        log_audit(action="LOCATION_ACCESS_CHANGED", request=request, user=request.user, tenant=tenant,
                  detail=f"{len(bulk)} grant(s) across {len(members)} user(s)")
        messages.success(request, "Location access updated.")
        return redirect("location_access")

    granted = set(UserLocationAccess.objects.filter(tenant=tenant).values_list("user_id", "location_id"))
    rows = []
    for m in members:
        is_admin = (m.role == roles_mod.ADMIN)
        rows.append({
            "user": m.user, "role": m.role, "is_admin": is_admin,
            "cells": [{"loc": loc, "checked": (m.user_id, loc.id) in granted} for loc in locations],
            "unrestricted": is_admin or not any((m.user_id, loc.id) in granted for loc in locations),
        })
    return render(request, "locations/location_access.html", {
        "tenant": tenant, "rows": rows, "locations": locations})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
@transaction.atomic
def site_access(request):
    """Admin matrix to grant users access to specific Sites. A user with no ticks
    may work in all sites (open by default); ticking any narrows them. Site access
    gates the global Company+Site context, separate from inventory-location access."""
    from core.models import OrgMembership, UserSiteAccess, Site
    tenant = _get_default_tenant(request)
    members = list(OrgMembership.objects.filter(tenant=tenant).select_related("user").order_by("user__username"))
    sites = list(Site.objects.filter(tenant=tenant).order_by("name"))

    if request.method == "POST":
        UserSiteAccess.objects.filter(tenant=tenant).delete()
        bulk = []
        for m in members:
            for site in sites:
                if request.POST.get(f"grant_{m.user_id}_{site.id}"):
                    bulk.append(UserSiteAccess(tenant=tenant, user=m.user, site=site))
        if bulk:
            UserSiteAccess.objects.bulk_create(bulk)
        log_audit(action="SITE_ACCESS_CHANGED", request=request, user=request.user, tenant=tenant,
                  detail=f"{len(bulk)} grant(s) across {len(members)} user(s)")
        messages.success(request, "Site access updated.")
        return redirect("site_access")

    granted = set(UserSiteAccess.objects.filter(tenant=tenant).values_list("user_id", "site_id"))
    rows = []
    for m in members:
        is_admin = (m.role == roles_mod.ADMIN)
        rows.append({
            "user": m.user, "role": m.role, "is_admin": is_admin,
            "cells": [{"site": site, "checked": (m.user_id, site.id) in granted} for site in sites],
            "unrestricted": is_admin or not any((m.user_id, site.id) in granted for site in sites),
        })
    return render(request, "locations/site_access.html", {
        "tenant": tenant, "rows": rows, "sites": sites})


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE, ROLE_READONLY])
def site_list(request):
    from core.models import Site
    tenant = _get_default_tenant(request)
    sites = Site.objects.filter(tenant=tenant).prefetch_related("locations").order_by("name")
    return render(request, "locations/site_list.html", {"tenant": tenant, "sites": sites})


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_WAREHOUSE])
def site_create(request):
    from core.forms import SiteForm
    tenant = _get_default_tenant(request)
    form = SiteForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        try:
            obj.save()
            messages.success(request, "Site created.")
            return redirect("site_list")
        except IntegrityError:
            form.add_error("name", "A site with this name already exists.")
    return render(request, "locations/site_form.html", {"tenant": tenant, "form": form, "mode": "create"})


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_WAREHOUSE])
def site_edit(request, site_id):
    from core.models import Site
    from core.forms import SiteForm
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Site, id=site_id, tenant=tenant)
    form = SiteForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        try:
            form.save()
            messages.success(request, "Site updated.")
            return redirect("site_list")
        except IntegrityError:
            form.add_error("name", "A site with this name already exists.")
    return render(request, "locations/site_form.html", {"tenant": tenant, "form": form, "mode": "edit"})


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_WAREHOUSE])
def site_delete(request, site_id):
    from core.models import Site
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Site, id=site_id, tenant=tenant)
    if request.method == "POST":
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant, detail=f"Site {obj.name}")
        obj.delete()  # locations' site FK is SET_NULL
        messages.success(request, "Site deleted.")
        return redirect("site_list")
    return render(request, "locations/site_delete.html", {"tenant": tenant, "site": obj})


@login_required
@role_required([ROLE_ADMIN, ROLE_READONLY])
def department_list(request):
    from core.models import Department
    tenant = _get_default_tenant(request)
    departments = (Department.objects.filter(tenant=tenant)
                   .select_related("site", "manager").prefetch_related("members")
                   .order_by("name"))
    return render(request, "departments/department_list.html", {"tenant": tenant, "departments": departments})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def department_create(request):
    from core.forms import DepartmentForm
    tenant = _get_default_tenant(request)
    form = DepartmentForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        try:
            with transaction.atomic():
                obj.save()
            messages.success(request, "Department created.")
            return redirect("department_list")
        except IntegrityError:
            form.add_error("name", "A department with this name already exists.")
    return render(request, "departments/department_form.html", {"tenant": tenant, "form": form, "mode": "create"})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def department_edit(request, department_id):
    from core.models import Department
    from core.forms import DepartmentForm
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Department, id=department_id, tenant=tenant)
    form = DepartmentForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        try:
            with transaction.atomic():
                form.save()
            messages.success(request, "Department updated.")
            return redirect("department_list")
        except IntegrityError:
            form.add_error("name", "A department with this name already exists.")
    return render(request, "departments/department_form.html", {"tenant": tenant, "form": form, "mode": "edit"})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def department_delete(request, department_id):
    from core.models import Department
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Department, id=department_id, tenant=tenant)
    if request.method == "POST":
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant,
                  detail=f"Department {obj.name}")
        obj.delete()  # members' department FK is SET_NULL
        messages.success(request, "Department deleted.")
        return redirect("department_list")
    return render(request, "departments/department_delete.html", {"tenant": tenant, "department": obj})


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE, ROLE_READONLY])
def bin_list(request):
    from core.models import Bin
    from core.access import accessible_location_ids
    tenant = _get_default_tenant(request)
    bins = Bin.objects.filter(tenant=tenant).select_related("location").order_by("location__name", "code")
    allowed = active_location_ids(request)  # scope to the selected site
    if allowed is not None:
        bins = bins.filter(location_id__in=allowed)
    return render(request, "locations/bin_list.html", {"tenant": tenant, "bins": bins})


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_WAREHOUSE])
def bin_create(request):
    from core.forms import BinForm
    tenant = _get_default_tenant(request)
    form = BinForm(request.POST or None)
    _scope_location_fields_by_site(form, request, "location")
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        try:
            obj.save()
            messages.success(request, "Bin created.")
            return redirect("bin_list")
        except IntegrityError:
            form.add_error("code", "A bin with this code already exists at that location.")
    return render(request, "locations/bin_form.html", {"tenant": tenant, "form": form, "mode": "create"})


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_WAREHOUSE])
def bin_edit(request, bin_id):
    from core.models import Bin
    from core.forms import BinForm
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Bin, id=bin_id, tenant=tenant)
    form = BinForm(request.POST or None, instance=obj)
    _scope_location_fields_by_site(form, request, "location")
    if request.method == "POST" and form.is_valid():
        try:
            form.save()
            messages.success(request, "Bin updated.")
            return redirect("bin_list")
        except IntegrityError:
            form.add_error("code", "A bin with this code already exists at that location.")
    return render(request, "locations/bin_form.html", {"tenant": tenant, "form": form, "mode": "edit"})


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_WAREHOUSE])
def bin_delete(request, bin_id):
    from core.models import Bin
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Bin, id=bin_id, tenant=tenant)
    if request.method == "POST":
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant, detail=f"Bin {obj}")
        obj.delete()
        messages.success(request, "Bin deleted.")
        return redirect("bin_list")
    return render(request, "locations/bin_delete.html", {"tenant": tenant, "bin": obj})


@login_required
@role_required([ROLE_ADMIN, ROLE_FINANCE])

def channel_list(request):
    tenant = _get_default_tenant(request)
    conns = ChannelConnection.objects.filter(tenant=tenant).order_by("channel", "name")
    return render(request, "channels/channel_list.html", {"tenant": tenant, "conns": conns})

@login_required
@role_required([ROLE_ADMIN, ROLE_FINANCE], [ROLE_ADMIN, ROLE_FINANCE])

def channel_create(request):
    tenant = _get_default_tenant(request)
    form = ChannelConnectionForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        obj.save()
        messages.success(request, "Connection saved.")
        return redirect("channel_list")
    return render(request, "channels/channel_form.html", {"tenant": tenant, "form": form, "mode": "create"})

@login_required
@role_required([ROLE_ADMIN, ROLE_FINANCE], [ROLE_ADMIN, ROLE_FINANCE])

def channel_edit(request, conn_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(ChannelConnection, id=conn_id, tenant=tenant)
    form = ChannelConnectionForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Connection updated.")
        return redirect("channel_list")
    return render(request, "channels/channel_form.html", {"tenant": tenant, "form": form, "mode": "edit"})

@login_required
@role_required([ROLE_ADMIN, ROLE_FINANCE], [ROLE_ADMIN, ROLE_FINANCE])

def channel_delete(request, conn_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(ChannelConnection, id=conn_id, tenant=tenant)
    if request.method == "POST":
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant,
                  detail=f"Channel connection {obj.name}")
        obj.delete()
        messages.success(request, "Connection deleted.")
        return redirect("channel_list")
    return render(request, "channels/channel_delete.html", {"tenant": tenant, "conn": obj})

@login_required
@role_required([ROLE_ADMIN, ROLE_SALES, ROLE_READONLY])

def sales_order_list(request):
    tenant = _get_default_tenant(request)
    orders = SalesOrder.objects.filter(tenant=tenant).order_by("-order_date")
    return render(request, "sales/sales_order_list.html", {"tenant": tenant, "orders": orders})

@transaction.atomic
@login_required
@role_required([ROLE_ADMIN, ROLE_SALES], [ROLE_ADMIN, ROLE_SALES])
def sales_order_create(request):
    tenant = _get_default_tenant(request)
    order = SalesOrder(tenant=tenant, currency_code=tenant.currency_code)

    if request.method == "POST":
        form = SalesOrderForm(request.POST, instance=order)
        formset = SalesOrderLineFormSet(request.POST, instance=order)
        if form.is_valid() and formset.is_valid():
            order = form.save(commit=False)
            order.tenant = tenant
            order.currency_code = tenant.currency_code
            order.save()
            formset.save()

            # Sync reservations for draft orders (components for kits)
            ref_id = f"{order.channel}:{order.order_number}"
            release_reservations(tenant=tenant, ref_type="SALES_ORDER", ref_id=ref_id)

            for line in order.lines.select_related("product").all():
                ship_loc = line.ship_from_location or order.ship_from_location
                if not ship_loc:
                    continue
                for comp, comp_qty in explode_product(line.product, Decimal(line.qty)):
                    reserve_stock(
                        tenant=tenant,
                        product=comp,
                        location=ship_loc,
                        qty=comp_qty,
                        ref_type="SALES_ORDER",
                        ref_id=ref_id,
                        lot_code=line.lot_code,
                        serial_number=line.serial_number,
                        expiry_date=line.expiry_date,
                    )

            action = form.cleaned_data.get("action") or "save"
            shortages = []
            if action == "post":
                shortages = _post_sales_order(order) or []
                if shortages:
                    sample = ", ".join([f"{s['sku']}@{s['location']} short {s['short_by']}" for s in shortages[:6]])
                    more = "" if len(shortages) <= 6 else f" (+{len(shortages)-6} more)"
                    messages.warning(request, "Posted with shortages (allowed): " + sample + more)
                else:
                    messages.success(request, "Sales order posted.")

            return redirect("sales_order_detail", order_id=order.id)
    else:
        form = SalesOrderForm(instance=order)
        formset = SalesOrderLineFormSet(instance=order)

    return render(request, "sales/sales_order_form.html", {
        "tenant": tenant, "form": form, "formset": formset, "mode": "create"
    })


@login_required
@role_required([ROLE_ADMIN, ROLE_SALES, ROLE_READONLY])

def sales_order_detail(request, order_id):
    tenant = _get_default_tenant(request)
    order = get_object_or_404(SalesOrder, id=order_id, tenant=tenant)
    subtotal = sum((l.line_total for l in order.lines.all()), Decimal("0.00"))
    return render(request, "sales/sales_order_detail.html", {
        "tenant": tenant, "order": order, "subtotal": subtotal
    })

@transaction.atomic
@login_required
@role_required([ROLE_ADMIN, ROLE_SALES], [ROLE_ADMIN, ROLE_SALES])
def sales_order_post(request, order_id):
    tenant = _get_default_tenant(request)
    order = get_object_or_404(SalesOrder, id=order_id, tenant=tenant)
    if request.method == "POST":
        shortages = _post_sales_order(order) or []
        if shortages:
            sample = ", ".join([f"{s['sku']}@{s['location']} short {s['short_by']}" for s in shortages[:6]])
            more = "" if len(shortages) <= 6 else f" (+{len(shortages)-6} more)"
            messages.warning(request, "Posted with shortages (allowed): " + sample + more)
        else:
            messages.success(request, "Sales order posted.")
        return redirect("sales_order_detail", order_id=order.id)
    return render(request, "sales/sales_order_post.html", {"tenant": tenant, "order": order})

def _post_sales_order(order: SalesOrder):
    """Post sales order:
    - releases existing reservations for the order
    - deducts inventory via movements (kit policy: deduct components only)
    - allows negative inventory but returns shortages for UI warnings (per your rule)
    """
    if order.status == SalesOrder.Status.POSTED:
        return []

    shortages = []
    ref_id = f"{order.channel}:{order.order_number}"

    # Consume the order's reservations: posting fulfils it, so the held stock
    # ships rather than being freed unfulfilled (reservation lifecycle fix).
    consume_reservations(tenant=order.tenant, ref_type="SALES_ORDER", ref_id=ref_id)

    cogs_total = Decimal("0.00")
    for line in order.lines.select_related("product").all():
        ship_loc = line.ship_from_location or order.ship_from_location
        if not ship_loc:
            continue

        qty = Decimal(line.qty)
        if qty <= 0:
            continue

        # Explode kits/bundles to components (recommended ERP approach)
        for comp, comp_qty in explode_product(line.product, qty):
            if comp_qty <= 0:
                continue

            # availability check (allow with warning)
            try:
                bal = InventoryBalance.objects.get(tenant=order.tenant, product=comp, location=ship_loc)
                available = (bal.on_hand or Decimal("0.00")) - (bal.reserved or Decimal("0.00"))
            except InventoryBalance.DoesNotExist:
                available = Decimal("0.00")

            if available < comp_qty:
                shortages.append({
                    "sku": comp.sku,
                    "location": ship_loc.name,
                    "required": str(comp_qty),
                    "available": str(available),
                    "short_by": str(comp_qty - available),
                })

            movement = apply_movement(
                tenant=order.tenant,
                product=comp,
                location=ship_loc,
                movement_type=InventoryMovement.MovementType.SALE,
                qty_delta=(comp_qty * Decimal("-1")),
                ref_type="SALES_ORDER",
                ref_id=ref_id,
                notes="Sales order posted",
                lot_code=line.lot_code,
                serial_number=line.serial_number,
                expiry_date=line.expiry_date,
            )
            # movement.value is negative for outbound; COGS is its absolute value.
            cogs_total += -(movement.value or Decimal("0.00"))

    # Expense cost of goods sold: DR COGS / CR Inventory.
    post_cogs(order.tenant, cogs_total, ref_id, entry_date=order.order_date.date())

    order.status = SalesOrder.Status.POSTED
    order.save()
    return shortages

@login_required
@role_required([ROLE_ADMIN])

def settings_tenant(request):
    tenant = _get_default_tenant(request)
    if not tenant:
        # your app already uses needs_setup state
        return redirect("/admin/")

    # Snapshot VAT settings before binding: a ModelForm mutates its instance
    # during is_valid(), so capture the originals from the DB first.
    vat_before = (tenant.vat_registered, tenant.vat_number)
    form = TenantSettingsForm(request.POST or None, request.FILES or None, instance=tenant)
    if request.method == "POST" and form.is_valid():
        obj = form.save()
        if (obj.vat_registered, obj.vat_number) != vat_before:
            log_audit(action="VAT_SETTINGS_CHANGED", request=request, user=request.user, tenant=tenant,
                      detail=f"VAT registered={obj.vat_registered}, number={obj.vat_number or '-'}")
        messages.success(request, "Company profile updated.")
        return redirect("settings_tenant")

    return render(request, "settings/tenant_settings.html", {
        "tenant": tenant,
        "form": form
    })


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
@transaction.atomic
def settings_group(request):
    """Manage the company group: create a group for this company, join an
    existing group the admin already belongs to, or leave the group."""
    from core.models import CompanyGroup, OrgMembership
    from core.access import group_companies
    tenant = _get_default_tenant(request)
    if request.method == "POST":
        op = request.POST.get("op")
        if op == "create":
            name = (request.POST.get("name") or "").strip()
            if not name:
                messages.error(request, "Enter a group name.")
            else:
                grp = CompanyGroup.objects.create(name=name)
                tenant.group = grp
                tenant.save(update_fields=["group"])
                log_audit(action="GROUP_CHANGED", request=request, user=request.user, tenant=tenant,
                          detail=f"Created group '{name}' and joined")
                messages.success(request, f"Group '{name}' created and {tenant.name} joined.")
        elif op == "join":
            gid = request.POST.get("group_id")
            grp = CompanyGroup.objects.filter(id=gid).first()
            if grp:
                tenant.group = grp
                tenant.save(update_fields=["group"])
                log_audit(action="GROUP_CHANGED", request=request, user=request.user, tenant=tenant,
                          detail=f"Joined group '{grp.name}'")
                messages.success(request, f"{tenant.name} joined group '{grp.name}'.")
        elif op == "leave":
            tenant.group = None
            tenant.save(update_fields=["group"])
            log_audit(action="GROUP_CHANGED", request=request, user=request.user, tenant=tenant, detail="Left group")
            messages.success(request, f"{tenant.name} left its group.")
        return redirect("settings_group")

    # Existing groups the admin can join: groups containing a company they belong to.
    my_tenant_ids = set(OrgMembership.objects.filter(user=request.user).values_list("tenant_id", flat=True))
    from core.models import Tenant as _T
    joinable = (CompanyGroup.objects
                .filter(companies__id__in=my_tenant_ids).distinct().order_by("name"))
    siblings = group_companies(request.user, tenant) if tenant.group_id else []
    return render(request, "settings/company_group.html", {
        "tenant": tenant, "joinable": [g for g in joinable if g.id != tenant.group_id],
        "siblings": siblings,
    })


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def uom_list(request):
    tenant = _get_default_tenant(request)
    uoms = UnitOfMeasure.objects.filter(tenant=tenant).order_by("code")
    return render(request, "uoms/uom_list.html", {"tenant": tenant, "uoms": uoms})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def uom_create(request):
    tenant = _get_default_tenant(request)
    form = UnitOfMeasureForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        try:
            obj.save()
            messages.success(request, "UOM created.")
            return redirect("uom_list")
        except IntegrityError:
            form.add_error("code", "UOM code already exists.")
    return render(request, "uoms/uom_form.html", {"tenant": tenant, "form": form, "mode": "create"})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def uom_edit(request, uom_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(UnitOfMeasure, id=uom_id, tenant=tenant)
    form = UnitOfMeasureForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        try:
            form.save()
            messages.success(request, "UOM updated.")
            return redirect("uom_list")
        except IntegrityError:
            form.add_error("code", "UOM code already exists.")
    return render(request, "uoms/uom_form.html", {"tenant": tenant, "form": form, "mode": "edit"})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def uom_delete(request, uom_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(UnitOfMeasure, id=uom_id, tenant=tenant)
    if request.method == "POST":
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant,
                  detail=f"UOM {obj}")
        obj.delete()
        messages.success(request, "UOM deleted.")
        return redirect("uom_list")
    return render(request, "uoms/uom_delete.html", {"tenant": tenant, "uom": obj})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def uom_conversion_list(request):
    tenant = _get_default_tenant(request)
    conversions = UOMConversion.objects.filter(tenant=tenant).select_related("product", "from_uom", "to_uom").order_by("product__sku", "from_uom__code")
    return render(request, "uoms/conversion_list.html", {"tenant": tenant, "conversions": conversions})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def uom_conversion_create(request):
    tenant = _get_default_tenant(request)
    form = UOMConversionForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        try:
            obj.save()
            messages.success(request, "Conversion saved.")
            return redirect("uom_conversion_list")
        except IntegrityError:
            form.add_error(None, "This conversion already exists.")
    return render(request, "uoms/conversion_form.html", {"tenant": tenant, "form": form, "mode": "create"})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def uom_conversion_edit(request, conv_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(UOMConversion, id=conv_id, tenant=tenant)
    form = UOMConversionForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        try:
            form.save()
            messages.success(request, "Conversion updated.")
            return redirect("uom_conversion_list")
        except IntegrityError:
            form.add_error(None, "This conversion already exists.")
    return render(request, "uoms/conversion_form.html", {"tenant": tenant, "form": form, "mode": "edit"})


@login_required
@role_required([ROLE_ADMIN], [ROLE_ADMIN])
def uom_conversion_delete(request, conv_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(UOMConversion, id=conv_id, tenant=tenant)
    if request.method == "POST":
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant,
                  detail=f"UOM conversion {obj}")
        obj.delete()
        messages.success(request, "Conversion deleted.")
        return redirect("uom_conversion_list")
    return render(request, "uoms/conversion_delete.html", {"tenant": tenant, "conv": obj})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_SALES], [ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_SALES])
def bom_list(request):
    tenant = _get_default_tenant(request)
    boms = BillOfMaterials.objects.filter(tenant=tenant).select_related("product").order_by("-is_active", "product__sku")
    return render(request, "boms/bom_list.html", {"tenant": tenant, "boms": boms})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
def bom_create(request):
    tenant = _get_default_tenant(request)
    form = BillOfMaterialsForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        obj.save()
        messages.success(request, "BOM created.")
        return redirect("bom_detail", bom_id=obj.id)
    return render(request, "boms/bom_form.html", {"tenant": tenant, "form": form, "mode": "create"})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_SALES], [ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_SALES])
def bom_detail(request, bom_id):
    tenant = _get_default_tenant(request)
    bom = get_object_or_404(BillOfMaterials, id=bom_id, tenant=tenant)
    if request.method == "POST":
        form = BillOfMaterialsForm(request.POST, instance=bom)
        formset = BOMLineFormSet(request.POST, instance=bom)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, "BOM updated.")
            return redirect("bom_detail", bom_id=bom.id)
    else:
        form = BillOfMaterialsForm(instance=bom)
        formset = BOMLineFormSet(instance=bom)
    return render(request, "boms/bom_detail.html", {"tenant": tenant, "bom": bom, "form": form, "formset": formset})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
def bom_delete(request, bom_id):
    tenant = _get_default_tenant(request)
    bom = get_object_or_404(BillOfMaterials, id=bom_id, tenant=tenant)
    if request.method == "POST":
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant,
                  detail=f"BOM {bom}")
        bom.delete()
        messages.success(request, "BOM deleted.")
        return redirect("bom_list")
    return render(request, "boms/bom_delete.html", {"tenant": tenant, "bom": bom})


# ---------------- Cycle Counts ----------------

from core.auth import role_required, ROLE_ADMIN, ROLE_WAREHOUSE, ROLE_FINANCE
from core.forms import CycleCountForm, CycleCountLineFormSet
from core.models import CycleCount, CycleCountLine, InventoryLotBalance

@role_required(read_groups=[ROLE_ADMIN, ROLE_WAREHOUSE, ROLE_FINANCE])
def cycle_count_list(request):
    tenant = _get_default_tenant(request)
    qs = CycleCount.objects.filter(tenant=tenant).select_related("location").order_by("-created_at")
    return render(request, "inventory/cycle_count_list.html", {"tenant": tenant, "cycle_counts": qs})

@role_required(read_groups=[ROLE_ADMIN, ROLE_WAREHOUSE], write_groups=[ROLE_ADMIN, ROLE_WAREHOUSE])
@transaction.atomic
def cycle_count_create(request):
    tenant = _get_default_tenant(request)
    cc = CycleCount(tenant=tenant)
    if request.method == "POST":
        form = CycleCountForm(request.POST, instance=cc)
        formset = CycleCountLineFormSet(request.POST, instance=cc)
        if form.is_valid() and formset.is_valid():
            cc = form.save(commit=False)
            cc.tenant = tenant
            cc.save()
            formset.save()
            return redirect("cycle_count_detail", cc_id=cc.id)
    else:
        form = CycleCountForm(instance=cc)
        formset = CycleCountLineFormSet(instance=cc)

    _scope_location_fields_by_site(form, request, "location")
    return render(request, "inventory/cycle_count_form.html", {
        "tenant": tenant, "form": form, "formset": formset
    })

@role_required(read_groups=[ROLE_ADMIN, ROLE_WAREHOUSE, ROLE_FINANCE])
def cycle_count_detail(request, cc_id):
    tenant = _get_default_tenant(request)
    cc = get_object_or_404(CycleCount, id=cc_id, tenant=tenant)
    return render(request, "inventory/cycle_count_detail.html", {"tenant": tenant, "cc": cc})

@role_required(read_groups=[ROLE_ADMIN, ROLE_WAREHOUSE], write_groups=[ROLE_ADMIN, ROLE_WAREHOUSE])
@transaction.atomic
def cycle_count_submit(request, cc_id):
    tenant = _get_default_tenant(request)
    cc = get_object_or_404(CycleCount, id=cc_id, tenant=tenant)
    if request.method != "POST":
        return redirect("cycle_count_detail", cc_id=cc.id)

    if cc.status != CycleCount.Status.DRAFT:
        return redirect("cycle_count_detail", cc_id=cc.id)

    # Snapshot system qty + calculate variance
    for line in cc.lines.select_related("product").all():
        if line.lot_code or line.serial_number or line.expiry_date:
            try:
                lb = InventoryLotBalance.objects.get(
                    tenant=tenant, product=line.product, location=cc.location,
                    lot_code=line.lot_code, serial_number=line.serial_number, expiry_date=line.expiry_date
                )
                system_qty = lb.on_hand
            except InventoryLotBalance.DoesNotExist:
                system_qty = Decimal("0.00")
        else:
            try:
                bal = InventoryBalance.objects.get(tenant=tenant, product=line.product, location=cc.location)
                system_qty = bal.on_hand
            except InventoryBalance.DoesNotExist:
                system_qty = Decimal("0.00")

        line.system_qty = system_qty
        line.variance_qty = Decimal(line.counted_qty) - system_qty
        line.save()

    cc.status = CycleCount.Status.SUBMITTED
    cc.save()
    return redirect("cycle_count_detail", cc_id=cc.id)

@role_required(read_groups=[ROLE_ADMIN, ROLE_FINANCE], write_groups=[ROLE_ADMIN, ROLE_FINANCE])
@transaction.atomic
def cycle_count_approve(request, cc_id):
    tenant = _get_default_tenant(request)
    cc = get_object_or_404(CycleCount, id=cc_id, tenant=tenant)
    if request.method == "POST" and cc.status == CycleCount.Status.SUBMITTED:
        cc.status = CycleCount.Status.APPROVED
        cc.save()
    return redirect("cycle_count_detail", cc_id=cc.id)

@role_required(read_groups=[ROLE_ADMIN, ROLE_FINANCE], write_groups=[ROLE_ADMIN, ROLE_FINANCE])
@transaction.atomic
def cycle_count_post(request, cc_id):
    tenant = _get_default_tenant(request)
    cc = get_object_or_404(CycleCount, id=cc_id, tenant=tenant)
    if request.method != "POST":
        return render(request, "inventory/cycle_count_post.html", {"tenant": tenant, "cc": cc})

    if cc.status != CycleCount.Status.APPROVED:
        return redirect("cycle_count_detail", cc_id=cc.id)

    # Staleness guard: the variance was snapshotted at submit/approval. If stock
    # has moved since, applying the frozen variance would post a wrong
    # correction (M5). Re-read the live system qty per line; if anything changed,
    # refresh the variances and bounce the count back for re-approval instead of
    # posting. Posting therefore only ever applies a variance that matches the
    # current book, making the resulting on-hand equal the counted quantity.
    def _live_system_qty(line):
        if line.lot_code or line.serial_number or line.expiry_date:
            lb = InventoryLotBalance.objects.filter(
                tenant=tenant, product=line.product, location=cc.location,
                lot_code=line.lot_code, serial_number=line.serial_number,
                expiry_date=line.expiry_date).first()
            return lb.on_hand if lb else Decimal("0.00")
        bal = InventoryBalance.objects.filter(
            tenant=tenant, product=line.product, location=cc.location).first()
        return bal.on_hand if bal else Decimal("0.00")

    lines = list(cc.lines.select_related("product").all())
    if any(_live_system_qty(line) != line.system_qty for line in lines):
        for line in lines:
            sys_qty = _live_system_qty(line)
            line.system_qty = sys_qty
            line.variance_qty = Decimal(line.counted_qty) - sys_qty
            line.save(update_fields=["system_qty", "variance_qty"])
        cc.status = CycleCount.Status.SUBMITTED
        cc.save(update_fields=["status"])
        messages.warning(request, "Stock changed since this count was approved - the variances "
                                  "were refreshed. Please review and re-approve before posting.")
        return redirect("cycle_count_detail", cc_id=cc.id)

    # Post variances as ADJUSTMENT movements, then book the net GL impact valued
    # identically to those movements (cycle-count variance valuation fix).
    from core.services.inventory import lot_layer_unit_cost
    from core.services.gl import post_cycle_count_adjustment
    net_value = Decimal("0.00")
    for line in lines:
        var = Decimal(line.variance_qty)
        if var == Decimal("0.00"):
            continue
        unit_cost = None
        # A positive (found) variance on a lot/serial item is valued at that
        # lot's existing layer cost, not the product average. A negative variance
        # already consumes the lot's own layers (lot-scoped FIFO), so its movement
        # value is the lot cost without extra handling.
        if var > 0 and (line.lot_code or line.serial_number or line.expiry_date):
            unit_cost = lot_layer_unit_cost(
                tenant, line.product, cc.location,
                lot_code=line.lot_code, serial_number=line.serial_number, expiry_date=line.expiry_date)
        movement = apply_movement(
            tenant=tenant,
            product=line.product,
            location=cc.location,
            movement_type="ADJUSTMENT",
            qty_delta=var,
            ref_type="CYCLE_COUNT",
            ref_id=str(cc.id),
            notes="Cycle count variance posted", user=request.user,
            lot_code=line.lot_code,
            serial_number=line.serial_number,
            expiry_date=line.expiry_date,
            unit_cost=unit_cost,
        )
        net_value += movement.value or Decimal("0.00")

    post_cycle_count_adjustment(tenant, cc, net_value, user=request.user)

    cc.status = CycleCount.Status.POSTED
    cc.save()
    return redirect("cycle_count_detail", cc_id=cc.id)


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE])
def transfer_list(request):
    from core.access import accessible_location_ids
    from django.db.models import Q
    tenant = _get_default_tenant(request)
    transfers = InventoryTransfer.objects.filter(tenant=tenant).order_by("-created_at")
    allowed = active_location_ids(request)  # scope to the selected site
    if allowed is not None:
        transfers = transfers.filter(Q(from_location_id__in=allowed) | Q(to_location_id__in=allowed))
    return render(request, "transfers/transfer_list.html", {"tenant": tenant, "transfers": transfers})


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE])
@transaction.atomic
def transfer_create(request):
    tenant = _get_default_tenant(request)
    transfer = InventoryTransfer(tenant=tenant, transfer_number="TR-" + timezone.now().strftime("%Y%m%d-%H%M%S-%f"))
    if request.method == "POST":
        form = InventoryTransferForm(request.POST, instance=transfer)
        formset = InventoryTransferLineFormSet(request.POST, instance=transfer)
        if form.is_valid() and formset.is_valid():
            transfer = form.save(commit=False)
            transfer.tenant = tenant
            transfer.save()
            formset.save()

            action = form.cleaned_data.get("action") or "save"
            if action == "post":
                _post_transfer(transfer, request)
            return redirect("transfer_detail", transfer_id=transfer.id)
    else:
        form = InventoryTransferForm(instance=transfer)
        formset = InventoryTransferLineFormSet(instance=transfer)

    # From-location is within the working site; to-location may be any accessible
    # location (cross-site transfers are allowed when permitted).
    _scope_location_fields_by_site(form, request, "from_location")
    _scope_location_fields(form, request, tenant, "to_location")
    return render(request, "transfers/transfer_form.html", {
        "tenant": tenant, "form": form, "formset": formset, "mode": "create"
    })


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE])
def transfer_detail(request, transfer_id):
    tenant = _get_default_tenant(request)
    transfer = get_object_or_404(InventoryTransfer, id=transfer_id, tenant=tenant)
    return render(request, "transfers/transfer_detail.html", {"tenant": tenant, "transfer": transfer})


@login_required
@role_required([ROLE_ADMIN, ROLE_WAREHOUSE])
@transaction.atomic
def transfer_post(request, transfer_id):
    tenant = _get_default_tenant(request)
    transfer = get_object_or_404(InventoryTransfer, id=transfer_id, tenant=tenant)
    if not _can_access_location(request, tenant, transfer.from_location_id, transfer.to_location_id):
        raise PermissionDenied("You do not have access to one of this transfer's locations.")
    if request.method == "POST":
        _post_transfer(transfer, request)
        return redirect("transfer_detail", transfer_id=transfer.id)
    return render(request, "transfers/transfer_post.html", {"tenant": tenant, "transfer": transfer})


def _post_transfer(transfer: InventoryTransfer, request=None):
    if transfer.status == InventoryTransfer.Status.POSTED:
        return
    user = request.user if request is not None else None
    # Post OUT then IN (lot-aware)
    for line in transfer.lines.select_related("product").all():
        qty = Decimal(line.qty)
        if qty <= 0:
            continue
        out_move = apply_movement(
            tenant=transfer.tenant,
            product=line.product,
            location=transfer.from_location,
            movement_type="TRANSFER_OUT",
            qty_delta=(qty * Decimal("-1")),
            ref_type="TRANSFER",
            ref_id=transfer.transfer_number,
            notes="Transfer out", user=user,
            lot_code=line.lot_code,
            serial_number=line.serial_number,
            expiry_date=line.expiry_date,
        )
        # Carry the exact cost relieved on the OUT side into the IN side so the
        # transfer is value-neutral. Without this, TRANSFER_IN re-layers at the
        # product's moving average and every internal transfer silently restates
        # total inventory value (C6).
        apply_movement(
            tenant=transfer.tenant,
            product=line.product,
            location=transfer.to_location,
            movement_type="TRANSFER_IN",
            qty_delta=qty,
            ref_type="TRANSFER",
            ref_id=transfer.transfer_number,
            notes="Transfer in", user=user,
            unit_cost=out_move.unit_cost,
            lot_code=line.lot_code,
            serial_number=line.serial_number,
            expiry_date=line.expiry_date,
        )
    transfer.status = InventoryTransfer.Status.POSTED
    transfer.posted_at = timezone.now()
    transfer.save()
    if request is not None:
        messages.success(request, f"Transfer {transfer.transfer_number} posted.")


@login_required
@role_required([ROLE_ADMIN, ROLE_FINANCE])
def invoice_list(request):
    tenant = _get_default_tenant(request)
    invoices = SupplierInvoice.objects.filter(tenant=tenant).order_by("-created_at")
    return render(request, "finance/invoice_list.html", {"tenant": tenant, "invoices": invoices})


@login_required
@role_required([ROLE_ADMIN, ROLE_FINANCE])
@transaction.atomic
def invoice_create(request):
    tenant = _get_default_tenant(request)
    inv = SupplierInvoice(tenant=tenant, currency_code=tenant.currency_code)
    if request.method == "POST":
        form = SupplierInvoiceForm(request.POST, request.FILES, instance=inv)
        formset = SupplierInvoiceLineFormSet(request.POST, instance=inv)
        if form.is_valid() and formset.is_valid():
            inv = form.save(commit=False)
            inv.tenant = tenant
            inv.save()
            formset.save()

            action = form.cleaned_data.get("action") or "save"
            if action == "submit":
                _run_3way_match(inv, request)
            elif action == "approve":
                inv.status = SupplierInvoice.Status.APPROVED
                inv.approved_by = request.user
                inv.approved_at = timezone.now()
                inv.save()
                messages.success(request, "Invoice approved.")
            elif action == "post":
                _run_3way_match(inv, request)
                inv.status = SupplierInvoice.Status.POSTED
                inv.save()
                messages.success(request, "Invoice posted.")
            return redirect("invoice_detail", invoice_id=inv.id)
    else:
        form = SupplierInvoiceForm(instance=inv)
        line_initial = [{"tax_code": tenant.default_tax_code}] if tenant.default_tax_code_id else None
        formset = SupplierInvoiceLineFormSet(instance=inv, initial=line_initial)

    return render(request, "finance/invoice_form.html", {"tenant": tenant, "form": form, "formset": formset})


@login_required
@role_required([ROLE_ADMIN, ROLE_FINANCE])
def invoice_detail(request, invoice_id):
    tenant = _get_default_tenant(request)
    inv = get_object_or_404(SupplierInvoice, id=invoice_id, tenant=tenant)
    match = _compute_3way(inv)
    return render(request, "finance/invoice_detail.html", {"tenant": tenant, "invoice": inv, "match": match})


def _compute_3way(inv: SupplierInvoice):
    # Return discrepancies list for UI
    discrepancies = []
    for line in inv.lines.select_related("product","po_line","receipt_line").all():
        po_qty = line.po_line.ordered_qty if line.po_line else None
        rec_qty = line.receipt_line.qty_received if line.receipt_line else None
        if rec_qty is not None and line.qty > rec_qty:
            discrepancies.append(f"{line.product.sku}: invoiced qty {line.qty} > received {rec_qty}")
        if po_qty is not None and line.qty > po_qty:
            discrepancies.append(f"{line.product.sku}: invoiced qty {line.qty} > ordered {po_qty}")
    return {"ok": len(discrepancies)==0, "discrepancies": discrepancies}


def _run_3way_match(inv: SupplierInvoice, request=None):
    m = _compute_3way(inv)
    if m["ok"]:
        inv.status = SupplierInvoice.Status.MATCHED
        inv.save()
        if request is not None:
            messages.success(request, "3-way match passed. Invoice marked MATCHED.")
    else:
        inv.status = SupplierInvoice.Status.DRAFT
        inv.save()
        if request is not None:
            messages.warning(request, "3-way match issues: " + "; ".join(m["discrepancies"][:5]))


@login_required
@role_required([ROLE_ADMIN, ROLE_SALES, ROLE_WAREHOUSE, ROLE_READONLY])
def return_list(request):
    tenant = _get_default_tenant(request)
    rmas = ReturnAuthorization.objects.filter(tenant=tenant).order_by("-created_at")
    return render(request, "returns/return_list.html", {"tenant": tenant, "rmas": rmas})

@login_required
@role_required([ROLE_ADMIN, ROLE_SALES, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_SALES, ROLE_WAREHOUSE])
@transaction.atomic
def return_create(request):
    tenant = _get_default_tenant(request)
    rma = ReturnAuthorization(tenant=tenant)

    if request.method == "POST":
        form = ReturnAuthorizationForm(request.POST, instance=rma)
        formset = ReturnLineFormSet(request.POST, instance=rma)
        if form.is_valid() and formset.is_valid():
            rma = form.save(commit=False)
            rma.tenant = tenant
            rma.save()
            formset.save()

            action = form.cleaned_data.get("action") or "save"
            if action == "approve" and rma.status == ReturnAuthorization.Status.DRAFT:
                rma.status = ReturnAuthorization.Status.APPROVED
                rma.save()
                messages.success(request, "RMA approved.")
            elif action == "receive":
                _receive_rma(rma)
                messages.success(request, "Return received and restocked.")
            return redirect("return_detail", rma_id=rma.id)
    else:
        form = ReturnAuthorizationForm(instance=rma)
        formset = ReturnLineFormSet(instance=rma)

    return render(request, "returns/return_form.html", {
        "tenant": tenant, "form": form, "formset": formset, "mode": "create"
    })

@login_required
@role_required([ROLE_ADMIN, ROLE_SALES, ROLE_WAREHOUSE, ROLE_READONLY])
def return_detail(request, rma_id):
    tenant = _get_default_tenant(request)
    rma = get_object_or_404(ReturnAuthorization, id=rma_id, tenant=tenant)
    return render(request, "returns/return_detail.html", {"tenant": tenant, "rma": rma})

@login_required
@role_required([ROLE_ADMIN, ROLE_SALES, ROLE_WAREHOUSE], [ROLE_ADMIN, ROLE_SALES, ROLE_WAREHOUSE])
@transaction.atomic
def return_process(request, rma_id):
    tenant = _get_default_tenant(request)
    rma = get_object_or_404(ReturnAuthorization, id=rma_id, tenant=tenant)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "approve" and rma.status == ReturnAuthorization.Status.DRAFT:
            rma.status = ReturnAuthorization.Status.APPROVED
            rma.save()
            messages.success(request, "RMA approved.")
        elif action == "receive":
            _receive_rma(rma)
            messages.success(request, "Return received and restocked.")
        return redirect("return_detail", rma_id=rma.id)
    return redirect("return_detail", rma_id=rma.id)

def _receive_rma(rma: ReturnAuthorization):
    if rma.status in (ReturnAuthorization.Status.RECEIVED, ReturnAuthorization.Status.CLOSED):
        return
    # allow receiving from DRAFT/APPROVED
    ref_id = f"{rma.channel}:{rma.rma_number}"
    for line in rma.lines.select_related("product").all():
        qty = Decimal(line.qty)
        if qty <= 0:
            continue
        # If returned item is a kit, we restock components (consistent with 'deduct components only')
        for comp, comp_qty in explode_product(line.product, qty):
            apply_movement(
                tenant=rma.tenant,
                product=comp,
                location=rma.receive_location,
                movement_type=InventoryMovement.MovementType.RETURN,
                qty_delta=comp_qty,
                ref_type="RMA",
                ref_id=ref_id,
                notes="Return received", user=request.user,
                lot_code=line.lot_code,
                serial_number=line.serial_number,
                expiry_date=line.expiry_date,
            )
    rma.status = ReturnAuthorization.Status.RECEIVED
    rma.save()


# ============================
# VAT / Tax Codes
# ============================

@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def taxcode_list(request):
    tenant = _get_default_tenant(request)
    codes = TaxCode.objects.filter(tenant=tenant).order_by("code")
    return render(request, "tax/taxcode_list.html", {"tenant": tenant, "codes": codes})

@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def taxcode_create(request):
    tenant = _get_default_tenant(request)
    form = TaxCodeForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        obj.save()
        log_audit(action="VAT_RATE_CHANGED", request=request, user=request.user, tenant=tenant,
                  detail=f"Created tax code {obj.code} ({obj.get_kind_display()} @ {obj.rate})")
        return redirect("taxcode_list")
    return render(request, "tax/taxcode_form.html", {"tenant": tenant, "form": form, "mode": "create"})

@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def taxcode_edit(request, tax_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(TaxCode, id=tax_id, tenant=tenant)
    form = TaxCodeForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        form.save()
        log_audit(action="VAT_RATE_CHANGED", request=request, user=request.user, tenant=tenant,
                  detail=f"Updated tax code {obj.code} ({obj.get_kind_display()} @ {obj.rate})")
        return redirect("taxcode_list")
    return render(request, "tax/taxcode_form.html", {"tenant": tenant, "form": form, "mode": "edit"})

@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def taxcode_delete(request, tax_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(TaxCode, id=tax_id, tenant=tenant)
    if request.method == "POST":
        log_audit(action="RECORD_DELETED", request=request, user=request.user, tenant=tenant,
                  detail=f"Tax code {obj}")
        obj.delete()
        return redirect("taxcode_list")
    return render(request, "tax/taxcode_delete.html", {"tenant": tenant, "tax": obj})


# ============================
# Customers
# ============================

@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
def customer_list(request):
    tenant = _get_default_tenant(request)
    q = (request.GET.get("q") or "").strip()
    ctype = request.GET.get("type") or ""
    status = request.GET.get("status") or ""
    tag = (request.GET.get("tag") or "").strip()

    customers = Customer.objects.filter(tenant=tenant)
    if q:
        customers = customers.filter(
            Q(name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q)
            | Q(vat_number__icontains=q) | Q(company_number__icontains=q)
            | Q(contact_person__icontains=q))
    if ctype:
        customers = customers.filter(customer_type=ctype)
    if status:
        customers = customers.filter(status=status)
    if tag:
        customers = customers.filter(tags__icontains=tag)
    customers = customers.order_by("name")

    return render(request, "customers/customer_list.html", {
        "tenant": tenant, "customers": customers,
        "q": q, "type": ctype, "status": status, "tag": tag,
        "type_choices": Customer.Type.choices, "status_choices": Customer.Status.choices,
        "filtered": bool(q or ctype or status or tag),
    })


def _find_customer_duplicates(tenant, obj, exclude_id=None):
    """Possible duplicate customers matching email / phone / VAT / company number
    / name. Returns [{id, name, match}] (a different existing record per match)."""
    checks = [
        ("email", obj.email, "email"),
        ("phone", obj.phone, "phone"),
        ("vat_number", obj.vat_number, "VAT number"),
        ("company_number", obj.company_number, "company number"),
        ("name", obj.name, "name"),
    ]
    seen, out = set(), []
    for field, value, label in checks:
        value = (value or "").strip()
        if not value:
            continue
        qs = Customer.objects.filter(tenant=tenant, **{f"{field}__iexact": value})
        if exclude_id:
            qs = qs.exclude(id=exclude_id)
        for match in qs:
            if match.id not in seen:
                seen.add(match.id)
                out.append({"id": match.id, "name": match.name, "match": label})
    return out


@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
def customer_create(request):
    tenant = _get_default_tenant(request)
    form = CustomerForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        dups = _find_customer_duplicates(tenant, obj)
        if dups and request.POST.get("confirm_duplicate") != "1":
            return render(request, "customers/customer_form.html",
                          {"tenant": tenant, "form": form, "mode": "create", "duplicates": dups})
        try:
            obj.save()
        except IntegrityError:
            form.add_error("name", "A customer with this name already exists.")
            return render(request, "customers/customer_form.html", {"tenant": tenant, "form": form, "mode": "create"})
        return redirect("customer_detail", customer_id=obj.id)
    return render(request, "customers/customer_form.html", {"tenant": tenant, "form": form, "mode": "create"})

@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
def customer_edit(request, customer_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(Customer, id=customer_id, tenant=tenant)
    form = CustomerForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        edited = form.save(commit=False)
        dups = _find_customer_duplicates(tenant, edited, exclude_id=obj.id)
        if dups and request.POST.get("confirm_duplicate") != "1":
            return render(request, "customers/customer_form.html",
                          {"tenant": tenant, "form": form, "mode": "edit", "duplicates": dups})
        form.save()
        return redirect("customer_detail", customer_id=obj.id)
    return render(request, "customers/customer_form.html", {"tenant": tenant, "form": form, "mode": "edit"})


@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
def customer_detail(request, customer_id):
    """Customer profile: details, recent invoices/payments/credit notes/orders/
    quotes, outstanding balance, notes and an activity timeline."""
    tenant = _get_default_tenant(request)
    c = get_object_or_404(Customer, id=customer_id, tenant=tenant)

    invoices = list(CustomerInvoice.objects.filter(tenant=tenant, customer=c)
                    .prefetch_related("lines", "lines__tax_code", "payment_allocations", "credit_notes")
                    .order_by("-invoice_date", "-id"))
    payments = list(Payment.objects.filter(tenant=tenant, customer=c)
                    .order_by("-payment_date", "-id"))
    credit_notes = list(CreditNote.objects.filter(tenant=tenant, customer=c, kind=CreditNote.Kind.SALES)
                        .prefetch_related("lines", "lines__tax_code").order_by("-credit_note_date", "-id"))
    orders = list(CustomerOrder.objects.filter(tenant=tenant, customer=c).order_by("-order_date", "-id"))
    quotes = list(SalesQuote.objects.filter(tenant=tenant, customer=c).order_by("-quote_date", "-id"))

    # Activity timeline: merge dated events newest-first.
    timeline = []
    for q in quotes:
        timeline.append({"date": q.quote_date, "icon": "file-earmark-text", "kind": "Quote",
                         "text": f"Quote {q.quote_number} ({q.get_status_display()})",
                         "url": f"/quotes/{q.id}/"})
    for o in orders:
        timeline.append({"date": o.order_date, "icon": "cart-check", "kind": "Sales order",
                         "text": f"Sales order {o.order_number} ({o.get_status_display()})",
                         "url": f"/customer-orders/{o.id}/"})
    for inv in invoices:
        timeline.append({"date": inv.invoice_date, "icon": "receipt", "kind": "Invoice",
                         "text": f"Invoice {inv.invoice_number} ({inv.display_status}) - {inv.total}",
                         "url": f"/ar/invoices/{inv.id}/"})
    for p in payments:
        label = "Refund" if p.direction == Payment.Direction.REFUND else "Receipt"
        icon = "arrow-counterclockwise" if p.direction == Payment.Direction.REFUND else "cash-coin"
        timeline.append({"date": p.payment_date, "icon": icon, "kind": label,
                         "text": f"{label} {p.amount} ({p.get_method_display()})",
                         "url": f"/payments/{p.id}/"})
    for cn in credit_notes:
        timeline.append({"date": cn.credit_note_date, "icon": "arrow-return-left", "kind": "Credit note",
                         "text": f"Credit note {cn.credit_note_number} - {cn.total}",
                         "url": f"/credit-notes/{cn.id}/"})
    timeline.sort(key=lambda e: e["date"], reverse=True)

    sales_total = sum((inv.total for inv in invoices if inv.status in CustomerInvoice.ISSUED_STATES), Decimal("0.00"))

    return render(request, "customers/customer_detail.html", {
        "tenant": tenant, "c": c,
        "invoices": invoices[:10], "payments": payments[:10], "credit_notes": credit_notes[:10],
        "orders": orders[:10], "quotes": quotes[:10], "timeline": timeline[:40],
        "sales_total": sales_total, "invoice_count": len(invoices),
    })


# ============================
# Accounts Receivable Invoices
# ============================

@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
def ar_invoice_list(request):
    tenant = _get_default_tenant(request)
    invoices = (
        CustomerInvoice.objects.filter(tenant=tenant)
        .prefetch_related("lines", "lines__tax_code")
        .order_by("-invoice_date", "-id")
    )
    site = active_site_id(request)  # scope to the selected site
    if site:
        invoices = invoices.filter(site_id=site)
    return render(request, "ar/ar_invoice_list.html", {"tenant": tenant, "invoices": invoices})

@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def ar_invoice_create(request):
    tenant = _get_default_tenant(request)
    inv = CustomerInvoice(tenant=tenant, currency_code=tenant.currency_code)
    if request.method == "POST":
        form = CustomerInvoiceForm(request.POST, instance=inv)
        formset = CustomerInvoiceLineFormSet(request.POST, instance=inv)
        if form.is_valid() and formset.is_valid():
            inv = form.save(commit=False)
            inv.tenant = tenant
            inv.currency_code = tenant.currency_code
            inv.site = get_active_site(request)  # sale belongs to the working site
            if not inv.location_id:
                inv.location = _active_site_default_location(request, tenant)  # fulfil from the site's location
            # Auto-generate the invoice number when the admin left it blank.
            if not (inv.invoice_number or "").strip():
                from core.numbering import next_invoice_number
                inv.invoice_number = next_invoice_number(tenant)
            # Default due date from the company's payment terms when left blank.
            if not inv.due_date and inv.invoice_date and tenant.default_payment_terms_days:
                inv.due_date = inv.invoice_date + timezone.timedelta(days=tenant.default_payment_terms_days)
            inv.save()
            formset.save()

            action = form.cleaned_data.get("action") or "save"
            if action == "issue":
                ok, reason = inv.customer.credit_status(inv.total)
                if not ok:
                    messages.error(request, f"Saved as draft — not issued. {reason}")
                else:
                    post_customer_invoice(inv, user=request.user)

            return redirect("ar_invoice_detail", invoice_id=inv.id)
    else:
        from core.numbering import next_invoice_number
        initial = {"invoice_number": next_invoice_number(tenant)}
        if tenant.invoice_footer:
            initial["terms"] = tenant.invoice_footer
        form = CustomerInvoiceForm(instance=inv, initial=initial)
        # Pre-fill the first line's tax code with the company default.
        line_initial = [{"tax_code": tenant.default_tax_code}] if tenant.default_tax_code_id else None
        formset = CustomerInvoiceLineFormSet(instance=inv, initial=line_initial)

    return render(request, "ar/ar_invoice_form.html", {"tenant": tenant, "form": form, "formset": formset})

@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
def ar_invoice_detail(request, invoice_id):
    tenant = _get_default_tenant(request)
    inv = get_object_or_404(CustomerInvoice, id=invoice_id, tenant=tenant)
    can_delete = (inv.status == CustomerInvoice.Status.DRAFT and
                  (bool({ROLE_FINANCE, ROLE_ADMIN} & effective_groups(request)) or request.user.is_superuser))
    return render(request, "ar/ar_invoice_detail.html", {"tenant": tenant, "inv": inv, "perms_can_delete": can_delete})


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def ar_invoice_delete(request, invoice_id):
    """Soft-delete a DRAFT customer invoice (sensitive record). Posted invoices
    must be cancelled instead (they have a ledger entry), so deletion is blocked
    for them. The row is flagged, not removed, and the action is audited."""
    tenant = _get_default_tenant(request)
    inv = get_object_or_404(CustomerInvoice, id=invoice_id, tenant=tenant)
    if request.method != "POST":
        return redirect("ar_invoice_detail", invoice_id=inv.id)
    if inv.status != CustomerInvoice.Status.DRAFT:
        messages.error(request, "Only draft invoices can be deleted. Cancel a posted invoice instead.")
        return redirect("ar_invoice_detail", invoice_id=inv.id)
    inv.is_deleted = True
    inv.deleted_at = timezone.now()
    inv.deleted_by = request.user
    inv.save(update_fields=["is_deleted", "deleted_at", "deleted_by"])
    log_audit(action="INVOICE_DELETED", request=request, user=request.user, tenant=tenant,
              entity_type="CustomerInvoice", entity_id=inv.invoice_number,
              detail=f"Draft invoice {inv.invoice_number} ({inv.customer.name})")
    messages.success(request, f"Draft invoice {inv.invoice_number} deleted.")
    return redirect("ar_invoice_list")


@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
def ar_invoice_pdf(request, invoice_id):
    tenant = _get_default_tenant(request)
    inv = get_object_or_404(CustomerInvoice, id=invoice_id, tenant=tenant)
    from core.services.pdf import pdf_response
    return pdf_response(f"invoice-{inv.invoice_number}.pdf", "documents/invoice_pdf.html",
                        {"tenant": tenant, "inv": inv, "doc_title": "INVOICE", "number": inv.invoice_number,
                         "notes": inv.notes, "terms": inv.terms}, download=False)

@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def ar_invoice_issue(request, invoice_id):
    tenant = _get_default_tenant(request)
    inv = get_object_or_404(CustomerInvoice, id=invoice_id, tenant=tenant)
    if request.method == "POST":
        ok, reason = inv.customer.credit_status(inv.total)
        if not ok:
            messages.error(request, f"Cannot issue this invoice. {reason}")
            return redirect("ar_invoice_detail", invoice_id=inv.id)
        post_customer_invoice(inv, user=request.user)
        return redirect("ar_invoice_detail", invoice_id=inv.id)
    return render(request, "ar/ar_invoice_issue.html", {"tenant": tenant, "inv": inv})


@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def ar_invoice_send(request, invoice_id):
    """Email the invoice to the customer and mark it Sent. Issues it first if
    still a draft so a sent invoice is always on the ledger."""
    tenant = _get_default_tenant(request)
    inv = get_object_or_404(CustomerInvoice, id=invoice_id, tenant=tenant)
    if request.method == "POST":
        if inv.status == CustomerInvoice.Status.DRAFT:
            ok, reason = inv.customer.credit_status(inv.total)
            if not ok:
                messages.error(request, f"Cannot issue/send this invoice. {reason}")
                return redirect("ar_invoice_detail", invoice_id=inv.id)
            post_customer_invoice(inv, user=request.user)
        from core import notify
        from core.services.pdf import render_to_pdf
        pdf = render_to_pdf("documents/invoice_pdf.html", {
            "tenant": tenant, "inv": inv, "doc_title": "INVOICE", "number": inv.invoice_number,
            "notes": inv.notes, "terms": inv.terms})
        attachment = (f"invoice-{inv.invoice_number}.pdf", pdf, "application/pdf") if pdf else None
        sent = notify.notify_invoice(inv, request=request, attachment=attachment)
        inv.status = CustomerInvoice.Status.SENT
        inv.sent_at = timezone.now()
        inv.save(update_fields=["status", "sent_at"])
        log_audit(action="INVOICE_SENT", request=request, user=request.user, tenant=tenant,
                  detail=f"{inv.invoice_number} to {inv.customer.email or '(no email)'}")
        if sent:
            messages.success(request, f"Invoice {inv.invoice_number} emailed to {inv.customer.email}.")
        else:
            messages.warning(request, f"Invoice {inv.invoice_number} marked sent, but the customer has no email on file.")
    return redirect("ar_invoice_detail", invoice_id=inv.id)


@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def ar_invoice_cancel(request, invoice_id):
    """Cancel an unpaid invoice (reverses its GL entry if it was posted)."""
    tenant = _get_default_tenant(request)
    inv = get_object_or_404(CustomerInvoice, id=invoice_id, tenant=tenant)
    if request.method == "POST":
        if inv.amount_paid > Decimal("0.00"):
            messages.error(request, "This invoice has payments against it - raise a credit note or refund instead.")
        else:
            from core.models import JournalLine
            # Reverse the original journal entry if the invoice was posted.
            je = JournalEntry.objects.filter(tenant=tenant, ref_type="AR_INVOICE", ref_id=inv.invoice_number).order_by("-id").first()
            if je and inv.status in CustomerInvoice.ISSUED_STATES:
                rev = JournalEntry.objects.create(
                    tenant=tenant, entry_date=timezone.localdate(), ref_type="AR_INVOICE_CANCEL",
                    ref_id=inv.invoice_number, memo=f"Cancel invoice {inv.invoice_number}",
                    posted_by=request.user, posted_at=timezone.now())
                for l in je.lines.all():
                    JournalLine.objects.create(entry=rev, account=l.account, description="Cancellation",
                                               debit=l.credit, credit=l.debit)
            # Restore stock + reverse COGS so the ledger matches physical stock.
            if inv.status in CustomerInvoice.ISSUED_STATES:
                from core.services.gl import reverse_invoice_cogs
                reverse_invoice_cogs(inv, user=request.user)
            inv.status = CustomerInvoice.Status.CANCELLED
            inv.save(update_fields=["status"])
            log_audit(action="INVOICE_CANCELLED", request=request, user=request.user, tenant=tenant, detail=inv.invoice_number)
            messages.success(request, f"Invoice {inv.invoice_number} cancelled.")
    return redirect("ar_invoice_detail", invoice_id=inv.id)


# ============================
# Customer sales documents: Quotes -> Sales Orders -> Invoices
# ============================

SALES_DOC_ROLES = [ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN]
SALES_DOC_READ = [ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY]


def _copy_sales_lines(src, dest, line_model, fk_name):
    """Copy product/qty/price/discount/tax lines from one sales doc to another."""
    for l in src.lines.all():
        line_model.objects.create(**{
            fk_name: dest, "product": l.product, "description": l.description,
            "qty": l.qty, "unit_price": l.unit_price, "discount_pct": l.discount_pct,
            "tax_code": l.tax_code,
        })


# ---- Quotes ----

@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def quote_list(request):
    tenant = _get_default_tenant(request)
    quotes = SalesQuote.objects.filter(tenant=tenant).select_related("customer").prefetch_related("lines", "lines__tax_code").order_by("-quote_date", "-id")
    return render(request, "sales/quote_list.html", {"tenant": tenant, "quotes": quotes})


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def quote_create(request):
    tenant = _get_default_tenant(request)
    q = SalesQuote(tenant=tenant, currency_code=tenant.currency_code)
    if request.method == "POST":
        form = SalesQuoteForm(request.POST, instance=q)
        formset = SalesQuoteLineFormSet(request.POST, instance=q)
        if form.is_valid() and formset.is_valid():
            q = form.save(commit=False)
            q.tenant = tenant
            q.currency_code = tenant.currency_code
            if not (q.quote_number or "").strip():
                from core.numbering import next_quote_number
                q.quote_number = next_quote_number(tenant)
            q.save()
            formset.save()
            messages.success(request, f"Quote {q.quote_number} saved.")
            return redirect("quote_detail", quote_id=q.id)
    else:
        from core.numbering import next_quote_number
        initial = {"quote_number": next_quote_number(tenant)}
        if tenant.invoice_footer:
            initial["terms"] = tenant.invoice_footer
        form = SalesQuoteForm(instance=q, initial=initial)
        line_initial = [{"tax_code": tenant.default_tax_code}] if tenant.default_tax_code_id else None
        formset = SalesQuoteLineFormSet(instance=q, initial=line_initial)
    return render(request, "sales/doc_form.html", {
        "tenant": tenant, "form": form, "formset": formset,
        "doc_label": "Quote", "list_url": "/quotes/", "extra_fields": ["valid_until"]})


@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def quote_detail(request, quote_id):
    tenant = _get_default_tenant(request)
    q = get_object_or_404(SalesQuote, id=quote_id, tenant=tenant)
    return render(request, "sales/quote_detail.html", {"tenant": tenant, "q": q})


@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def quote_pdf(request, quote_id):
    tenant = _get_default_tenant(request)
    q = get_object_or_404(SalesQuote, id=quote_id, tenant=tenant)
    from core.services.pdf import pdf_response
    return pdf_response(f"quote-{q.quote_number}.pdf", "documents/quote_pdf.html",
                        {"tenant": tenant, "q": q, "doc_title": "QUOTE", "number": q.quote_number,
                         "notes": q.notes, "terms": q.terms}, download=False)


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def quote_send(request, quote_id):
    tenant = _get_default_tenant(request)
    q = get_object_or_404(SalesQuote, id=quote_id, tenant=tenant)
    if request.method == "POST":
        from core import notify
        from core.services.pdf import render_to_pdf
        pdf = render_to_pdf("documents/quote_pdf.html", {
            "tenant": tenant, "q": q, "doc_title": "QUOTE", "number": q.quote_number,
            "notes": q.notes, "terms": q.terms})
        attachment = (f"quote-{q.quote_number}.pdf", pdf, "application/pdf") if pdf else None
        sent = notify.notify_sales_document(q, "Quote", q.quote_number, request=request, attachment=attachment)
        if q.status == SalesQuote.Status.DRAFT:
            q.status = SalesQuote.Status.SENT
        q.sent_at = timezone.now()
        q.save(update_fields=["status", "sent_at"])
        log_audit(action="QUOTE_SENT", request=request, user=request.user, tenant=tenant, detail=q.quote_number)
        messages.success(request, f"Quote {q.quote_number} {'emailed' if sent else 'marked sent (no customer email)'}.")
    return redirect("quote_detail", quote_id=q.id)


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def quote_status(request, quote_id, to):
    tenant = _get_default_tenant(request)
    q = get_object_or_404(SalesQuote, id=quote_id, tenant=tenant)
    mapping = {"accept": SalesQuote.Status.ACCEPTED, "decline": SalesQuote.Status.DECLINED,
               "cancel": SalesQuote.Status.CANCELLED}
    if request.method == "POST" and to in mapping:
        q.status = mapping[to]
        q.save(update_fields=["status"])
        messages.success(request, f"Quote {q.quote_number} marked {q.get_status_display()}.")
    return redirect("quote_detail", quote_id=q.id)


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def quote_to_order(request, quote_id):
    tenant = _get_default_tenant(request)
    q = get_object_or_404(SalesQuote, id=quote_id, tenant=tenant)
    if request.method == "POST":
        from core.numbering import next_order_number
        order = CustomerOrder.objects.create(
            tenant=tenant, customer=q.customer, order_number=next_order_number(tenant),
            currency_code=q.currency_code, notes=q.notes, terms=q.terms, quote=q,
            status=CustomerOrder.Status.CONFIRMED)
        _copy_sales_lines(q, order, CustomerOrderLine, "order")
        q.status = SalesQuote.Status.CONVERTED
        q.save(update_fields=["status"])
        log_audit(action="QUOTE_CONVERTED", request=request, user=request.user, tenant=tenant,
                  detail=f"{q.quote_number} -> order {order.order_number}")
        messages.success(request, f"Quote {q.quote_number} converted to sales order {order.order_number}.")
        return redirect("corder_detail", order_id=order.id)
    return redirect("quote_detail", quote_id=q.id)


def _invoice_from_lines(tenant, customer, currency, notes, terms, src, line_attr, user):
    """Create a draft CustomerInvoice copying a quote/order's lines."""
    from core.numbering import next_invoice_number
    inv = CustomerInvoice.objects.create(
        tenant=tenant, customer=customer, invoice_number=next_invoice_number(tenant),
        currency_code=currency, notes=notes, terms=terms,
        site=getattr(src, "site", None),  # carry the order's site
        location=getattr(src, "location", None))  # and its fulfilling location
    if tenant.default_payment_terms_days:
        inv.due_date = inv.invoice_date + timezone.timedelta(days=tenant.default_payment_terms_days)
    setattr(inv, line_attr, src)
    inv.save()
    for l in src.lines.all():
        CustomerInvoiceLine.objects.create(
            invoice=inv, product=l.product, description=l.description, qty=l.qty,
            unit_price=l.unit_price, discount_pct=l.discount_pct, tax_code=l.tax_code)
    return inv


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def quote_to_invoice(request, quote_id):
    tenant = _get_default_tenant(request)
    q = get_object_or_404(SalesQuote, id=quote_id, tenant=tenant)
    if request.method == "POST":
        inv = _invoice_from_lines(tenant, q.customer, q.currency_code, q.notes, q.terms, q, "source_quote", request.user)
        q.status = SalesQuote.Status.CONVERTED
        q.save(update_fields=["status"])
        log_audit(action="QUOTE_CONVERTED", request=request, user=request.user, tenant=tenant,
                  detail=f"{q.quote_number} -> invoice {inv.invoice_number}")
        messages.success(request, f"Quote {q.quote_number} converted to invoice {inv.invoice_number} (draft).")
        return redirect("ar_invoice_detail", invoice_id=inv.id)
    return redirect("quote_detail", quote_id=q.id)


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def quote_delete(request, quote_id):
    tenant = _get_default_tenant(request)
    q = get_object_or_404(SalesQuote, id=quote_id, tenant=tenant)
    if request.method == "POST":
        if q.status == SalesQuote.Status.CONVERTED:
            messages.error(request, "Converted quotes can't be deleted - they're linked to an order or invoice.")
            return redirect("quote_detail", quote_id=q.id)
        number = q.quote_number
        q.delete()
        log_audit(action="QUOTE_DELETED", request=request, user=request.user, tenant=tenant, detail=number)
        messages.success(request, f"Quote {number} deleted.")
        return redirect("quote_list")
    return redirect("quote_detail", quote_id=q.id)


# ---- Customer sales orders ----

@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def corder_list(request):
    tenant = _get_default_tenant(request)
    orders = CustomerOrder.objects.filter(tenant=tenant).select_related("customer").prefetch_related("lines", "lines__tax_code").order_by("-order_date", "-id")
    site = active_site_id(request)  # scope to the selected site
    if site:
        orders = orders.filter(site_id=site)
    return render(request, "sales/corder_list.html", {"tenant": tenant, "orders": orders})


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def corder_create(request):
    tenant = _get_default_tenant(request)
    o = CustomerOrder(tenant=tenant, currency_code=tenant.currency_code)
    if request.method == "POST":
        form = CustomerOrderForm(request.POST, instance=o)
        formset = CustomerOrderLineFormSet(request.POST, instance=o)
        if form.is_valid() and formset.is_valid():
            o = form.save(commit=False)
            o.tenant = tenant
            o.currency_code = tenant.currency_code
            o.site = get_active_site(request)  # sale belongs to the working site
            if not o.location_id:
                o.location = _active_site_default_location(request, tenant)  # fulfil from the site's location
            if not (o.order_number or "").strip():
                from core.numbering import next_order_number
                o.order_number = next_order_number(tenant)
            o.save()
            formset.save()
            messages.success(request, f"Sales order {o.order_number} saved.")
            return redirect("corder_detail", order_id=o.id)
    else:
        from core.numbering import next_order_number
        initial = {"order_number": next_order_number(tenant)}
        if tenant.invoice_footer:
            initial["terms"] = tenant.invoice_footer
        form = CustomerOrderForm(instance=o, initial=initial)
        line_initial = [{"tax_code": tenant.default_tax_code}] if tenant.default_tax_code_id else None
        formset = CustomerOrderLineFormSet(instance=o, initial=line_initial)
    return render(request, "sales/doc_form.html", {
        "tenant": tenant, "form": form, "formset": formset,
        "doc_label": "Sales order", "list_url": "/customer-orders/", "extra_fields": []})


@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def corder_detail(request, order_id):
    tenant = _get_default_tenant(request)
    o = get_object_or_404(CustomerOrder, id=order_id, tenant=tenant)
    return render(request, "sales/corder_detail.html", {"tenant": tenant, "o": o})


@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def corder_pdf(request, order_id):
    tenant = _get_default_tenant(request)
    o = get_object_or_404(CustomerOrder, id=order_id, tenant=tenant)
    from core.services.pdf import pdf_response
    return pdf_response(f"sales-order-{o.order_number}.pdf", "documents/order_pdf.html",
                        {"tenant": tenant, "o": o, "doc_title": "SALES ORDER", "number": o.order_number,
                         "notes": o.notes, "terms": o.terms}, download=False)


def _order_stock_location(tenant):
    return (Location.objects.filter(tenant=tenant, type=Location.Type.WAREHOUSE).order_by("id").first()
            or Location.objects.filter(tenant=tenant).order_by("id").first())


def _active_site_default_location(request, tenant):
    """A sensible default fulfilling/receiving location under the selected site:
    first active stock-holding location in that site, else any in it, else the
    company's first location."""
    site = get_active_site(request)
    if site is not None:
        loc = (Location.objects.filter(tenant=tenant, site=site, is_active=True, holds_stock=True).order_by("id").first()
               or Location.objects.filter(tenant=tenant, site=site).order_by("id").first())
        if loc is not None:
            return loc
    return _order_stock_location(tenant)


def _po_destination(po):
    """Receiving location for a PO's goods: the PO's own receiving_location when
    set, otherwise the org's first location (legacy fallback)."""
    return po.receiving_location or Location.objects.filter(tenant=po.tenant).order_by("id").first()


def _reserve_customer_order(o):
    """Reserve stock for a confirmed order's product lines (re-syncs existing
    reservations so an edit reflects the latest lines)."""
    from core.services.inventory import reserve_stock, release_reservations
    tenant = o.tenant
    loc = o.location or _order_stock_location(tenant)  # reserve from the order's own location when set
    release_reservations(tenant=tenant, ref_type="CUSTOMER_ORDER", ref_id=str(o.id))
    if loc is None:
        return
    for line in o.lines.all():
        if line.product_id and line.qty and line.qty > 0:
            reserve_stock(tenant=tenant, product=line.product, location=loc, qty=line.qty,
                          ref_type="CUSTOMER_ORDER", ref_id=str(o.id))


def _release_customer_order(o):
    from core.services.inventory import release_reservations
    release_reservations(tenant=o.tenant, ref_type="CUSTOMER_ORDER", ref_id=str(o.id))


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def corder_status(request, order_id, to):
    tenant = _get_default_tenant(request)
    o = get_object_or_404(CustomerOrder, id=order_id, tenant=tenant)
    mapping = {"confirm": CustomerOrder.Status.CONFIRMED, "cancel": CustomerOrder.Status.CANCELLED}
    if request.method == "POST" and to in mapping:
        if to == "confirm":
            ok, reason = o.customer.credit_status(o.total)
            if o.customer.status == Customer.Status.ON_HOLD:
                # Hard block: an on-hold customer cannot have orders confirmed.
                messages.error(request, f"Cannot confirm order. {reason}")
                return redirect("corder_detail", order_id=o.id)
            if not ok:
                # Over the limit: warn but allow (no receivable until invoiced).
                messages.warning(request, reason)
        o.status = mapping[to]
        o.save(update_fields=["status"])
        # Reserve stock on confirm; release it on cancel.
        if to == "confirm":
            _reserve_customer_order(o)
        else:
            _release_customer_order(o)
        messages.success(request, f"Sales order {o.order_number} marked {o.get_status_display()}.")
    return redirect("corder_detail", order_id=o.id)


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def corder_to_invoice(request, order_id):
    tenant = _get_default_tenant(request)
    o = get_object_or_404(CustomerOrder, id=order_id, tenant=tenant)
    if request.method == "POST":
        inv = _invoice_from_lines(tenant, o.customer, o.currency_code, o.notes, o.terms, o, "source_order", request.user)
        o.status = CustomerOrder.Status.INVOICED
        o.save(update_fields=["status"])
        # Consume reservations: the order is now fulfilled into an invoice (the
        # invoice deducts actual stock on issue), so the hold ships rather than
        # being freed unfulfilled (reservation lifecycle fix).
        consume_reservations(tenant=tenant, ref_type="CUSTOMER_ORDER", ref_id=str(o.id))
        log_audit(action="ORDER_INVOICED", request=request, user=request.user, tenant=tenant,
                  detail=f"{o.order_number} -> invoice {inv.invoice_number}")
        messages.success(request, f"Sales order {o.order_number} converted to invoice {inv.invoice_number} (draft).")
        return redirect("ar_invoice_detail", invoice_id=inv.id)
    return redirect("corder_detail", order_id=o.id)


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def corder_delete(request, order_id):
    tenant = _get_default_tenant(request)
    o = get_object_or_404(CustomerOrder, id=order_id, tenant=tenant)
    if request.method == "POST":
        if o.status == CustomerOrder.Status.INVOICED:
            messages.error(request, "Invoiced sales orders can't be deleted - they're linked to an invoice.")
            return redirect("corder_detail", order_id=o.id)
        number = o.order_number
        _release_customer_order(o)  # free any reservations before removing the order
        o.delete()
        log_audit(action="ORDER_DELETED", request=request, user=request.user, tenant=tenant, detail=number)
        messages.success(request, f"Sales order {number} deleted.")
        return redirect("corder_list")
    return redirect("corder_detail", order_id=o.id)


# ---- Recurring invoices ----

@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def recurring_list(request):
    tenant = _get_default_tenant(request)
    templates = RecurringInvoice.objects.filter(tenant=tenant).select_related("customer").prefetch_related("lines", "lines__tax_code").order_by("-is_active", "next_run_date")
    due = [t for t in templates if t.is_active and t.next_run_date <= timezone.localdate()]
    return render(request, "sales/recurring_list.html", {
        "tenant": tenant, "templates": templates, "due_count": len(due), "today": timezone.localdate()})


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def recurring_create(request):
    tenant = _get_default_tenant(request)
    t = RecurringInvoice(tenant=tenant, currency_code=tenant.currency_code)
    if request.method == "POST":
        form = RecurringInvoiceForm(request.POST, instance=t)
        formset = RecurringInvoiceLineFormSet(request.POST, instance=t)
        if form.is_valid() and formset.is_valid():
            t = form.save(commit=False)
            t.tenant = tenant
            t.currency_code = tenant.currency_code
            if not t.next_run_date:
                t.next_run_date = t.start_date
            t.save()
            formset.save()
            messages.success(request, f"Recurring invoice '{t.name}' saved.")
            return redirect("recurring_detail", template_id=t.id)
    else:
        initial = {}
        if tenant.invoice_footer:
            initial["terms"] = tenant.invoice_footer
        form = RecurringInvoiceForm(instance=t, initial=initial)
        line_initial = [{"tax_code": tenant.default_tax_code}] if tenant.default_tax_code_id else None
        formset = RecurringInvoiceLineFormSet(instance=t, initial=line_initial)
    return render(request, "sales/doc_form.html", {
        "tenant": tenant, "form": form, "formset": formset,
        "doc_label": "Recurring invoice", "list_url": "/recurring-invoices/"})


@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def recurring_detail(request, template_id):
    tenant = _get_default_tenant(request)
    t = get_object_or_404(RecurringInvoice, id=template_id, tenant=tenant)
    generated = CustomerInvoice.objects.filter(tenant=tenant, customer=t.customer).order_by("-invoice_date")[:50]
    return render(request, "sales/recurring_detail.html", {"tenant": tenant, "t": t, "generated": generated})


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def recurring_toggle(request, template_id):
    tenant = _get_default_tenant(request)
    t = get_object_or_404(RecurringInvoice, id=template_id, tenant=tenant)
    if request.method == "POST":
        t.is_active = not t.is_active
        t.save(update_fields=["is_active"])
        messages.success(request, f"'{t.name}' {'resumed' if t.is_active else 'paused'}.")
    return redirect("recurring_detail", template_id=t.id)


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def recurring_generate(request, template_id):
    tenant = _get_default_tenant(request)
    t = get_object_or_404(RecurringInvoice, id=template_id, tenant=tenant)
    if request.method == "POST":
        from core.services import recurring
        created = recurring.generate_for_template(t, user=request.user)
        log_audit(action="RECURRING_GENERATED", request=request, user=request.user, tenant=tenant,
                  detail=f"{t.name}: {len(created)} invoice(s)")
        messages.success(request, f"Generated {len(created)} invoice(s) from '{t.name}'." if created else "Nothing due to generate yet.")
    return redirect("recurring_detail", template_id=t.id)


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def recurring_run_due(request):
    tenant = _get_default_tenant(request)
    if request.method == "POST":
        from core.services import recurring
        created = recurring.generate_due(tenant=tenant, user=request.user)
        log_audit(action="RECURRING_GENERATED", request=request, user=request.user, tenant=tenant,
                  detail=f"run due: {len(created)} invoice(s)")
        messages.success(request, f"Generated {len(created)} invoice(s) from due recurring templates." if created else "No recurring invoices were due.")
    return redirect("recurring_list")


# ---- Draft document editing ----

@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def ar_invoice_edit(request, invoice_id):
    tenant = _get_default_tenant(request)
    inv = get_object_or_404(CustomerInvoice, id=invoice_id, tenant=tenant)
    if inv.status != CustomerInvoice.Status.DRAFT:
        messages.error(request, "Only draft invoices can be edited.")
        return redirect("ar_invoice_detail", invoice_id=inv.id)
    if request.method == "POST":
        form = CustomerInvoiceForm(request.POST, instance=inv)
        formset = CustomerInvoiceLineFormSet(request.POST, instance=inv)
        if form.is_valid() and formset.is_valid():
            inv = form.save(commit=False)
            inv.tenant = tenant
            if not (inv.invoice_number or "").strip():
                from core.numbering import next_invoice_number
                inv.invoice_number = next_invoice_number(tenant)
            inv.save()
            formset.save()
            if (form.cleaned_data.get("action") or "save") == "issue":
                post_customer_invoice(inv, user=request.user)
            messages.success(request, f"Invoice {inv.invoice_number} updated.")
            return redirect("ar_invoice_detail", invoice_id=inv.id)
    else:
        form = CustomerInvoiceForm(instance=inv)
        formset = CustomerInvoiceLineFormSet(instance=inv)
    return render(request, "ar/ar_invoice_form.html", {"tenant": tenant, "form": form, "formset": formset, "is_edit": True})


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def quote_edit(request, quote_id):
    tenant = _get_default_tenant(request)
    q = get_object_or_404(SalesQuote, id=quote_id, tenant=tenant)
    if q.status not in (SalesQuote.Status.DRAFT, SalesQuote.Status.SENT):
        messages.error(request, "Only draft or sent quotes can be edited.")
        return redirect("quote_detail", quote_id=q.id)
    if request.method == "POST":
        form = SalesQuoteForm(request.POST, instance=q)
        formset = SalesQuoteLineFormSet(request.POST, instance=q)
        if form.is_valid() and formset.is_valid():
            q = form.save(commit=False)
            q.tenant = tenant
            q.save()
            formset.save()
            messages.success(request, f"Quote {q.quote_number} updated.")
            return redirect("quote_detail", quote_id=q.id)
    else:
        form = SalesQuoteForm(instance=q)
        formset = SalesQuoteLineFormSet(instance=q)
    return render(request, "sales/doc_form.html", {
        "tenant": tenant, "form": form, "formset": formset,
        "doc_label": "Quote", "list_url": "/quotes/", "is_edit": True})


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def corder_edit(request, order_id):
    tenant = _get_default_tenant(request)
    o = get_object_or_404(CustomerOrder, id=order_id, tenant=tenant)
    if o.status not in (CustomerOrder.Status.DRAFT, CustomerOrder.Status.CONFIRMED):
        messages.error(request, "Invoiced or cancelled orders cannot be edited.")
        return redirect("corder_detail", order_id=o.id)
    if request.method == "POST":
        form = CustomerOrderForm(request.POST, instance=o)
        formset = CustomerOrderLineFormSet(request.POST, instance=o)
        if form.is_valid() and formset.is_valid():
            o = form.save(commit=False)
            o.tenant = tenant
            o.save()
            formset.save()
            # Keep reservations in step with the edited lines on confirmed orders.
            if o.status == CustomerOrder.Status.CONFIRMED:
                _reserve_customer_order(o)
            messages.success(request, f"Sales order {o.order_number} updated.")
            return redirect("corder_detail", order_id=o.id)
    else:
        form = CustomerOrderForm(instance=o)
        formset = CustomerOrderLineFormSet(instance=o)
    return render(request, "sales/doc_form.html", {
        "tenant": tenant, "form": form, "formset": formset,
        "doc_label": "Sales order", "list_url": "/customer-orders/", "is_edit": True})


@role_required(SALES_DOC_ROLES, write_groups=SALES_DOC_ROLES)
@transaction.atomic
def recurring_edit(request, template_id):
    tenant = _get_default_tenant(request)
    t = get_object_or_404(RecurringInvoice, id=template_id, tenant=tenant)
    if request.method == "POST":
        form = RecurringInvoiceForm(request.POST, instance=t)
        formset = RecurringInvoiceLineFormSet(request.POST, instance=t)
        if form.is_valid() and formset.is_valid():
            t = form.save(commit=False)
            t.tenant = tenant
            if not t.next_run_date:
                t.next_run_date = t.start_date
            t.save()
            formset.save()
            messages.success(request, f"Recurring invoice '{t.name}' updated.")
            return redirect("recurring_detail", template_id=t.id)
    else:
        form = RecurringInvoiceForm(instance=t)
        formset = RecurringInvoiceLineFormSet(instance=t)
    return render(request, "sales/doc_form.html", {
        "tenant": tenant, "form": form, "formset": formset,
        "doc_label": "Recurring invoice", "list_url": "/recurring-invoices/", "is_edit": True})


# ---- Sales reports ----

def _sales_period(request, tenant):
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    if not date_from and not date_to:
        date_from, date_to = reports_service.current_financial_year(tenant)
    elif not date_to:
        date_to = timezone.localdate()
    return date_from, date_to


def _sales_location_filter(request, tenant):
    """Resolve the location filter for sales reports.

    Returns (location_ids, selected_id, locations) where location_ids is what to
    pass to the report service (None = no restriction), selected_id is the chosen
    location for the dropdown, and locations is the list to populate it. Honours
    the user's per-location access (a restricted user can't widen past their grants)."""
    from core.access import accessible_location_ids, accessible_locations
    allowed = active_location_ids(request)  # scope to the selected site
    locations = list(accessible_locations(request.user, tenant).order_by("name"))
    selected_id = None
    raw = (request.GET.get("location") or "").strip()
    if raw.isdigit():
        sid = int(raw)
        if allowed is None or sid in allowed:
            selected_id = sid
    if selected_id is not None:
        return [selected_id], selected_id, locations
    return allowed, None, locations


@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def sales_reports_index(request):
    tenant = _get_default_tenant(request)
    return render(request, "sales/reports_index.html", {"tenant": tenant})


@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def report_sales_history(request):
    tenant = _get_default_tenant(request)
    date_from, date_to = _sales_period(request, tenant)
    loc_ids, selected_loc, locations = _sales_location_filter(request, tenant)
    from core.services import sales_reports
    data = sales_reports.sales_history(tenant, date_from, date_to, location_ids=loc_ids)
    return render(request, "sales/report_history.html", {
        "tenant": tenant, "data": data, "date_from": date_from, "date_to": date_to,
        "locations": locations, "selected_loc": selected_loc,
        "export_kind": "sales-history", "export_qs": f"?from={date_from}&to={date_to}"})


@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def report_sales_by_product(request):
    tenant = _get_default_tenant(request)
    date_from, date_to = _sales_period(request, tenant)
    loc_ids, selected_loc, locations = _sales_location_filter(request, tenant)
    from core.services import sales_reports
    data = sales_reports.sales_by_product(tenant, date_from, date_to, location_ids=loc_ids)
    return render(request, "sales/report_grouped.html", {
        "tenant": tenant, "data": data, "date_from": date_from, "date_to": date_to,
        "title": "Sales by Product", "label": "Product", "show_qty": True,
        "locations": locations, "selected_loc": selected_loc,
        "export_kind": "sales-by-product", "export_qs": f"?from={date_from}&to={date_to}"})


@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def report_sales_by_customer(request):
    tenant = _get_default_tenant(request)
    date_from, date_to = _sales_period(request, tenant)
    loc_ids, selected_loc, locations = _sales_location_filter(request, tenant)
    from core.services import sales_reports
    data = sales_reports.sales_by_customer(tenant, date_from, date_to, location_ids=loc_ids)
    return render(request, "sales/report_grouped.html", {
        "tenant": tenant, "data": data, "date_from": date_from, "date_to": date_to,
        "title": "Sales by Customer", "label": "Customer", "show_count": True,
        "locations": locations, "selected_loc": selected_loc,
        "export_kind": "sales-by-customer", "export_qs": f"?from={date_from}&to={date_to}"})


@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def report_sales_by_channel(request):
    tenant = _get_default_tenant(request)
    date_from, date_to = _sales_period(request, tenant)
    from core.services import sales_reports
    data = sales_reports.sales_by_channel(tenant, date_from, date_to)
    return render(request, "sales/report_grouped.html", {
        "tenant": tenant, "data": data, "date_from": date_from, "date_to": date_to,
        "title": "Sales by Channel", "label": "Channel", "show_count": True, "channel_mode": True,
        "export_kind": "sales-by-channel", "export_qs": f"?from={date_from}&to={date_to}"})


@role_required(SALES_DOC_READ, write_groups=SALES_DOC_ROLES)
def report_profitability(request):
    """Gross-margin report: revenue - COGS by product and by customer."""
    tenant = _get_default_tenant(request)
    date_from, date_to = _sales_period(request, tenant)
    from core.services import sales_reports
    data = sales_reports.profitability(tenant, date_from, date_to)
    return render(request, "sales/report_profitability.html", {
        "tenant": tenant, "data": data, "date_from": date_from, "date_to": date_to})


@role_required([ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_FINANCE, ROLE_READONLY],
               write_groups=[ROLE_ADMIN, ROLE_PROCUREMENT, ROLE_FINANCE])
def report_supplier_scorecard(request):
    """Supplier performance: spend, on-time delivery and price variance."""
    tenant = _get_default_tenant(request)
    date_from, date_to = _sales_period(request, tenant)
    from core.services import purchasing
    data = purchasing.supplier_scorecard(tenant, date_from, date_to)
    return render(request, "reports/supplier_scorecard.html", {
        "tenant": tenant, "data": data, "date_from": date_from, "date_to": date_to})


# ============================
# General Ledger
# ============================

@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def gl_account_list(request):
    tenant = _get_default_tenant(request)
    accounts = GLAccount.objects.filter(tenant=tenant).order_by("code")
    return render(request, "gl/gl_account_list.html", {"tenant": tenant, "accounts": accounts})

@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def gl_account_create(request):
    tenant = _get_default_tenant(request)
    form = GLAccountForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.tenant = tenant
        obj.save()
        return redirect("gl_account_list")
    return render(request, "gl/gl_account_form.html", {"tenant": tenant, "form": form, "mode": "create"})

@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def gl_account_edit(request, account_id):
    tenant = _get_default_tenant(request)
    obj = get_object_or_404(GLAccount, id=account_id, tenant=tenant)
    form = GLAccountForm(request.POST or None, instance=obj)
    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect("gl_account_list")
    return render(request, "gl/gl_account_form.html", {"tenant": tenant, "form": form, "mode": "edit"})

@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def journal_list(request):
    tenant = _get_default_tenant(request)
    entries = JournalEntry.objects.filter(tenant=tenant).order_by("-entry_date", "-id")[:200]
    return render(request, "gl/journal_list.html", {"tenant": tenant, "entries": entries})

@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def journal_detail(request, je_id):
    tenant = _get_default_tenant(request)
    je = get_object_or_404(JournalEntry, id=je_id, tenant=tenant)
    return render(request, "gl/journal_detail.html", {"tenant": tenant, "je": je})


# ============================
# AP invoice posting (3-way match -> post to GL)
# ============================

@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def invoice_post(request, invoice_id):
    tenant = _get_default_tenant(request)
    inv = get_object_or_404(SupplierInvoice, id=invoice_id, tenant=tenant)
    if request.method == "POST":
        post_supplier_invoice(inv, user=request.user)
        return redirect("invoice_detail", invoice_id=inv.id)
    return render(request, "finance/invoice_post.html", {"tenant": tenant, "inv": inv})


@login_required
@role_required([ROLE_ADMIN, ROLE_PROCUREMENT], [ROLE_ADMIN, ROLE_PROCUREMENT])
def shipment_new(request, po_id):
    tenant = _get_default_tenant(request)
    po = get_object_or_404(PurchaseOrder, id=po_id, tenant=tenant)

    if po.status == PurchaseOrder.Status.DRAFT:
        messages.error(request, "Submit the PO before creating shipments.")
        return redirect("po_detail", po_id=po.id)

    dests = Location.objects.filter(tenant=tenant).order_by("name")
    if request.method == "POST":
        dest_id = request.POST.get("destination_id")
        dest = get_object_or_404(Location, id=dest_id, tenant=tenant)
        shipment = Shipment.objects.create(
            tenant=tenant,
            po=po,
            from_supplier=po.supplier,
            destination=dest,
            carrier=(request.POST.get("carrier") or "").strip() or None,
            tracking_number=(request.POST.get("tracking_number") or "").strip() or None,
            eta=(request.POST.get("eta") or None),
            status=Shipment.Status.CREATED,
        )
        # Create empty lines for this shipment (expected 0 by default)
        for pol in po.lines.all():
            ShipmentLine.objects.get_or_create(
                shipment=shipment,
                po_line=pol,
                defaults={"expected_qty": Decimal("0.00")},
            )
        messages.success(request, "Shipment created. Allocate quantities to it.")
        return redirect("shipment_detail", shipment_id=shipment.id)

    return render(request, "shipments/shipment_new.html", {"tenant": tenant, "po": po, "dests": dests})


# ============================
# Financial Reports
# ============================

def _parse_date(raw):
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return timezone.datetime.fromisoformat(raw).date()
    except (ValueError, TypeError):
        return None


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def reports_index(request):
    tenant = _get_default_tenant(request)
    return render(request, "reports/index.html", {"tenant": tenant})


def _report_site_filter(request, tenant):
    """Resolve an optional Site filter for finance reports. Returns
    (site_ids, selected_id, sites). Default (no ?site=) is company-wide
    (site_ids=None) - a controlled consolidated report view. A chosen, accessible
    site narrows to it. `sites` populates the dropdown."""
    from core.access import accessible_site_ids
    sites = list(selectable_sites(request.user, tenant)) if tenant is not None else []
    raw = (request.GET.get("site") or "").strip()
    if raw.isdigit():
        sid = int(raw)
        allowed = accessible_site_ids(request.user, tenant)
        if allowed is None or sid in allowed:
            return [sid], sid, sites
    return None, None, sites


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def report_trial_balance(request):
    tenant = _get_default_tenant(request)
    as_of = _parse_date(request.GET.get("as_of")) or timezone.localdate()
    site_ids, sel_site, sites = _report_site_filter(request, tenant)
    data = reports_service.trial_balance(tenant, date_to=as_of, site_ids=site_ids)
    return render(request, "reports/trial_balance.html", {
        "tenant": tenant, "as_of": as_of, "data": data,
        "sites": sites, "selected_site": sel_site,
        "export_kind": "trial-balance", "export_qs": f"?as_of={as_of}",
    })


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def report_pnl(request):
    tenant = _get_default_tenant(request)
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    # Default to the company's current financial year when no dates are given.
    if not date_from and not date_to:
        date_from, date_to = reports_service.current_financial_year(tenant)
    elif not date_to:
        date_to = timezone.localdate()
    site_ids, sel_site, sites = _report_site_filter(request, tenant)
    data = reports_service.profit_and_loss(tenant, date_from=date_from, date_to=date_to, site_ids=site_ids)
    return render(request, "reports/pnl.html", {
        "tenant": tenant, "date_from": date_from, "date_to": date_to, "data": data,
        "sites": sites, "selected_site": sel_site,
        "export_kind": "profit-and-loss", "export_qs": f"?from={date_from}&to={date_to}",
    })


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def report_balance_sheet(request):
    tenant = _get_default_tenant(request)
    as_of = _parse_date(request.GET.get("as_of")) or timezone.localdate()
    site_ids, sel_site, sites = _report_site_filter(request, tenant)
    data = reports_service.balance_sheet(tenant, as_of=as_of, site_ids=site_ids)
    return render(request, "reports/balance_sheet.html", {
        "tenant": tenant, "as_of": as_of, "data": data,
        "sites": sites, "selected_site": sel_site, "site_filtered": bool(site_ids),
        "export_kind": "balance-sheet", "export_qs": f"?as_of={as_of}",
    })


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def consolidated_reports(request):
    """Group-level consolidated P&L / balance sheet / stock across the companies
    in this tenant's group that the user belongs to."""
    from core.access import group_companies
    tenant = _get_default_tenant(request)
    companies = group_companies(request.user, tenant)
    data = reports_service.consolidated(companies)
    return render(request, "reports/consolidated.html", {
        "tenant": tenant, "data": data,
        "group": getattr(tenant, "group", None), "company_count": len(companies),
    })


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def intercompany_list(request):
    """List and raise inter-company sales between companies in the group."""
    from core.access import group_companies
    from core.models import InterCompanyTransaction
    from core.services.intercompany import create_intercompany_sale
    tenant = _get_default_tenant(request)
    companies = group_companies(request.user, tenant)
    targets = [c for c in companies if c.id != tenant.id]

    if request.method == "POST":
        if not tenant.group_id:
            messages.error(request, "Set up a company group first.")
            return redirect("intercompany_list")
        to_id = request.POST.get("to_tenant")
        target = next((c for c in targets if str(c.id) == str(to_id)), None)
        amount = request.POST.get("amount") or "0"
        desc = (request.POST.get("description") or "").strip()
        try:
            from decimal import Decimal as _D
            if target is None:
                raise ValueError("Choose a company in your group to sell to.")
            ict = create_intercompany_sale(tenant, target, _D(amount), desc, user=request.user)
            log_audit(action="INTERCOMPANY_SALE", request=request, user=request.user, tenant=tenant,
                      entity_type="InterCompanyTransaction", entity_id=ict.id,
                      detail=f"{tenant.name} -> {target.name} {ict.amount}")
            messages.success(request, f"Inter-company sale to {target.name} posted ({ict.amount}).")
        except (ValueError, InvalidOperation) as e:
            messages.error(request, str(e) or "Invalid amount.")
        return redirect("intercompany_list")

    company_ids = [c.id for c in companies]
    txns = (InterCompanyTransaction.objects
            .filter(from_tenant_id__in=company_ids)
            .select_related("from_tenant", "to_tenant", "customer_invoice", "expense")
            .order_by("-created_at")[:100])
    return render(request, "reports/intercompany.html", {
        "tenant": tenant, "targets": targets, "txns": txns, "group": getattr(tenant, "group", None)})


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def report_cash_flow(request):
    tenant = _get_default_tenant(request)
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    if not date_from and not date_to:
        date_from, date_to = reports_service.current_financial_year(tenant)
    elif not date_to:
        date_to = timezone.localdate()
    data = reports_service.cash_flow_summary(tenant, date_from=date_from, date_to=date_to)
    return render(request, "reports/cash_flow.html", {
        "tenant": tenant, "date_from": date_from, "date_to": date_to, "data": data,
        "export_kind": "cash-flow", "export_qs": f"?from={date_from}&to={date_to}",
    })


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def report_aged_receivables(request):
    tenant = _get_default_tenant(request)
    as_of = _parse_date(request.GET.get("as_of")) or timezone.localdate()
    data = reports_service.aged_receivables(tenant, as_of=as_of)
    return render(request, "reports/aged.html", {
        "tenant": tenant, "as_of": as_of, "data": data,
        "title": "Aged Debtors (Receivables)", "party_label": "Customer",
        "export_kind": "aged-receivables", "export_qs": f"?as_of={as_of}",
    })


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_WAREHOUSE, ROLE_PROCUREMENT, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def report_stock_valuation(request):
    from core.access import accessible_location_ids
    tenant = _get_default_tenant(request)
    allowed = active_location_ids(request)  # scope to the selected site
    data = reports_service.stock_valuation(tenant, location_ids=allowed)
    return render(request, "reports/stock_valuation.html", {"tenant": tenant, "data": data})


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_WAREHOUSE, ROLE_PROCUREMENT, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def report_inventory_analytics(request):
    """Inventory valuation depth (per location / lot) + turnover KPIs."""
    from core.access import accessible_location_ids
    tenant = _get_default_tenant(request)
    date_from, date_to = _sales_period(request, tenant)
    allowed = active_location_ids(request)  # scope to the selected site
    data = reports_service.inventory_analytics(tenant, date_from, date_to, location_ids=allowed)
    return render(request, "reports/inventory_analytics.html", {
        "tenant": tenant, "data": data, "date_from": date_from, "date_to": date_to})


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def report_aged_payables(request):
    tenant = _get_default_tenant(request)
    as_of = _parse_date(request.GET.get("as_of")) or timezone.localdate()
    data = reports_service.aged_payables(tenant, as_of=as_of)
    return render(request, "reports/aged.html", {
        "tenant": tenant, "as_of": as_of, "data": data,
        "title": "Aged Creditors (Payables)", "party_label": "Supplier",
        "export_kind": "aged-payables", "export_qs": f"?as_of={as_of}",
    })


# ============================
# Payments (AR receipts / AP payments) + bank reconciliation
# ============================

def _allocate_fifo(payment, open_invoices, invoice_field):
    """Allocate the payment amount across open invoices oldest-first."""
    remaining = payment.amount
    for inv in open_invoices:
        if remaining <= Decimal("0.00"):
            break
        outstanding = inv.outstanding
        if outstanding <= Decimal("0.00"):
            continue
        alloc = min(remaining, outstanding)
        PaymentAllocation.objects.create(payment=payment, amount=alloc, **{invoice_field: inv})
        remaining -= alloc


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def payment_list(request):
    tenant = _get_default_tenant(request)
    payments = Payment.objects.filter(tenant=tenant).select_related("customer", "supplier").order_by("-payment_date", "-id")
    return render(request, "payments/payment_list.html", {"tenant": tenant, "payments": payments})


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def receipt_create(request):
    tenant = _get_default_tenant(request)
    form = ReceiptForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        payment = form.save(commit=False)
        payment.tenant = tenant
        payment.direction = Payment.Direction.RECEIPT
        payment.currency_code = tenant.currency_code
        payment.save()
        open_invoices = (CustomerInvoice.objects
                         .filter(tenant=tenant, customer=payment.customer, status=CustomerInvoice.Status.ISSUED)
                         .order_by("invoice_date", "id"))
        _allocate_fifo(payment, open_invoices, "customer_invoice")
        post_payment(payment, user=request.user)
        messages.success(request, f"Receipt of {payment.amount} recorded and allocated.")
        return redirect("payment_detail", payment_id=payment.id)
    return render(request, "payments/payment_form.html", {"tenant": tenant, "form": form, "mode": "receipt"})


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def supplier_payment_create(request):
    tenant = _get_default_tenant(request)
    form = SupplierPaymentForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        payment = form.save(commit=False)
        payment.tenant = tenant
        payment.direction = Payment.Direction.PAYMENT
        payment.currency_code = tenant.currency_code
        payment.save()
        open_invoices = (SupplierInvoice.objects
                         .filter(tenant=tenant, supplier=payment.supplier, status=SupplierInvoice.Status.POSTED)
                         .order_by("invoice_date", "id"))
        _allocate_fifo(payment, open_invoices, "supplier_invoice")
        post_payment(payment, user=request.user)
        messages.success(request, f"Payment of {payment.amount} recorded and allocated.")
        return redirect("payment_detail", payment_id=payment.id)
    return render(request, "payments/payment_form.html", {"tenant": tenant, "form": form, "mode": "payment"})


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def refund_create(request):
    """Record money refunded to a customer (cash out)."""
    tenant = _get_default_tenant(request)
    form = RefundForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        payment = form.save(commit=False)
        payment.tenant = tenant
        payment.direction = Payment.Direction.REFUND
        payment.currency_code = tenant.currency_code
        payment.save()
        post_payment(payment, user=request.user)
        log_audit(action="REFUND_RECORDED", request=request, user=request.user, tenant=tenant,
                  detail=f"{payment.amount} to {payment.party_name}")
        messages.success(request, f"Refund of {payment.amount} to {payment.party_name} recorded.")
        return redirect("payment_detail", payment_id=payment.id)
    return render(request, "payments/payment_form.html", {"tenant": tenant, "form": form, "mode": "refund"})


def _statement_context(request, customer_id):
    tenant = _get_default_tenant(request)
    customer = get_object_or_404(Customer, id=customer_id, tenant=tenant)
    from core.services import statements
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    if not date_from or not date_to:
        date_from, date_to = statements.default_period(tenant)
    data = statements.customer_statement(tenant, customer, date_from, date_to)
    return tenant, customer, data


@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
def customer_statement(request, customer_id):
    tenant, customer, data = _statement_context(request, customer_id)
    return render(request, "sales/statement.html", {"tenant": tenant, "customer": customer, "data": data})


@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
def customer_statement_pdf(request, customer_id):
    tenant, customer, data = _statement_context(request, customer_id)
    from core.services.pdf import pdf_response
    return pdf_response(f"statement-{customer.name}.pdf", "documents/statement_pdf.html",
                        {"tenant": tenant, "customer": customer, "data": data,
                         "doc_title": "STATEMENT", "number": customer.name}, download=False)


@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
def customer_statement_email(request, customer_id):
    from django.conf import settings
    tenant, customer, data = _statement_context(request, customer_id)
    if request.method == "POST":
        if not customer.email:
            messages.warning(request, f"{customer.name} has no email address on file.")
        else:
            from core.services.pdf import render_to_pdf
            from django.core.mail import EmailMessage
            pdf = render_to_pdf("documents/statement_pdf.html", {
                "tenant": tenant, "customer": customer, "data": data,
                "doc_title": "STATEMENT", "number": customer.name})
            msg = EmailMessage(
                subject=f"Statement of account from {tenant.name}",
                body=(f"Dear {customer.name},\n\nPlease find your statement of account attached.\n"
                      f"Balance due: {tenant.currency_code} {data['closing']:.2f}\n\nThank you.\n"),
                from_email=settings.DEFAULT_FROM_EMAIL, to=[customer.email])
            if pdf:
                msg.attach(f"statement-{customer.name}.pdf", pdf, "application/pdf")
            msg.send(fail_silently=True)
            log_audit(action="STATEMENT_SENT", request=request, user=request.user, tenant=tenant, detail=customer.name)
            messages.success(request, f"Statement emailed to {customer.email}.")
    return redirect("customer_statement", customer_id=customer.id)


@role_required([ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_SALES, ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def ar_invoice_refund(request, invoice_id):
    """Refund a paid invoice: record a customer refund for the amount paid and
    mark the invoice Refunded."""
    tenant = _get_default_tenant(request)
    inv = get_object_or_404(CustomerInvoice, id=invoice_id, tenant=tenant)
    if request.method == "POST":
        amount = inv.amount_paid
        if amount <= Decimal("0.00"):
            messages.error(request, "There is nothing paid on this invoice to refund.")
        else:
            payment = Payment.objects.create(
                tenant=tenant, direction=Payment.Direction.REFUND, customer=inv.customer,
                amount=amount, method=Payment.Method.BANK, currency_code=tenant.currency_code,
                reference=f"Refund {inv.invoice_number}")
            post_payment(payment, user=request.user)
            inv.status = CustomerInvoice.Status.REFUNDED
            inv.save(update_fields=["status"])
            log_audit(action="REFUND_RECORDED", request=request, user=request.user, tenant=tenant,
                      detail=f"{amount} for invoice {inv.invoice_number}")
            messages.success(request, f"Refunded {amount} for invoice {inv.invoice_number}.")
    return redirect("ar_invoice_detail", invoice_id=inv.id)


# ============================
# Expenses
# ============================

# Staff who may record/submit expenses (everyone but read-only); approving and
# direct posting stay with Finance/Admin.
EXPENSE_STAFF = [ROLE_ADMIN, ROLE_FINANCE, ROLE_PROCUREMENT, ROLE_WAREHOUSE, ROLE_SALES]
EXPENSE_APPROVERS = [ROLE_FINANCE, ROLE_ADMIN]


def _can_approve_expenses(request):
    return bool(set(EXPENSE_APPROVERS) & effective_groups(request)) or request.user.is_superuser


@role_required(EXPENSE_STAFF + [ROLE_READONLY], write_groups=EXPENSE_STAFF)
def expense_list(request):
    tenant = _get_default_tenant(request)
    expenses = Expense.objects.filter(tenant=tenant).select_related("category", "supplier", "tax_code").order_by("-expense_date", "-id")
    total = sum((e.total for e in expenses), Decimal("0.00"))
    return render(request, "expenses/expense_list.html", {
        "tenant": tenant, "expenses": expenses, "total": total,
        "can_approve": _can_approve_expenses(request)})


def _notify_expense_submitted(request, tenant, expense):
    """Notify expense approvers that ``expense`` is awaiting their approval.
    Shared by the create form and the draft-submit action so both paths emit
    the same notification."""
    from core import notify
    from django.urls import reverse
    notify.notify_roles(
        tenant, [roles_mod.ADMIN, roles_mod.FINANCE, roles_mod.ACCOUNTANT, roles_mod.MANAGER],
        exclude_user=request.user, category="APPROVAL_REQUEST", actor=request.user,
        url=reverse("expense_detail", kwargs={"expense_id": expense.id}),
        title=f"Expense awaiting approval: {expense.payee}",
        message=f"{request.user.username} submitted an expense of {expense.currency_code} {expense.total} for approval.",
        request=request,
    )


@role_required(EXPENSE_STAFF, write_groups=EXPENSE_STAFF)
@transaction.atomic
def expense_create(request):
    tenant = _get_default_tenant(request)
    initial = {}
    if tenant.default_tax_code_id:
        initial["tax_code"] = tenant.default_tax_code
    form = ExpenseForm(request.POST or None, request.FILES or None, initial=initial)
    is_approver = _can_approve_expenses(request)
    if request.method == "POST" and form.is_valid():
        expense = form.save(commit=False)
        expense.tenant = tenant
        expense.site = get_active_site(request)  # attribute the cost to the working site
        expense.currency_code = tenant.currency_code
        expense.save()
        action = request.POST.get("action") or "save"
        threshold = tenant.expense_approval_threshold or Decimal("0.00")
        needs_approval = bool(threshold and threshold > 0 and expense.total >= threshold)
        if action == "post" and is_approver and not needs_approval:
            post_expense(expense, user=request.user)
            messages.success(request, f"Expense recorded and posted ({expense.total}).")
        elif action in ("post", "submit"):
            expense.status = Expense.Status.SUBMITTED
            expense.submitted_by = request.user
            expense.save(update_fields=["status", "submitted_by"])
            log_audit(action="expense_submit", request=request, user=request.user, tenant=tenant,
                      detail=f"{expense.payee} {expense.total}")
            _notify_expense_submitted(request, tenant, expense)
            if action == "post" and is_approver and needs_approval:
                messages.warning(request, f"Expense submitted — approval required (total {expense.total} ≥ threshold {threshold}).")
            else:
                messages.success(request, "Expense submitted for approval.")
        else:
            messages.success(request, "Expense saved as draft.")
        return redirect("expense_detail", expense_id=expense.id)
    return render(request, "expenses/expense_form.html", {
        "tenant": tenant, "form": form, "can_approve": is_approver})


@role_required(EXPENSE_STAFF + [ROLE_READONLY], write_groups=EXPENSE_STAFF)
def expense_detail(request, expense_id):
    tenant = _get_default_tenant(request)
    expense = get_object_or_404(Expense, id=expense_id, tenant=tenant)
    je = JournalEntry.objects.filter(tenant=tenant, ref_type="EXPENSE", ref_id=str(expense.id)).prefetch_related("lines", "lines__account").order_by("-id").first()
    return render(request, "expenses/expense_detail.html", {
        "tenant": tenant, "expense": expense, "je": je,
        "can_approve": _can_approve_expenses(request)})


@role_required(EXPENSE_STAFF, write_groups=EXPENSE_STAFF)
@transaction.atomic
def expense_submit(request, expense_id):
    tenant = _get_default_tenant(request)
    expense = get_object_or_404(Expense, id=expense_id, tenant=tenant)
    if request.method == "POST" and expense.status == Expense.Status.DRAFT:
        expense.status = Expense.Status.SUBMITTED
        expense.submitted_by = request.user
        expense.save(update_fields=["status", "submitted_by"])
        log_audit(action="expense_submit", request=request, user=request.user, tenant=tenant,
                  detail=f"{expense.payee} {expense.total}")
        _notify_expense_submitted(request, tenant, expense)
        messages.success(request, "Expense submitted for approval.")
    return redirect("expense_detail", expense_id=expense.id)


@role_required(EXPENSE_APPROVERS, write_groups=EXPENSE_APPROVERS)
@transaction.atomic
def expense_approve(request, expense_id):
    tenant = _get_default_tenant(request)
    expense = get_object_or_404(Expense, id=expense_id, tenant=tenant)
    if request.method != "POST":
        return redirect("expense_detail", expense_id=expense.id)
    if expense.status != Expense.Status.SUBMITTED:
        messages.info(request, "Only submitted expenses can be approved.")
        return redirect("expense_detail", expense_id=expense.id)
    post_expense(expense, user=request.user)  # sets POSTED + posts the GL entry
    expense.approved_by = request.user
    expense.approved_at = timezone.now()
    expense.save(update_fields=["approved_by", "approved_at"])
    log_audit(action="expense_approve", request=request, user=request.user, tenant=tenant,
              detail=f"{expense.payee} {expense.total}")
    from core import notify
    from django.urls import reverse
    notify.notify_user(
        expense.submitted_by, tenant=tenant, category="APPROVAL_RESULT", actor=request.user,
        url=reverse("expense_detail", kwargs={"expense_id": expense.id}),
        title=f"Expense approved: {expense.payee}",
        message=f"Your expense of {expense.currency_code} {expense.total} was approved by {request.user.username}.",
        request=request,
    )
    messages.success(request, f"Expense approved and posted ({expense.total}).")
    return redirect("expense_detail", expense_id=expense.id)


@role_required(EXPENSE_APPROVERS, write_groups=EXPENSE_APPROVERS)
@transaction.atomic
def expense_reject(request, expense_id):
    tenant = _get_default_tenant(request)
    expense = get_object_or_404(Expense, id=expense_id, tenant=tenant)
    if request.method != "POST":
        return redirect("expense_detail", expense_id=expense.id)
    if expense.status != Expense.Status.SUBMITTED:
        messages.info(request, "Only submitted expenses can be rejected.")
        return redirect("expense_detail", expense_id=expense.id)
    expense.status = Expense.Status.REJECTED
    expense.rejected_reason = (request.POST.get("reason") or "").strip() or None
    expense.approved_by = request.user
    expense.approved_at = timezone.now()
    expense.save(update_fields=["status", "rejected_reason", "approved_by", "approved_at"])
    log_audit(action="expense_reject", request=request, user=request.user, tenant=tenant,
              detail=f"{expense.payee}: {expense.rejected_reason or ''}")
    from core import notify
    from django.urls import reverse
    notify.notify_user(
        expense.submitted_by, tenant=tenant, category="APPROVAL_RESULT", actor=request.user,
        url=reverse("expense_detail", kwargs={"expense_id": expense.id}),
        title=f"Expense rejected: {expense.payee}",
        message=f"Your expense of {expense.currency_code} {expense.total} was rejected by {request.user.username}."
                + (f" Reason: {expense.rejected_reason}" if expense.rejected_reason else ""),
        request=request,
    )
    messages.success(request, "Expense rejected.")
    return redirect("expense_detail", expense_id=expense.id)


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def expense_post(request, expense_id):
    tenant = _get_default_tenant(request)
    expense = get_object_or_404(Expense, id=expense_id, tenant=tenant)
    if request.method == "POST" and expense.status != Expense.Status.POSTED:
        post_expense(expense, user=request.user)
        messages.success(request, "Expense posted to the ledger.")
    return redirect("expense_detail", expense_id=expense.id)


# ============================
# Credit notes
# ============================

@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def credit_note_list(request):
    tenant = _get_default_tenant(request)
    notes = CreditNote.objects.filter(tenant=tenant).select_related("customer", "supplier").prefetch_related("lines", "lines__tax_code").order_by("-credit_note_date", "-id")
    return render(request, "credit_notes/credit_note_list.html", {"tenant": tenant, "notes": notes})


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def credit_note_create(request):
    tenant = _get_default_tenant(request)
    cn = CreditNote(tenant=tenant, currency_code=tenant.currency_code)
    if request.method == "POST":
        form = CreditNoteForm(request.POST, instance=cn)
        formset = CreditNoteLineFormSet(request.POST, instance=cn)
        if form.is_valid() and formset.is_valid():
            cn = form.save(commit=False)
            cn.tenant = tenant
            cn.currency_code = tenant.currency_code
            cn.save()
            formset.save()
            if (request.POST.get("action") or "save") == "post":
                post_credit_note(cn, user=request.user)
                messages.success(request, f"Credit note {cn.credit_note_number} posted ({cn.total}).")
            else:
                messages.success(request, "Credit note saved as draft.")
            return redirect("credit_note_detail", note_id=cn.id)
    else:
        line_initial = [{"tax_code": tenant.default_tax_code}] if tenant.default_tax_code_id else None
        form = CreditNoteForm(instance=cn)
        formset = CreditNoteLineFormSet(instance=cn, initial=line_initial)
    return render(request, "credit_notes/credit_note_form.html", {"tenant": tenant, "form": form, "formset": formset})


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def credit_note_detail(request, note_id):
    tenant = _get_default_tenant(request)
    cn = get_object_or_404(CreditNote, id=note_id, tenant=tenant)
    je = JournalEntry.objects.filter(tenant=tenant, ref_type="CREDIT_NOTE", ref_id=str(cn.id)).prefetch_related("lines", "lines__account").order_by("-id").first()
    return render(request, "credit_notes/credit_note_detail.html", {"tenant": tenant, "cn": cn, "je": je})


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def credit_note_pdf(request, note_id):
    tenant = _get_default_tenant(request)
    cn = get_object_or_404(CreditNote, id=note_id, tenant=tenant)
    from core.services.pdf import pdf_response
    return pdf_response(f"credit-note-{cn.credit_note_number}.pdf", "documents/credit_note_pdf.html",
                        {"tenant": tenant, "cn": cn, "doc_title": "CREDIT NOTE", "number": cn.credit_note_number,
                         "notes": cn.reason}, download=False)


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def credit_note_post(request, note_id):
    tenant = _get_default_tenant(request)
    cn = get_object_or_404(CreditNote, id=note_id, tenant=tenant)
    if request.method == "POST" and cn.status != CreditNote.Status.POSTED:
        post_credit_note(cn, user=request.user)
        messages.success(request, "Credit note posted to the ledger.")
    return redirect("credit_note_detail", note_id=cn.id)


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def payment_detail(request, payment_id):
    tenant = _get_default_tenant(request)
    payment = get_object_or_404(Payment, id=payment_id, tenant=tenant)
    allocations = payment.allocations.select_related("customer_invoice", "supplier_invoice").all()
    can_delete = bool({ROLE_FINANCE, ROLE_ADMIN} & effective_groups(request)) or request.user.is_superuser
    return render(request, "payments/payment_detail.html", {
        "tenant": tenant, "payment": payment, "allocations": allocations,
        "perms_can_delete": can_delete,
    })


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def payment_delete(request, payment_id):
    """Soft-delete a payment (sensitive record): reverse its ledger entry, free
    the invoices it settled, flag it deleted, and audit it. The row survives for
    the audit trail (visible via all_objects)."""
    tenant = _get_default_tenant(request)
    payment = get_object_or_404(Payment, id=payment_id, tenant=tenant)
    if request.method != "POST":
        return redirect("payment_detail", payment_id=payment.id)

    # Re-open any customer invoices this payment had fully settled, then drop the
    # allocations so their outstanding balances recompute.
    for alloc in payment.allocations.select_related("customer_invoice").all():
        inv = alloc.customer_invoice
        if inv and inv.status == CustomerInvoice.Status.PAID:
            inv.status = CustomerInvoice.Status.SENT
            inv.save(update_fields=["status"])
    payment.allocations.all().delete()

    if payment.status == Payment.Status.POSTED:
        from core.services.gl import reverse_payment
        reverse_payment(payment, user=request.user)

    payment.is_deleted = True
    payment.deleted_at = timezone.now()
    payment.deleted_by = request.user
    payment.save(update_fields=["is_deleted", "deleted_at", "deleted_by"])
    log_audit(action="PAYMENT_DELETED", request=request, user=request.user, tenant=tenant,
              entity_type="Payment", entity_id=payment.id, old_value=payment.amount,
              detail=f"{payment.get_direction_display()} {payment.amount} ({payment.party_name})")
    messages.success(request, "Payment deleted and its ledger entry reversed.")
    return redirect("payment_list")


# ============================
# Bank transactions + reconciliation
# ============================

def _signed_payment(p):
    return p.amount if p.direction == Payment.Direction.RECEIPT else -p.amount


def _recon_rows(tenant, txns):
    """For each bank transaction, the internal records (payments / paid expenses)
    it could match: same signed amount, not already matched to another line."""
    matched_pids = set(BankTransaction.objects.filter(tenant=tenant, matched_payment__isnull=False).values_list("matched_payment_id", flat=True))
    matched_eids = set(BankTransaction.objects.filter(tenant=tenant, matched_expense__isnull=False).values_list("matched_expense_id", flat=True))
    payments = list(Payment.objects.filter(tenant=tenant, status=Payment.Status.POSTED).select_related("customer", "supplier"))
    expenses = list(Expense.objects.filter(tenant=tenant, status=Expense.Status.POSTED, paid=True))
    rows = []
    for t in txns:
        cands = []
        for p in payments:
            if p.id in matched_pids and p.id != t.matched_payment_id:
                continue
            if _signed_payment(p) == t.amount:
                cands.append({"value": f"payment:{p.id}", "label": f"Payment - {p.party_name} {p.amount} ({p.payment_date})", "selected": p.id == t.matched_payment_id})
        for e in expenses:
            if e.id in matched_eids and e.id != t.matched_expense_id:
                continue
            if -e.total == t.amount:
                cands.append({"value": f"expense:{e.id}", "label": f"Expense - {e.payee} {e.total} ({e.expense_date})", "selected": e.id == t.matched_expense_id})
        rows.append({"txn": t, "candidates": cands})
    return rows


def _apply_match(t, sel):
    if sel == "" or sel is None:
        t.matched_payment = None
        t.matched_expense = None
        t.is_reconciled = False
        t.reconciled_at = None
    elif sel.startswith("payment:"):
        t.matched_payment_id = int(sel.split(":")[1])
        t.matched_expense = None
        t.is_reconciled = True
        t.reconciled_at = timezone.now()
    elif sel.startswith("expense:"):
        t.matched_expense_id = int(sel.split(":")[1])
        t.matched_payment = None
        t.is_reconciled = True
        t.reconciled_at = timezone.now()
    t.save()
    if t.matched_payment_id:
        Payment.objects.filter(id=t.matched_payment_id).update(is_reconciled=True, reconciled_at=timezone.now())


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def bank_transactions_list(request):
    tenant = _get_default_tenant(request)
    txns = BankTransaction.objects.filter(tenant=tenant)
    total_in = sum((t.amount for t in txns if t.amount > 0), Decimal("0.00"))
    total_out = sum((t.amount for t in txns if t.amount < 0), Decimal("0.00"))
    return render(request, "bank/bank_transactions.html", {
        "tenant": tenant, "txns": txns, "total_in": total_in,
        "total_out": total_out, "net": total_in + total_out,
    })


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def bank_transaction_add(request):
    tenant = _get_default_tenant(request)
    form = BankTransactionForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        t = form.save(commit=False)
        t.tenant = tenant
        t.save()
        messages.success(request, "Bank transaction added.")
        return redirect("bank_transactions_list")
    return render(request, "bank/bank_transaction_form.html", {"tenant": tenant, "form": form})


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def bank_transaction_import(request):
    """Import a bank statement CSV with columns: date, description, amount."""
    tenant = _get_default_tenant(request)
    summary = None
    if request.method == "POST" and request.FILES.get("file"):
        import csv as _csv, io
        raw = request.FILES["file"].read().decode("utf-8-sig", errors="replace")
        reader = _csv.DictReader(io.StringIO(raw))
        created, errors = 0, []
        for n, row in enumerate(reader, start=2):
            row = {(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}
            try:
                d = _parse_date(row.get("date"))
                if not d:
                    raise ValueError("invalid or missing date")
                amount = Decimal((row.get("amount") or "0").replace(",", ""))
                desc = row.get("description") or "(no description)"
            except (InvalidOperation, ValueError) as exc:
                errors.append((n, str(exc)))
                continue
            BankTransaction.objects.create(tenant=tenant, txn_date=d, description=desc[:255],
                                           amount=amount, reference=(row.get("reference") or "")[:100] or None)
            created += 1
        summary = {"created": created, "errors": errors}
        if created:
            messages.success(request, f"Imported {created} bank transaction(s).")
    return render(request, "bank/bank_import.html", {"tenant": tenant, "summary": summary})


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def bank_reconciliation(request):
    tenant = _get_default_tenant(request)
    txns = list(BankTransaction.objects.filter(tenant=tenant).select_related("matched_payment", "matched_expense"))

    if request.method == "POST":
        action = request.POST.get("action") or "save"
        if action == "auto":
            matched = 0
            for r in _recon_rows(tenant, [t for t in txns if not t.is_reconciled]):
                if len(r["candidates"]) == 1:
                    _apply_match(r["txn"], r["candidates"][0]["value"])
                    matched += 1
            messages.success(request, f"Auto-matched {matched} transaction(s).")
        else:
            for t in txns:
                _apply_match(t, request.POST.get(f"match_{t.id}", ""))
            messages.success(request, "Reconciliation saved.")
        return redirect("bank_reconciliation")

    rows = _recon_rows(tenant, txns)
    statement_balance = sum((t.amount for t in txns), Decimal("0.00"))
    cleared = sum((t.amount for t in txns if t.is_reconciled), Decimal("0.00"))
    bank_acc = GLAccount.objects.filter(tenant=tenant, code="1050").first()
    book_balance = Decimal("0.00")
    if bank_acc:
        book_balance = reports_service.account_balances(tenant).get(bank_acc, {}).get("balance", Decimal("0.00"))
    return render(request, "payments/bank_reconciliation.html", {
        "tenant": tenant, "rows": rows,
        "statement_balance": statement_balance, "cleared": cleared,
        "uncleared": statement_balance - cleared, "book_balance": book_balance,
        "difference": book_balance - cleared,
        "unreconciled_count": sum(1 for t in txns if not t.is_reconciled),
    })


# ============================
# VAT return (MTD 9-box)
# ============================

@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def vat_index(request):
    tenant = _get_default_tenant(request)
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    preview = None
    if date_from and date_to:
        preview = vat_service.compute_vat_return(tenant, date_from, date_to)
    returns = VatReturn.objects.filter(tenant=tenant)
    return render(request, "vat/index.html", {
        "tenant": tenant, "returns": returns, "preview": preview,
        "date_from": date_from, "date_to": date_to,
    })


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def vat_save(request):
    tenant = _get_default_tenant(request)
    if request.method != "POST":
        return redirect("vat_index")
    date_from = _parse_date(request.POST.get("from"))
    date_to = _parse_date(request.POST.get("to"))
    if not (date_from and date_to) or date_to < date_from:
        messages.error(request, "Please provide a valid period (from / to).")
        return redirect("vat_index")
    vr = vat_service.save_vat_return(tenant, date_from, date_to)
    log_audit(action="VAT_RETURN_SAVED", request=request, user=request.user, tenant=tenant,
              detail=f"{date_from} to {date_to}: net VAT {vr.box5_net_vat}")
    messages.success(request, "VAT return saved as draft.")
    return redirect("vat_detail", vr_id=vr.id)


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def vat_detail(request, vr_id):
    tenant = _get_default_tenant(request)
    vr = get_object_or_404(VatReturn, id=vr_id, tenant=tenant)
    return render(request, "vat/detail.html", {"tenant": tenant, "vr": vr})


@role_required([ROLE_FINANCE, ROLE_ADMIN], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
@transaction.atomic
def vat_submit(request, vr_id):
    tenant = _get_default_tenant(request)
    vr = get_object_or_404(VatReturn, id=vr_id, tenant=tenant)
    if request.method == "POST":
        vat_service.submit_vat_return(vr, user=request.user)
        log_audit(action="VAT_RETURN_SUBMITTED", request=request, user=request.user, tenant=tenant,
                  detail=f"{vr.period_from} to {vr.period_to} ({vr.hmrc_reference})")
        messages.warning(
            request,
            "Return marked submitted locally. Live HMRC MTD filing is not yet "
            "connected (needs HMRC credentials) - reference is a local stub.",
        )
    return redirect("vat_detail", vr_id=vr.id)


@role_required([ROLE_FINANCE, ROLE_ADMIN, ROLE_READONLY], write_groups=[ROLE_FINANCE, ROLE_ADMIN])
def vat_records(request):
    """Digital VAT records: every VAT-bearing transaction line in the period
    (the audit trail behind the return)."""
    tenant = _get_default_tenant(request)
    date_from = _parse_date(request.GET.get("from"))
    date_to = _parse_date(request.GET.get("to"))
    if not date_from and not date_to:
        date_from, date_to = reports_service.current_financial_year(tenant)
    elif not date_to:
        date_to = timezone.localdate()
    records = vat_service.vat_transactions(tenant, date_from, date_to)
    return render(request, "vat/records.html", {
        "tenant": tenant, "records": records, "date_from": date_from, "date_to": date_to,
        "export_qs": f"?from={date_from}&to={date_to}",
    })


# ============================
# Notifications & email log
# ============================

@login_required
def notifications_list(request):
    """Every notification for the signed-in user (newest first)."""
    from core.models import Notification
    notes = Notification.objects.filter(recipient=request.user)[:200]
    unread = sum(1 for n in notes if not n.is_read)
    return render(request, "notifications/list.html", {"notes": notes, "unread": unread})


@login_required
def notification_open(request, note_id):
    """Mark a notification read and bounce to its linked page (or the list)."""
    from core.models import Notification
    note = get_object_or_404(Notification, id=note_id, recipient=request.user)
    note.mark_read()
    return redirect(note.url or "notifications_list")


@login_required
def notification_mark_all_read(request):
    from core.models import Notification
    if request.method == "POST":
        Notification.objects.filter(recipient=request.user, is_read=False).update(
            is_read=True, read_at=timezone.now())
        messages.success(request, "All notifications marked as read.")
    return redirect("notifications_list")


@login_required
def notification_preferences(request):
    """Per-category in-app / email channel toggles for the signed-in user."""
    from core.models import Notification, NotificationPreference
    tenant = _get_default_tenant(request)
    categories = Notification.Category.choices
    if request.method == "POST":
        for code, _label in categories:
            in_app = bool(request.POST.get(f"in_app_{code}"))
            email = bool(request.POST.get(f"email_{code}"))
            NotificationPreference.objects.update_or_create(
                user=request.user, tenant=tenant, category=code,
                defaults={"in_app": in_app, "email": email},
            )
        messages.success(request, "Notification preferences saved.")
        return redirect("notification_preferences")
    existing = {p.category: p for p in NotificationPreference.objects.filter(
        user=request.user, tenant=tenant)}
    rows = []
    for code, label in categories:
        p = existing.get(code)
        rows.append({"code": code, "label": label,
                     "in_app": p.in_app if p else True,
                     "email": p.email if p else True})
    return render(request, "notifications/preferences.html", {"rows": rows})


@role_required([ROLE_ADMIN, ROLE_FINANCE], [ROLE_ADMIN, ROLE_FINANCE])
def email_log(request):
    """Audit trail of outbound emails for this organisation."""
    from core.models import EmailLog
    tenant = _get_default_tenant(request)
    logs = EmailLog.objects.filter(tenant=tenant)[:300]
    return render(request, "notifications/email_log.html", {"logs": logs})

