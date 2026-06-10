from decimal import Decimal

from django.contrib.auth.models import User, Group
from django.test import TestCase, Client

from core.models import (
    Tenant, Location, Supplier, Product, PurchaseOrder, PurchaseOrderLine,
    Shipment, ShipmentLine, InventoryBalance, InventoryMovement, GoodsReceipt,
    UserProfile, Customer, CustomerInvoice, CustomerInvoiceLine, TaxCode,
)
from core.services.gl import post_customer_invoice


class _CtxClient(Client):
    """Test client that auto-selects the user's default company + first
    accessible site on login, so the mandatory post-login context gate is
    satisfied. Mirrors the real auto-selection for single-context users; tests
    that exercise the gate/selection itself opt out via ``client_class = Client``.
    """

    def login(self, **credentials):
        ok = super().login(**credentials)
        if ok:
            self._select_default_context()
        return ok

    def _select_default_context(self):
        from core.access import (get_memberships, selectable_sites,
                                 SESSION_TENANT_KEY, SESSION_SITE_KEY)
        uid = self.session.get("_auth_user_id")
        if not uid:
            return
        u = User.objects.filter(pk=uid).first()
        if u is None:
            return
        memberships = get_memberships(u)
        if memberships:
            m = next((x for x in memberships if x.is_default), memberships[0])
            tenant = m.tenant
        else:
            prof = UserProfile.objects.filter(user=u).first()
            tenant = prof.tenant if prof else Tenant.objects.order_by("id").first()
        if tenant is None:
            return
        site = selectable_sites(u, tenant).first()
        s = self.session
        s[SESSION_TENANT_KEY] = tenant.id
        if site:
            s[SESSION_SITE_KEY] = site.id
        s.save()


# Apply suite-wide: every TestCase's self.client auto-selects a context on login.
TestCase.client_class = _CtxClient


class ReceivingFlowTests(TestCase):
    """Locks in the fix for the previously-broken PO receiving flow."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name="Acme")
        self.loc = Location.objects.create(tenant=self.tenant, name="Main WH")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Sup")
        self.product = Product.objects.create(tenant=self.tenant, sku="SKU-001", name="Widget")
        self.po = PurchaseOrder.objects.create(
            tenant=self.tenant, po_number="PO-1", supplier=self.supplier,
            status=PurchaseOrder.Status.SUBMITTED,
        )
        self.pol = PurchaseOrderLine.objects.create(
            po=self.po, product=self.product, ordered_qty=Decimal("10"), unit_cost=Decimal("2.50"),
        )
        self.shipment = Shipment.objects.create(
            tenant=self.tenant, po=self.po, from_supplier=self.supplier, destination=self.loc,
        )
        self.sl = ShipmentLine.objects.create(
            shipment=self.shipment, po_line=self.pol, expected_qty=Decimal("10"),
        )

        self.user = User.objects.create_user("wh", password="pw")
        wh_group, _ = Group.objects.get_or_create(name="Warehouse")
        self.user.groups.add(wh_group)
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="wh", password="pw")

    def test_receive_updates_inventory_and_ledger(self):
        resp = self.client.post(
            f"/po/{self.po.id}/receive/",
            {"grn_number": "GRN-TEST", f"recv_{self.sl.id}": "4"},
        )
        self.assertEqual(resp.status_code, 302)

        bal = InventoryBalance.objects.get(tenant=self.tenant, product=self.product, location=self.loc)
        self.assertEqual(bal.on_hand, Decimal("4.00"))

        mv = InventoryMovement.objects.get(tenant=self.tenant, product=self.product)
        self.assertEqual(mv.movement_type, InventoryMovement.MovementType.RECEIVE)
        self.assertEqual(mv.qty_delta, Decimal("4.00"))

        self.pol.refresh_from_db()
        self.assertEqual(self.pol.received_qty, Decimal("4.00"))
        self.po.refresh_from_db()
        self.assertEqual(self.po.status, PurchaseOrder.Status.PARTIALLY_RECEIVED)

        # Receipt capitalizes stock: DR Inventory / CR GRNI at 4 x 2.50 = 10.00
        from core.models import JournalEntry
        je = JournalEntry.objects.get(tenant=self.tenant, ref_type="GRN")
        self.assertEqual(je.total_debit, je.total_credit)
        self.assertEqual(je.total_debit, Decimal("10.00"))

    def test_over_receipt_is_rolled_back(self):
        self.client.post(f"/po/{self.po.id}/receive/", {f"recv_{self.sl.id}": "999"})
        # Whole transaction rolls back: no balance, no movement, no orphan GRN.
        self.assertFalse(InventoryBalance.objects.filter(tenant=self.tenant).exists())
        self.assertFalse(InventoryMovement.objects.filter(tenant=self.tenant).exists())
        self.assertFalse(GoodsReceipt.objects.filter(tenant=self.tenant).exists())

    def test_receive_requires_login(self):
        self.client.logout()
        resp = self.client.post(f"/po/{self.po.id}/receive/", {f"recv_{self.sl.id}": "1"})
        self.assertIn(resp.status_code, (302, 403))  # redirected to login or forbidden
        self.assertFalse(InventoryMovement.objects.filter(tenant=self.tenant).exists())


class TemplateRenderTests(TestCase):
    """Smoke-test that the redesigned templates render without errors."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name="Acme UI")
        self.user = User.objects.create_user("u", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Read-only")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)

    def test_login_page_renders(self):
        resp = self.client.get("/login/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "SwifPro")  # wordmark
        self.assertContains(resp, "Business intelligence")

    def test_landing_page_renders(self):
        self.client.login(username="u", password="pw")
        resp = self.client.get("/", follow=True)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "SwifPro BI")
        self.assertContains(resp, "Dashboard")

    def test_po_list_renders(self):
        self.user.groups.add(Group.objects.get_or_create(name="Admin")[0])
        self.client.login(username="u", password="pw")
        resp = self.client.get("/po/")
        self.assertEqual(resp.status_code, 200)


class PoAmendTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Amend Co")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="S")
        self.product = Product.objects.create(tenant=self.tenant, sku="SKU-1", name="P")
        self.po = PurchaseOrder.objects.create(
            tenant=self.tenant, po_number="PO-9", supplier=self.supplier,
            status=PurchaseOrder.Status.SUBMITTED, version=1,
        )
        PurchaseOrderLine.objects.create(po=self.po, product=self.product, ordered_qty=Decimal("5"), unit_cost=Decimal("1"))
        self.user = User.objects.create_user("p", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Admin")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="p", password="pw")

    def test_amend_creates_versioned_po(self):
        resp = self.client.post(f"/po/{self.po.id}/amend/", {"reason": "price change"})
        self.assertEqual(resp.status_code, 302)
        new = PurchaseOrder.objects.get(tenant=self.tenant, version=2)
        self.assertEqual(new.po_number, "PO-9-v2")
        self.assertTrue(new.is_current)
        self.po.refresh_from_db()
        self.assertFalse(self.po.is_current)


class FormTenantScopeTests(TestCase):
    def test_supplier_dropdown_scoped_to_current_tenant(self):
        from core.current import set_current_tenant, clear_current_tenant
        from core.forms import PurchaseOrderForm

        t_a = Tenant.objects.create(name="Tenant A")
        t_b = Tenant.objects.create(name="Tenant B")
        Supplier.objects.create(tenant=t_a, name="A Supplier")
        Supplier.objects.create(tenant=t_b, name="B Supplier")

        set_current_tenant(t_a)
        try:
            names = list(PurchaseOrderForm().fields["supplier"].queryset.values_list("name", flat=True))
        finally:
            clear_current_tenant()
        self.assertEqual(names, ["A Supplier"])


class LocationProfileTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Loc Co")
        Location.objects.filter(tenant=self.tenant).delete()  # drop auto-seeded Main Location; this class manages its own
        self.user = User.objects.create_user("locu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="locu", password="pw")

    def test_storage_unit_type_available(self):
        from core.models import Location
        self.assertIn("STORAGE", dict(Location.Type.choices))
        self.assertEqual(dict(Location.Type.choices)["STORAGE"], "Storage room")
        # New inventory-location types from the spec are available too.
        for t in ["SHOP_FLOOR", "BACK_ROOM", "COLD_STORAGE", "DAMAGED"]:
            self.assertIn(t, dict(Location.Type.choices))

    def test_create_with_all_fields(self):
        from core.models import Location
        resp = self.client.post("/locations/new/", {
            "name": "Camden Shop", "type": "STORE", "address": "1 High St\nLondon",
            "contact_person": "Sam", "phone": "+44 20 7946 0000", "email": "shop@x.example",
            "opening_hours": "Mon-Fri 9-5", "is_active": "on", "holds_stock": "on",
        })
        self.assertEqual(resp.status_code, 302)
        loc = Location.objects.get(tenant=self.tenant, name="Camden Shop")
        self.assertEqual(loc.type, "STORE")
        self.assertEqual(loc.contact_person, "Sam")
        self.assertEqual(loc.email, "shop@x.example")
        self.assertEqual(loc.opening_hours, "Mon-Fri 9-5")
        self.assertTrue(loc.is_active)
        self.assertTrue(loc.holds_stock)

    def test_inactive_or_nonstock_location_excluded_from_stock_form(self):
        from core.current import set_current_tenant, clear_current_tenant
        from core.forms import StockAdjustmentForm
        from core.models import Location
        Location.objects.create(tenant=self.tenant, name="Live WH", type="WAREHOUSE", is_active=True, holds_stock=True)
        Location.objects.create(tenant=self.tenant, name="Closed WH", type="WAREHOUSE", is_active=False, holds_stock=True)
        Location.objects.create(tenant=self.tenant, name="Head Office", type="OFFICE", is_active=True, holds_stock=False)
        set_current_tenant(self.tenant)
        try:
            names = set(StockAdjustmentForm().fields["location"].queryset.values_list("name", flat=True))
        finally:
            clear_current_tenant()
        self.assertEqual(names, {"Live WH"})


class CompanyGroupTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, CompanyGroup
        self.grp = CompanyGroup.objects.create(name="Acme Holdings")
        self.t1 = Tenant.objects.create(name="Acme UK", group=self.grp)
        self.t2 = Tenant.objects.create(name="Acme EU", group=self.grp)
        self.t3 = Tenant.objects.create(name="Other Co")  # no group
        self.user = User.objects.create_user("grpu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t1, role="ADMIN", is_default=True)
        OrgMembership.objects.create(user=self.user, tenant=self.t2, role="ADMIN")

    def test_group_companies_limited_to_membership(self):
        from core.access import group_companies
        from core.models import Tenant
        Tenant.objects.create(name="Acme US", group=self.grp)  # user not a member -> excluded
        names = {t.name for t in group_companies(self.user, self.t1)}
        self.assertEqual(names, {"Acme UK", "Acme EU"})

    def test_no_group_returns_self(self):
        from core.access import group_companies
        self.assertEqual([t.name for t in group_companies(self.user, self.t3)], ["Other Co"])

    def test_create_group_via_view(self):
        self.client.login(username="grpu", password="pw")
        resp = self.client.post("/settings/group/", {"op": "create", "name": "NewGroup"})
        self.assertEqual(resp.status_code, 302)
        self.t1.refresh_from_db()
        self.assertEqual(self.t1.group.name, "NewGroup")

    def test_group_page_renders(self):
        self.client.login(username="grpu", password="pw")
        self.assertEqual(self.client.get("/settings/group/").status_code, 200)

    def test_consolidated_sums_across_companies(self):
        from core.services import reports
        from core.services.gl import post_customer_invoice
        from core.models import Customer, CustomerInvoice, CustomerInvoiceLine, TaxCode
        for t, net in [(self.t1, "200"), (self.t2, "300")]:
            std = TaxCode.objects.get(tenant=t, code="STD")
            c = Customer.objects.create(tenant=t, name=f"C-{t.id}")
            inv = CustomerInvoice.objects.create(tenant=t, customer=c, invoice_number=f"INV-{t.id}")
            CustomerInvoiceLine.objects.create(invoice=inv, description="X", qty=Decimal("1"),
                                               unit_price=Decimal(net), tax_code=std)
            post_customer_invoice(inv)
        data = reports.consolidated([self.t1, self.t2])
        self.assertEqual(len(data["rows"]), 2)
        self.assertEqual(data["totals"]["revenue"], Decimal("500.00"))

    def test_consolidated_page_renders(self):
        self.client.login(username="grpu", password="pw")
        resp = self.client.get("/reports/consolidated/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["company_count"], 2)


class InterCompanyTradingTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, CompanyGroup
        self.grp = CompanyGroup.objects.create(name="Holdings")
        self.a = Tenant.objects.create(name="Company A", group=self.grp)
        self.b = Tenant.objects.create(name="Company B", group=self.grp)
        self.user = User.objects.create_user("icu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.a, role="ADMIN", is_default=True)
        OrgMembership.objects.create(user=self.user, tenant=self.b, role="ADMIN")

    def _bal(self, tenant, code):
        from core.models import GLAccount, JournalLine
        from django.db.models import Sum
        acc = GLAccount.objects.get(tenant=tenant, code=code)
        agg = JournalLine.objects.filter(account=acc).aggregate(d=Sum("debit"), c=Sum("credit"))
        return (agg["d"] or Decimal("0")) - (agg["c"] or Decimal("0"))

    def test_sale_posts_ar_in_seller_and_ap_in_buyer(self):
        from core.services.intercompany import create_intercompany_sale
        ict = create_intercompany_sale(self.a, self.b, Decimal("500"), "Mgmt fee", user=self.user)
        self.assertEqual(self._bal(self.a, "1100"), Decimal("500.00"))   # AR in A
        self.assertEqual(self._bal(self.a, "4000"), Decimal("-500.00"))  # sales in A
        self.assertTrue(ict.customer_invoice.is_intercompany)
        self.assertEqual(self._bal(self.b, "6000"), Decimal("500.00"))   # expense in B
        self.assertEqual(self._bal(self.b, "2000"), Decimal("-500.00"))  # AP in B
        self.assertTrue(ict.expense.is_intercompany)

    def test_consolidation_eliminates_intragroup(self):
        from core.services.intercompany import create_intercompany_sale
        from core.services import reports
        create_intercompany_sale(self.a, self.b, Decimal("500"), "fee", user=self.user)
        data = reports.consolidated([self.a, self.b])
        self.assertEqual(data["totals"]["eliminations"], Decimal("500.00"))
        self.assertEqual(data["totals"]["revenue"] - data["totals"]["net_revenue"], Decimal("500.00"))

    def test_create_via_view_and_validation(self):
        from core.models import InterCompanyTransaction
        self.client.login(username="icu", password="pw")
        resp = self.client.post("/intercompany/", {"to_tenant": self.b.id, "amount": "250", "description": "svc"})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(InterCompanyTransaction.objects.filter(from_tenant=self.a, to_tenant=self.b).count(), 1)
        self.client.post("/intercompany/", {"to_tenant": self.b.id, "amount": "0"})  # rejected
        self.assertEqual(InterCompanyTransaction.objects.count(), 1)


class SiteTierTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Site Co")
        self.user = User.objects.create_user("siteu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="siteu", password="pw")

    def test_create_site_and_assign_location(self):
        from core.models import Site, Location
        resp = self.client.post("/sites/new/", {"name": "Manchester Plant", "code": "MCR", "is_active": "on"})
        self.assertEqual(resp.status_code, 302)
        site = Site.objects.get(tenant=self.tenant, name="Manchester Plant")
        resp = self.client.post("/locations/new/", {
            "name": "MCR Warehouse", "site": site.id, "type": "WAREHOUSE",
            "is_active": "on", "holds_stock": "on",
        })
        self.assertEqual(resp.status_code, 302)
        loc = Location.objects.get(tenant=self.tenant, name="MCR Warehouse")
        self.assertEqual(loc.site_id, site.id)

    def test_site_list_renders_with_locations(self):
        from core.models import Site, Location
        site = Site.objects.create(tenant=self.tenant, name="HQ")
        Location.objects.create(tenant=self.tenant, name="HQ WH", type="WAREHOUSE", site=site)
        resp = self.client.get("/sites/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "HQ WH")

    def test_site_scoped_to_tenant_in_location_form(self):
        from core.current import set_current_tenant, clear_current_tenant
        from core.forms import LocationForm
        from core.models import Site, Tenant as T
        Site.objects.create(tenant=self.tenant, name="Mine")
        other = T.objects.create(name="Other")
        Site.objects.create(tenant=other, name="Theirs")
        set_current_tenant(self.tenant)
        try:
            names = set(LocationForm().fields["site"].queryset.values_list("name", flat=True))
        finally:
            clear_current_tenant()
        self.assertEqual(names, {"Main Site", "Mine"})  # tenant's sites only (auto + own), not "Theirs"


class BinTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, Location
        self.tenant = Tenant.objects.create(name="Bin Co")
        self.wh = Location.objects.create(tenant=self.tenant, name="WH", type="WAREHOUSE", is_active=True, holds_stock=True)
        self.user = User.objects.create_user("binu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="binu", password="pw")

    def test_create_bin(self):
        from core.models import Bin
        resp = self.client.post("/bins/new/", {"location": self.wh.id, "code": "A-01", "is_active": "on"})
        self.assertEqual(resp.status_code, 302)
        b = Bin.objects.get(tenant=self.tenant)
        self.assertEqual(b.code, "A-01")
        self.assertEqual(b.location_id, self.wh.id)

    def test_adjustment_records_bin_on_movement(self):
        from core.models import Product, Bin, StockAdjustment, InventoryMovement
        p = Product.objects.create(tenant=self.tenant, sku="B1", name="P")
        b = Bin.objects.create(tenant=self.tenant, location=self.wh, code="A-02")
        resp = self.client.post("/inventory/adjustments/new/", {
            "product": p.id, "location": self.wh.id, "bin": b.id,
            "reason": "ADJUSTMENT", "qty_delta": "5",
        })
        self.assertEqual(resp.status_code, 302)
        adj = StockAdjustment.objects.get(tenant=self.tenant)
        self.assertEqual(adj.bin_id, b.id)
        mv = InventoryMovement.objects.get(tenant=self.tenant, product=p)
        self.assertEqual(mv.bin_id, b.id)

    def test_bin_must_match_location(self):
        from core.models import Product, Bin, Location, StockAdjustment
        p = Product.objects.create(tenant=self.tenant, sku="B2", name="P")
        other = Location.objects.create(tenant=self.tenant, name="WH2", type="WAREHOUSE", is_active=True, holds_stock=True)
        b = Bin.objects.create(tenant=self.tenant, location=other, code="X-01")
        resp = self.client.post("/inventory/adjustments/new/", {
            "product": p.id, "location": self.wh.id, "bin": b.id,
            "reason": "ADJUSTMENT", "qty_delta": "5",
        })
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Bin must belong to the chosen location")
        self.assertEqual(StockAdjustment.objects.filter(tenant=self.tenant).count(), 0)


class LocationAccessTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, Location, InventoryBalance, Product
        self.tenant = Tenant.objects.create(name="Acc Co")
        self.wh1 = Location.objects.create(tenant=self.tenant, name="WH1", type="WAREHOUSE")
        self.wh2 = Location.objects.create(tenant=self.tenant, name="WH2", type="WAREHOUSE")
        self.p = Product.objects.create(tenant=self.tenant, sku="A1", name="P")
        InventoryBalance.objects.create(tenant=self.tenant, product=self.p, location=self.wh1, on_hand=Decimal("10"))
        InventoryBalance.objects.create(tenant=self.tenant, product=self.p, location=self.wh2, on_hand=Decimal("5"))
        self.admin = User.objects.create_user("accadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.wh = User.objects.create_user("accwh", password="pw")
        OrgMembership.objects.create(user=self.wh, tenant=self.tenant, role="WAREHOUSE", is_default=True)

    def test_unrestricted_by_default(self):
        from core.access import accessible_location_ids
        self.assertIsNone(accessible_location_ids(self.wh, self.tenant))  # no grants -> all

    def test_admin_always_unrestricted(self):
        from core.access import accessible_location_ids
        from core.models import UserLocationAccess
        UserLocationAccess.objects.create(tenant=self.tenant, user=self.admin, location=self.wh1)
        self.assertIsNone(accessible_location_ids(self.admin, self.tenant))  # admin -> all despite a grant

    def test_grant_restricts_inventory_and_reports(self):
        from core.models import UserLocationAccess
        from core.services import reports
        from core.access import accessible_location_ids
        UserLocationAccess.objects.create(tenant=self.tenant, user=self.wh, location=self.wh1)
        self.client.login(username="accwh", password="pw")
        resp = self.client.get("/inventory/")
        self.assertEqual(resp.status_code, 200)
        locs = {b.location.name for b in resp.context["balances"]}
        self.assertEqual(locs, {"WH1"})
        allowed = accessible_location_ids(self.wh, self.tenant)
        data = reports.stock_valuation(self.tenant, location_ids=allowed)
        qty = sum(r["qty"] for r in data["rows"])
        self.assertEqual(qty, Decimal("10"))

    def test_access_matrix_admin_only_and_saves(self):
        from core.models import UserLocationAccess
        self.client.login(username="accwh", password="pw")
        self.assertEqual(self.client.get("/locations/access/").status_code, 403)
        self.client.login(username="accadmin", password="pw")
        self.assertEqual(self.client.get("/locations/access/").status_code, 200)
        resp = self.client.post("/locations/access/", {f"grant_{self.wh.id}_{self.wh1.id}": "on"})
        self.assertEqual(resp.status_code, 302)
        grants = set(UserLocationAccess.objects.filter(tenant=self.tenant, user=self.wh).values_list("location_id", flat=True))
        self.assertEqual(grants, {self.wh1.id})


class GLBalanceTests(TestCase):
    def test_supplier_invoice_journal_balances_with_vat(self):
        from core.models import (
            GoodsReceipt, GoodsReceiptLine, Location, SupplierInvoice, SupplierInvoiceLine,
        )
        from core.services.gl import post_supplier_invoice

        tenant = Tenant.objects.create(name="AP Co")
        supplier = Supplier.objects.create(tenant=tenant, name="Sup")
        product = Product.objects.create(tenant=tenant, sku="SKU-AP", name="P")
        loc = Location.objects.create(tenant=tenant, name="WH")
        po = PurchaseOrder.objects.create(tenant=tenant, po_number="PO-AP", supplier=supplier)
        grn = GoodsReceipt.objects.create(tenant=tenant, po=po, grn_number="GRN-AP", received_to=loc, status=GoodsReceipt.Status.POSTED)
        inv = SupplierInvoice.objects.create(tenant=tenant, supplier=supplier, po=po, receipt=grn, invoice_number="SINV-1")
        std = TaxCode.objects.get(tenant=tenant, code="STD")
        SupplierInvoiceLine.objects.create(invoice=inv, product=product, qty=Decimal("10"), unit_cost=Decimal("5.00"), tax_code=std)

        je = post_supplier_invoice(inv)
        self.assertEqual(je.total_debit, je.total_credit)
        self.assertEqual(je.total_credit, Decimal("60.00"))  # 50 net + 20% VAT input

    def test_customer_invoice_journal_balances(self):
        tenant = Tenant.objects.create(name="Acme2")  # signal bootstraps GL accounts + tax codes
        customer = Customer.objects.create(tenant=tenant, name="Cust")
        inv = CustomerInvoice.objects.create(tenant=tenant, customer=customer, invoice_number="INV-1")
        std = TaxCode.objects.get(tenant=tenant, code="STD")
        CustomerInvoiceLine.objects.create(
            invoice=inv, description="Item", qty=Decimal("2"), unit_price=Decimal("100.00"), tax_code=std,
        )

        je = post_customer_invoice(inv)

        self.assertEqual(je.total_debit, je.total_credit)
        self.assertEqual(je.total_debit, Decimal("240.00"))  # 200 net + 20% VAT
        inv.refresh_from_db()
        self.assertEqual(inv.status, "ISSUED")


class RoleDashboardTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Role Co")
        self.user = User.objects.create_user("salesuser", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="SALES", is_default=True)

    def test_login_redirects_to_role_dashboard(self):
        self.client.login(username="salesuser", password="pw")
        resp = self.client.get("/")
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, "/dashboard/sales")

    def test_role_dashboard_renders(self):
        self.client.login(username="salesuser", password="pw")
        resp = self.client.get("/dashboard/sales")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Sales Dashboard")

    def test_cross_role_dashboard_is_forbidden_and_audited(self):
        from core.models import AuditLog
        self.client.login(username="salesuser", password="pw")
        resp = self.client.get("/dashboard/finance")
        self.assertEqual(resp.status_code, 403)
        self.assertContains(resp, "Access denied", status_code=403)
        self.assertTrue(AuditLog.objects.filter(action="ACCESS_DENIED").exists())

    def test_admin_can_view_any_dashboard(self):
        from core.models import OrgMembership
        admin = User.objects.create_user("owneruser", password="pw")
        OrgMembership.objects.create(user=admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="owneruser", password="pw")
        self.assertEqual(self.client.get("/dashboard/warehouse").status_code, 200)
        self.assertEqual(self.client.get("/dashboard/finance").status_code, 200)

    def test_multi_org_redirects_to_picker(self):
        from core.models import OrgMembership
        t2 = Tenant.objects.create(name="Org Two")
        OrgMembership.objects.create(user=self.user, tenant=t2, role="ACCOUNTANT")
        self.client.login(username="salesuser", password="pw")
        # Simulate a fresh login with no context chosen yet (the auto-context
        # client pre-selects one; clear it to exercise the multi-company gate).
        s = self.client.session
        s.pop("active_tenant_id", None)
        s.pop("active_location_id", None)
        s.save()
        resp = self.client.get("/")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/select-org/", resp.url)

    def test_login_is_audited(self):
        from core.models import AuditLog
        self.client.post("/login/", {"username": "salesuser", "password": "pw"})
        self.assertTrue(AuditLog.objects.filter(action="LOGIN", username="salesuser").exists())

    def test_logout_via_post(self):
        self.client.login(username="salesuser", password="pw")
        resp = self.client.post("/logout/")
        self.assertEqual(resp.status_code, 302)
        # session cleared -> protected page now redirects to login
        self.assertEqual(self.client.get("/dashboard/sales").status_code, 302)

    def test_role_landing_override(self):
        self.tenant.role_landing = {"SALES": "reports_index"}
        self.tenant.save()
        self.client.login(username="salesuser", password="pw")
        resp = self.client.get("/")
        self.assertEqual(resp.url, "/reports/")


class DashboardKpiTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="KPI Co")
        self.admin = User.objects.create_user("kpiadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)

    def test_each_role_dashboard_has_kpis(self):
        self.client.login(username="kpiadmin", password="pw")
        for path in ["admin", "finance", "sales", "warehouse", "purchasing", "accountant", "read-only"]:
            resp = self.client.get(f"/dashboard/{path}")
            self.assertEqual(resp.status_code, 200, path)
            self.assertTrue(len(resp.context["kpis"]) >= 3, f"{path} has too few KPIs")

    def test_kpis_reflect_data(self):
        from core.models import Product, InventoryBalance, Location, PurchaseOrder, PurchaseOrderLine, Supplier
        sup = Supplier.objects.create(tenant=self.tenant, name="S")
        loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        p = Product.objects.create(tenant=self.tenant, sku="K1", name="P", reorder_level=Decimal("10"), is_active=True)
        InventoryBalance.objects.create(tenant=self.tenant, product=p, location=loc, on_hand=Decimal("2"))
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-K", supplier=sup,
                                          status=PurchaseOrder.Status.SENT)
        PurchaseOrderLine.objects.create(po=po, product=p, ordered_qty=Decimal("5"), unit_cost=Decimal("1"))
        self.client.login(username="kpiadmin", password="pw")
        kpis = {k["label"]: k["value"] for k in self.client.get("/dashboard/purchasing").context["kpis"]}
        self.assertEqual(kpis["Open purchase orders"], 1)
        self.assertEqual(kpis["Low-stock items"], 1)


class CompanyProfileTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Profile Co")
        self.user = User.objects.create_user("padmin", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="padmin", password="pw")

    def _base_post(self, **overrides):
        data = {
            "name": "Profile Co", "legal_name": "Profile Co Ltd", "trading_name": "Profile",
            "business_type": "LTD", "company_number": "12345678", "utr_number": "1234567890",
            "vat_number": "GB123456789",
            "address_line1": "1 High St", "address_city": "Manchester", "address_postcode": "M1 2AB",
            "address_country": "United Kingdom",
            "billing_same_as_business": "on", "billing_country": "United Kingdom",
            "email": "ops@profile.test", "phone": "+44 20 7946 0000", "website": "https://profile.test",
            "currency_code": "GBP", "country": "United Kingdom", "timezone": "Europe/London",
            "financial_year_start_month": "4", "default_payment_terms_days": "30",
            "po_approval_threshold": "0",
        }
        data.update(overrides)
        return data

    def test_settings_page_renders(self):
        resp = self.client.get("/settings/tenant/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Company Profile")

    def test_save_full_profile(self):
        resp = self.client.post("/settings/tenant/", self._base_post())
        self.assertEqual(resp.status_code, 302)
        self.tenant.refresh_from_db()
        self.assertEqual(self.tenant.business_type, "LTD")
        self.assertEqual(self.tenant.legal_name, "Profile Co Ltd")
        self.assertEqual(self.tenant.financial_year_start_month, 4)
        self.assertEqual(self.tenant.default_payment_terms_days, 30)

    def test_invalid_company_number_rejected(self):
        resp = self.client.post("/settings/tenant/", self._base_post(company_number="ABC"))
        self.assertEqual(resp.status_code, 200)  # re-rendered with errors
        self.assertContains(resp, "valid UK company number")

    def test_invalid_vat_number_rejected(self):
        resp = self.client.post("/settings/tenant/", self._base_post(vat_number="12"))
        self.assertContains(resp, "valid UK VAT number")

    def test_vat_required_when_registered(self):
        resp = self.client.post("/settings/tenant/", self._base_post(vat_registered="on", vat_number=""))
        self.assertContains(resp, "VAT number is required")

    def test_required_fields_enforced(self):
        # Blank legal name / address must be rejected (re-rendered, not saved).
        resp = self.client.post("/settings/tenant/", self._base_post(legal_name="", address_line1=""))
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(resp.context["form"].is_valid())


class UserManagementTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Members Co")
        self.admin = User.objects.create_user("mgadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.bob = User.objects.create_user("bob", password="pw")
        self.bob_m = OrgMembership.objects.create(user=self.bob, tenant=self.tenant, role="SALES", is_default=True)
        self.client.login(username="mgadmin", password="pw")

    def test_members_list_renders(self):
        resp = self.client.get("/users/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "bob")

    def test_change_role_audited(self):
        from core.models import OrgMembership, AuditLog
        resp = self.client.post(f"/users/{self.bob_m.id}/role/", {"role": "WAREHOUSE"})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(OrgMembership.objects.get(id=self.bob_m.id).role, "WAREHOUSE")
        self.assertTrue(AuditLog.objects.filter(action="ROLE_CHANGED").exists())

    def test_deactivate_and_remove(self):
        from core.models import OrgMembership, AuditLog
        self.client.post(f"/users/{self.bob_m.id}/active/")
        self.bob.refresh_from_db()
        self.assertFalse(self.bob.is_active)
        self.assertTrue(AuditLog.objects.filter(action="USER_DEACTIVATED").exists())
        self.client.post(f"/users/{self.bob_m.id}/remove/")
        self.assertFalse(OrgMembership.objects.filter(id=self.bob_m.id).exists())
        self.assertTrue(AuditLog.objects.filter(action="USER_REMOVED").exists())

    def test_cannot_remove_last_admin(self):
        from core.models import OrgMembership
        admin_m = OrgMembership.objects.get(user=self.admin, tenant=self.tenant)
        resp = self.client.post(f"/users/{admin_m.id}/remove/")
        self.assertTrue(OrgMembership.objects.filter(id=admin_m.id).exists())  # blocked

    def test_non_admin_blocked(self):
        c = Client(); c.login(username="bob", password="pw")
        self.assertEqual(c.get("/users/").status_code, 403)


class AuditTrailPhase3Tests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, Product
        self.tenant = Tenant.objects.create(name="Audit Co")
        self.admin = User.objects.create_user("auadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.sales = User.objects.create_user("ausales", password="pw")
        OrgMembership.objects.create(user=self.sales, tenant=self.tenant, role="SALES", is_default=True)
        self.product = Product.objects.create(tenant=self.tenant, sku="SKU-DEL", name="Doomed")

    def test_record_delete_audited(self):
        from core.models import AuditLog
        self.client.login(username="auadmin", password="pw")
        resp = self.client.post(f"/products/{self.product.id}/delete/")
        self.assertEqual(resp.status_code, 302)
        log = AuditLog.objects.filter(action="RECORD_DELETED").first()
        self.assertIsNotNone(log)
        self.assertIn("SKU-DEL", log.detail)

    def test_data_export_csv_and_audit(self):
        from core.models import AuditLog
        self.client.login(username="auadmin", password="pw")
        resp = self.client.get("/export/products.csv")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "text/csv")
        self.assertIn("SKU-DEL", resp.content.decode())
        self.assertTrue(AuditLog.objects.filter(action="DATA_EXPORTED").exists())

    def test_data_export_blocked_without_permission(self):
        self.client.login(username="ausales", password="pw")  # SALES lacks export_data
        self.assertEqual(self.client.get("/export/products.csv").status_code, 403)

    def test_audit_log_export_admin_only(self):
        self.client.login(username="ausales", password="pw")
        self.assertEqual(self.client.get("/audit/export.csv").status_code, 403)
        self.client.login(username="auadmin", password="pw")
        resp = self.client.get("/audit/export.csv")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("timestamp", resp.content.decode())

    def test_password_change_audited(self):
        from core.models import AuditLog
        self.client.login(username="auadmin", password="pw")
        resp = self.client.post("/account/password/", {
            "old_password": "pw",
            "new_password1": "Str0ng-Pass-99",
            "new_password2": "Str0ng-Pass-99",
        })
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(AuditLog.objects.filter(action="PASSWORD_CHANGED").exists())
        self.assertTrue(self.client.login(username="auadmin", password="Str0ng-Pass-99"))


class AuditSoftDeleteTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, GLAccount
        self.tenant = Tenant.objects.create(name="SoftDel Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.admin = User.objects.create_user("sdadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="sdadmin", password="pw")

    def _issued_invoice(self, number, net="200"):
        from core.models import CustomerInvoice, CustomerInvoiceLine, Customer
        from core.services.gl import post_customer_invoice
        cust = Customer.objects.create(tenant=self.tenant, name=f"C-{number}")
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=cust, invoice_number=number)
        CustomerInvoiceLine.objects.create(invoice=inv, description="X", qty=Decimal("1"),
                                           unit_price=Decimal(net), tax_code=self.std)
        post_customer_invoice(inv)
        return inv

    def test_product_cost_change_audited(self):
        from core.models import Product, AuditLog
        p = Product.objects.create(tenant=self.tenant, sku="PC1", name="W", standard_cost=Decimal("5.00"))
        resp = self.client.post(f"/products/{p.id}/edit/", {
            "sku": "PC1", "name": "W", "product_type": "STOCK", "uom": "each",
            "cost_method": "AVERAGE", "standard_cost": "8.50", "sales_price": "0", "reorder_level": "0",
        })
        self.assertEqual(resp.status_code, 302)
        log = AuditLog.objects.get(action="PRODUCT_COST_CHANGED")
        self.assertEqual(log.entity_id, "PC1")
        self.assertEqual(log.old_value, "5.00")
        self.assertEqual(log.new_value, "8.50")

    def test_draft_invoice_soft_delete(self):
        from core.models import CustomerInvoice, Customer, AuditLog
        cust = Customer.objects.create(tenant=self.tenant, name="Draft Co")
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=cust, invoice_number="INV-D1")
        resp = self.client.post(f"/ar/invoices/{inv.id}/delete/")
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(CustomerInvoice.objects.filter(id=inv.id).exists())          # hidden
        self.assertTrue(CustomerInvoice.all_objects.filter(id=inv.id, is_deleted=True).exists())  # kept
        self.assertTrue(AuditLog.objects.filter(action="INVOICE_DELETED").exists())

    def test_posted_invoice_cannot_be_deleted(self):
        from core.models import CustomerInvoice
        inv = self._issued_invoice("INV-P1")
        self.client.post(f"/ar/invoices/{inv.id}/delete/")
        self.assertTrue(CustomerInvoice.objects.filter(id=inv.id).exists())  # still there, not deleted

    def test_payment_soft_delete_reverses_gl_and_reopens_invoice(self):
        from core.models import CustomerInvoice, Payment, PaymentAllocation, JournalEntry, AuditLog
        from core.services.gl import post_payment
        inv = self._issued_invoice("INV-PAY")  # total 240
        pay = Payment.objects.create(tenant=self.tenant, customer=inv.customer,
                                     direction=Payment.Direction.RECEIPT, amount=Decimal("240"),
                                     payment_date="2026-06-01")
        PaymentAllocation.objects.create(payment=pay, customer_invoice=inv, amount=Decimal("240"))
        post_payment(pay)
        inv.refresh_from_db()
        self.assertEqual(inv.status, CustomerInvoice.Status.PAID)

        resp = self.client.post(f"/payments/{pay.id}/delete/")
        self.assertEqual(resp.status_code, 302)
        # payment hidden but retained
        self.assertFalse(Payment.objects.filter(id=pay.id).exists())
        self.assertTrue(Payment.all_objects.filter(id=pay.id, is_deleted=True).exists())
        # reversing JE posted + invoice re-opened with full outstanding
        self.assertTrue(JournalEntry.objects.filter(tenant=self.tenant, ref_type="PAYMENT_REVERSAL", ref_id=str(pay.id)).exists())
        inv.refresh_from_db()
        self.assertEqual(inv.status, CustomerInvoice.Status.SENT)
        self.assertEqual(inv.outstanding, Decimal("240.00"))
        self.assertTrue(AuditLog.objects.filter(action="PAYMENT_DELETED").exists())


class AuditLogStructuredTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="AuditS Co")
        self.admin = User.objects.create_user("asadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.sales = User.objects.create_user("assales", password="pw")
        OrgMembership.objects.create(user=self.sales, tenant=self.tenant, role="SALES", is_default=True)

    def test_log_audit_captures_structured_fields_and_user_agent(self):
        from core.audit import log_audit
        from core.models import AuditLog

        class FakeReq:
            path = "/x/"
            META = {"REMOTE_ADDR": "10.0.0.5", "HTTP_USER_AGENT": "Mozilla/5.0 TestBrowser"}
        log_audit(action="UPDATE", request=FakeReq(), user=self.admin, tenant=self.tenant,
                  entity_type="Product", entity_id="42", old_value="5.00", new_value="7.50",
                  detail="cost change")
        log = AuditLog.objects.get(action="UPDATE")
        self.assertEqual(log.entity_type, "Product")
        self.assertEqual(log.entity_id, "42")
        self.assertEqual(log.old_value, "5.00")
        self.assertEqual(log.new_value, "7.50")
        self.assertEqual(log.ip, "10.0.0.5")
        self.assertEqual(log.user_agent, "Mozilla/5.0 TestBrowser")
        self.assertEqual(log.change_summary, "5.00 → 7.50")

    def test_audit_log_is_immutable(self):
        from core.models import AuditLog
        log = AuditLog.objects.create(tenant=self.tenant, action="LOGIN", username="x")
        log.action = "TAMPERED"
        with self.assertRaises(ValueError):
            log.save()
        log.refresh_from_db()
        self.assertEqual(log.action, "LOGIN")

    def test_viewer_filters_by_action_and_is_admin_only(self):
        from core.audit import log_audit
        log_audit(action="LOGIN", user=self.admin, tenant=self.tenant)
        log_audit(action="DATA_EXPORTED", user=self.admin, tenant=self.tenant)
        # Non-admin cannot view.
        self.client.login(username="assales", password="pw")
        self.assertEqual(self.client.get("/audit/").status_code, 403)
        # Admin can, and the action filter narrows results.
        self.client.login(username="asadmin", password="pw")
        resp = self.client.get("/audit/?action=DATA_EXPORTED")
        self.assertEqual(resp.status_code, 200)
        actions = {l.action for l in resp.context["logs"]}
        self.assertEqual(actions, {"DATA_EXPORTED"})


class UserPermissionOverrideTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, UserPermissionOverride
        self.tenant = Tenant.objects.create(name="Grants Co")
        self.admin = User.objects.create_user("gradmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.wh = User.objects.create_user("grwh", password="pw")
        self.wh_m = OrgMembership.objects.create(user=self.wh, tenant=self.tenant, role="WAREHOUSE", is_default=True)

    def test_effective_permissions_grant_and_revoke(self):
        from core import permissions as P
        # baseline: WAREHOUSE lacks view_finance_reports, has manage_inventory
        base = P.role_permissions("WAREHOUSE")
        self.assertNotIn(P.VIEW_FINANCE_REPORTS, base)
        eff = P.effective_permissions("WAREHOUSE", {P.VIEW_FINANCE_REPORTS: P.GRANT, P.MANAGE_INVENTORY: P.REVOKE})
        self.assertIn(P.VIEW_FINANCE_REPORTS, eff)
        self.assertNotIn(P.MANAGE_INVENTORY, eff)

    def test_admin_always_full_regardless_of_overrides(self):
        from core import permissions as P
        eff = P.effective_permissions("ADMIN", {P.MANAGE_USERS: P.REVOKE})
        self.assertEqual(eff, set(P.ALL_PERMISSIONS))

    def test_grant_enables_gated_view(self):
        from core.models import UserPermissionOverride
        from core import permissions as P
        # WAREHOUSE lacks export_data -> export blocked
        self.client.login(username="grwh", password="pw")
        self.assertEqual(self.client.get("/export/products.csv").status_code, 403)
        # grant export_data -> now allowed
        UserPermissionOverride.objects.create(tenant=self.tenant, user=self.wh,
                                              permission=P.EXPORT_DATA, effect=UserPermissionOverride.GRANT)
        self.assertEqual(self.client.get("/export/products.csv").status_code, 200)

    def test_editor_saves_overrides_and_audits(self):
        from core.models import UserPermissionOverride, AuditLog
        from core import permissions as P
        self.client.login(username="gradmin", password="pw")
        # WAREHOUSE baseline perms that should stay ticked, plus grant export_data
        data = {f"perm_{c}": "on" for c in P.role_permissions("WAREHOUSE")}
        data[f"perm_{P.EXPORT_DATA}"] = "on"
        resp = self.client.post(f"/users/{self.wh_m.id}/permissions/", data)
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(UserPermissionOverride.objects.filter(
            tenant=self.tenant, user=self.wh, permission=P.EXPORT_DATA,
            effect=UserPermissionOverride.GRANT).exists())
        self.assertTrue(AuditLog.objects.filter(action="PERMISSION_CHANGED").exists())

    def test_role_change_clears_overrides_by_default(self):
        from core.models import UserPermissionOverride
        from core import permissions as P
        UserPermissionOverride.objects.create(tenant=self.tenant, user=self.wh,
                                              permission=P.EXPORT_DATA, effect=UserPermissionOverride.GRANT)
        self.client.login(username="gradmin", password="pw")
        self.client.post(f"/users/{self.wh_m.id}/role/", {"role": "SALES"})
        self.assertFalse(UserPermissionOverride.objects.filter(tenant=self.tenant, user=self.wh).exists())

    def test_role_change_keeps_overrides_when_policy_on(self):
        from core.models import UserPermissionOverride
        from core import permissions as P
        self.tenant.keep_permissions_on_role_change = True
        self.tenant.save()
        # A meaningful grant (SALES lacks export_data) and a redundant one
        # (SALES already has manage_customers, so granting it is redundant).
        UserPermissionOverride.objects.create(tenant=self.tenant, user=self.wh,
                                              permission=P.EXPORT_DATA, effect=UserPermissionOverride.GRANT)
        UserPermissionOverride.objects.create(tenant=self.tenant, user=self.wh,
                                              permission=P.MANAGE_CUSTOMERS, effect=UserPermissionOverride.GRANT)
        self.client.login(username="gradmin", password="pw")
        self.client.post(f"/users/{self.wh_m.id}/role/", {"role": "SALES"})
        remaining = set(UserPermissionOverride.objects.filter(tenant=self.tenant, user=self.wh)
                        .values_list("permission", flat=True))
        self.assertEqual(remaining, {P.EXPORT_DATA})  # redundant grant pruned, meaningful one kept

    def test_policy_toggle_saves_and_audits(self):
        from core.models import AuditLog
        self.client.login(username="gradmin", password="pw")
        resp = self.client.post("/team/permissions/", {"keep_permissions_on_role_change": "on"})
        self.assertEqual(resp.status_code, 302)
        self.tenant.refresh_from_db()
        self.assertTrue(self.tenant.keep_permissions_on_role_change)
        self.assertTrue(AuditLog.objects.filter(action="SETTINGS_CHANGED").exists())

    def test_reset_to_role_default(self):
        from core.models import UserPermissionOverride
        from core import permissions as P
        UserPermissionOverride.objects.create(tenant=self.tenant, user=self.wh,
                                              permission=P.EXPORT_DATA, effect=UserPermissionOverride.GRANT)
        self.client.login(username="gradmin", password="pw")
        self.client.post(f"/users/{self.wh_m.id}/permissions/", {"reset": "1"})
        self.assertFalse(UserPermissionOverride.objects.filter(tenant=self.tenant, user=self.wh).exists())


class PermissionMatrixTests(TestCase):
    def test_matrix_helpers(self):
        from core import permissions as P
        self.assertTrue(P.role_has_permission("ADMIN", P.DELETE_RECORDS))
        self.assertTrue(P.role_has_permission("ACCOUNTANT", P.VIEW_FINANCE_REPORTS))
        self.assertFalse(P.role_has_permission("READONLY", P.MANAGE_INVOICES))
        self.assertFalse(P.role_has_permission("SALES", P.EXPORT_DATA))
        self.assertTrue(P.role_has_permission("ADMIN", P.MANAGE_USERS))

    def test_matrix_page_admin_only(self):
        from core.models import OrgMembership
        t = Tenant.objects.create(name="Perm Co")
        admin = User.objects.create_user("permadmin", password="pw")
        OrgMembership.objects.create(user=admin, tenant=t, role="ADMIN", is_default=True)
        sales = User.objects.create_user("permsales", password="pw")
        OrgMembership.objects.create(user=sales, tenant=t, role="SALES", is_default=True)

        c = Client(); c.login(username="permadmin", password="pw")
        resp = c.get("/team/permissions/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Roles &amp; Permissions")
        self.assertContains(resp, "Approve transactions")

        c2 = Client(); c2.login(username="permsales", password="pw")
        self.assertEqual(c2.get("/team/permissions/").status_code, 403)


class TeamInviteTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Invite Co")
        self.admin = User.objects.create_user("invadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="invadmin", password="pw")

    def test_invite_creates_user_and_membership_and_emails(self):
        from django.core import mail
        from core.models import OrgMembership
        from django.contrib.auth.models import User as U
        resp = self.client.post("/team/invite/", {"name": "Pat Jones", "email": "pat@team.test", "role": "WAREHOUSE"})
        self.assertEqual(resp.status_code, 302)
        u = U.objects.get(email="pat@team.test")
        self.assertTrue(OrgMembership.objects.filter(user=u, tenant=self.tenant, role="WAREHOUSE").exists())
        self.assertTrue(any("pat@team.test" in m.to for m in mail.outbox))

    def test_invite_requires_admin(self):
        from core.models import OrgMembership
        u = User.objects.create_user("notadmin", password="pw")
        OrgMembership.objects.create(user=u, tenant=self.tenant, role="SALES", is_default=True)
        c = Client(); c.login(username="notadmin", password="pw")
        self.assertEqual(c.get("/team/invite/").status_code, 403)


class CompanyDefaultsTests(TestCase):
    def setUp(self):
        from datetime import date
        from core.models import OrgMembership, Customer, TaxCode
        self.date = date
        self.tenant = Tenant.objects.create(name="Defaults Co", default_payment_terms_days=14,
                                            financial_year_start_month=4)
        self.tenant.default_tax_code = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.tenant.save()
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.user = User.objects.create_user("dadmin", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="dadmin", password="pw")

    def test_due_date_defaults_from_payment_terms(self):
        from core.models import CustomerInvoice
        resp = self.client.post("/ar/invoices/new/", {
            "customer": self.customer.id, "invoice_number": "INV-D1",
            "invoice_date": "2026-06-01", "action": "save",
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "0",
            "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-description": "Item", "lines-0-qty": "1", "lines-0-unit_price": "100",
        })
        self.assertEqual(resp.status_code, 302)
        inv = CustomerInvoice.objects.get(tenant=self.tenant, invoice_number="INV-D1")
        self.assertEqual(inv.due_date, self.date(2026, 6, 15))  # 1 June + 14 days

    def test_supplier_invoice_line_defaults_tax(self):
        resp = self.client.get("/invoices/new/")
        self.assertEqual(resp.status_code, 200)
        fs = resp.context["formset"]
        self.assertEqual(fs.forms[0].initial.get("tax_code"), self.tenant.default_tax_code)

    def test_financial_year_helper(self):
        from core.services import reports
        start, end = reports.current_financial_year(self.tenant, today=self.date(2026, 6, 1))
        self.assertEqual(start, self.date(2026, 4, 1))
        self.assertEqual(end, self.date(2027, 3, 31))

    def test_invoice_shows_branding(self):
        from core.models import CustomerInvoice
        self.tenant.legal_name = "Defaults Co Ltd"
        self.tenant.invoice_footer = "Thanks for your business"
        self.tenant.save()
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-B1")
        resp = self.client.get(f"/ar/invoices/{inv.id}/")
        self.assertContains(resp, "Defaults Co Ltd")
        self.assertContains(resp, "Thanks for your business")


class CsvImportTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Import Co")
        self.user = User.objects.create_user("iadmin", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="iadmin", password="pw")

    def _csv(self, text):
        from django.core.files.uploadedfile import SimpleUploadedFile
        return SimpleUploadedFile("data.csv", text.encode("utf-8"), content_type="text/csv")

    def test_import_products_creates_and_reports_errors(self):
        from core.models import Product
        csv_text = "sku,name,standard_cost\nSKU-A,Widget A,2.50\nSKU-B,Gadget B,4\n,Missing SKU,1\n"
        resp = self.client.post("/products/import/", {"file": self._csv(csv_text)})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(Product.objects.filter(tenant=self.tenant).count(), 2)
        p = Product.objects.get(tenant=self.tenant, sku="SKU-A")
        from decimal import Decimal
        self.assertEqual(p.standard_cost, Decimal("2.50"))
        self.assertContains(resp, "Skipped: 1")

    def test_import_is_upsert(self):
        from core.models import Product
        self.client.post("/products/import/", {"file": self._csv("sku,name\nSKU-X,First\n")})
        self.client.post("/products/import/", {"file": self._csv("sku,name\nSKU-X,Updated Name\n")})
        self.assertEqual(Product.objects.filter(tenant=self.tenant, sku="SKU-X").count(), 1)
        self.assertEqual(Product.objects.get(tenant=self.tenant, sku="SKU-X").name, "Updated Name")

    def test_import_customers_and_suppliers(self):
        from core.models import Customer, Supplier
        self.client.post("/customers/import/", {"file": self._csv("name,email\nAcme Retail,ar@acme.test\n")})
        self.client.post("/suppliers/import/", {"file": self._csv("name,currency_code\nGlobex,USD\n")})
        self.assertTrue(Customer.objects.filter(tenant=self.tenant, name="Acme Retail").exists())
        self.assertEqual(Supplier.objects.get(tenant=self.tenant, name="Globex").currency_code, "USD")

    def test_template_download(self):
        resp = self.client.get("/import/products/template.csv")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "text/csv")
        self.assertIn("sku", resp.content.decode())

    def test_import_requires_permission(self):
        # A sales user cannot import products (procurement/admin only).
        from core.models import OrgMembership
        u = User.objects.create_user("salesimp", password="pw")
        OrgMembership.objects.create(user=u, tenant=self.tenant, role="SALES", is_default=True)
        c = Client(); c.login(username="salesimp", password="pw")
        self.assertEqual(c.get("/products/import/").status_code, 403)


class OnboardingTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Onboard Co")
        self.user = User.objects.create_user("oadmin", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="oadmin", password="pw")

    def test_onboarding_page_renders_with_steps(self):
        resp = self.client.get("/onboarding/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Let's get you set up")
        self.assertContains(resp, "First location")

    def test_step_completion_detected(self):
        from core.models import Location
        Location.objects.create(tenant=self.tenant, name="HQ")
        resp = self.client.get("/onboarding/")
        # location step now done -> at least one "Done" badge
        self.assertContains(resp, "Done")

    def test_finish_sets_flag(self):
        resp = self.client.post("/onboarding/finish/")
        self.assertEqual(resp.status_code, 302)
        self.tenant.refresh_from_db()
        self.assertTrue(self.tenant.onboarding_complete)

    def test_create_new_organisation(self):
        from core.models import Tenant as T, OrgMembership
        resp = self.client.post("/onboarding/new-organisation/", {
            "name": "Fresh Org", "business_type": "LTD", "currency_code": "GBP", "country": "United Kingdom",
        })
        self.assertEqual(resp.status_code, 302)
        org = T.objects.get(name="Fresh Org")
        self.assertTrue(OrgMembership.objects.filter(user=self.user, tenant=org, role="ADMIN").exists())

    def test_dashboard_banner_when_not_onboarded(self):
        resp = self.client.get("/dashboard/admin")
        self.assertContains(resp, "Finish setting up")


class AccessRequestTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Req Co")
        self.admin = User.objects.create_user("adminu", password="pw", email="admin@req.test")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)

    def test_public_can_submit_request(self):
        from core.models import AccessRequest
        resp = self.client.post("/request-access/", {
            "name": "Jane Doe", "employee_id": "E123", "email": "jane@acme.test", "team": "Sales",
        })
        self.assertEqual(resp.status_code, 302)
        r = AccessRequest.objects.get(email="jane@acme.test")
        self.assertEqual(r.status, "PENDING")
        self.assertEqual(r.name, "Jane Doe")

    def test_request_form_public_and_get(self):
        self.assertEqual(self.client.get("/request-access/").status_code, 200)

    def test_list_requires_admin(self):
        staff = User.objects.create_user("sales2", password="pw")
        from core.models import OrgMembership
        OrgMembership.objects.create(user=staff, tenant=self.tenant, role="SALES", is_default=True)
        self.client.login(username="sales2", password="pw")
        self.assertEqual(self.client.get("/access-requests/").status_code, 403)

    def test_admin_approve_creates_account(self):
        from core.models import AccessRequest, OrgMembership
        from django.contrib.auth.models import User as U
        req = AccessRequest.objects.create(name="Bob Lee", email="bob@acme.test", team="Warehouse")
        self.client.login(username="adminu", password="pw")
        resp = self.client.post(f"/access-requests/{req.id}/action/", {"action": "approve", "role": "WAREHOUSE"})
        self.assertEqual(resp.status_code, 302)
        req.refresh_from_db()
        self.assertEqual(req.status, "APPROVED")
        self.assertIsNotNone(req.created_user)
        # The created user has a Warehouse membership in this tenant.
        self.assertTrue(OrgMembership.objects.filter(user=req.created_user, tenant=self.tenant, role="WAREHOUSE").exists())
        self.assertEqual(U.objects.filter(username=req.created_user.username).count(), 1)

    def test_admin_reject(self):
        from core.models import AccessRequest
        req = AccessRequest.objects.create(name="Eve", email="eve@acme.test")
        self.client.login(username="adminu", password="pw")
        self.client.post(f"/access-requests/{req.id}/action/", {"action": "reject"})
        req.refresh_from_db()
        self.assertEqual(req.status, "REJECTED")
        self.assertIsNone(req.created_user)

    def test_submit_emails_admin(self):
        from django.core import mail
        self.client.post("/request-access/", {"name": "Jane", "email": "jane@acme.test", "team": "Sales"})
        self.assertTrue(any("admin@req.test" in m.to for m in mail.outbox))

    def test_approve_emails_applicant_with_credentials(self):
        from django.core import mail
        from core.models import AccessRequest
        req = AccessRequest.objects.create(name="Bob", email="bob@acme.test", tenant=self.tenant)
        self.client.login(username="adminu", password="pw")
        self.client.post(f"/access-requests/{req.id}/action/", {"action": "approve", "role": "SALES"})
        applicant_mails = [m for m in mail.outbox if "bob@acme.test" in m.to]
        self.assertTrue(applicant_mails)
        self.assertIn("Temporary password", applicant_mails[-1].body)


class NotificationTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Notify Co")
        self.admin = User.objects.create_user("nadmin", password="pw", email="nadmin@notify.test")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.member = User.objects.create_user("nmember", password="pw", email="nmember@notify.test")
        OrgMembership.objects.create(user=self.member, tenant=self.tenant, role="SALES", is_default=True)

    def test_notify_user_creates_in_app_and_email(self):
        from django.core import mail
        from core import notify
        from core.models import Notification, EmailLog
        note = notify.notify_user(self.member, tenant=self.tenant, category="GENERAL",
                                  title="Hello", message="A message", url="/dashboard/")
        self.assertIsNotNone(note)
        self.assertTrue(Notification.objects.filter(recipient=self.member, title="Hello").exists())
        self.assertTrue(any("nmember@notify.test" in m.to for m in mail.outbox))
        self.assertTrue(EmailLog.objects.filter(to_email="nmember@notify.test", status="SENT").exists())

    def test_email_preference_suppresses_email_only(self):
        from django.core import mail
        from core import notify
        from core.models import Notification, NotificationPreference
        NotificationPreference.objects.create(user=self.member, tenant=self.tenant,
                                              category="GENERAL", in_app=True, email=False)
        notify.notify_user(self.member, tenant=self.tenant, category="GENERAL", title="No email")
        self.assertTrue(Notification.objects.filter(recipient=self.member, title="No email").exists())
        self.assertFalse(any("nmember@notify.test" in m.to for m in mail.outbox))

    def test_in_app_preference_suppresses_notification(self):
        from core import notify
        from core.models import Notification, NotificationPreference
        NotificationPreference.objects.create(user=self.member, tenant=self.tenant,
                                              category="GENERAL", in_app=False, email=True)
        note = notify.notify_user(self.member, tenant=self.tenant, category="GENERAL", title="Email only")
        self.assertIsNone(note)
        self.assertFalse(Notification.objects.filter(recipient=self.member, title="Email only").exists())

    def test_open_marks_read_and_redirects(self):
        from core import notify
        from core.models import Notification
        note = notify.notify_user(self.member, tenant=self.tenant, category="GENERAL",
                                  title="Open me", url="/dashboard/")
        self.client.login(username="nmember", password="pw")
        resp = self.client.get(f"/notifications/{note.id}/open/")
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, "/dashboard/")
        note.refresh_from_db()
        self.assertTrue(note.is_read)

    def test_mark_all_read(self):
        from core import notify
        from core.models import Notification
        for i in range(3):
            notify.notify_user(self.member, tenant=self.tenant, category="GENERAL", title=f"N{i}")
        self.client.login(username="nmember", password="pw")
        self.client.post("/notifications/read-all/")
        self.assertEqual(Notification.objects.filter(recipient=self.member, is_read=False).count(), 0)

    def test_bell_context_shows_unread_count(self):
        from core import notify
        notify.notify_user(self.member, tenant=self.tenant, category="GENERAL", title="Unread one")
        self.client.login(username="nmember", password="pw")
        resp = self.client.get("/dashboard/sales")
        self.assertEqual(resp.context["unread_notifications"], 1)
        self.assertEqual(len(resp.context["recent_notifications"]), 1)

    def test_email_log_requires_admin_or_finance(self):
        self.client.login(username="nmember", password="pw")
        self.assertEqual(self.client.get("/email-log/").status_code, 403)
        self.client.login(username="nadmin", password="pw")
        self.assertEqual(self.client.get("/email-log/").status_code, 200)

    def test_only_recipient_can_open_notification(self):
        from core import notify
        note = notify.notify_user(self.admin, tenant=self.tenant, category="GENERAL", title="Admin only")
        self.client.login(username="nmember", password="pw")
        self.assertEqual(self.client.get(f"/notifications/{note.id}/open/").status_code, 404)

    def test_expense_create_with_submit_notifies_approvers(self):
        """Submitting an expense straight from the create form (action=submit)
        must notify approvers, just like the draft-then-submit path."""
        from core.models import Notification, GLAccount, TaxCode
        category = GLAccount.objects.get(tenant=self.tenant, code="6100")
        std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.client.login(username="nmember", password="pw")
        resp = self.client.post("/expenses/new/", {
            "expense_date": "2026-05-30", "payee": "Cab Co", "category": category.id,
            "description": "Taxi", "net_amount": "50.00", "tax_code": std.id,
            "method": "BANK", "reference": "T-1", "action": "submit",
        })
        self.assertEqual(resp.status_code, 302)
        notes = Notification.objects.filter(recipient=self.admin, category="APPROVAL_REQUEST")
        self.assertTrue(notes.exists(), "approver should be notified when an expense is submitted via the create form")
        self.assertIn("Cab Co", notes.first().title)


class PerOrgEnforcementTests(TestCase):
    """A multi-org user's module access must follow their ACTIVE org's role."""
    def setUp(self):
        from core.models import OrgMembership
        self.org_a = Tenant.objects.create(name="Org A")
        self.org_b = Tenant.objects.create(name="Org B")
        self.user = User.objects.create_user("multi", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.org_a, role="SALES", is_default=True)
        OrgMembership.objects.create(user=self.user, tenant=self.org_b, role="FINANCE")
        self.client.login(username="multi", password="pw")

    def _activate(self, tenant):
        s = self.client.session
        s["active_tenant_id"] = tenant.id
        s.save()

    def test_sales_org_blocked_from_finance_module(self):
        self._activate(self.org_a)  # SALES here
        self.assertEqual(self.client.get("/invoices/").status_code, 403)  # supplier invoices = finance/admin

    def test_finance_org_allowed_finance_module(self):
        self._activate(self.org_b)  # FINANCE here
        self.assertEqual(self.client.get("/invoices/").status_code, 200)

    def test_sales_org_allowed_sales_module(self):
        self._activate(self.org_a)
        self.assertEqual(self.client.get("/sales-orders/").status_code, 200)

    def test_finance_org_blocked_from_sales_only_module(self):
        self._activate(self.org_b)  # FINANCE -> no Sales group
        self.assertEqual(self.client.get("/sales-orders/").status_code, 403)

    def test_group_only_user_still_enforced(self):
        # Legacy user with a Django group but no membership keeps working.
        u = User.objects.create_user("legacy", password="pw")
        u.groups.add(Group.objects.get_or_create(name="Finance")[0])
        c = Client()
        c.login(username="legacy", password="pw")
        self.assertEqual(c.get("/invoices/").status_code, 200)


class PaymentsTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Pay Co")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-P1")
        CustomerInvoiceLine.objects.create(invoice=self.inv, description="Item", qty=Decimal("2"), unit_price=Decimal("100.00"), tax_code=self.std)
        post_customer_invoice(self.inv)  # total 240, status ISSUED

        self.user = User.objects.create_user("pay", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Finance")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="pay", password="pw")

    def test_full_receipt_settles_invoice_and_clears_ar(self):
        from core.services import reports
        from core.models import GLAccount, Payment

        resp = self.client.post("/payments/receipts/new/", {
            "customer": self.customer.id, "payment_date": "2026-05-30",
            "amount": "240.00", "method": "BANK", "reference": "FPS-1",
        })
        self.assertEqual(resp.status_code, 302)

        payment = Payment.objects.get(tenant=self.tenant)
        self.assertEqual(payment.status, "POSTED")
        self.assertEqual(payment.allocated, Decimal("240.00"))

        self.inv.refresh_from_db()
        self.assertEqual(self.inv.status, "PAID")

        # AR account nets to zero (invoice DR 240, receipt CR 240).
        balances = reports.account_balances(self.tenant)
        ar = GLAccount.objects.get(tenant=self.tenant, code="1100")
        bank = GLAccount.objects.get(tenant=self.tenant, code="1050")
        self.assertEqual(balances[ar]["balance"], Decimal("0.00"))
        self.assertEqual(balances[bank]["balance"], Decimal("240.00"))

    def test_partial_receipt_leaves_invoice_open(self):
        from core.models import Payment
        self.client.post("/payments/receipts/new/", {
            "customer": self.customer.id, "payment_date": "2026-05-30",
            "amount": "100.00", "method": "BANK", "reference": "FPS-2",
        })
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.status, "ISSUED")
        self.assertEqual(self.inv.outstanding, Decimal("140.00"))

    def test_bank_reconciliation_matches_statement_line(self):
        from core.models import Payment, BankTransaction
        self.client.post("/payments/receipts/new/", {
            "customer": self.customer.id, "payment_date": "2026-05-30",
            "amount": "240.00", "method": "BANK",
        })
        payment = Payment.objects.get(tenant=self.tenant)
        self.assertFalse(payment.is_reconciled)

        # Statement line of the same amount auto-matches to the receipt.
        BankTransaction.objects.create(tenant=self.tenant, description="FPS CREDIT", amount=Decimal("240.00"))
        resp = self.client.post("/bank/reconcile/", {"action": "auto"})
        self.assertEqual(resp.status_code, 302)
        payment.refresh_from_db()
        self.assertTrue(payment.is_reconciled)

    def test_payment_pages_render(self):
        for path in ["/payments/", "/payments/receipts/new/", "/payments/payments/new/", "/bank/reconcile/"]:
            self.assertEqual(self.client.get(path).status_code, 200, path)


class CostingTests(TestCase):
    def setUp(self):
        from core.models import Location
        self.tenant = Tenant.objects.create(name="Cost Co")
        self.product = Product.objects.create(tenant=self.tenant, sku="SKU-C", name="P")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH")

    def test_moving_average_and_outbound_valuation(self):
        from core.services.inventory import apply_movement
        from core.services import reports

        apply_movement(tenant=self.tenant, product=self.product, location=self.loc,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="1",
                       unit_cost=Decimal("2.00"))
        self.product.refresh_from_db()
        self.assertEqual(self.product.average_cost, Decimal("2.0000"))

        apply_movement(tenant=self.tenant, product=self.product, location=self.loc,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="2",
                       unit_cost=Decimal("4.00"))
        self.product.refresh_from_db()
        self.assertEqual(self.product.average_cost, Decimal("3.0000"))  # (20+40)/20

        sale = apply_movement(tenant=self.tenant, product=self.product, location=self.loc,
                              movement_type="SALE", qty_delta=Decimal("-5"), ref_type="T", ref_id="3")
        self.assertEqual(sale.unit_cost, Decimal("3.0000"))
        self.assertEqual(sale.value, Decimal("-15.00"))

        val = reports.stock_valuation(self.tenant)
        self.assertEqual(val["total"], Decimal("45.00"))  # 15 on hand x 3.00

    def test_sales_order_posts_cogs_journal(self):
        from core.services.inventory import apply_movement
        from core.views import _post_sales_order
        from core.models import SalesOrder, SalesOrderLine, JournalEntry

        apply_movement(tenant=self.tenant, product=self.product, location=self.loc,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="1",
                       unit_cost=Decimal("3.00"))
        order = SalesOrder.objects.create(tenant=self.tenant, order_number="SO-C1", ship_from_location=self.loc)
        SalesOrderLine.objects.create(order=order, product=self.product, qty=Decimal("4"), unit_price=Decimal("10.00"))

        _post_sales_order(order)

        je = JournalEntry.objects.get(tenant=self.tenant, ref_type="COGS")
        self.assertEqual(je.total_debit, je.total_credit)
        self.assertEqual(je.total_debit, Decimal("12.00"))  # 4 x 3.00


class FifoCostingTests(TestCase):
    def setUp(self):
        from core.models import Location
        self.tenant = Tenant.objects.create(name="FIFO Co")
        self.product = Product.objects.create(tenant=self.tenant, sku="SKU-F", name="P", cost_method=Product.CostMethod.FIFO)
        self.loc = Location.objects.create(tenant=self.tenant, name="WH")

    def test_fifo_consumes_oldest_layers_first(self):
        from core.services.inventory import apply_movement
        from core.services import reports

        # Layer 1: 10 @ 2.00, Layer 2: 10 @ 5.00
        apply_movement(tenant=self.tenant, product=self.product, location=self.loc,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="1", unit_cost=Decimal("2.00"))
        apply_movement(tenant=self.tenant, product=self.product, location=self.loc,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="2", unit_cost=Decimal("5.00"))

        # Sell 12: 10 @ 2.00 + 2 @ 5.00 = 30.00 COGS (NOT average 3.50*12=42).
        sale = apply_movement(tenant=self.tenant, product=self.product, location=self.loc,
                              movement_type="SALE", qty_delta=Decimal("-12"), ref_type="T", ref_id="3")
        self.assertEqual(sale.value, Decimal("-30.00"))

        # Remaining 8 @ 5.00 = 40.00 on the valuation report.
        val = reports.stock_valuation(self.tenant)
        self.assertEqual(val["total"], Decimal("40.00"))


class CriticalInventoryFixTests(TestCase):
    """Regression tests for the Critical inventory-costing fixes C5/C6/C7/C10."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name="Crit Co")
        self.wh_a = Location.objects.create(tenant=self.tenant, name="WH-A")
        self.wh_b = Location.objects.create(tenant=self.tenant, name="WH-B")

    # ---- C5: FIFO layers are scoped per location ----
    def test_c5_fifo_outbound_only_consumes_its_own_location_layers(self):
        from core.services.inventory import apply_movement
        from core.models import InventoryCostLayer
        p = Product.objects.create(tenant=self.tenant, sku="SKU-C5", name="P",
                                   cost_method=Product.CostMethod.FIFO)
        # 10 @ 2.00 into A, 10 @ 9.00 into B.
        apply_movement(tenant=self.tenant, product=p, location=self.wh_a,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="a1",
                       unit_cost=Decimal("2.00"))
        apply_movement(tenant=self.tenant, product=p, location=self.wh_b,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="b1",
                       unit_cost=Decimal("9.00"))
        # Sell 5 from B -> must cost 5 @ 9.00 = 45 (B's own layer), NOT A's 2.00.
        sale = apply_movement(tenant=self.tenant, product=p, location=self.wh_b,
                              movement_type="SALE", qty_delta=Decimal("-5"), ref_type="T", ref_id="s1")
        self.assertEqual(sale.value, Decimal("-45.00"))
        # A's layer is untouched; B's layer reduced to 5 remaining.
        self.assertEqual(InventoryCostLayer.objects.get(product=p, location=self.wh_a).qty_remaining, Decimal("10.00"))
        self.assertEqual(InventoryCostLayer.objects.get(product=p, location=self.wh_b).qty_remaining, Decimal("5.00"))

    # ---- C6: FIFO transfer is value-neutral (re-layers at consumed cost) ----
    def test_c6_fifo_transfer_preserves_total_value(self):
        from core.services.inventory import apply_movement
        from core.services import reports
        from core.models import InventoryTransfer, InventoryTransferLine
        p = Product.objects.create(tenant=self.tenant, sku="SKU-C6", name="P",
                                   cost_method=Product.CostMethod.FIFO)
        # Two layers into A at different costs: 10 @ 2.00 and 10 @ 8.00 -> value 100.
        apply_movement(tenant=self.tenant, product=p, location=self.wh_a,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="a1",
                       unit_cost=Decimal("2.00"))
        apply_movement(tenant=self.tenant, product=p, location=self.wh_a,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="a2",
                       unit_cost=Decimal("8.00"))
        before = reports.stock_valuation(self.tenant)["total"]
        self.assertEqual(before, Decimal("100.00"))

        # Transfer 15 from A to B (consumes 10@2 + 5@8 = 60 on the OUT side).
        tr = InventoryTransfer.objects.create(tenant=self.tenant, transfer_number="TR-1",
                                              from_location=self.wh_a, to_location=self.wh_b)
        InventoryTransferLine.objects.create(transfer=tr, product=p, qty=Decimal("15"))
        from core.views import _post_transfer
        _post_transfer(tr)

        # Total inventory value unchanged by an internal transfer.
        after = reports.stock_valuation(self.tenant)["total"]
        self.assertEqual(after, Decimal("100.00"))
        # B now holds a layer worth exactly the cost relieved on the OUT side (60).
        from core.models import InventoryCostLayer
        b_val = sum((l.qty_remaining * l.unit_cost for l in
                     InventoryCostLayer.objects.filter(product=p, location=self.wh_b)), Decimal("0.00"))
        self.assertEqual(b_val.quantize(Decimal("0.01")), Decimal("60.00"))

    # ---- C7: cancelling an issued invoice reverses COGS + restores stock ----
    def test_c7_cancel_issued_invoice_restores_stock_and_reverses_cogs(self):
        from core.services.inventory import apply_movement
        from core.services.gl import reverse_invoice_cogs
        from core.models import JournalEntry, GLAccount
        # Tenant whose signal seeds GL accounts.
        t = Tenant.objects.create(name="Crit AR Co")
        loc = Location.objects.create(tenant=t, name="WH")
        p = Product.objects.create(tenant=t, sku="SKU-C7", name="P")
        apply_movement(tenant=t, product=p, location=loc, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="r1", unit_cost=Decimal("3.00"))
        cust = Customer.objects.create(tenant=t, name="Cust")
        inv = CustomerInvoice.objects.create(tenant=t, customer=cust, invoice_number="INV-C7", location=loc)
        std = TaxCode.objects.get(tenant=t, code="STD")
        CustomerInvoiceLine.objects.create(invoice=inv, product=p, description="P",
                                           qty=Decimal("4"), unit_price=Decimal("10.00"), tax_code=std)
        post_customer_invoice(inv)

        bal = InventoryBalance.objects.get(tenant=t, product=p, location=loc)
        self.assertEqual(bal.on_hand, Decimal("6.00"))  # 10 - 4 sold
        cogs = GLAccount.objects.get(tenant=t, code="5000")
        self.assertEqual(_account_balance(t, "5000"), Decimal("12.00"))  # 4 @ 3.00

        # Cancel -> stock back to 10, COGS net zero.
        reverse_invoice_cogs(inv, user=None)
        bal.refresh_from_db()
        self.assertEqual(bal.on_hand, Decimal("10.00"))
        self.assertEqual(_account_balance(t, "5000"), Decimal("0.00"))

        # Idempotent: a second call does nothing.
        reverse_invoice_cogs(inv, user=None)
        bal.refresh_from_db()
        self.assertEqual(bal.on_hand, Decimal("10.00"))
        self.assertEqual(_account_balance(t, "5000"), Decimal("0.00"))

    # ---- C10: moving-average reads on-hand under the product lock ----
    def test_c10_sequential_receipts_keep_correct_average(self):
        from core.services.inventory import apply_movement
        p = Product.objects.create(tenant=self.tenant, sku="SKU-C10", name="P")
        apply_movement(tenant=self.tenant, product=p, location=self.wh_a,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="1",
                       unit_cost=Decimal("2.00"))
        # Second receipt at a different location must still fold into the
        # company-wide average using the up-to-date prior on-hand (20 total).
        apply_movement(tenant=self.tenant, product=p, location=self.wh_b,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="2",
                       unit_cost=Decimal("4.00"))
        p.refresh_from_db()
        self.assertEqual(p.average_cost, Decimal("3.0000"))  # (20+40)/20


def _account_balance(tenant, code):
    """Net debit balance (debits - credits) for a GL account, across all entries."""
    from core.models import JournalLine
    from django.db.models import Sum
    agg = JournalLine.objects.filter(entry__tenant=tenant, account__code=code).aggregate(
        d=Sum("debit"), c=Sum("credit"))
    return (agg["d"] or Decimal("0.00")) - (agg["c"] or Decimal("0.00"))


class HighInventoryFixTests(TestCase):
    """Regression tests for the High inventory findings H1/H7/H8/H14."""

    # ---- H1: post_cogs is idempotent on (tenant, ref) ----
    def test_h1_post_cogs_idempotent_on_ref(self):
        from core.services.gl import post_cogs
        from core.models import JournalEntry
        t = Tenant.objects.create(name="H1 Co")  # signal seeds GL accounts
        je1 = post_cogs(t, Decimal("50.00"), "SO-1")
        je2 = post_cogs(t, Decimal("50.00"), "SO-1")
        self.assertEqual(je1.id, je2.id)
        self.assertEqual(
            JournalEntry.objects.filter(tenant=t, ref_type="COGS", ref_id="SO-1").count(), 1)
        self.assertEqual(_account_balance(t, "5000"), Decimal("50.00"))  # COGS booked once

    # ---- H7: optional negative-stock block ----
    def test_h7_block_negative_stock_rejects_oversell(self):
        from core.services.inventory import apply_movement
        from django.core.exceptions import ValidationError
        t = Tenant.objects.create(name="H7 Co", block_negative_stock=True)
        loc = Location.objects.create(tenant=t, name="WH")
        p = Product.objects.create(tenant=t, sku="SKU-H7", name="P")
        apply_movement(tenant=t, product=p, location=loc, movement_type="RECEIVE",
                       qty_delta=Decimal("5"), ref_type="T", ref_id="1", unit_cost=Decimal("2.00"))
        with self.assertRaises(ValidationError):
            apply_movement(tenant=t, product=p, location=loc, movement_type="SALE",
                           qty_delta=Decimal("-10"), ref_type="T", ref_id="2")
        # The rejected movement rolled back: on-hand untouched, no SALE recorded.
        bal = InventoryBalance.objects.get(tenant=t, product=p, location=loc)
        self.assertEqual(bal.on_hand, Decimal("5.00"))
        self.assertFalse(InventoryMovement.objects.filter(tenant=t, movement_type="SALE").exists())

    def test_h7_default_tenant_still_allows_negative(self):
        from core.services.inventory import apply_movement
        t = Tenant.objects.create(name="H7 Allow Co")  # block_negative_stock defaults False
        loc = Location.objects.create(tenant=t, name="WH")
        p = Product.objects.create(tenant=t, sku="SKU-H7B", name="P")
        apply_movement(tenant=t, product=p, location=loc, movement_type="SALE",
                       qty_delta=Decimal("-3"), ref_type="T", ref_id="1")
        bal = InventoryBalance.objects.get(tenant=t, product=p, location=loc)
        self.assertEqual(bal.on_hand, Decimal("-3.00"))

    # ---- H8: GRNI clears against the bill; price variance to PPV ----
    def test_h8_grni_clears_with_price_variance(self):
        from core.models import (PurchaseOrder, PurchaseOrderLine, GoodsReceipt, GoodsReceiptLine,
                                  SupplierInvoice, SupplierInvoiceLine)
        from core.services.gl import post_inventory_receipt, post_supplier_invoice
        t = Tenant.objects.create(name="H8 Co")
        supplier = Supplier.objects.create(tenant=t, name="S")
        p = Product.objects.create(tenant=t, sku="SKU-H8", name="P")
        loc = Location.objects.create(tenant=t, name="WH")
        po = PurchaseOrder.objects.create(tenant=t, po_number="PO-H8", supplier=supplier)
        pol = PurchaseOrderLine.objects.create(po=po, product=p, ordered_qty=Decimal("10"), unit_cost=Decimal("5.00"))
        grn = GoodsReceipt.objects.create(tenant=t, po=po, grn_number="GRN-H8", received_to=loc,
                                          status=GoodsReceipt.Status.POSTED)
        GoodsReceiptLine.objects.create(receipt=grn, po_line=pol, product=p,
                                        qty_received=Decimal("10"), unit_cost=Decimal("5.00"))
        # Received goods value = 50: receipt credits GRNI 50.
        post_inventory_receipt(t, Decimal("50.00"), grn.grn_number)
        self.assertEqual(_account_balance(t, "2100"), Decimal("-50.00"))  # GRNI credit

        # Supplier bills 10 @ 6.00 = 60 net -> 10 unfavourable price variance.
        std = TaxCode.objects.get(tenant=t, code="STD")
        inv = SupplierInvoice.objects.create(tenant=t, supplier=supplier, po=po, receipt=grn,
                                             invoice_number="SINV-H8")
        SupplierInvoiceLine.objects.create(invoice=inv, product=p, qty=Decimal("10"),
                                           unit_cost=Decimal("6.00"), tax_code=std)
        je = post_supplier_invoice(inv)

        self.assertEqual(je.total_debit, je.total_credit)
        self.assertEqual(_account_balance(t, "2100"), Decimal("0.00"))   # GRNI fully cleared
        self.assertEqual(_account_balance(t, "5100"), Decimal("10.00"))  # PPV holds the variance
        self.assertEqual(_account_balance(t, "2000"), Decimal("-72.00"))  # AP = 60 + 20% VAT

    # ---- H14: per-location analytics value FIFO from layers ----
    def test_h14_by_location_fifo_uses_layer_cost(self):
        from django.utils import timezone
        from core.services.inventory import apply_movement
        from core.services import reports
        t = Tenant.objects.create(name="H14 Co")
        a = Location.objects.create(tenant=t, name="A")
        b = Location.objects.create(tenant=t, name="B")
        p = Product.objects.create(tenant=t, sku="SKU-H14", name="P", cost_method=Product.CostMethod.FIFO)
        apply_movement(tenant=t, product=p, location=a, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="a", unit_cost=Decimal("2.00"))
        apply_movement(tenant=t, product=p, location=b, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="b", unit_cost=Decimal("8.00"))
        # Company average is now 5.00; per-location FIFO must NOT use that.
        today = timezone.localdate()
        res = reports.inventory_analytics(t, today, today)
        by_loc = {e["location"].id: e["value"] for e in res["by_location"]}
        self.assertEqual(by_loc[a.id], Decimal("20.00"))  # 10 @ 2.00, not 10 @ 5.00
        self.assertEqual(by_loc[b.id], Decimal("80.00"))  # 10 @ 8.00, not 10 @ 5.00
        # Per-location total reconciles to the grand current_value.
        self.assertEqual(sum(by_loc.values()), res["current_value"])


class MediumInventoryFixTests(TestCase):
    """Regression tests for the Medium inventory findings M5/M6/M7/M13/M14."""

    # ---- M14: release survives a missing balance row ----
    def test_m14_release_survives_missing_balance_row(self):
        from core.services.inventory import release_reservations
        from core.models import InventoryReservation
        t = Tenant.objects.create(name="M14 Co")
        loc = Location.objects.create(tenant=t, name="WH")
        p = Product.objects.create(tenant=t, sku="SKU-M14", name="P")
        r = InventoryReservation.objects.create(
            tenant=t, product=p, location=loc, qty=Decimal("5"),
            status=InventoryReservation.Status.ACTIVE, ref_type="X", ref_id="1")
        # No InventoryBalance row exists for (p, loc): release must not crash.
        release_reservations(tenant=t, ref_type="X", ref_id="1")
        r.refresh_from_db()
        self.assertEqual(r.status, InventoryReservation.Status.RELEASED)

    # ---- M7: ATP blocks over-reservation under strict control ----
    def test_m7_atp_blocks_over_reservation(self):
        from core.services.inventory import reserve_stock
        from core.models import InventoryBalance, InventoryReservation
        from django.core.exceptions import ValidationError
        t = Tenant.objects.create(name="M7 Co", block_negative_stock=True)
        loc = Location.objects.create(tenant=t, name="WH")
        p = Product.objects.create(tenant=t, sku="SKU-M7", name="P")
        InventoryBalance.objects.create(tenant=t, product=p, location=loc,
                                        on_hand=Decimal("5"), reserved=Decimal("0"))
        with self.assertRaises(ValidationError):
            reserve_stock(tenant=t, product=p, location=loc, qty=Decimal("10"),
                          ref_type="X", ref_id="1")
        bal = InventoryBalance.objects.get(tenant=t, product=p, location=loc)
        self.assertEqual(bal.reserved, Decimal("0.00"))  # rolled back
        self.assertFalse(InventoryReservation.objects.filter(tenant=t).exists())
        # Within available is still fine.
        reserve_stock(tenant=t, product=p, location=loc, qty=Decimal("3"),
                      ref_type="X", ref_id="2")
        bal.refresh_from_db()
        self.assertEqual(bal.reserved, Decimal("3.00"))

    def test_m7_over_reservation_allowed_by_default(self):
        from core.services.inventory import reserve_stock
        from core.models import InventoryBalance
        t = Tenant.objects.create(name="M7 Allow Co")  # strict control off
        loc = Location.objects.create(tenant=t, name="WH")
        p = Product.objects.create(tenant=t, sku="SKU-M7B", name="P")
        InventoryBalance.objects.create(tenant=t, product=p, location=loc,
                                        on_hand=Decimal("5"), reserved=Decimal("0"))
        reserve_stock(tenant=t, product=p, location=loc, qty=Decimal("10"),
                      ref_type="X", ref_id="1")  # over-reserve permitted
        self.assertEqual(InventoryBalance.objects.get(tenant=t, product=p, location=loc).reserved,
                         Decimal("10.00"))

    # ---- M6: a serial/lot can't be issued without on-hand stock ----
    def test_m6_serial_not_in_stock_cannot_be_issued(self):
        from core.services.inventory import apply_movement
        from django.core.exceptions import ValidationError
        t = Tenant.objects.create(name="M6 Co", block_negative_stock=True)
        loc = Location.objects.create(tenant=t, name="WH")
        p = Product.objects.create(tenant=t, sku="SKU-M6", name="P")
        # Receive serial SN-1 (on-hand at location is now 1).
        apply_movement(tenant=t, product=p, location=loc, movement_type="RECEIVE",
                       qty_delta=Decimal("1"), ref_type="T", ref_id="1",
                       unit_cost=Decimal("2.00"), serial_number="SN-1")
        # Issuing a different serial passes the location on-hand check (1 -> 0)
        # but must fail the lot/serial check: SN-2 has none on hand.
        with self.assertRaises(ValidationError):
            apply_movement(tenant=t, product=p, location=loc, movement_type="SALE",
                           qty_delta=Decimal("-1"), ref_type="T", ref_id="2",
                           serial_number="SN-2")

    # ---- M13: a bad landed-cost amount rolls the whole receipt back ----
    def test_m13_bad_landed_cost_rolls_back_receipt(self):
        t = Tenant.objects.create(name="M13 Co")
        loc = Location.objects.create(tenant=t, name="WH")
        supplier = Supplier.objects.create(tenant=t, name="S")
        product = Product.objects.create(tenant=t, sku="SKU-M13", name="P")
        po = PurchaseOrder.objects.create(tenant=t, po_number="PO-M13", supplier=supplier,
                                          status=PurchaseOrder.Status.SUBMITTED)
        pol = PurchaseOrderLine.objects.create(po=po, product=product,
                                               ordered_qty=Decimal("10"), unit_cost=Decimal("3.00"))
        shipment = Shipment.objects.create(tenant=t, po=po, from_supplier=supplier, destination=loc)
        sl = ShipmentLine.objects.create(shipment=shipment, po_line=pol, expected_qty=Decimal("10"))
        user = User.objects.create_user("m13u", password="pw")
        user.groups.add(Group.objects.get_or_create(name="Warehouse")[0])
        UserProfile.objects.create(user=user, tenant=t)
        self.client.login(username="m13u", password="pw")

        from django.urls import reverse
        resp = self.client.post(reverse("receive_po", args=[po.id]), {
            "grn_number": "GRN-M13", f"recv_{sl.id}": "5",
            "landed_cost_name": "Freight", "landed_cost_amount": "not-a-number",
        })
        self.assertEqual(resp.status_code, 302)
        # Whole receipt rolled back: no GRN, no movement, no stock, no landed cost.
        self.assertFalse(GoodsReceipt.objects.filter(tenant=t).exists())
        self.assertFalse(InventoryMovement.objects.filter(tenant=t).exists())
        self.assertFalse(InventoryBalance.objects.filter(tenant=t, on_hand__gt=0).exists())

    # ---- M5: a stale cycle count bounces for re-approval instead of mis-posting ----
    def test_m5_stale_cycle_count_refreshes_and_bounces(self):
        from core.models import (CycleCount, CycleCountLine, InventoryBalance,
                                  InventoryMovement, OrgMembership)
        from django.urls import reverse
        t = Tenant.objects.create(name="M5 Co")
        loc = Location.objects.create(tenant=t, name="WH")
        p = Product.objects.create(tenant=t, sku="SKU-M5", name="P")
        bal = InventoryBalance.objects.create(tenant=t, product=p, location=loc,
                                              on_hand=Decimal("10"), reserved=Decimal("0"))
        cc = CycleCount.objects.create(tenant=t, location=loc, status=CycleCount.Status.APPROVED)
        line = CycleCountLine.objects.create(cycle_count=cc, product=p, system_qty=Decimal("10"),
                                             counted_qty=Decimal("12"), variance_qty=Decimal("2"))
        user = User.objects.create_user("m5u", password="pw")
        OrgMembership.objects.create(user=user, tenant=t, role="ADMIN", is_default=True)
        self.client.login(username="m5u", password="pw")

        # Stock moves to 7 after approval -> the frozen variance is now stale.
        bal.on_hand = Decimal("7"); bal.save()
        self.client.post(reverse("cycle_count_post", args=[cc.id]))

        cc.refresh_from_db(); line.refresh_from_db()
        self.assertEqual(cc.status, CycleCount.Status.SUBMITTED)   # bounced, not posted
        self.assertEqual(line.system_qty, Decimal("7.00"))          # refreshed to live
        self.assertEqual(line.variance_qty, Decimal("5.00"))        # 12 - 7
        self.assertFalse(InventoryMovement.objects.filter(tenant=t, ref_type="CYCLE_COUNT").exists())

        # Re-approve (no further movement) and post -> book reaches the counted 12.
        cc.status = CycleCount.Status.APPROVED; cc.save()
        self.client.post(reverse("cycle_count_post", args=[cc.id]))
        bal.refresh_from_db()
        self.assertEqual(bal.on_hand, Decimal("12.00"))
        self.assertTrue(InventoryMovement.objects.filter(tenant=t, ref_type="CYCLE_COUNT").exists())


class StablePoReceiptTests(TestCase):
    """M17: GoodsReceiptLines are the source of truth for received/open qty;
    quantities must not drift across PO amendments/versions."""

    def setUp(self):
        self.tenant = Tenant.objects.create(name="Stable PO Co")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="S")
        self.product = Product.objects.create(tenant=self.tenant, sku="SKU-STB", name="P")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH")
        self.po = PurchaseOrder.objects.create(
            tenant=self.tenant, po_number="PO-STB", supplier=self.supplier,
            status=PurchaseOrder.Status.SUBMITTED, version=1)
        self.line = PurchaseOrderLine.objects.create(
            po=self.po, product=self.product, ordered_qty=Decimal("10"), unit_cost=Decimal("3.00"))
        self.user = User.objects.create_user("stbu", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Admin")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="stbu", password="pw")
        self._grn_seq = 0

    def _post_receipt(self, po_line, qty):
        """Create a POSTED goods receipt against a (possibly old) PO line and
        re-derive received_qty, mirroring what receive_po does."""
        from core.models import GoodsReceipt, GoodsReceiptLine
        from core.services.purchasing import sync_po_line_received
        self._grn_seq += 1
        grn = GoodsReceipt.objects.create(
            tenant=self.tenant, po=po_line.po, grn_number=f"GRN-STB-{self._grn_seq}",
            received_to=self.loc, status=GoodsReceipt.Status.POSTED)
        GoodsReceiptLine.objects.create(
            receipt=grn, po_line=po_line, root_line_id=(po_line.root_line_id or po_line.id),
            product=po_line.product, qty_received=Decimal(qty), unit_cost=po_line.unit_cost)
        sync_po_line_received(po_line)
        return grn

    def _current_line(self):
        from core.models import PurchaseOrderLine
        return PurchaseOrderLine.objects.get(po__tenant=self.tenant, po__is_current=True,
                                             product=self.product)

    def test_amend_after_partial_receipt_keeps_open_qty(self):
        # Receive 4 of 10, then amend.
        self._post_receipt(self.line, "4")
        self.line.refresh_from_db()
        self.assertEqual(self.line.received_qty, Decimal("4.00"))
        self.assertEqual(self.line.open_qty, Decimal("6.00"))

        resp = self.client.post(f"/po/{self.po.id}/amend/", {"reason": "price change"})
        self.assertEqual(resp.status_code, 302)
        cur = self._current_line()
        self.assertEqual(cur.po.version, 2)
        # The new version inherits the received total; open is still 6 (not 10).
        self.assertEqual(cur.received_qty, Decimal("4.00"))
        self.assertEqual(cur.open_qty, Decimal("6.00"))

    def test_receipt_against_old_version_reflects_in_current(self):
        self._post_receipt(self.line, "4")
        self.client.post(f"/po/{self.po.id}/amend/", {"reason": "amend"})
        cur = self._current_line()

        # A late receipt posts against the ORIGINAL (superseded) PO line.
        self._post_receipt(self.line, "3")

        cur.refresh_from_db()
        self.line.refresh_from_db()
        # Current version shows the full 7 received / 3 open.
        self.assertEqual(cur.received_qty, Decimal("7.00"))
        self.assertEqual(cur.open_qty, Decimal("3.00"))
        # No drift: the old version reports the same stable total.
        self.assertEqual(self.line.received_qty, Decimal("7.00"))

    def test_multiple_versions_no_drift(self):
        self._post_receipt(self.line, "2")
        self.client.post(f"/po/{self.po.id}/amend/", {"reason": "v2"})
        v2 = self._current_line()
        self._post_receipt(self.line, "2")  # against v1 line
        self.client.post(f"/po/{v2.po.id}/amend/", {"reason": "v3"})
        v3 = self._current_line()
        self._post_receipt(v2, "3")  # against v2 line

        from core.models import PurchaseOrderLine
        totals = {l.po.version: l.received_qty for l in
                  PurchaseOrderLine.objects.filter(po__tenant=self.tenant, product=self.product)
                  .select_related("po")}
        # Every version reports the same stable received total (2+2+3 = 7).
        self.assertEqual(set(totals.values()), {Decimal("7.00")})
        v3.refresh_from_db()
        self.assertEqual(v3.open_qty, Decimal("3.00"))  # 10 ordered - 7 received


class LotScopedCostingTests(TestCase):
    """M6: lot-tracked FIFO stock is costed from the issued lot's own layer
    (specific identification / FEFO), not the global FIFO queue or average."""

    def setUp(self):
        import datetime
        self.tenant = Tenant.objects.create(name="Lot Cost Co")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH")
        self.product = Product.objects.create(
            tenant=self.tenant, sku="SKU-LOT", name="P",
            cost_method=Product.CostMethod.FIFO, track_lots=True, track_expiry=True)
        self.early = datetime.date(2026, 1, 1)
        self.late = datetime.date(2027, 1, 1)

    def _receive(self, lot, expiry, qty, cost, ref):
        from core.services.inventory import apply_movement
        return apply_movement(
            tenant=self.tenant, product=self.product, location=self.loc,
            movement_type="RECEIVE", qty_delta=Decimal(qty), ref_type="T", ref_id=ref,
            unit_cost=Decimal(cost), lot_code=lot, expiry_date=expiry)

    def test_fefo_selects_earliest_expiring_lot(self):
        from core.services.inventory import select_fefo_lots
        # LATE lot received first, EARLY lot second (so FIFO order != FEFO order).
        self._receive("LATE", self.late, "5", "2.00", "r1")
        self._receive("EARLY", self.early, "5", "9.00", "r2")
        picks = select_fefo_lots(tenant=self.tenant, product=self.product, location=self.loc, qty=Decimal("3"))
        self.assertEqual(picks[0][0].lot_code, "EARLY")  # earliest expiry first
        self.assertEqual(picks[0][1], Decimal("3.00"))

    def test_cogs_follows_issued_lot_not_global_fifo(self):
        from core.services.inventory import apply_movement
        from core.models import InventoryIssueCost
        # LATE lot is older (would be picked by global FIFO) and cheaper (2.00);
        # EARLY lot is newer but pricier (9.00). Company average would be 5.50.
        self._receive("LATE", self.late, "5", "2.00", "r1")
        self._receive("EARLY", self.early, "5", "9.00", "r2")

        # Issue the EARLY lot (as FEFO would): COGS must be 9.00/unit, not 2.00
        # (global FIFO) and not 5.50 (average).
        sale = apply_movement(
            tenant=self.tenant, product=self.product, location=self.loc,
            movement_type="SALE", qty_delta=Decimal("-2"), ref_type="T", ref_id="s1",
            lot_code="EARLY", expiry_date=self.early)
        self.assertEqual(sale.value, Decimal("-18.00"))     # 2 @ 9.00
        self.assertEqual(sale.unit_cost, Decimal("9.0000"))

        # Issue-cost trail links the consumption to the EARLY lot's layer.
        ic = InventoryIssueCost.objects.get(movement=sale)
        self.assertEqual(ic.lot_code, "EARLY")
        self.assertEqual(ic.unit_cost, Decimal("9.0000"))
        self.assertEqual(ic.total_cost, Decimal("18.00"))
        self.assertEqual(ic.cost_layer.lot_code, "EARLY")

        # Issuing the other lot is costed at its own 2.00.
        sale2 = apply_movement(
            tenant=self.tenant, product=self.product, location=self.loc,
            movement_type="SALE", qty_delta=Decimal("-1"), ref_type="T", ref_id="s2",
            lot_code="LATE", expiry_date=self.late)
        self.assertEqual(sale2.value, Decimal("-2.00"))     # 1 @ 2.00

    def test_non_lot_issue_unchanged_global_fifo(self):
        # Regression: with no lot specified, costing is the original global FIFO.
        from core.services.inventory import apply_movement
        self._receive("LATE", self.late, "5", "2.00", "r1")
        self._receive("EARLY", self.early, "5", "9.00", "r2")
        sale = apply_movement(
            tenant=self.tenant, product=self.product, location=self.loc,
            movement_type="SALE", qty_delta=Decimal("-6"), ref_type="T", ref_id="s1")
        # Oldest-first across all layers: 5 @ 2.00 + 1 @ 9.00 = 19.00.
        self.assertEqual(sale.value, Decimal("-19.00"))


class LotValuationReportTests(TestCase):
    """inventory_analytics values each lot from its own cost layer, not the
    product moving-average."""

    def test_lot_detail_uses_lot_layer_cost(self):
        import datetime
        from django.utils import timezone
        from core.services.inventory import apply_movement
        from core.services import reports
        t = Tenant.objects.create(name="LotVal Co")
        loc = Location.objects.create(tenant=t, name="WH")
        p = Product.objects.create(tenant=t, sku="SKU-LV", name="P",
                                   cost_method=Product.CostMethod.FIFO, track_lots=True, track_expiry=True)
        apply_movement(tenant=t, product=p, location=loc, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="a", unit_cost=Decimal("2.00"),
                       lot_code="A", expiry_date=datetime.date(2026, 6, 1))
        apply_movement(tenant=t, product=p, location=loc, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="b", unit_cost=Decimal("8.00"),
                       lot_code="B", expiry_date=datetime.date(2026, 12, 1))
        # Company average is 5.00; lot valuation must NOT use it.
        today = timezone.localdate()
        res = reports.inventory_analytics(t, today, today)
        lots = {l["lot"]: l["value"] for l in res["lots"]}
        self.assertEqual(lots["A"], Decimal("20.00"))  # 10 @ 2.00
        self.assertEqual(lots["B"], Decimal("80.00"))  # 10 @ 8.00, not 10 @ 5.00 = 50
        # And the per-location FIFO total stays consistent.
        self.assertEqual(res["current_value"], Decimal("100.00"))


class InventoryGlReconciliationTests(TestCase):
    """Inventory subledger (movement value) vs GL inventory control account."""

    def setUp(self):
        from core.models import GLAccount
        self.t = Tenant.objects.create(name="Recon Co")  # signal seeds GL accounts
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.p = Product.objects.create(tenant=self.t, sku="SKU-R", name="P")
        self.inv_acc = GLAccount.objects.get(tenant=self.t, code="1000")

    def _mv(self, value, created=None):
        value = Decimal(value)
        m = InventoryMovement.objects.create(
            tenant=self.t, product=self.p, location=self.loc, site_id=self.loc.site_id,
            movement_type=("RECEIVE" if value > 0 else "SALE"),
            qty_delta=(Decimal("1") if value > 0 else Decimal("-1")),
            unit_cost=abs(value), value=value, ref_type="T", ref_id="x")
        if created is not None:
            InventoryMovement.objects.filter(id=m.id).update(created_at=created)
        return m

    def _gl(self, debit="0", credit="0", entry_date=None):
        from core.models import JournalEntry, JournalLine
        from django.utils import timezone
        je = JournalEntry.objects.create(tenant=self.t, entry_date=entry_date or timezone.localdate(),
                                         ref_type="T", ref_id="x", memo="t")
        JournalLine.objects.create(entry=je, account=self.inv_acc,
                                   debit=Decimal(debit), credit=Decimal(credit))
        return je

    def test_matched_postings_reconcile(self):
        from core.services import reports
        self._mv("50"); self._gl(debit="50")
        rec = reports.inventory_gl_reconciliation(self.t)
        self.assertEqual(rec["closing_subledger"], Decimal("50.00"))
        self.assertEqual(rec["closing_gl"], Decimal("50.00"))
        self.assertEqual(rec["variance"], Decimal("0.00"))
        self.assertTrue(rec["balanced"])

    def test_missing_gl_posting_flagged(self):
        from core.services import reports
        self._mv("50")  # movement but no GL entry
        rec = reports.inventory_gl_reconciliation(self.t)
        self.assertEqual(rec["closing_subledger"], Decimal("50.00"))
        self.assertEqual(rec["closing_gl"], Decimal("0.00"))
        self.assertEqual(rec["variance"], Decimal("50.00"))
        self.assertFalse(rec["balanced"])
        # Periodic check flags it.
        flagged = reports.check_inventory_gl_variance(tenant=self.t)
        self.assertEqual(len(flagged), 1)

    def test_duplicate_gl_posting_flagged(self):
        from core.services import reports
        self._mv("50"); self._gl(debit="50"); self._gl(debit="50")  # double-posted GL
        rec = reports.inventory_gl_reconciliation(self.t)
        self.assertEqual(rec["closing_gl"], Decimal("100.00"))
        self.assertEqual(rec["variance"], Decimal("-50.00"))

    def test_backdated_movement_lands_in_opening(self):
        import datetime
        from django.utils import timezone
        from core.services import reports
        old = timezone.make_aware(datetime.datetime(2026, 1, 1, 12, 0, 0))
        self._mv("50", created=old)
        self._gl(debit="50", entry_date=datetime.date(2026, 1, 1))
        # As-of everything: reconciles.
        self.assertEqual(reports.inventory_gl_reconciliation(self.t)["variance"], Decimal("0.00"))
        # With a window starting after the backdate, it sits in opening, not the period.
        rec = reports.inventory_gl_reconciliation(self.t, date_from=datetime.date(2026, 6, 1))
        self.assertEqual(rec["opening_subledger"], Decimal("50.00"))
        self.assertEqual(rec["movement_debits"], Decimal("0.00"))
        self.assertEqual(rec["movement_ids"], [])
        self.assertTrue(rec["balanced"])


class CycleCountGlValuationTests(TestCase):
    """Cycle-count variance posts to the GL, valued at the issued lot's cost."""

    def setUp(self):
        import datetime
        from core.models import OrgMembership
        self.t = Tenant.objects.create(name="CCGL Co")
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.p = Product.objects.create(tenant=self.t, sku="SKU-CC", name="P",
                                        cost_method=Product.CostMethod.FIFO, track_lots=True, track_expiry=True)
        self.eA = datetime.date(2026, 6, 1)
        self.eB = datetime.date(2026, 12, 1)
        from core.services.inventory import apply_movement
        apply_movement(tenant=self.t, product=self.p, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="a", unit_cost=Decimal("2.00"),
                       lot_code="A", expiry_date=self.eA)
        apply_movement(tenant=self.t, product=self.p, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="b", unit_cost=Decimal("8.00"),
                       lot_code="B", expiry_date=self.eB)
        self.user = User.objects.create_user("ccu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t, role="ADMIN", is_default=True)
        self.client.login(username="ccu", password="pw")

    def test_variance_valued_per_lot_and_posted_to_gl(self):
        from django.urls import reverse
        from core.models import CycleCount, CycleCountLine
        cc = CycleCount.objects.create(tenant=self.t, location=self.loc, status=CycleCount.Status.APPROVED)
        # Shrink each lot by 1: lot A (cost 2) and lot B (cost 8).
        CycleCountLine.objects.create(cycle_count=cc, product=self.p, lot_code="A", expiry_date=self.eA,
                                      system_qty=Decimal("10"), counted_qty=Decimal("9"), variance_qty=Decimal("-1"))
        CycleCountLine.objects.create(cycle_count=cc, product=self.p, lot_code="B", expiry_date=self.eB,
                                      system_qty=Decimal("10"), counted_qty=Decimal("9"), variance_qty=Decimal("-1"))
        self.client.post(reverse("cycle_count_post", args=[cc.id]))

        # Each variance movement is valued at its own lot's cost (not the 5.00 avg).
        mv = {m.lot_code: m.value for m in
              InventoryMovement.objects.filter(tenant=self.t, ref_type="CYCLE_COUNT")}
        self.assertEqual(mv["A"], Decimal("-2.00"))   # 1 @ 2.00
        self.assertEqual(mv["B"], Decimal("-8.00"))   # 1 @ 8.00, NOT 1 @ 5.00 average
        # GL inventory control credited by exactly the movement total (2 + 8 = 10):
        # the cycle count now posts to the GL, valued the same as the movements.
        self.assertEqual(_account_balance(self.t, "1000"), Decimal("-10.00"))
        self.assertEqual(sum(mv.values()), Decimal("-10.00"))


class ReservationLifecycleTests(TestCase):
    """Reservations transition ACTIVE -> CONSUMED on fulfilment (partial-aware),
    -> RELEASED on cancellation, and ATP reflects only active holds."""

    def setUp(self):
        self.t = Tenant.objects.create(name="Resv LC Co")
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.p = Product.objects.create(tenant=self.t, sku="SKU-RL", name="P")
        InventoryBalance.objects.create(tenant=self.t, product=self.p, location=self.loc,
                                        on_hand=Decimal("50"), reserved=Decimal("0"))

    def _reserve(self, qty, ref="ORD-1"):
        from core.services.inventory import reserve_stock
        reserve_stock(tenant=self.t, product=self.p, location=self.loc, qty=Decimal(qty),
                      ref_type="ORDER", ref_id=ref)

    def _bal(self):
        return InventoryBalance.objects.get(tenant=self.t, product=self.p, location=self.loc)

    def test_full_fulfilment_consumes(self):
        from core.services.inventory import consume_reservations
        from core.models import InventoryReservation
        self._reserve("8")
        self.assertEqual(self._bal().reserved, Decimal("8.00"))
        consumed = consume_reservations(tenant=self.t, ref_type="ORDER", ref_id="ORD-1")
        self.assertEqual(consumed, Decimal("8.00"))
        self.assertEqual(self._bal().reserved, Decimal("0.00"))  # ATP restored
        r = InventoryReservation.objects.get(tenant=self.t, ref_id="ORD-1")
        self.assertEqual(r.status, InventoryReservation.Status.CONSUMED)

    def test_partial_fulfilment_splits(self):
        from core.services.inventory import consume_reservations
        from core.models import InventoryReservation
        self._reserve("8")
        consumed = consume_reservations(tenant=self.t, ref_type="ORDER", ref_id="ORD-1", qty=Decimal("3"))
        self.assertEqual(consumed, Decimal("3.00"))
        self.assertEqual(self._bal().reserved, Decimal("5.00"))  # 5 still held
        active = InventoryReservation.objects.get(tenant=self.t, ref_id="ORD-1",
                                                  status=InventoryReservation.Status.ACTIVE)
        self.assertEqual(active.qty, Decimal("5.00"))
        consumed_row = InventoryReservation.objects.get(tenant=self.t, ref_id="ORD-1",
                                                        status=InventoryReservation.Status.CONSUMED)
        self.assertEqual(consumed_row.qty, Decimal("3.00"))

    def test_cancellation_releases(self):
        from core.services.inventory import release_reservations
        from core.models import InventoryReservation
        self._reserve("8")
        release_reservations(tenant=self.t, ref_type="ORDER", ref_id="ORD-1")
        self.assertEqual(self._bal().reserved, Decimal("0.00"))
        self.assertEqual(InventoryReservation.objects.get(tenant=self.t, ref_id="ORD-1").status,
                         InventoryReservation.Status.RELEASED)

    def test_atp_excludes_consumed(self):
        from core.services.inventory import consume_reservations
        self._reserve("20")
        b = self._bal()
        self.assertEqual(b.on_hand - b.reserved, Decimal("30.00"))  # 50 - 20 held
        consume_reservations(tenant=self.t, ref_type="ORDER", ref_id="ORD-1", qty=Decimal("20"))
        b = self._bal()
        self.assertEqual(b.on_hand - b.reserved, Decimal("50.00"))  # hold cleared

    def test_expire_stale_reservations(self):
        import datetime
        from django.utils import timezone
        from core.services.inventory import expire_stale_reservations
        from core.models import InventoryReservation
        self._reserve("4")
        # Backdate the reservation so it looks stale.
        InventoryReservation.objects.filter(tenant=self.t, ref_id="ORD-1").update(
            created_at=timezone.make_aware(datetime.datetime(2026, 1, 1, 0, 0, 0)))
        n = expire_stale_reservations(
            tenant=self.t, older_than=timezone.make_aware(datetime.datetime(2026, 3, 1, 0, 0, 0)))
        self.assertEqual(n, 1)
        self.assertEqual(self._bal().reserved, Decimal("0.00"))


class CycleCountValuationCatchupTests(TestCase):
    """Catch-up reconciliation for historical cycle-count valuation drift."""

    def setUp(self):
        import datetime
        self.t = Tenant.objects.create(name="Catchup Co")  # signal seeds GL accounts
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.p = Product.objects.create(tenant=self.t, sku="SKU-CU", name="P",
                                        cost_method=Product.CostMethod.FIFO, track_lots=True, track_expiry=True)
        self.eA = datetime.date(2026, 6, 1)
        self.eB = datetime.date(2026, 12, 1)

    def _receive(self, lot, expiry, qty, cost, gl=False, ref="r"):
        from core.services.inventory import apply_movement
        from core.services.gl import post_inventory_receipt
        apply_movement(tenant=self.t, product=self.p, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal(qty), ref_type="T", ref_id=ref, unit_cost=Decimal(cost),
                       lot_code=lot, expiry_date=expiry)
        if gl:
            post_inventory_receipt(self.t, Decimal(qty) * Decimal(cost), ref)

    def _hist_cc(self, lot, expiry, qty_delta, value, ref_id="hist", created=None):
        """A pre-existing cycle-count movement created directly (simulating data
        posted before lot-scoped costing/GL), optionally backdated."""
        m = InventoryMovement.objects.create(
            tenant=self.t, site_id=self.loc.site_id, product=self.p, location=self.loc,
            movement_type="ADJUSTMENT", qty_delta=Decimal(qty_delta), unit_cost=None,
            value=Decimal(value), ref_type="CYCLE_COUNT", ref_id=ref_id,
            lot_code=lot, expiry_date=expiry)
        if created is not None:
            InventoryMovement.objects.filter(id=m.id).update(created_at=created)
            m.refresh_from_db()
        return m

    def _cc_gl(self, value, ref_id="hist"):
        from core.models import GLAccount, JournalEntry, JournalLine
        from django.utils import timezone
        inv = GLAccount.objects.get(tenant=self.t, code="1000")
        adj = GLAccount.objects.get(tenant=self.t, code="5200")
        je = JournalEntry.objects.create(tenant=self.t, entry_date=timezone.localdate(),
                                         ref_type="CYCLE_COUNT", ref_id=ref_id, memo="hist")
        amt = abs(Decimal(value))
        if Decimal(value) < 0:
            JournalLine.objects.create(entry=je, account=adj, debit=amt, credit=Decimal("0.00"))
            JournalLine.objects.create(entry=je, account=inv, debit=Decimal("0.00"), credit=amt)
        else:
            JournalLine.objects.create(entry=je, account=inv, debit=amt, credit=Decimal("0.00"))
            JournalLine.objects.create(entry=je, account=adj, debit=Decimal("0.00"), credit=amt)
        return je

    def test_no_variance_when_already_lot_correct_with_gl(self):
        from core.services import inventory_corrections as recon
        self._receive("B", self.eB, "10", "8.00")
        # Movement valued at lot cost (8) AND already has a GL -> nothing to do.
        self._hist_cc("B", self.eB, "-2", "-16.00")
        self._cc_gl("-16.00")
        self.assertEqual(recon.find_drift(self.t), [])

    def test_variance_detected_dry_run_makes_no_changes(self):
        from core.models import CycleCountValuationCorrection
        from core.services import inventory_corrections as recon
        self._receive("A", self.eA, "10", "2.00")
        self._receive("B", self.eB, "10", "8.00")
        self._hist_cc("B", self.eB, "-2", "-10.00")  # valued at average 5, not lot 8; no GL
        rows = recon.find_drift(self.t)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["expected_value"], Decimal("-16.00"))  # 2 @ lot cost 8
        self.assertEqual(rows[0]["variance"], Decimal("-6.00"))
        self.assertEqual(rows[0]["valuation_source"], CycleCountValuationCorrection.Source.LOT_LAYER)
        # Dry-run / find only: nothing posted.
        self.assertEqual(CycleCountValuationCorrection.objects.count(), 0)

    def test_correction_applied_once_and_idempotent(self):
        from core.models import CycleCountValuationCorrection
        from core.services import inventory_corrections as recon
        self._receive("A", self.eA, "10", "2.00")
        self._receive("B", self.eB, "10", "8.00")
        m = self._hist_cc("B", self.eB, "-2", "-10.00")

        summary = recon.apply_corrections(self.t)
        self.assertEqual(summary["corrected_count"], 1)
        corr = CycleCountValuationCorrection.objects.get(original_movement=m)
        self.assertEqual(corr.expected_value, Decimal("-16.00"))
        self.assertEqual(corr.variance, Decimal("-6.00"))
        self.assertIsNotNone(corr.correction_journal)
        self.assertIsNotNone(corr.reval_movement)
        self.assertEqual(corr.reval_movement.value, Decimal("-6.00"))  # subledger top-up
        # Re-run: no double correction.
        summary2 = recon.apply_corrections(self.t)
        self.assertEqual(summary2["corrected_count"], 0)
        self.assertEqual(CycleCountValuationCorrection.objects.count(), 1)

    def test_legacy_lot_without_layer_uses_product_average_fallback(self):
        from core.models import CycleCountValuationCorrection
        from core.services import inventory_corrections as recon
        self._receive("A", self.eA, "10", "2.00")
        self._receive("B", self.eB, "10", "8.00")  # company average = 5.00
        # Lot C was never received -> no cost layer. Movement valued at average 5.
        self._hist_cc("C", self.eA, "-1", "-5.00", ref_id="histC")
        rows = recon.find_drift(self.t)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["valuation_source"], CycleCountValuationCorrection.Source.PRODUCT_AVERAGE)
        self.assertEqual(rows[0]["expected_value"], Decimal("-5.00"))  # fallback = average
        self.assertEqual(rows[0]["variance"], Decimal("0.00"))         # value already average

    def test_gl_reconciliation_improves_after_correction(self):
        from core.services import inventory_corrections as recon
        from core.services import reports
        # Clean baseline: receipts post matching GL so subledger == GL.
        self._receive("A", self.eA, "10", "2.00", gl=True, ref="ra")
        self._receive("B", self.eB, "10", "8.00", gl=True, ref="rb")
        self.assertEqual(reports.inventory_gl_reconciliation(self.t)["variance"], Decimal("0.00"))

        # Historical cycle count: subledger gets -10 (average), no GL -> drift.
        self._hist_cc("B", self.eB, "-2", "-10.00")
        before = reports.inventory_gl_reconciliation(self.t)["variance"]
        self.assertEqual(before, Decimal("-10.00"))

        recon.apply_corrections(self.t)
        after = reports.inventory_gl_reconciliation(self.t)["variance"]
        self.assertEqual(after, Decimal("0.00"))  # subledger and GL both at lot-correct value
        self.assertLess(abs(after), abs(before))

    def test_closed_period_correction_posts_to_current_period(self):
        import datetime
        from django.utils import timezone
        from core.models import CycleCountValuationCorrection
        from core.services import inventory_corrections as recon
        self._receive("A", self.eA, "10", "2.00")
        self._receive("B", self.eB, "10", "8.00")
        old = timezone.make_aware(datetime.datetime(2026, 1, 15, 9, 0, 0))
        self._hist_cc("B", self.eB, "-2", "-10.00", created=old)

        posting = datetime.date(2026, 7, 1)
        recon.apply_corrections(self.t, lock_date=datetime.date(2026, 3, 31),
                                posting_date=posting)
        corr = CycleCountValuationCorrection.objects.get()
        self.assertTrue(corr.posted_to_current_period)
        self.assertEqual(corr.correction_journal.entry_date, posting)  # not the Jan movement date
        self.assertIn("closed period", corr.note)

    def test_closed_period_block_skips_correction(self):
        import datetime
        from django.utils import timezone
        from core.models import CycleCountValuationCorrection
        from core.services import inventory_corrections as recon
        self._receive("B", self.eB, "10", "8.00")
        old = timezone.make_aware(datetime.datetime(2026, 1, 15, 9, 0, 0))
        self._hist_cc("B", self.eB, "-2", "-10.00", created=old)

        summary = recon.apply_corrections(self.t, lock_date=datetime.date(2026, 3, 31),
                                          block_closed=True)
        self.assertEqual(summary["corrected_count"], 0)
        self.assertEqual(summary["blocked_count"], 1)
        self.assertEqual(CycleCountValuationCorrection.objects.count(), 0)


class UomConversionServiceTests(TestCase):
    """Conversion engine: stock is kept in base units; lines may use other UOMs."""

    def setUp(self):
        from core.models import UnitOfMeasure
        self.t = Tenant.objects.create(name="UOM Svc Co")
        self.each = UnitOfMeasure.objects.create(tenant=self.t, code="EACH")
        self.case = UnitOfMeasure.objects.create(tenant=self.t, code="CASE")
        self.p = Product.objects.create(tenant=self.t, sku="U1", name="P", base_uom=self.each)

    def test_identity_when_no_or_base_uom(self):
        from core.services.uom import to_base_qty
        self.assertEqual(to_base_qty(self.p, Decimal("5"), None), Decimal("5.00"))
        self.assertEqual(to_base_qty(self.p, Decimal("5"), self.each), Decimal("5.00"))

    def test_global_conversion_qty_and_cost(self):
        from core.models import UOMConversion
        from core.services.uom import to_base_qty, base_unit_cost
        UOMConversion.objects.create(tenant=self.t, product=None, from_uom=self.case,
                                     to_uom=self.each, multiplier=Decimal("12"))
        self.assertEqual(to_base_qty(self.p, Decimal("5"), self.case), Decimal("60.00"))
        # Money preserved: 24/case -> 2/each.
        self.assertEqual(base_unit_cost(self.p, Decimal("24.00"), self.case), Decimal("2.0000"))

    def test_product_specific_overrides_global(self):
        from core.models import UOMConversion
        from core.services.uom import to_base_qty
        UOMConversion.objects.create(tenant=self.t, product=None, from_uom=self.case,
                                     to_uom=self.each, multiplier=Decimal("12"))
        UOMConversion.objects.create(tenant=self.t, product=self.p, from_uom=self.case,
                                     to_uom=self.each, multiplier=Decimal("6"))
        self.assertEqual(to_base_qty(self.p, Decimal("2"), self.case), Decimal("12.00"))  # uses 6

    def test_reverse_rule_is_inverted(self):
        from core.models import UOMConversion
        from core.services.uom import to_base_qty
        # Only the base->case rule exists: 1 EACH = 0.5 CASE => 1 CASE = 2 EACH.
        UOMConversion.objects.create(tenant=self.t, product=None, from_uom=self.each,
                                     to_uom=self.case, multiplier=Decimal("0.5"))
        self.assertEqual(to_base_qty(self.p, Decimal("3"), self.case), Decimal("6.00"))

    def test_missing_rule_raises(self):
        from django.core.exceptions import ValidationError
        from core.services.uom import to_base_qty
        with self.assertRaises(ValidationError):
            to_base_qty(self.p, Decimal("5"), self.case)  # no conversion configured


class UomPurchasingTests(TestCase):
    """Buying in a purchase UOM stores/cost stock in the product's base unit."""

    def setUp(self):
        from core.models import UnitOfMeasure, UOMConversion
        self.tenant = Tenant.objects.create(name="UOM Buy Co")
        self.each = UnitOfMeasure.objects.create(tenant=self.tenant, code="EACH")
        self.case = UnitOfMeasure.objects.create(tenant=self.tenant, code="CASE")
        UOMConversion.objects.create(tenant=self.tenant, product=None, from_uom=self.case,
                                     to_uom=self.each, multiplier=Decimal("12"))
        self.loc = Location.objects.create(tenant=self.tenant, name="WH")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="S")
        self.product = Product.objects.create(tenant=self.tenant, sku="SKU-UB", name="P", base_uom=self.each)
        self.po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-UB", supplier=self.supplier,
                                               status=PurchaseOrder.Status.SUBMITTED)
        # 5 CASE @ 24/CASE.
        self.pol = PurchaseOrderLine.objects.create(po=self.po, product=self.product, uom=self.case,
                                                    ordered_qty=Decimal("5"), unit_cost=Decimal("24.00"))
        self.shipment = Shipment.objects.create(tenant=self.tenant, po=self.po, from_supplier=self.supplier, destination=self.loc)
        self.sl = ShipmentLine.objects.create(shipment=self.shipment, po_line=self.pol, expected_qty=Decimal("5"))
        self.user = User.objects.create_user("ubu", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Warehouse")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="ubu", password="pw")

    def test_receive_in_cases_stores_base_eaches(self):
        from core.models import JournalEntry
        from django.urls import reverse
        resp = self.client.post(reverse("receive_po", args=[self.po.id]),
                                {"grn_number": "GRN-UB", f"recv_{self.sl.id}": "5"})
        self.assertEqual(resp.status_code, 302)
        # 5 CASE x 12 = 60 EACH on hand, costed at 2.00/each.
        bal = InventoryBalance.objects.get(tenant=self.tenant, product=self.product, location=self.loc)
        self.assertEqual(bal.on_hand, Decimal("60.00"))
        self.product.refresh_from_db()
        self.assertEqual(self.product.average_cost, Decimal("2.0000"))
        # Money is unchanged: GRN capitalises 5 x 24 = 120 and balances.
        je = JournalEntry.objects.get(tenant=self.tenant, ref_type="GRN")
        self.assertEqual(je.total_debit, je.total_credit)
        self.assertEqual(je.total_debit, Decimal("120.00"))
        # PO received/open stay in the purchase UOM (cases).
        self.pol.refresh_from_db()
        self.assertEqual(self.pol.received_qty, Decimal("5.00"))
        self.assertEqual(self.pol.open_qty, Decimal("0.00"))


class UomSalesTests(TestCase):
    """Selling in a UOM relieves stock and books COGS in the base unit."""

    def setUp(self):
        from core.models import UnitOfMeasure, UOMConversion
        self.t = Tenant.objects.create(name="UOM Sell Co")
        self.each = UnitOfMeasure.objects.create(tenant=self.t, code="EACH")
        self.case = UnitOfMeasure.objects.create(tenant=self.t, code="CASE")
        UOMConversion.objects.create(tenant=self.t, product=None, from_uom=self.case,
                                     to_uom=self.each, multiplier=Decimal("12"))
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.product = Product.objects.create(tenant=self.t, sku="SKU-US", name="P", base_uom=self.each)

    def test_invoice_in_cases_relieves_base_eaches_and_costs_at_base(self):
        from core.services.inventory import apply_movement
        from core.models import JournalEntry, GLAccount, Customer
        # 100 EACH on hand @ 3.00 (base).
        apply_movement(tenant=self.t, product=self.product, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("100"), ref_type="T", ref_id="r", unit_cost=Decimal("3.00"))
        cust = Customer.objects.create(tenant=self.t, name="C")
        inv = CustomerInvoice.objects.create(tenant=self.t, customer=cust, invoice_number="INV-US", location=self.loc)
        std = TaxCode.objects.get(tenant=self.t, code="STD")
        # Sell 2 CASE = 24 EACH.
        CustomerInvoiceLine.objects.create(invoice=inv, product=self.product, description="P",
                                           qty=Decimal("2"), uom=self.case, unit_price=Decimal("50.00"), tax_code=std)
        post_customer_invoice(inv)

        bal = InventoryBalance.objects.get(tenant=self.t, product=self.product, location=self.loc)
        self.assertEqual(bal.on_hand, Decimal("76.00"))  # 100 - 24 EACH relieved
        cogs = JournalEntry.objects.get(tenant=self.t, ref_type="COGS")
        self.assertEqual(cogs.total_debit, Decimal("72.00"))  # 24 EACH @ 3.00


class NearExpiryReportTests(TestCase):
    """Near-expiry lot report (visibility only)."""

    def setUp(self):
        import datetime
        self.t = Tenant.objects.create(name="Expiry Co")
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.p = Product.objects.create(tenant=self.t, sku="SKU-EX", name="P",
                                        cost_method=Product.CostMethod.FIFO, track_lots=True, track_expiry=True)
        self.today = datetime.date(2026, 6, 1)

    def _recv(self, lot, expiry, qty, cost):
        from core.services.inventory import apply_movement
        apply_movement(tenant=self.t, product=self.p, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal(qty), ref_type="T", ref_id=lot, unit_cost=Decimal(cost),
                       lot_code=lot, expiry_date=expiry)

    def _issue(self, lot, expiry, qty):
        from core.services.inventory import apply_movement
        apply_movement(tenant=self.t, product=self.p, location=self.loc, movement_type="SALE",
                       qty_delta=Decimal(qty), ref_type="T", ref_id=lot + "s",
                       lot_code=lot, expiry_date=expiry)

    def test_statuses_and_window(self):
        import datetime
        from core.services import reports
        self._recv("NEAR", self.today + datetime.timedelta(days=20), "10", "8.00")
        self._recv("FAR", self.today + datetime.timedelta(days=200), "10", "2.00")
        self._recv("OLD", self.today - datetime.timedelta(days=5), "10", "5.00")
        rows = {r["lot_code"]: r for r in reports.near_expiry_lots(self.t, days=30, today=self.today)}
        self.assertEqual(rows["NEAR"]["status"], "near_expiry")   # within window
        self.assertEqual(rows["OLD"]["status"], "expired")        # past expiry
        self.assertNotIn("FAR", rows)                             # outside window, default view
        self.assertEqual(rows["NEAR"]["days_until"], 20)
        # status='all' brings the OK lot in.
        all_rows = {r["lot_code"]: r for r in reports.near_expiry_lots(self.t, days=30, today=self.today, status="all")}
        self.assertEqual(all_rows["FAR"]["status"], "okay")

    def test_value_uses_lot_cost_layer(self):
        import datetime
        from core.services import reports
        # Company average becomes 5.00, but NEAR's own layer cost is 8.00.
        self._recv("NEAR", self.today + datetime.timedelta(days=10), "10", "8.00")
        self._recv("OTHER", self.today + datetime.timedelta(days=10), "10", "2.00")
        rows = {r["lot_code"]: r for r in reports.near_expiry_lots(self.t, days=30, today=self.today)}
        self.assertEqual(rows["NEAR"]["value"], Decimal("80.00"))   # 10 @ 8, not 10 @ 5 avg
        self.assertEqual(rows["NEAR"]["valuation_source"], "lot_layer")

    def test_zero_balance_excluded_unless_audit(self):
        import datetime
        from core.services import reports
        exp = self.today + datetime.timedelta(days=10)
        self._recv("Z", exp, "5", "3.00")
        self._issue("Z", exp, "-5")  # depletes the lot
        default = {r["lot_code"] for r in reports.near_expiry_lots(self.t, days=30, today=self.today)}
        self.assertNotIn("Z", default)
        audit = {r["lot_code"] for r in reports.near_expiry_lots(self.t, days=30, today=self.today, include_zero=True)}
        self.assertIn("Z", audit)


class LotTraceTests(TestCase):
    """Lot traceability across receipt -> transfer -> issue, with costing trail."""

    def setUp(self):
        self.t = Tenant.objects.create(name="Trace Co")
        self.a = Location.objects.create(tenant=self.t, name="A", type=Location.Type.WAREHOUSE)
        self.b = Location.objects.create(tenant=self.t, name="B", type=Location.Type.WAREHOUSE)
        self.p = Product.objects.create(tenant=self.t, sku="SKU-TRC", name="P",
                                        cost_method=Product.CostMethod.FIFO, track_lots=True)

    def test_trace_receipt_transfer_issue_with_costs(self):
        from core.models import (InventoryTransfer, InventoryTransferLine, GoodsReceipt,
                                  PurchaseOrder, Supplier)
        from core.services.inventory import apply_movement
        from core.views import _post_transfer
        from core.services import reports
        # Receipt against a GRN (with PO + supplier) so the source is traceable.
        supplier = Supplier.objects.create(tenant=self.t, name="Acme Supplies")
        po = PurchaseOrder.objects.create(tenant=self.t, po_number="PO-TRC", supplier=supplier)
        GoodsReceipt.objects.create(tenant=self.t, po=po, grn_number="GRN-TRC", received_to=self.a,
                                    status=GoodsReceipt.Status.POSTED)
        apply_movement(tenant=self.t, product=self.p, location=self.a, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="GRN", ref_id="GRN-TRC",
                       unit_cost=Decimal("4.00"), lot_code="L1")
        # Transfer 4 A->B (one-step), carrying the lot.
        tr = InventoryTransfer.objects.create(tenant=self.t, transfer_number="TR-TRC",
                                              from_location=self.a, to_location=self.b)
        InventoryTransferLine.objects.create(transfer=tr, product=self.p, qty=Decimal("4"), lot_code="L1")
        _post_transfer(tr)
        # Issue 3 from B.
        apply_movement(tenant=self.t, product=self.p, location=self.b, movement_type="SALE",
                       qty_delta=Decimal("-3"), ref_type="ORDER", ref_id="SO-1", lot_code="L1")

        data = reports.lot_trace(self.t, self.p.id, "L1")
        self.assertIsNotNone(data)
        types = [row["m"].movement_type for row in data["movements"]]
        for mt in ("RECEIVE", "TRANSFER_OUT", "TRANSFER_IN", "SALE"):
            self.assertIn(mt, types)
        # Receipt is linked to its GRN / PO / supplier.
        recv = next(r for r in data["movements"] if r["m"].movement_type == "RECEIVE")
        self.assertEqual(recv["grn"].grn_number, "GRN-TRC")
        self.assertEqual(recv["po"].po_number, "PO-TRC")
        self.assertEqual(recv["supplier"].name, "Acme Supplies")
        # Issue carries the InventoryIssueCost linkage.
        sale = next(r for r in data["movements"] if r["m"].movement_type == "SALE")
        self.assertTrue(sale["issue_costs"])
        self.assertEqual(sale["issue_costs"][0].unit_cost, Decimal("4.0000"))
        # Current balances by location are present.
        bal = {b.location.name: b.on_hand for b in data["balances"]}
        self.assertEqual(bal["A"], Decimal("6.00"))   # 10 received - 4 transferred
        self.assertEqual(bal["B"], Decimal("1.00"))   # 4 in - 3 issued

    def test_trace_respects_tenant_isolation(self):
        from core.services.inventory import apply_movement
        from core.services import reports
        apply_movement(tenant=self.t, product=self.p, location=self.a, movement_type="RECEIVE",
                       qty_delta=Decimal("5"), ref_type="T", ref_id="r", unit_cost=Decimal("4.00"), lot_code="L1")
        other = Tenant.objects.create(name="Other Co")
        # Another tenant cannot trace this tenant's product/lot.
        self.assertIsNone(reports.lot_trace(other, self.p.id, "L1"))


class UomNavAndSeedTests(TestCase):
    """UOM master-data pages are reachable from the nav, and the seed is idempotent."""

    def test_admin_sidebar_has_uom_links(self):
        from core.roles import sidebar_for_role, ADMIN, SALES
        admin_urls = [u for (_, items) in sidebar_for_role(ADMIN) for (_, u, _) in items]
        self.assertIn("/uoms/", admin_urls)
        self.assertIn("/uom-conversions/", admin_urls)
        # Admin-only (matches the view permissions): a non-admin doesn't see them.
        sales_urls = [u for (_, items) in sidebar_for_role(SALES) for (_, u, _) in items]
        self.assertNotIn("/uoms/", sales_urls)

    def test_seed_uom_demo_is_idempotent(self):
        from django.core.management import call_command
        from core.models import UnitOfMeasure, UOMConversion
        t = Tenant.objects.create(name="Seed Co")
        call_command("seed_uom_demo", "--tenant", "Seed Co")
        call_command("seed_uom_demo", "--tenant", "Seed Co")  # second run = no dupes
        self.assertEqual(UnitOfMeasure.objects.filter(tenant=t, code="CASE").count(), 1)
        self.assertEqual(UnitOfMeasure.objects.filter(tenant=t, code="EA").count(), 1)
        conv = UOMConversion.objects.get(tenant=t, from_uom__code="CASE")
        self.assertEqual(conv.to_uom.code, "EA")
        self.assertEqual(conv.multiplier, Decimal("12"))


class UiPassRenderTests(TestCase):
    """Smoke-test the UI-pass templates render (UOM dropdowns, transfer states,
    bin section, reports help)."""

    def setUp(self):
        from core.models import OrgMembership
        self.t = Tenant.objects.create(name="UI Pass Co")
        self.user = User.objects.create_user("uip", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t, role="ADMIN", is_default=True)
        self.client.login(username="uip", password="pw")

    def test_form_and_report_pages_render(self):
        from django.urls import reverse
        for name in ["po_create", "quote_create", "corder_create", "sales_order_create",
                     "ar_invoice_create", "inventory_list", "reports_index",
                     "report_near_expiry", "report_lot_trace"]:
            resp = self.client.get(reverse(name))
            self.assertEqual(resp.status_code, 200, f"{name} did not render: {resp.status_code}")
        # Reports page documents the control commands.
        resp = self.client.get(reverse("reports_index"))
        self.assertContains(resp, "reconcile_cycle_count_valuation")
        self.assertContains(resp, "check_inventory_gl")

    def test_inventory_list_shows_bin_section(self):
        from django.urls import reverse
        from core.models import Bin
        from core.services.inventory import apply_movement
        loc = Location.objects.create(tenant=self.t, name="WH")
        b = Bin.objects.create(tenant=self.t, location=loc, code="A1")
        p = Product.objects.create(tenant=self.t, sku="SKU-UIB", name="P")
        apply_movement(tenant=self.t, product=p, location=loc, movement_type="RECEIVE",
                       qty_delta=Decimal("5"), ref_type="T", ref_id="1", unit_cost=Decimal("2.00"), bin=b)
        resp = self.client.get(reverse("inventory_list"))
        self.assertContains(resp, "By bin")
        self.assertContains(resp, "A1")

    def test_transfer_detail_renders_in_each_state(self):
        from core.models import InventoryTransfer, InventoryTransferLine
        from core.services.inventory import apply_movement
        from core.views import _dispatch_transfer
        a = Location.objects.create(tenant=self.t, name="A", type=Location.Type.WAREHOUSE)
        b = Location.objects.create(tenant=self.t, name="B", type=Location.Type.WAREHOUSE)
        p = Product.objects.create(tenant=self.t, sku="SKU-UIT", name="P")
        apply_movement(tenant=self.t, product=p, location=a, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="1", unit_cost=Decimal("2.00"))
        tr = InventoryTransfer.objects.create(tenant=self.t, transfer_number="TR-UI",
                                              from_location=a, to_location=b)
        InventoryTransferLine.objects.create(transfer=tr, product=p, qty=Decimal("4"))
        # Draft shows the Dispatch action.
        resp = self.client.get(f"/transfers/{tr.id}/")
        self.assertContains(resp, "Dispatch")
        # Dispatched shows the in-transit banner and the receive form.
        _dispatch_transfer(tr)
        resp = self.client.get(f"/transfers/{tr.id}/")
        self.assertContains(resp, "In transit")
        self.assertContains(resp, "Receive into")


class BinBalanceTests(TestCase):
    """Bin-level on-hand tracking (a sub-balance of the location balance)."""

    def setUp(self):
        from core.models import Bin
        self.t = Tenant.objects.create(name="Bin Co")
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.bin_a = Bin.objects.create(tenant=self.t, location=self.loc, code="A1")
        self.bin_b = Bin.objects.create(tenant=self.t, location=self.loc, code="B2")
        self.p = Product.objects.create(tenant=self.t, sku="SKU-BIN", name="P")

    def test_movements_track_per_bin_and_roll_up_to_location(self):
        from core.services.inventory import apply_movement, bin_balances
        from core.models import InventoryBalance, InventoryBinBalance
        apply_movement(tenant=self.t, product=self.p, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="1", unit_cost=Decimal("2.00"), bin=self.bin_a)
        apply_movement(tenant=self.t, product=self.p, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("4"), ref_type="T", ref_id="2", unit_cost=Decimal("2.00"), bin=self.bin_b)
        # Location balance is the total; bins hold the split.
        self.assertEqual(InventoryBalance.objects.get(tenant=self.t, product=self.p, location=self.loc).on_hand,
                         Decimal("14.00"))
        self.assertEqual(InventoryBinBalance.objects.get(tenant=self.t, bin=self.bin_a).on_hand, Decimal("10.00"))
        self.assertEqual(InventoryBinBalance.objects.get(tenant=self.t, bin=self.bin_b).on_hand, Decimal("4.00"))
        # Issue from bin A only.
        apply_movement(tenant=self.t, product=self.p, location=self.loc, movement_type="SALE",
                       qty_delta=Decimal("-3"), ref_type="T", ref_id="3", bin=self.bin_a)
        self.assertEqual(InventoryBinBalance.objects.get(tenant=self.t, bin=self.bin_a).on_hand, Decimal("7.00"))
        self.assertEqual(InventoryBinBalance.objects.get(tenant=self.t, bin=self.bin_b).on_hand, Decimal("4.00"))
        # Helper: only positive bin balances, filterable.
        rows = list(bin_balances(self.t, location=self.loc))
        self.assertEqual(len(rows), 2)
        self.assertEqual({r.bin.code for r in rows}, {"A1", "B2"})

    def test_strict_control_blocks_bin_oversell(self):
        from core.services.inventory import apply_movement
        from django.core.exceptions import ValidationError
        self.t.block_negative_stock = True
        self.t.save(update_fields=["block_negative_stock"])
        apply_movement(tenant=self.t, product=self.p, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("5"), ref_type="T", ref_id="1", unit_cost=Decimal("2.00"), bin=self.bin_a)
        # Bin B has nothing; issuing from it is rejected even though the location
        # has stock (in bin A).
        with self.assertRaises(ValidationError):
            apply_movement(tenant=self.t, product=self.p, location=self.loc, movement_type="SALE",
                           qty_delta=Decimal("-1"), ref_type="T", ref_id="2", bin=self.bin_b)

    def test_no_bin_movements_leave_bin_balances_empty(self):
        from core.services.inventory import apply_movement
        from core.models import InventoryBinBalance
        apply_movement(tenant=self.t, product=self.p, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="1", unit_cost=Decimal("2.00"))
        self.assertEqual(InventoryBinBalance.objects.filter(tenant=self.t).count(), 0)


class InTransitTransferTests(TestCase):
    """Two-step dispatch -> receive transfers: value-neutral, partial receipt,
    in-transit GL, shortage write-off, cancellation."""

    def setUp(self):
        from core.models import InventoryTransfer, InventoryTransferLine
        from core.services.inventory import apply_movement
        from core.services.gl import post_inventory_receipt
        self.t = Tenant.objects.create(name="Transit Co")
        self.a = Location.objects.create(tenant=self.t, name="WH-A")
        self.b = Location.objects.create(tenant=self.t, name="WH-B")
        self.p = Product.objects.create(tenant=self.t, sku="SKU-TR", name="P",
                                        cost_method=Product.CostMethod.FIFO)
        # 100 @ 2.00 at A, with a matching GL receipt so the baseline reconciles.
        apply_movement(tenant=self.t, product=self.p, location=self.a, movement_type="RECEIVE",
                       qty_delta=Decimal("100"), ref_type="T", ref_id="r", unit_cost=Decimal("2.00"))
        post_inventory_receipt(self.t, Decimal("200.00"), "r")
        self.tr = InventoryTransfer.objects.create(tenant=self.t, transfer_number="TR-1",
                                                   from_location=self.a, to_location=self.b)
        self.line = InventoryTransferLine.objects.create(transfer=self.tr, product=self.p, qty=Decimal("10"))

    def _on_hand(self, loc):
        from core.models import InventoryBalance
        b = InventoryBalance.objects.filter(tenant=self.t, product=self.p, location=loc).first()
        return b.on_hand if b else Decimal("0.00")

    def _recon(self):
        from core.services import reports
        return reports.inventory_gl_reconciliation(self.t)["variance"]

    def test_dispatch_then_full_receive_is_value_neutral(self):
        from core.models import InventoryTransfer
        from core.views import _dispatch_transfer, _receive_transfer
        self.assertEqual(self._recon(), Decimal("0.00"))

        _dispatch_transfer(self.tr)
        self.tr.refresh_from_db(); self.line.refresh_from_db()
        self.assertEqual(self.tr.status, InventoryTransfer.Status.DISPATCHED)
        self.assertEqual(self._on_hand(self.a), Decimal("90.00"))   # relieved into transit
        self.assertEqual(self._on_hand(self.b), Decimal("0.00"))    # not arrived yet
        self.assertEqual(self.line.in_transit_qty, Decimal("10.00"))
        self.assertEqual(_account_balance(self.t, "1010"), Decimal("20.00"))  # in-transit asset
        self.assertEqual(self._recon(), Decimal("0.00"))            # control stays balanced

        _receive_transfer(self.tr)
        self.tr.refresh_from_db()
        self.assertEqual(self.tr.status, InventoryTransfer.Status.RECEIVED)
        self.assertEqual(self._on_hand(self.a), Decimal("90.00"))
        self.assertEqual(self._on_hand(self.b), Decimal("10.00"))   # arrived, value-neutral
        self.assertEqual(_account_balance(self.t, "1010"), Decimal("0.00"))  # transit cleared
        self.assertEqual(self._recon(), Decimal("0.00"))

    def test_partial_receipt_keeps_remainder_in_transit(self):
        from core.models import InventoryTransfer
        from core.views import _dispatch_transfer, _receive_transfer
        _dispatch_transfer(self.tr)
        _receive_transfer(self.tr, receipts={self.line.id: Decimal("6")})
        self.tr.refresh_from_db(); self.line.refresh_from_db()
        self.assertEqual(self.tr.status, InventoryTransfer.Status.DISPATCHED)  # still open
        self.assertEqual(self._on_hand(self.b), Decimal("6.00"))
        self.assertEqual(self.line.in_transit_qty, Decimal("4.00"))
        self.assertEqual(_account_balance(self.t, "1010"), Decimal("8.00"))   # 4 @ 2 still in transit
        self.assertEqual(self._recon(), Decimal("0.00"))

    def test_close_short_writes_off_in_transit_loss(self):
        from core.models import InventoryTransfer
        from core.views import _dispatch_transfer, _receive_transfer
        _dispatch_transfer(self.tr)
        _receive_transfer(self.tr, receipts={self.line.id: Decimal("6")}, close_short=True)
        self.tr.refresh_from_db(); self.line.refresh_from_db()
        self.assertEqual(self.tr.status, InventoryTransfer.Status.RECEIVED)
        self.assertEqual(self.line.in_transit_qty, Decimal("0.00"))           # closed
        self.assertEqual(_account_balance(self.t, "1010"), Decimal("0.00"))   # transit cleared
        self.assertEqual(_account_balance(self.t, "5200"), Decimal("8.00"))   # 4 @ 2 lost in transit
        self.assertEqual(self._recon(), Decimal("0.00"))

    def test_cancel_dispatched_returns_stock_to_source(self):
        from core.models import InventoryTransfer
        from core.views import _dispatch_transfer, _cancel_dispatched_transfer
        _dispatch_transfer(self.tr)
        _cancel_dispatched_transfer(self.tr)
        self.tr.refresh_from_db()
        self.assertEqual(self.tr.status, InventoryTransfer.Status.CANCELLED)
        self.assertEqual(self._on_hand(self.a), Decimal("100.00"))   # all returned
        self.assertEqual(_account_balance(self.t, "1010"), Decimal("0.00"))
        self.assertEqual(self._recon(), Decimal("0.00"))


class LandedCostTests(TestCase):
    def setUp(self):
        from core.models import Location
        self.tenant = Tenant.objects.create(name="Landed Co")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="S")
        self.product = Product.objects.create(tenant=self.tenant, sku="SKU-L", name="P")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH")
        self.po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-L", supplier=self.supplier, status=PurchaseOrder.Status.SUBMITTED)
        self.pol = PurchaseOrderLine.objects.create(po=self.po, product=self.product, ordered_qty=Decimal("10"), unit_cost=Decimal("10.00"))
        self.shipment = Shipment.objects.create(tenant=self.tenant, po=self.po, from_supplier=self.supplier, destination=self.loc)
        self.sl = ShipmentLine.objects.create(shipment=self.shipment, po_line=self.pol, expected_qty=Decimal("10"))
        self.user = User.objects.create_user("wh2", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Warehouse")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="wh2", password="pw")

    def test_landed_cost_capitalized_and_balanced(self):
        from core.models import JournalEntry, GLAccount
        from core.services import reports

        # Receive 10 @ 10.00 goods (100) + 20.00 freight.
        resp = self.client.post(f"/po/{self.po.id}/receive/", {
            "grn_number": "GRN-L",
            f"recv_{self.sl.id}": "10",
            "landed_cost_name": "Freight",
            "landed_cost_amount": "20.00",
        })
        self.assertEqual(resp.status_code, 302)

        # Inventory capitalized at 120 (100 goods + 20 landed); journal balances.
        je = JournalEntry.objects.get(tenant=self.tenant, ref_type="GRN")
        self.assertEqual(je.total_debit, je.total_credit)
        self.assertEqual(je.total_debit, Decimal("120.00"))

        # Average cost now 12.00 -> valuation 120.
        self.product.refresh_from_db()
        self.assertEqual(self.product.average_cost, Decimal("12.0000"))
        self.assertEqual(reports.stock_valuation(self.tenant)["total"], Decimal("120.00"))


class StandardCostingTests(TestCase):
    def setUp(self):
        from core.models import Location
        self.tenant = Tenant.objects.create(name="Std Co")
        # Standard cost 10.00, but we'll buy at 12.00 -> unfavourable variance.
        self.product = Product.objects.create(
            tenant=self.tenant, sku="SKU-S", name="P",
            cost_method=Product.CostMethod.STANDARD, standard_cost=Decimal("10.00"),
        )
        self.loc = Location.objects.create(tenant=self.tenant, name="WH")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="S")
        self.po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-S", supplier=self.supplier, status=PurchaseOrder.Status.SUBMITTED)
        self.pol = PurchaseOrderLine.objects.create(po=self.po, product=self.product, ordered_qty=Decimal("10"), unit_cost=Decimal("12.00"))
        self.shipment = Shipment.objects.create(tenant=self.tenant, po=self.po, from_supplier=self.supplier, destination=self.loc)
        self.sl = ShipmentLine.objects.create(shipment=self.shipment, po_line=self.pol, expected_qty=Decimal("10"))
        self.user = User.objects.create_user("wh3", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Warehouse")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="wh3", password="pw")

    def test_receipt_values_at_standard_and_books_variance(self):
        from core.models import JournalEntry, GLAccount
        from core.services import reports

        # Receive 10 @ actual 12.00; standard is 10.00.
        resp = self.client.post(f"/po/{self.po.id}/receive/", {
            "grn_number": "GRN-S", f"recv_{self.sl.id}": "10",
        })
        self.assertEqual(resp.status_code, 302)

        # Inventory carried at standard: 10 x 10.00 = 100.00
        self.assertEqual(reports.stock_valuation(self.tenant)["total"], Decimal("100.00"))

        je = JournalEntry.objects.get(tenant=self.tenant, ref_type="GRN")
        self.assertEqual(je.total_debit, je.total_credit)            # balanced
        self.assertEqual(je.total_credit, Decimal("120.00"))         # GRNI at actual 120
        # Variance = 120 actual - 100 standard = 20 unfavourable (DR PPV).
        ppv = GLAccount.objects.get(tenant=self.tenant, code="5100")
        ppv_line = je.lines.get(account=ppv)
        self.assertEqual(ppv_line.debit, Decimal("20.00"))

    def test_sale_cogs_at_standard(self):
        from core.services.inventory import apply_movement
        from core.views import _post_sales_order
        from core.models import SalesOrder, SalesOrderLine, JournalEntry

        apply_movement(tenant=self.tenant, product=self.product, location=self.loc,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="T", ref_id="1", unit_cost=Decimal("12.00"))
        order = SalesOrder.objects.create(tenant=self.tenant, order_number="SO-S1", ship_from_location=self.loc)
        SalesOrderLine.objects.create(order=order, product=self.product, qty=Decimal("4"), unit_price=Decimal("30.00"))
        _post_sales_order(order)

        je = JournalEntry.objects.get(tenant=self.tenant, ref_type="COGS")
        self.assertEqual(je.total_debit, Decimal("40.00"))  # 4 x standard 10.00


class VatReturnTests(TestCase):
    def setUp(self):
        from datetime import date
        from core.models import GoodsReceipt, Location, SupplierInvoice, SupplierInvoiceLine
        self.date = date
        self.tenant = Tenant.objects.create(name="VAT Co")
        std = TaxCode.objects.get(tenant=self.tenant, code="STD")

        # Sales: net 200, output VAT 40
        customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        ci = CustomerInvoice.objects.create(tenant=self.tenant, customer=customer, invoice_number="CINV-1", invoice_date=date(2026, 5, 15))
        CustomerInvoiceLine.objects.create(invoice=ci, description="X", qty=Decimal("2"), unit_price=Decimal("100.00"), tax_code=std)
        post_customer_invoice(ci)

        # Purchases: net 50, input VAT 10
        supplier = Supplier.objects.create(tenant=self.tenant, name="Sup")
        product = Product.objects.create(tenant=self.tenant, sku="SKU-V", name="P")
        loc = Location.objects.create(tenant=self.tenant, name="WH")
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-V", supplier=supplier)
        grn = GoodsReceipt.objects.create(tenant=self.tenant, po=po, grn_number="GRN-V", received_to=loc, status=GoodsReceipt.Status.POSTED)
        si = SupplierInvoice.objects.create(tenant=self.tenant, supplier=supplier, po=po, receipt=grn, invoice_number="SINV-V", invoice_date=date(2026, 5, 16), status="POSTED")
        SupplierInvoiceLine.objects.create(invoice=si, product=product, qty=Decimal("10"), unit_cost=Decimal("5.00"), tax_code=std)

        self.user = User.objects.create_user("vat", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Finance")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="vat", password="pw")

    def test_compute_boxes(self):
        from core.services import vat
        b = vat.compute_vat_return(self.tenant, self.date(2026, 5, 1), self.date(2026, 5, 31))
        self.assertEqual(b["box1_vat_due_sales"], Decimal("40.00"))
        self.assertEqual(b["box4_vat_reclaimed"], Decimal("10.00"))
        self.assertEqual(b["box5_net_vat"], Decimal("30.00"))
        self.assertEqual(b["box6_total_sales_ex_vat"], Decimal("200.00"))
        self.assertEqual(b["box7_total_purchases_ex_vat"], Decimal("50.00"))

    def test_save_and_submit(self):
        from core.services import vat
        from core.models import VatReturn
        vr = vat.save_vat_return(self.tenant, self.date(2026, 5, 1), self.date(2026, 5, 31))
        self.assertEqual(vr.box5_net_vat, Decimal("30.00"))
        vat.submit_vat_return(vr)
        vr.refresh_from_db()
        self.assertEqual(vr.status, VatReturn.Status.SUBMITTED)
        self.assertTrue(vr.hmrc_reference.startswith("LOCAL-STUB"))

    def test_vat_pages_render(self):
        self.assertEqual(self.client.get("/vat/").status_code, 200)
        self.assertEqual(self.client.get("/vat/?from=2026-05-01&to=2026-05-31").status_code, 200)


class FinancialReportsTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Reports Co")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-R1")
        std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        CustomerInvoiceLine.objects.create(invoice=inv, description="Item", qty=Decimal("2"), unit_price=Decimal("100.00"), tax_code=std)
        post_customer_invoice(inv)  # DR AR 240 / CR Sales 200 / CR VAT 40

        self.user = User.objects.create_user("fin", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Finance")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="fin", password="pw")

    def test_trial_balance_is_balanced(self):
        from core.services import reports
        tb = reports.trial_balance(self.tenant)
        self.assertTrue(tb["balanced"])
        self.assertEqual(tb["total_debit"], Decimal("240.00"))

    def test_pnl_net_profit(self):
        from core.services import reports
        pnl = reports.profit_and_loss(self.tenant)
        self.assertEqual(pnl["income_total"], Decimal("200.00"))
        self.assertEqual(pnl["expense_total"], Decimal("0.00"))
        self.assertEqual(pnl["net_profit"], Decimal("200.00"))

    def test_balance_sheet_balances(self):
        from core.services import reports
        bs = reports.balance_sheet(self.tenant)
        self.assertEqual(bs["asset_total"], Decimal("240.00"))          # AR
        self.assertEqual(bs["liability_total"], Decimal("40.00"))       # VAT output
        self.assertEqual(bs["retained_earnings"], Decimal("200.00"))    # net income
        self.assertTrue(bs["balanced"])

    def test_aged_receivables_lists_issued_invoice(self):
        from core.services import reports
        ar = reports.aged_receivables(self.tenant)
        self.assertEqual(ar["total"], Decimal("240.00"))
        self.assertEqual(len(ar["rows"]), 1)

    def test_report_pages_render(self):
        for path in ["/reports/", "/reports/trial-balance/", "/reports/profit-and-loss/",
                     "/reports/balance-sheet/", "/reports/cash-flow/",
                     "/reports/aged-receivables/", "/reports/aged-payables/"]:
            resp = self.client.get(path)
            self.assertEqual(resp.status_code, 200, f"{path} -> {resp.status_code}")


class FinanceExportTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="FX Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-X1")
        CustomerInvoiceLine.objects.create(invoice=inv, description="Item", qty=Decimal("2"), unit_price=Decimal("100.00"), tax_code=self.std)
        post_customer_invoice(inv)
        self.admin = User.objects.create_user("xadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.wh = User.objects.create_user("xwh", password="pw")
        OrgMembership.objects.create(user=self.wh, tenant=self.tenant, role="WAREHOUSE", is_default=True)

    def test_each_kind_exports_csv(self):
        self.client.login(username="xadmin", password="pw")
        kinds = ["trial-balance", "profit-and-loss", "balance-sheet", "cash-flow",
                 "aged-receivables", "aged-payables", "journal", "expenses",
                 "payments", "invoices", "bills", "credit-notes", "bank-transactions"]
        for k in kinds:
            resp = self.client.get(f"/finance/export/{k}.csv")
            self.assertEqual(resp.status_code, 200, f"{k} -> {resp.status_code}")
            self.assertEqual(resp["Content-Type"], "text/csv")

    def test_invoices_export_has_data(self):
        self.client.login(username="xadmin", password="pw")
        body = self.client.get("/finance/export/invoices.csv").content.decode()
        self.assertIn("INV-X1", body)
        self.assertIn("Outstanding", body)

    def test_xlsx_export_returns_workbook(self):
        import io
        from openpyxl import load_workbook
        self.client.login(username="xadmin", password="pw")
        resp = self.client.get("/finance/export/invoices.csv?format=xlsx")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn(".xlsx", resp["Content-Disposition"])
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "Number")  # header row
        self.assertTrue(any("INV-X1" in str(c.value) for c in ws["A"]))

    def test_xlsx_export_audited(self):
        from core.models import AuditLog
        self.client.login(username="xadmin", password="pw")
        self.client.get("/finance/export/trial-balance.csv?format=xlsx")
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="DATA_EXPORTED").exists())

    def test_export_audited(self):
        from core.models import AuditLog
        self.client.login(username="xadmin", password="pw")
        self.client.get("/finance/export/trial-balance.csv")
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="DATA_EXPORTED").exists())

    def test_export_blocked_without_permission(self):
        self.client.login(username="xwh", password="pw")  # WAREHOUSE lacks export_data
        self.assertEqual(self.client.get("/finance/export/trial-balance.csv").status_code, 403)

    def test_unknown_kind_404(self):
        self.client.login(username="xadmin", password="pw")
        self.assertEqual(self.client.get("/finance/export/nonsense.csv").status_code, 404)


class CashFlowAndGrossProfitTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="CF Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-CF1")
        CustomerInvoiceLine.objects.create(invoice=self.inv, description="Item", qty=Decimal("2"), unit_price=Decimal("100.00"), tax_code=self.std)
        post_customer_invoice(self.inv)  # income 200
        self.user = User.objects.create_user("cf", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Finance")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="cf", password="pw")

    def test_cogs_account_is_cogs_type_and_gross_profit(self):
        from core.models import GLAccount
        from core.services import reports
        from core.services.gl import post_cogs
        cogs_acc = GLAccount.objects.get(tenant=self.tenant, code="5000")
        self.assertEqual(cogs_acc.type, "COGS")
        post_cogs(self.tenant, Decimal("80.00"), "SALE-1")  # DR COGS 80 / CR inventory 80
        pnl = reports.profit_and_loss(self.tenant)
        self.assertEqual(pnl["income_total"], Decimal("200.00"))
        self.assertEqual(pnl["cogs_total"], Decimal("80.00"))
        self.assertEqual(pnl["gross_profit"], Decimal("120.00"))
        self.assertEqual(pnl["net_profit"], Decimal("120.00"))

    def test_cash_flow_summary_reflects_receipt(self):
        from core.models import Payment, PaymentAllocation, GLAccount
        from core.services.gl import post_payment
        from core.services import reports
        p = Payment.objects.create(tenant=self.tenant, direction=Payment.Direction.RECEIPT,
                                   customer=self.customer, amount=Decimal("240.00"), method="BANK")
        PaymentAllocation.objects.create(payment=p, customer_invoice=self.inv, amount=Decimal("240.00"))
        post_payment(p)
        import datetime
        cf = reports.cash_flow_summary(self.tenant, date_from=datetime.date(2000, 1, 1), date_to=datetime.date(2100, 1, 1))
        self.assertEqual(cf["cash_in"], Decimal("240.00"))
        self.assertEqual(cf["cash_out"], Decimal("0.00"))
        self.assertEqual(cf["net"], Decimal("240.00"))
        ar = GLAccount.objects.get(tenant=self.tenant, code="1100")
        ar_row = next((r for r in cf["rows"] if r["account"] == ar), None)
        self.assertIsNotNone(ar_row)
        self.assertEqual(ar_row["amount"], Decimal("240.00"))


class ExpenseTests(TestCase):
    def setUp(self):
        from core.models import GLAccount
        self.tenant = Tenant.objects.create(name="Exp Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.rent = GLAccount.objects.get(tenant=self.tenant, code="6100")
        self.user = User.objects.create_user("exp", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Finance")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="exp", password="pw")

    def _post(self, **overrides):
        data = {
            "expense_date": "2026-05-30", "payee": "Landlord", "category": self.rent.id,
            "description": "Rent", "net_amount": "1000.00", "tax_code": self.std.id,
            "method": "BANK", "reference": "R-1", "paid": "on", "action": "post",
        }
        data.update(overrides)
        return self.client.post("/expenses/new/", data)

    def test_paid_expense_posts_balanced_je(self):
        from core.models import Expense, GLAccount
        from core.services import reports
        resp = self._post()
        self.assertEqual(resp.status_code, 302)
        e = Expense.objects.get(tenant=self.tenant)
        self.assertEqual(e.status, "POSTED")
        self.assertEqual(e.total, Decimal("1200.00"))
        balances = reports.account_balances(self.tenant)
        self.assertEqual(balances[self.rent]["balance"], Decimal("1000.00"))
        vat_in = GLAccount.objects.get(tenant=self.tenant, code="1300")
        bank = GLAccount.objects.get(tenant=self.tenant, code="1050")
        self.assertEqual(balances[vat_in]["balance"], Decimal("200.00"))
        self.assertEqual(balances[bank]["balance"], Decimal("-1200.00"))  # cash out

    def test_unpaid_expense_credits_accounts_payable(self):
        from core.models import GLAccount
        from core.services import reports
        self._post(paid="")  # unchecked -> owed
        balances = reports.account_balances(self.tenant)
        ap = GLAccount.objects.get(tenant=self.tenant, code="2000")
        bank = GLAccount.objects.get(tenant=self.tenant, code="1050")
        self.assertEqual(ap["balance"] if isinstance(ap, dict) else balances[ap]["balance"], Decimal("1200.00"))
        self.assertEqual(balances[bank]["balance"], Decimal("0.00"))

    def test_expense_shows_in_pnl(self):
        from core.services import reports
        self._post()
        pnl = reports.profit_and_loss(self.tenant)
        self.assertEqual(pnl["expense_total"], Decimal("1000.00"))

    def test_draft_then_post(self):
        from core.models import Expense, JournalEntry
        resp = self._post(action="save")
        e = Expense.objects.get(tenant=self.tenant)
        self.assertEqual(e.status, "DRAFT")
        self.assertFalse(JournalEntry.objects.filter(tenant=self.tenant, ref_type="EXPENSE").exists())
        self.client.post(f"/expenses/{e.id}/post/")
        e.refresh_from_db()
        self.assertEqual(e.status, "POSTED")
        self.assertTrue(JournalEntry.objects.filter(tenant=self.tenant, ref_type="EXPENSE").exists())

    def test_category_dropdown_excludes_non_expense_accounts(self):
        resp = self.client.get("/expenses/new/")
        self.assertEqual(resp.status_code, 200)
        form = resp.context["form"]
        codes = {a.code for a in form.fields["category"].queryset}
        self.assertIn("6100", codes)
        self.assertNotIn("1050", codes)  # bank is not an expense category
        self.assertNotIn("4000", codes)  # sales income is not an expense category


class ExpenseEntryCompletenessTests(TestCase):
    def setUp(self):
        from core.models import GLAccount, OrgMembership
        self.tenant = Tenant.objects.create(name="ExpC Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.rent = GLAccount.objects.get(tenant=self.tenant, code="6100")
        self.user = User.objects.create_user("expc", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="FINANCE", is_default=True)
        self.client.login(username="expc", password="pw")

    def _data(self, **overrides):
        data = {
            "expense_date": "2026-06-01", "payee": "X", "category": self.rent.id,
            "net_amount": "100.00", "tax_code": self.std.id, "method": "CARD",
            "action": "save",
        }
        data.update(overrides)
        return data

    def test_new_categories_seeded(self):
        from core.models import GLAccount
        for code, name in [("6150", "Repairs & Maintenance"), ("6250", "Insurance"), ("6450", "Meals & Entertainment")]:
            acc = GLAccount.objects.get(tenant=self.tenant, code=code)
            self.assertEqual(acc.name, name)
            self.assertEqual(acc.type, "EXPENSE")

    def test_reimbursable_flag_saved(self):
        from core.models import Expense
        self.client.post("/expenses/new/", self._data(reimbursable="on"))
        e = Expense.objects.get(tenant=self.tenant)
        self.assertTrue(e.reimbursable)

    def test_receipt_upload_stored_tenant_scoped(self):
        from core.models import Expense
        from django.core.files.uploadedfile import SimpleUploadedFile
        pdf = SimpleUploadedFile("receipt.pdf", b"%PDF-1.4 test", content_type="application/pdf")
        self.client.post("/expenses/new/", self._data(receipt=pdf))
        e = Expense.objects.get(tenant=self.tenant)
        self.assertTrue(e.receipt)
        self.assertIn(f"expense_receipts/{self.tenant.id}/", e.receipt.name)

    def test_receipt_rejects_bad_type(self):
        from core.models import Expense
        from django.core.files.uploadedfile import SimpleUploadedFile
        bad = SimpleUploadedFile("malware.exe", b"MZ", content_type="application/octet-stream")
        resp = self.client.post("/expenses/new/", self._data(receipt=bad))
        self.assertEqual(resp.status_code, 200)  # re-render with error
        self.assertContains(resp, "Receipt must be a PDF or an image")
        self.assertEqual(Expense.objects.filter(tenant=self.tenant).count(), 0)

    def test_supplier_history_shows_expense(self):
        from core.models import Expense, Supplier
        sup = Supplier.objects.create(tenant=self.tenant, name="Acme")
        Expense.objects.create(tenant=self.tenant, expense_date="2026-06-01", payee="Acme",
                               supplier=sup, category=self.rent, net_amount=Decimal("50"))
        resp = self.client.get(f"/suppliers/{sup.id}/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context["expenses"]), 1)
        self.assertTrue(any(t["kind"] == "Expense" for t in resp.context["timeline"]))


class ExpenseApprovalWorkflowTests(TestCase):
    def setUp(self):
        from core.models import GLAccount, OrgMembership
        self.tenant = Tenant.objects.create(name="ExpW Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.rent = GLAccount.objects.get(tenant=self.tenant, code="6100")
        self.staff = User.objects.create_user("staff", password="pw")
        OrgMembership.objects.create(user=self.staff, tenant=self.tenant, role="SALES", is_default=True)
        self.fin = User.objects.create_user("fin", password="pw")
        OrgMembership.objects.create(user=self.fin, tenant=self.tenant, role="FINANCE", is_default=True)

    def _data(self, **o):
        d = {"expense_date": "2026-06-01", "payee": "Cafe", "category": self.rent.id,
             "net_amount": "100.00", "tax_code": self.std.id, "method": "CARD", "action": "submit"}
        d.update(o)
        return d

    def test_staff_submits_for_approval(self):
        from core.models import Expense, JournalEntry
        self.client.login(username="staff", password="pw")
        resp = self.client.post("/expenses/new/", self._data())
        self.assertEqual(resp.status_code, 302)
        e = Expense.objects.get(tenant=self.tenant)
        self.assertEqual(e.status, Expense.Status.SUBMITTED)
        self.assertEqual(e.submitted_by, self.staff)
        self.assertFalse(JournalEntry.objects.filter(tenant=self.tenant, ref_type="EXPENSE").exists())

    def test_staff_cannot_approve(self):
        from core.models import Expense
        e = Expense.objects.create(tenant=self.tenant, expense_date="2026-06-01", payee="X",
                                   category=self.rent, net_amount=Decimal("100"),
                                   status=Expense.Status.SUBMITTED, submitted_by=self.staff)
        self.client.login(username="staff", password="pw")
        resp = self.client.post(f"/expenses/{e.id}/approve/")
        self.assertEqual(resp.status_code, 403)
        e.refresh_from_db()
        self.assertEqual(e.status, Expense.Status.SUBMITTED)

    def test_finance_approves_and_posts(self):
        from core.models import Expense, JournalEntry
        e = Expense.objects.create(tenant=self.tenant, expense_date="2026-06-01", payee="X",
                                   category=self.rent, net_amount=Decimal("100"), tax_code=self.std,
                                   status=Expense.Status.SUBMITTED, submitted_by=self.staff)
        self.client.login(username="fin", password="pw")
        resp = self.client.post(f"/expenses/{e.id}/approve/")
        self.assertEqual(resp.status_code, 302)
        e.refresh_from_db()
        self.assertEqual(e.status, Expense.Status.POSTED)
        self.assertEqual(e.approved_by, self.fin)
        self.assertIsNotNone(e.approved_at)
        self.assertTrue(JournalEntry.objects.filter(tenant=self.tenant, ref_type="EXPENSE", ref_id=str(e.id)).exists())

    def test_finance_rejects(self):
        from core.models import Expense, JournalEntry
        e = Expense.objects.create(tenant=self.tenant, expense_date="2026-06-01", payee="X",
                                   category=self.rent, net_amount=Decimal("100"),
                                   status=Expense.Status.SUBMITTED, submitted_by=self.staff)
        self.client.login(username="fin", password="pw")
        self.client.post(f"/expenses/{e.id}/reject/", {"reason": "No receipt"})
        e.refresh_from_db()
        self.assertEqual(e.status, Expense.Status.REJECTED)
        self.assertEqual(e.rejected_reason, "No receipt")
        self.assertFalse(JournalEntry.objects.filter(tenant=self.tenant, ref_type="EXPENSE").exists())

    def test_finance_direct_post_below_threshold(self):
        from core.models import Expense
        self.client.login(username="fin", password="pw")
        self.client.post("/expenses/new/", self._data(action="post"))
        e = Expense.objects.get(tenant=self.tenant)
        self.assertEqual(e.status, Expense.Status.POSTED)

    def test_threshold_forces_approval_even_for_finance(self):
        from core.models import Expense
        self.tenant.expense_approval_threshold = Decimal("50.00")
        self.tenant.save()
        self.client.login(username="fin", password="pw")
        self.client.post("/expenses/new/", self._data(action="post", net_amount="100.00"))
        e = Expense.objects.get(tenant=self.tenant)
        self.assertEqual(e.status, Expense.Status.SUBMITTED)  # 120 total >= 50 -> needs approval


class BankTransactionTests(TestCase):
    def setUp(self):
        self.tenant = Tenant.objects.create(name="Bank Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-B1")
        CustomerInvoiceLine.objects.create(invoice=inv, description="Item", qty=Decimal("2"), unit_price=Decimal("100.00"), tax_code=self.std)
        post_customer_invoice(inv)
        from core.models import Payment, PaymentAllocation
        from core.services.gl import post_payment
        self.payment = Payment.objects.create(tenant=self.tenant, direction=Payment.Direction.RECEIPT,
                                              customer=self.customer, amount=Decimal("240.00"), method="BANK", reference="FPS-9")
        PaymentAllocation.objects.create(payment=self.payment, customer_invoice=inv, amount=Decimal("240.00"))
        post_payment(self.payment)
        self.user = User.objects.create_user("bk", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Finance")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="bk", password="pw")

    def test_import_creates_transactions(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from core.models import BankTransaction
        csv = b"date,description,amount,reference\n2026-06-01,FPS CREDIT,240.00,FPS-9\n2026-06-02,BANK FEE,-5.00,\n"
        resp = self.client.post("/bank/transactions/import/", {"file": SimpleUploadedFile("s.csv", csv, content_type="text/csv")})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(BankTransaction.objects.filter(tenant=self.tenant).count(), 2)

    def test_auto_match_reconciles_exact_amount(self):
        from core.models import BankTransaction
        t = BankTransaction.objects.create(tenant=self.tenant, description="FPS CREDIT", amount=Decimal("240.00"))
        resp = self.client.post("/bank/reconcile/", {"action": "auto"})
        self.assertEqual(resp.status_code, 302)
        t.refresh_from_db()
        self.assertTrue(t.is_reconciled)
        self.assertEqual(t.matched_payment_id, self.payment.id)
        self.payment.refresh_from_db()
        self.assertTrue(self.payment.is_reconciled)

    def test_manual_match_and_unmatch(self):
        from core.models import BankTransaction
        t = BankTransaction.objects.create(tenant=self.tenant, description="FPS CREDIT", amount=Decimal("240.00"))
        self.client.post("/bank/reconcile/", {f"match_{t.id}": f"payment:{self.payment.id}"})
        t.refresh_from_db()
        self.assertTrue(t.is_reconciled)
        self.client.post("/bank/reconcile/", {f"match_{t.id}": ""})
        t.refresh_from_db()
        self.assertFalse(t.is_reconciled)
        self.assertIsNone(t.matched_payment_id)

    def test_reconcile_page_renders_with_book_balance(self):
        from core.models import BankTransaction
        BankTransaction.objects.create(tenant=self.tenant, description="FPS CREDIT", amount=Decimal("240.00"))
        resp = self.client.get("/bank/reconcile/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["book_balance"], Decimal("240.00"))
        self.assertEqual(resp.context["cleared"], Decimal("0.00"))


class CreditNoteTests(TestCase):
    def setUp(self):
        from core.models import GLAccount
        self.tenant = Tenant.objects.create(name="CN Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Supp")
        self.inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-CN1")
        CustomerInvoiceLine.objects.create(invoice=self.inv, description="Item", qty=Decimal("2"), unit_price=Decimal("100.00"), tax_code=self.std)
        post_customer_invoice(self.inv)  # total 240
        self.user = User.objects.create_user("cn", password="pw")
        self.user.groups.add(Group.objects.get_or_create(name="Finance")[0])
        UserProfile.objects.create(user=self.user, tenant=self.tenant)
        self.client.login(username="cn", password="pw")

    def test_sales_credit_note_reduces_invoice_and_ar(self):
        from core.models import CreditNote, CreditNoteLine, GLAccount
        from core.services.gl import post_credit_note
        from core.services import reports
        cn = CreditNote.objects.create(tenant=self.tenant, kind=CreditNote.Kind.SALES,
                                       credit_note_number="CN-1", customer=self.customer, customer_invoice=self.inv)
        CreditNoteLine.objects.create(credit_note=cn, description="Refund", qty=Decimal("2"), unit_amount=Decimal("100.00"), tax_code=self.std)
        post_credit_note(cn, user=self.user)
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.outstanding, Decimal("0.00"))
        self.assertEqual(self.inv.status, "PAID")
        balances = reports.account_balances(self.tenant)
        ar = GLAccount.objects.get(tenant=self.tenant, code="1100")
        self.assertEqual(balances[ar]["balance"], Decimal("0.00"))  # 240 invoice - 240 credit

    def test_sales_credit_drops_invoice_from_aged_receivables(self):
        from core.models import CreditNote, CreditNoteLine
        from core.services.gl import post_credit_note
        from core.services import reports
        self.assertEqual(reports.aged_receivables(self.tenant)["total"], Decimal("240.00"))
        cn = CreditNote.objects.create(tenant=self.tenant, kind=CreditNote.Kind.SALES,
                                       credit_note_number="CN-2", customer=self.customer, customer_invoice=self.inv)
        CreditNoteLine.objects.create(credit_note=cn, description="Refund", qty=Decimal("2"), unit_amount=Decimal("100.00"), tax_code=self.std)
        post_credit_note(cn, user=self.user)
        self.assertEqual(reports.aged_receivables(self.tenant)["total"], Decimal("0.00"))

    def test_purchase_credit_note_posts_balanced_je(self):
        from core.models import CreditNote, CreditNoteLine, GLAccount, JournalEntry
        from core.services.gl import post_credit_note
        from core.services import reports
        acc = GLAccount.objects.get(tenant=self.tenant, code="6900")
        cn = CreditNote.objects.create(tenant=self.tenant, kind=CreditNote.Kind.PURCHASE,
                                       credit_note_number="CN-P1", supplier=self.supplier)
        CreditNoteLine.objects.create(credit_note=cn, description="Overcharge", qty=Decimal("1"), unit_amount=Decimal("100.00"), tax_code=self.std, account=acc)
        je = post_credit_note(cn, user=self.user)
        self.assertEqual(je.total_debit, je.total_credit)  # balanced
        self.assertEqual(je.total_debit, Decimal("120.00"))
        balances = reports.account_balances(self.tenant)
        ap = GLAccount.objects.get(tenant=self.tenant, code="2000")
        self.assertEqual(balances[ap]["balance"], Decimal("-120.00"))  # payable reduced

    def test_create_view_posts_sales_credit(self):
        from core.models import CreditNote
        resp = self.client.post("/credit-notes/new/", {
            "kind": "SALES", "credit_note_number": "CN-V1", "credit_note_date": "2026-06-01",
            "customer": self.customer.id, "customer_invoice": self.inv.id, "action": "post",
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "0",
            "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-description": "Refund", "lines-0-qty": "1", "lines-0-unit_amount": "50",
            "lines-0-tax_code": self.std.id,
        })
        self.assertEqual(resp.status_code, 302)
        cn = CreditNote.objects.get(tenant=self.tenant, credit_note_number="CN-V1")
        self.assertEqual(cn.status, "POSTED")
        self.assertEqual(cn.total, Decimal("60.00"))


class VatTaxCodeTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="VAT Codes Co")
        self.admin = User.objects.create_user("vcadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="vcadmin", password="pw")

    def test_all_five_default_codes_with_treatments(self):
        codes = {c.code: c for c in TaxCode.objects.filter(tenant=self.tenant)}
        self.assertEqual(codes["STD"].kind, "STANDARD")
        self.assertEqual(codes["RED"].kind, "REDUCED")
        self.assertEqual(codes["RED"].rate, Decimal("0.0500"))
        self.assertEqual(codes["ZERO"].kind, "ZERO")
        self.assertEqual(codes["EXEMPT"].kind, "EXEMPT")
        self.assertEqual(codes["OS"].kind, "OUTSIDE")

    def test_in_vat_boxes_excludes_outside_scope(self):
        codes = {c.code: c for c in TaxCode.objects.filter(tenant=self.tenant)}
        self.assertTrue(codes["STD"].in_vat_boxes)
        self.assertTrue(codes["ZERO"].in_vat_boxes)
        self.assertTrue(codes["EXEMPT"].in_vat_boxes)
        self.assertFalse(codes["OS"].in_vat_boxes)

    def test_create_tax_code_audited(self):
        from core.models import AuditLog
        resp = self.client.post("/tax-codes/new/", {
            "code": "CUSTOM", "name": "Custom", "rate": "0.10", "kind": "REDUCED", "is_active": "on",
        })
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(TaxCode.objects.filter(tenant=self.tenant, code="CUSTOM").exists())
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="VAT_RATE_CHANGED").exists())


class VatEngineTests(TestCase):
    def setUp(self):
        from core.models import GLAccount
        self.tenant = Tenant.objects.create(name="VAT Engine Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.os = TaxCode.objects.get(tenant=self.tenant, code="OS")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Supp")
        self.expense_acc = GLAccount.objects.get(tenant=self.tenant, code="6900")

        # Sales invoice: 1000 @ standard + 500 outside-scope.
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-V1")
        CustomerInvoiceLine.objects.create(invoice=inv, description="Std", qty=Decimal("1"), unit_price=Decimal("1000.00"), tax_code=self.std)
        CustomerInvoiceLine.objects.create(invoice=inv, description="OutOfScope", qty=Decimal("1"), unit_price=Decimal("500.00"), tax_code=self.os)
        post_customer_invoice(inv)

        # Expense: 200 @ standard (input VAT 40).
        from core.models import Expense
        from core.services.gl import post_expense
        e = Expense.objects.create(tenant=self.tenant, payee="Office", category=self.expense_acc,
                                   net_amount=Decimal("200.00"), tax_code=self.std, paid=True)
        post_expense(e)

        # Sales credit note: 100 @ standard (reduces output).
        from core.models import CreditNote, CreditNoteLine
        from core.services.gl import post_credit_note
        cn = CreditNote.objects.create(tenant=self.tenant, kind=CreditNote.Kind.SALES,
                                       credit_note_number="CN-V1", customer=self.customer)
        CreditNoteLine.objects.create(credit_note=cn, description="Refund", qty=Decimal("1"), unit_amount=Decimal("100.00"), tax_code=self.std)
        post_credit_note(cn)

        # Purchase credit note: 50 @ standard (reduces input).
        cnp = CreditNote.objects.create(tenant=self.tenant, kind=CreditNote.Kind.PURCHASE,
                                        credit_note_number="CN-VP1", supplier=self.supplier)
        CreditNoteLine.objects.create(credit_note=cnp, description="Return", qty=Decimal("1"), unit_amount=Decimal("50.00"), tax_code=self.std, account=self.expense_acc)
        post_credit_note(cnp)

    def _summary(self):
        import datetime
        from core.services import vat
        return vat.vat_summary(self.tenant, datetime.date(2000, 1, 1), datetime.date(2100, 1, 1))

    def test_summary_five_totals(self):
        s = self._summary()
        self.assertEqual(s["vat_on_sales"], Decimal("180.00"))        # 200 invoice - 20 credit
        self.assertEqual(s["total_sales_ex_vat"], Decimal("900.00"))  # 1000 - 100 (OS 500 excluded)
        self.assertEqual(s["vat_reclaimable"], Decimal("30.00"))      # 40 expense - 10 purchase credit
        self.assertEqual(s["total_purchases_ex_vat"], Decimal("150.00"))  # 200 - 50
        self.assertEqual(s["net_vat"], Decimal("150.00"))            # 180 - 30

    def test_outside_scope_excluded_from_box6(self):
        import datetime
        from core.services import vat
        boxes = vat.compute_vat_return(self.tenant, datetime.date(2000, 1, 1), datetime.date(2100, 1, 1))
        self.assertEqual(boxes["box6_total_sales_ex_vat"], Decimal("900.00"))
        self.assertEqual(boxes["box1_vat_due_sales"], Decimal("180.00"))
        self.assertEqual(boxes["box4_vat_reclaimed"], Decimal("30.00"))
        self.assertEqual(boxes["box5_net_vat"], Decimal("150.00"))

    def test_expense_included_in_input_vat(self):
        # Records should include the expense as a PURCHASE with 40 VAT.
        import datetime
        from core.services import vat
        recs = vat.vat_transactions(self.tenant, datetime.date(2000, 1, 1), datetime.date(2100, 1, 1))
        exp = [r for r in recs if r["doc_type"] == "Expense"]
        self.assertEqual(len(exp), 1)
        self.assertEqual(exp[0]["vat"], Decimal("40.00"))

    def test_breakdown_has_rate_groups(self):
        s = self._summary()
        keys = {(b["direction"], b["treatment"]) for b in s["breakdown"]}
        self.assertIn(("SALES", "Standard rate"), keys)
        self.assertIn(("SALES", "Outside the scope of VAT"), keys)
        self.assertIn(("PURCHASE", "Standard rate"), keys)


class VatUiAuditExportTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="VAT UI Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-U1")
        CustomerInvoiceLine.objects.create(invoice=inv, description="Item", qty=Decimal("1"), unit_price=Decimal("1000.00"), tax_code=self.std)
        post_customer_invoice(inv)
        self.admin = User.objects.create_user("vuadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="vuadmin", password="pw")

    def test_vat_records_page_renders(self):
        resp = self.client.get("/vat/records/?from=2000-01-01&to=2100-01-01")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "INV-U1")

    def test_save_return_audited(self):
        from core.models import AuditLog
        resp = self.client.post("/vat/save/", {"from": "2000-01-01", "to": "2100-01-01"})
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="VAT_RETURN_SAVED").exists())

    def test_submit_return_audited(self):
        from core.models import AuditLog, VatReturn
        self.client.post("/vat/save/", {"from": "2000-01-01", "to": "2100-01-01"})
        vr = VatReturn.objects.get(tenant=self.tenant)
        self.client.post(f"/vat/{vr.id}/submit/")
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="VAT_RETURN_SUBMITTED").exists())

    def test_vat_exports(self):
        for k in ("vat-return", "vat-transactions"):
            resp = self.client.get(f"/finance/export/{k}.csv?from=2000-01-01&to=2100-01-01")
            self.assertEqual(resp.status_code, 200, k)
            self.assertEqual(resp["Content-Type"], "text/csv")
        # the return export contains the box labels
        body = self.client.get("/finance/export/vat-return.csv?from=2000-01-01&to=2100-01-01").content.decode()
        self.assertIn("Net VAT to pay / reclaim", body)

    def test_vat_settings_change_audited(self):
        from core.models import AuditLog
        data = {
            "name": self.tenant.name, "legal_name": "VAT UI Co Ltd", "business_type": "LTD",
            "vat_number": "GB123456789", "vat_registered": "on",
            "address_line1": "1 High St", "address_city": "Manchester", "address_postcode": "M1 2AB",
            "address_country": "United Kingdom", "billing_same_as_business": "on", "billing_country": "United Kingdom",
            "email": "ops@vatui.test", "phone": "+44 20 7946 0000",
            "currency_code": "GBP", "country": "United Kingdom", "timezone": "Europe/London",
            "financial_year_start_month": "4", "default_payment_terms_days": "30", "po_approval_threshold": "0",
        }
        resp = self.client.post("/settings/tenant/", data)
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="VAT_SETTINGS_CHANGED").exists())


class InvoiceCompletenessTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Inv Co", default_payment_terms_days=30)
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.tenant.default_tax_code = self.std
        self.tenant.save()
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust", email="cust@example.com")
        self.user = User.objects.create_user("invu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="invu", password="pw")

    def _create(self, number="", action="save", discount="0"):
        return self.client.post("/ar/invoices/new/", {
            "customer": self.customer.id, "invoice_number": number,
            "invoice_date": "2026-06-01", "action": action,
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "0",
            "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-description": "Widget", "lines-0-qty": "2", "lines-0-unit_price": "100",
            "lines-0-discount_pct": discount, "lines-0-tax_code": self.std.id,
        })

    def test_invoice_number_auto_generated_when_blank(self):
        self._create(number="")
        inv = CustomerInvoice.objects.get(tenant=self.tenant)
        self.assertEqual(inv.invoice_number, "INV-0001")

    def test_admin_can_override_invoice_number(self):
        self._create(number="CUSTOM-9")
        self.assertTrue(CustomerInvoice.objects.filter(tenant=self.tenant, invoice_number="CUSTOM-9").exists())

    def test_line_discount_applied_to_totals(self):
        self._create(discount="10")  # 2 x 100 = 200, less 10% = 180; VAT 36; total 216
        inv = CustomerInvoice.objects.get(tenant=self.tenant)
        self.assertEqual(inv.subtotal, Decimal("180.00"))
        self.assertEqual(inv.tax_total, Decimal("36.00"))
        self.assertEqual(inv.total, Decimal("216.00"))

    def test_due_date_defaulted_from_terms(self):
        self._create()
        inv = CustomerInvoice.objects.get(tenant=self.tenant)
        self.assertEqual(inv.due_date.isoformat(), "2026-07-01")  # +30 days

    def test_send_marks_sent_and_audits(self):
        from core.models import AuditLog
        self._create(action="issue")
        inv = CustomerInvoice.objects.get(tenant=self.tenant)
        resp = self.client.post(f"/ar/invoices/{inv.id}/send/")
        self.assertEqual(resp.status_code, 302)
        inv.refresh_from_db()
        self.assertEqual(inv.status, "SENT")
        self.assertIsNotNone(inv.sent_at)
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="INVOICE_SENT").exists())

    def test_cancel_reverses_gl(self):
        from core.services import reports
        from core.models import GLAccount
        self._create(action="issue")
        inv = CustomerInvoice.objects.get(tenant=self.tenant)
        self.client.post(f"/ar/invoices/{inv.id}/cancel/")
        inv.refresh_from_db()
        self.assertEqual(inv.status, "CANCELLED")
        ar = GLAccount.objects.get(tenant=self.tenant, code="1100")
        self.assertEqual(reports.account_balances(self.tenant)[ar]["balance"], Decimal("0.00"))

    def test_partial_payment_shows_partially_paid_and_stays_in_aged(self):
        from core.models import Payment, PaymentAllocation
        from core.services.gl import post_payment
        from core.services import reports
        self._create(action="issue")  # total 240
        inv = CustomerInvoice.objects.get(tenant=self.tenant)
        p = Payment.objects.create(tenant=self.tenant, direction=Payment.Direction.RECEIPT,
                                   customer=self.customer, amount=Decimal("100.00"), method="BANK")
        PaymentAllocation.objects.create(payment=p, customer_invoice=inv, amount=Decimal("100.00"))
        post_payment(p)
        inv.refresh_from_db()
        self.assertEqual(inv.display_status, "Partially paid")
        self.assertEqual(reports.aged_receivables(self.tenant)["total"], Decimal("140.00"))

    def test_overdue_display(self):
        self._create(action="issue")
        inv = CustomerInvoice.objects.get(tenant=self.tenant)
        inv.due_date = inv.invoice_date.replace(year=2000)
        inv.save()
        self.assertEqual(inv.display_status, "Overdue")


class PdfGenerationTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="PDF Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust", email="c@example.com")
        self.inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-PDF1")
        CustomerInvoiceLine.objects.create(invoice=self.inv, description="Widget", qty=Decimal("2"), unit_price=Decimal("100.00"), tax_code=self.std)
        post_customer_invoice(self.inv)
        self.user = User.objects.create_user("pdfu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="pdfu", password="pw")

    def test_invoice_pdf_renders(self):
        resp = self.client.get(f"/ar/invoices/{self.inv.id}/pdf/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/pdf")
        self.assertTrue(resp.content[:5] == b"%PDF-")

    def test_credit_note_pdf_renders(self):
        from core.models import CreditNote, CreditNoteLine
        from core.services.gl import post_credit_note
        cn = CreditNote.objects.create(tenant=self.tenant, kind=CreditNote.Kind.SALES,
                                       credit_note_number="CN-PDF1", customer=self.customer)
        CreditNoteLine.objects.create(credit_note=cn, description="Refund", qty=Decimal("1"), unit_amount=Decimal("50.00"), tax_code=self.std)
        post_credit_note(cn)
        resp = self.client.get(f"/credit-notes/{cn.id}/pdf/")
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.content[:5] == b"%PDF-")

    def test_send_attaches_pdf(self):
        from django.core import mail
        with self.settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"):
            self.client.post(f"/ar/invoices/{self.inv.id}/send/")
            self.assertEqual(len(mail.outbox), 1)
            self.assertTrue(mail.outbox[0].attachments)
            fname, content, mime = mail.outbox[0].attachments[0]
            self.assertEqual(mime, "application/pdf")


class SalesDocumentFlowTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Flow Co", default_payment_terms_days=30)
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.tenant.default_tax_code = self.std
        self.tenant.save()
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust", email="c@example.com")
        self.user = User.objects.create_user("flowu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="flowu", password="pw")

    def _make_quote(self, number=""):
        self.client.post("/quotes/new/", {
            "customer": self.customer.id, "quote_number": number, "quote_date": "2026-06-01",
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "0", "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-description": "Widget", "lines-0-qty": "2", "lines-0-unit_price": "100",
            "lines-0-discount_pct": "0", "lines-0-tax_code": self.std.id,
        })
        from core.models import SalesQuote
        return SalesQuote.objects.get(tenant=self.tenant)

    def test_quote_auto_numbered(self):
        q = self._make_quote()
        self.assertEqual(q.quote_number, "QUO-0001")
        self.assertEqual(q.total, Decimal("240.00"))

    def test_quote_to_order_copies_lines(self):
        from core.models import CustomerOrder
        q = self._make_quote()
        resp = self.client.post(f"/quotes/{q.id}/to-order/")
        self.assertEqual(resp.status_code, 302)
        q.refresh_from_db()
        self.assertEqual(q.status, "CONVERTED")
        order = CustomerOrder.objects.get(tenant=self.tenant)
        self.assertEqual(order.quote_id, q.id)
        self.assertEqual(order.lines.count(), 1)
        self.assertEqual(order.total, Decimal("240.00"))

    def test_quote_to_invoice(self):
        q = self._make_quote()
        self.client.post(f"/quotes/{q.id}/to-invoice/")
        inv = CustomerInvoice.objects.get(tenant=self.tenant)
        self.assertEqual(inv.source_quote_id, q.id)
        self.assertEqual(inv.lines.count(), 1)
        self.assertEqual(inv.invoice_number, "INV-0001")
        self.assertIsNotNone(inv.due_date)

    def test_order_to_invoice(self):
        from core.models import CustomerOrder
        q = self._make_quote()
        self.client.post(f"/quotes/{q.id}/to-order/")
        order = CustomerOrder.objects.get(tenant=self.tenant)
        self.client.post(f"/customer-orders/{order.id}/to-invoice/")
        order.refresh_from_db()
        self.assertEqual(order.status, "INVOICED")
        inv = CustomerInvoice.objects.get(tenant=self.tenant)
        self.assertEqual(inv.source_order_id, order.id)
        self.assertEqual(inv.total, Decimal("240.00"))

    def test_quote_pdf_and_send(self):
        from django.core import mail
        q = self._make_quote()
        self.assertTrue(self.client.get(f"/quotes/{q.id}/pdf/").content[:5] == b"%PDF-")
        with self.settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"):
            self.client.post(f"/quotes/{q.id}/send/")
            self.assertEqual(len(mail.outbox), 1)
            self.assertTrue(mail.outbox[0].attachments)
        q.refresh_from_db()
        self.assertEqual(q.status, "SENT")

    def test_quote_accept(self):
        q = self._make_quote()
        self.client.post(f"/quotes/{q.id}/status/accept/")
        q.refresh_from_db()
        self.assertEqual(q.status, "ACCEPTED")


class RecurringInvoiceTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, RecurringInvoice, RecurringInvoiceLine
        import datetime
        self.tenant = Tenant.objects.create(name="Rec Co", default_payment_terms_days=14)
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.tmpl = RecurringInvoice.objects.create(
            tenant=self.tenant, customer=self.customer, name="Monthly retainer",
            frequency="MONTHLY", interval=1, start_date=datetime.date(2026, 1, 1),
            next_run_date=datetime.date(2026, 1, 1), auto_issue=True)
        RecurringInvoiceLine.objects.create(template=self.tmpl, description="Retainer",
                                            qty=Decimal("1"), unit_price=Decimal("500.00"), tax_code=self.std)
        self.user = User.objects.create_user("recu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="recu", password="pw")

    def test_add_months_clamps_day(self):
        import datetime
        from core.services.recurring import add_months
        self.assertEqual(add_months(datetime.date(2026, 1, 31), 1), datetime.date(2026, 2, 28))
        self.assertEqual(add_months(datetime.date(2026, 11, 15), 3), datetime.date(2027, 2, 15))

    def test_generate_catches_up_and_advances(self):
        import datetime
        from core.services import recurring
        # From Jan 1 to Mar 15 2026 -> Jan, Feb, Mar = 3 monthly invoices.
        created = recurring.generate_for_template(self.tmpl, today=datetime.date(2026, 3, 15))
        self.assertEqual(len(created), 3)
        self.tmpl.refresh_from_db()
        self.assertEqual(self.tmpl.occurrences, 3)
        self.assertEqual(self.tmpl.next_run_date, datetime.date(2026, 4, 1))
        # All issued + numbered + due date set from terms.
        for inv in created:
            self.assertEqual(inv.status, "ISSUED")
            self.assertTrue(inv.invoice_number.startswith("INV-"))
            self.assertIsNotNone(inv.due_date)
        self.assertEqual(created[0].total, Decimal("600.00"))  # 500 + 20% VAT

    def test_max_occurrences_deactivates(self):
        import datetime
        from core.services import recurring
        self.tmpl.max_occurrences = 2
        self.tmpl.save()
        created = recurring.generate_for_template(self.tmpl, today=datetime.date(2026, 12, 31))
        self.assertEqual(len(created), 2)
        self.tmpl.refresh_from_db()
        self.assertFalse(self.tmpl.is_active)

    def test_draft_mode_does_not_post(self):
        import datetime
        from core.services import recurring
        self.tmpl.auto_issue = False
        self.tmpl.save()
        created = recurring.generate_for_template(self.tmpl, today=datetime.date(2026, 1, 1))
        self.assertEqual(created[0].status, "DRAFT")

    def test_generate_now_view(self):
        resp = self.client.post(f"/recurring-invoices/{self.tmpl.id}/generate/")
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(CustomerInvoice.objects.filter(tenant=self.tenant).exists())

    def test_create_view_auto_next_run(self):
        from core.models import RecurringInvoice
        resp = self.client.post("/recurring-invoices/new/", {
            "name": "Weekly", "customer": self.customer.id, "frequency": "WEEKLY", "interval": "1",
            "start_date": "2026-02-01", "next_run_date": "", "auto_issue": "on",
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "0", "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-description": "Svc", "lines-0-qty": "1", "lines-0-unit_price": "10", "lines-0-discount_pct": "0", "lines-0-tax_code": self.std.id,
        })
        self.assertEqual(resp.status_code, 302)
        t = RecurringInvoice.objects.get(tenant=self.tenant, name="Weekly")
        self.assertEqual(t.next_run_date.isoformat(), "2026-02-01")  # defaulted from start


class RefundTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Refund Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-R1")
        CustomerInvoiceLine.objects.create(invoice=self.inv, description="Item", qty=Decimal("2"), unit_price=Decimal("100.00"), tax_code=self.std)
        post_customer_invoice(self.inv)  # total 240
        from core.models import Payment, PaymentAllocation
        from core.services.gl import post_payment
        p = Payment.objects.create(tenant=self.tenant, direction=Payment.Direction.RECEIPT,
                                   customer=self.customer, amount=Decimal("240.00"), method="BANK")
        PaymentAllocation.objects.create(payment=p, customer_invoice=self.inv, amount=Decimal("240.00"))
        post_payment(p)
        self.user = User.objects.create_user("refu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="refu", password="pw")

    def test_invoice_refund_sets_status_and_reverses_cash(self):
        from core.models import Payment, GLAccount
        from core.services import reports
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.status, "PAID")
        resp = self.client.post(f"/ar/invoices/{self.inv.id}/refund/")
        self.assertEqual(resp.status_code, 302)
        self.inv.refresh_from_db()
        self.assertEqual(self.inv.status, "REFUNDED")
        self.assertEqual(self.inv.display_status, "Refunded")
        refund = Payment.objects.get(tenant=self.tenant, direction="REFUND")
        self.assertEqual(refund.amount, Decimal("240.00"))
        # Bank nets to zero: +240 receipt then -240 refund.
        bank = GLAccount.objects.get(tenant=self.tenant, code="1050")
        self.assertEqual(reports.account_balances(self.tenant)[bank]["balance"], Decimal("0.00"))

    def test_standalone_refund_records_and_audits(self):
        from core.models import AuditLog, Payment
        resp = self.client.post("/payments/refunds/new/", {
            "customer": self.customer.id, "payment_date": "2026-06-01",
            "amount": "50.00", "method": "BANK", "reference": "RFD-1",
        })
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(Payment.objects.filter(tenant=self.tenant, direction="REFUND", amount=Decimal("50.00")).exists())
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="REFUND_RECORDED").exists())


class CustomerStatementTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, Payment, PaymentAllocation
        from core.services.gl import post_payment
        import datetime
        self.tenant = Tenant.objects.create(name="Stmt Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust", email="c@example.com")
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-S1",
                                             invoice_date=datetime.date(2026, 3, 1))
        CustomerInvoiceLine.objects.create(invoice=inv, description="Item", qty=Decimal("1"), unit_price=Decimal("100.00"), tax_code=self.std)
        post_customer_invoice(inv)  # 120
        p = Payment.objects.create(tenant=self.tenant, direction=Payment.Direction.RECEIPT, customer=self.customer,
                                   amount=Decimal("50.00"), method="BANK", payment_date=datetime.date(2026, 3, 10))
        PaymentAllocation.objects.create(payment=p, customer_invoice=inv, amount=Decimal("50.00"))
        post_payment(p)
        self.user = User.objects.create_user("stmtu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="stmtu", password="pw")

    def test_statement_running_balance(self):
        import datetime
        from core.services import statements
        data = statements.customer_statement(self.tenant, self.customer,
                                             datetime.date(2026, 1, 1), datetime.date(2026, 12, 31))
        self.assertEqual(data["opening"], Decimal("0.00"))
        self.assertEqual(len(data["rows"]), 2)            # invoice + receipt
        self.assertEqual(data["rows"][-1]["balance"], Decimal("70.00"))  # 120 - 50
        self.assertEqual(data["closing"], Decimal("70.00"))

    def test_opening_balance_excludes_pre_period(self):
        import datetime
        from core.services import statements
        # Period starting after both transactions -> they fall into opening.
        data = statements.customer_statement(self.tenant, self.customer,
                                             datetime.date(2026, 6, 1), datetime.date(2026, 12, 31))
        self.assertEqual(data["opening"], Decimal("70.00"))
        self.assertEqual(len(data["rows"]), 0)
        self.assertEqual(data["closing"], Decimal("70.00"))

    def test_statement_page_and_pdf(self):
        self.assertEqual(self.client.get(f"/customers/{self.customer.id}/statement/").status_code, 200)
        pdf = self.client.get(f"/customers/{self.customer.id}/statement/pdf/")
        self.assertTrue(pdf.content[:5] == b"%PDF-")

    def test_statement_email(self):
        from django.core import mail
        from core.models import AuditLog
        with self.settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"):
            self.client.post(f"/customers/{self.customer.id}/statement/email/")
            self.assertEqual(len(mail.outbox), 1)
            self.assertTrue(mail.outbox[0].attachments)
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="STATEMENT_SENT").exists())


class SalesReportsTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        import datetime
        self.tenant = Tenant.objects.create(name="SR Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.c1 = Customer.objects.create(tenant=self.tenant, name="Alpha")
        self.c2 = Customer.objects.create(tenant=self.tenant, name="Beta")
        self.p1 = Product.objects.create(tenant=self.tenant, sku="SKU-1", name="Widget")
        self.p2 = Product.objects.create(tenant=self.tenant, sku="SKU-2", name="Gadget")
        self._inv(self.c1, [(self.p1, 2, 100), (self.p2, 1, 50)], "INV-1", datetime.date(2026, 3, 1))
        self._inv(self.c2, [(self.p1, 1, 100)], "INV-2", datetime.date(2026, 3, 2))
        self.user = User.objects.create_user("sru", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="sru", password="pw")

    def _inv(self, customer, lines, number, date):
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=customer, invoice_number=number, invoice_date=date)
        for prod, qty, price in lines:
            CustomerInvoiceLine.objects.create(invoice=inv, product=prod, qty=Decimal(qty), unit_price=Decimal(price), tax_code=self.std)
        post_customer_invoice(inv)
        return inv

    def _range(self):
        import datetime
        return datetime.date(2026, 1, 1), datetime.date(2026, 12, 31)

    def test_history_totals(self):
        from core.services import sales_reports
        d = sales_reports.sales_history(self.tenant, *self._range())
        self.assertEqual(len(d["rows"]), 2)
        self.assertEqual(d["net_total"], Decimal("350.00"))   # 250 + 100
        self.assertEqual(d["grand_total"], Decimal("420.00"))  # +20% VAT

    def test_by_product(self):
        from core.services import sales_reports
        d = sales_reports.sales_by_product(self.tenant, *self._range())
        by_sku = {r["key"]: r for r in d["rows"]}
        self.assertEqual(by_sku["SKU-1"]["qty"], Decimal("3.00"))   # 2 + 1
        self.assertEqual(by_sku["SKU-1"]["net"], Decimal("300.00"))
        self.assertEqual(by_sku["SKU-2"]["net"], Decimal("50.00"))

    def test_by_customer(self):
        from core.services import sales_reports
        d = sales_reports.sales_by_customer(self.tenant, *self._range())
        by_name = {r["name"]: r for r in d["rows"]}
        self.assertEqual(by_name["Alpha"]["total"], Decimal("300.00"))  # (250)*1.2
        self.assertEqual(by_name["Beta"]["total"], Decimal("120.00"))

    def test_by_channel_direct(self):
        from core.services import sales_reports
        d = sales_reports.sales_by_channel(self.tenant, *self._range())
        direct = [r for r in d["rows"] if r["channel"].startswith("Direct")][0]
        self.assertEqual(direct["count"], 2)
        self.assertEqual(direct["total"], Decimal("420.00"))

    def test_report_pages_and_exports(self):
        for path in ["/sales/reports/", "/sales/reports/history/", "/sales/reports/by-product/",
                     "/sales/reports/by-customer/", "/sales/reports/by-channel/"]:
            self.assertEqual(self.client.get(path).status_code, 200, path)
        for k in ["sales-history", "sales-by-product", "sales-by-customer", "sales-by-channel"]:
            r = self.client.get(f"/finance/export/{k}.csv?from=2026-01-01&to=2026-12-31")
            self.assertEqual(r.status_code, 200, k)
            self.assertEqual(r["Content-Type"], "text/csv")


class InvoiceInventoryCogsTests(TestCase):
    def setUp(self):
        from core.models import Location, InventoryMovement, InventoryBalance
        from core.services.inventory import apply_movement
        self.tenant = Tenant.objects.create(name="COGS Co")
        Location.objects.filter(tenant=self.tenant).delete()  # drop auto-seeded Main Location; this class manages its own
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.loc = Location.objects.create(tenant=self.tenant, name="Main WH", type=Location.Type.WAREHOUSE)
        self.product = Product.objects.create(tenant=self.tenant, sku="SKU-C1", name="Widget",
                                              cost_method=Product.CostMethod.AVERAGE)
        # Opening stock: 100 @ 4.00 (sets moving-average cost).
        apply_movement(tenant=self.tenant, product=self.product, location=self.loc,
                       movement_type=InventoryMovement.MovementType.RECEIVE, qty_delta=Decimal("100"),
                       ref_type="SEED", ref_id="OPEN", unit_cost=Decimal("4.00"))

    def _invoice(self, with_product=True):
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-C1")
        if with_product:
            CustomerInvoiceLine.objects.create(invoice=inv, product=self.product, qty=Decimal("10"),
                                               unit_price=Decimal("25.00"), tax_code=self.std)
        else:
            CustomerInvoiceLine.objects.create(invoice=inv, description="Consulting", qty=Decimal("1"),
                                               unit_price=Decimal("250.00"), tax_code=self.std)
        post_customer_invoice(inv)
        return inv

    def test_stock_deducted_and_cogs_posted(self):
        from core.models import InventoryBalance, JournalEntry, GLAccount
        from core.services import reports
        self._invoice(with_product=True)
        bal = InventoryBalance.objects.get(tenant=self.tenant, product=self.product, location=self.loc)
        self.assertEqual(bal.on_hand, Decimal("90.00"))  # 100 - 10
        cogs_je = JournalEntry.objects.filter(tenant=self.tenant, ref_type="COGS", ref_id="INV-C1").first()
        self.assertIsNotNone(cogs_je)
        self.assertEqual(cogs_je.total_debit, Decimal("40.00"))  # 10 @ 4.00
        # P&L: revenue 250, COGS 40, gross profit 210.
        pnl = reports.profit_and_loss(self.tenant)
        self.assertEqual(pnl["cogs_total"], Decimal("40.00"))
        self.assertEqual(pnl["gross_profit"], Decimal("210.00"))

    def test_service_line_does_not_touch_stock(self):
        from core.models import JournalEntry, InventoryBalance
        self._invoice(with_product=False)
        self.assertFalse(JournalEntry.objects.filter(tenant=self.tenant, ref_type="COGS").exists())
        bal = InventoryBalance.objects.get(tenant=self.tenant, product=self.product, location=self.loc)
        self.assertEqual(bal.on_hand, Decimal("100.00"))  # untouched

    def test_reissue_does_not_double_post_cogs(self):
        from core.models import JournalEntry
        inv = self._invoice(with_product=True)
        post_customer_invoice(inv)  # idempotent re-call
        self.assertEqual(JournalEntry.objects.filter(tenant=self.tenant, ref_type="COGS", ref_id="INV-C1").count(), 1)

    def test_no_location_skips_cogs(self):
        from core.models import JournalEntry
        # A separate tenant with a product but no stock location at all.
        t2 = Tenant.objects.create(name="NoLoc Co")
        std2 = TaxCode.objects.get(tenant=t2, code="STD")
        cust2 = Customer.objects.create(tenant=t2, name="C2")
        prod2 = Product.objects.create(tenant=t2, sku="SKU-N1", name="Svc")
        inv = CustomerInvoice.objects.create(tenant=t2, customer=cust2, invoice_number="INV-N1")
        CustomerInvoiceLine.objects.create(invoice=inv, product=prod2, qty=Decimal("5"),
                                           unit_price=Decimal("25.00"), tax_code=std2)
        post_customer_invoice(inv)
        self.assertFalse(JournalEntry.objects.filter(tenant=t2, ref_type="COGS").exists())


class SalesEditTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, SalesQuote, SalesQuoteLine
        self.tenant = Tenant.objects.create(name="Edit Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.user = User.objects.create_user("editu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="editu", password="pw")

    def _formset(self, prefix_id, desc, qty, price, extra=None):
        data = {
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "1",
            "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-id": str(prefix_id), "lines-0-description": desc, "lines-0-qty": str(qty),
            "lines-0-unit_price": str(price), "lines-0-discount_pct": "0", "lines-0-tax_code": self.std.id,
        }
        if extra:
            data.update(extra)
        return data

    def test_edit_draft_invoice(self):
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-E1")
        line = CustomerInvoiceLine.objects.create(invoice=inv, description="Old", qty=Decimal("1"), unit_price=Decimal("10"), tax_code=self.std)
        data = {"customer": self.customer.id, "invoice_number": "INV-E1", "invoice_date": "2026-06-01", "action": "save"}
        data.update(self._formset(line.id, "New widget", 3, 50))
        resp = self.client.post(f"/ar/invoices/{inv.id}/edit/", data)
        self.assertEqual(resp.status_code, 302)
        inv.refresh_from_db()
        self.assertEqual(inv.subtotal, Decimal("150.00"))
        self.assertEqual(inv.lines.first().description, "New widget")

    def test_cannot_edit_issued_invoice(self):
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-E2")
        CustomerInvoiceLine.objects.create(invoice=inv, description="X", qty=Decimal("1"), unit_price=Decimal("10"), tax_code=self.std)
        post_customer_invoice(inv)
        resp = self.client.get(f"/ar/invoices/{inv.id}/edit/")
        self.assertEqual(resp.status_code, 302)  # redirected away

    def test_edit_draft_quote(self):
        from core.models import SalesQuote, SalesQuoteLine
        q = SalesQuote.objects.create(tenant=self.tenant, customer=self.customer, quote_number="QUO-E1")
        line = SalesQuoteLine.objects.create(quote=q, description="Old", qty=Decimal("1"), unit_price=Decimal("10"), tax_code=self.std)
        data = {"customer": self.customer.id, "quote_number": "QUO-E1", "quote_date": "2026-06-01"}
        data.update(self._formset(line.id, "Revised", 2, 75))
        resp = self.client.post(f"/quotes/{q.id}/edit/", data)
        self.assertEqual(resp.status_code, 302)
        q.refresh_from_db()
        self.assertEqual(q.total, Decimal("180.00"))  # 150 + 20% VAT

    def test_edit_recurring_template(self):
        from core.models import RecurringInvoice, RecurringInvoiceLine
        import datetime
        t = RecurringInvoice.objects.create(tenant=self.tenant, customer=self.customer, name="Old name",
                                            frequency="MONTHLY", interval=1, start_date=datetime.date(2026, 1, 1),
                                            next_run_date=datetime.date(2026, 1, 1))
        line = RecurringInvoiceLine.objects.create(template=t, description="Svc", qty=Decimal("1"), unit_price=Decimal("100"), tax_code=self.std)
        data = {"name": "New name", "customer": self.customer.id, "frequency": "MONTHLY", "interval": "1",
                "start_date": "2026-01-01", "next_run_date": "2026-01-01", "auto_issue": "on"}
        data.update(self._formset(line.id, "Svc", 1, 150))
        resp = self.client.post(f"/recurring-invoices/{t.id}/edit/", data)
        self.assertEqual(resp.status_code, 302)
        t.refresh_from_db()
        self.assertEqual(t.name, "New name")
        self.assertEqual(t.total, Decimal("180.00"))


class SalesHousekeepingTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, SalesQuote, SalesQuoteLine, RecurringInvoice, RecurringInvoiceLine
        import datetime
        self.tenant = Tenant.objects.create(name="HK Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.user = User.objects.create_user("hku", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="hku", password="pw")

    def test_expire_quotes(self):
        from core.models import SalesQuote
        from core.services import housekeeping
        import datetime
        past = SalesQuote.objects.create(tenant=self.tenant, customer=self.customer, quote_number="QUO-X1",
                                         status=SalesQuote.Status.SENT, valid_until=datetime.date(2020, 1, 1))
        future = SalesQuote.objects.create(tenant=self.tenant, customer=self.customer, quote_number="QUO-X2",
                                           status=SalesQuote.Status.SENT, valid_until=datetime.date(2999, 1, 1))
        n = housekeeping.expire_quotes(self.tenant)
        self.assertEqual(n, 1)
        past.refresh_from_db(); future.refresh_from_db()
        self.assertEqual(past.status, "EXPIRED")
        self.assertEqual(future.status, "SENT")

    def test_run_for_tenant_throttled_once_per_day(self):
        from core.services import housekeeping
        r1 = housekeeping.run_for_tenant(self.tenant)
        self.assertIsNotNone(r1)
        r2 = housekeeping.run_for_tenant(self.tenant)  # same day -> skipped
        self.assertIsNone(r2)

    def test_run_for_tenant_generates_due_recurring(self):
        from core.models import RecurringInvoice, RecurringInvoiceLine, CustomerInvoice
        from core.services import housekeeping
        import datetime
        from django.utils import timezone
        from core.services.recurring import add_months
        start = add_months(timezone.localdate(), -2)
        t = RecurringInvoice.objects.create(tenant=self.tenant, customer=self.customer, name="Retainer",
                                            frequency="MONTHLY", interval=1, start_date=start, next_run_date=start, auto_issue=True)
        RecurringInvoiceLine.objects.create(template=t, description="Svc", qty=Decimal("1"), unit_price=Decimal("100"), tax_code=self.std)
        res = housekeeping.run_for_tenant(self.tenant, force=True)
        self.assertGreaterEqual(res["generated"], 2)
        self.assertTrue(CustomerInvoice.objects.filter(tenant=self.tenant).exists())

    def test_quote_delete(self):
        from core.models import SalesQuote, AuditLog
        q = SalesQuote.objects.create(tenant=self.tenant, customer=self.customer, quote_number="QUO-D1", status=SalesQuote.Status.DRAFT)
        resp = self.client.post(f"/quotes/{q.id}/delete/")
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(SalesQuote.objects.filter(id=q.id).exists())
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="QUOTE_DELETED").exists())

    def test_converted_quote_cannot_be_deleted(self):
        from core.models import SalesQuote
        q = SalesQuote.objects.create(tenant=self.tenant, customer=self.customer, quote_number="QUO-D2", status=SalesQuote.Status.CONVERTED)
        self.client.post(f"/quotes/{q.id}/delete/")
        self.assertTrue(SalesQuote.objects.filter(id=q.id).exists())  # blocked

    def test_order_delete_and_invoiced_block(self):
        from core.models import CustomerOrder
        draft = CustomerOrder.objects.create(tenant=self.tenant, customer=self.customer, order_number="SO-D1", status=CustomerOrder.Status.DRAFT)
        self.client.post(f"/customer-orders/{draft.id}/delete/")
        self.assertFalse(CustomerOrder.objects.filter(id=draft.id).exists())
        invoiced = CustomerOrder.objects.create(tenant=self.tenant, customer=self.customer, order_number="SO-D2", status=CustomerOrder.Status.INVOICED)
        self.client.post(f"/customer-orders/{invoiced.id}/delete/")
        self.assertTrue(CustomerOrder.objects.filter(id=invoiced.id).exists())  # blocked

    def test_management_command(self):
        from django.core.management import call_command
        from core.models import SalesQuote
        import datetime
        SalesQuote.objects.create(tenant=self.tenant, customer=self.customer, quote_number="QUO-C1",
                                  status=SalesQuote.Status.SENT, valid_until=datetime.date(2020, 1, 1))
        call_command("run_sales_housekeeping")
        self.assertEqual(SalesQuote.objects.get(quote_number="QUO-C1").status, "EXPIRED")


class CustomerRecordTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Cust Co")
        self.user = User.objects.create_user("custu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="custu", password="pw")

    def test_create_with_all_fields(self):
        resp = self.client.post("/customers/new/", {
            "name": "Acme Wholesale", "customer_type": "WHOLESALE", "status": "ACTIVE",
            "contact_person": "Jo Bloggs", "email": "jo@acme.example", "phone": "+44 20 7946 0000",
            "vat_number": "GB123456789", "company_number": "12345678",
            "billing_address": "1 High St", "shipping_address": "Dock 4",
            "payment_terms_days": "45", "credit_limit": "2500.00", "tags": "VIP, Reseller",
            "notes": "Top account",
        })
        self.assertEqual(resp.status_code, 302)
        c = Customer.objects.get(tenant=self.tenant, name="Acme Wholesale")
        self.assertEqual(c.customer_type, "WHOLESALE")
        self.assertEqual(c.payment_terms_days, 45)
        self.assertEqual(c.credit_limit, Decimal("2500.00"))
        self.assertEqual(c.tag_list, ["VIP", "Reseller"])

    def test_outstanding_balance_and_credit_limit(self):
        from core.services.gl import post_customer_invoice
        std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        c = Customer.objects.create(tenant=self.tenant, name="Limited Co", credit_limit=Decimal("100.00"))
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=c, invoice_number="INV-CL1")
        CustomerInvoiceLine.objects.create(invoice=inv, description="X", qty=Decimal("1"), unit_price=Decimal("200.00"), tax_code=std)
        post_customer_invoice(inv)  # 240 outstanding
        self.assertEqual(c.outstanding_balance, Decimal("240.00"))
        self.assertEqual(c.available_credit, Decimal("-140.00"))
        self.assertTrue(c.is_over_limit)

    def test_import_export_round_trip(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv = (b"name,customer_type,contact_person,email,phone,vat_number,company_number,"
               b"billing_address,shipping_address,payment_terms_days,tags\n"
               b"Imported Ltd,TRADE,Sam,sam@imp.example,123,GB1,1122,Addr,Ship,14,Trade\n")
        resp = self.client.post("/customers/import/", {"file": SimpleUploadedFile("c.csv", csv, content_type="text/csv")})
        self.assertEqual(resp.status_code, 200)
        c = Customer.objects.get(tenant=self.tenant, name="Imported Ltd")
        self.assertEqual(c.customer_type, "TRADE")
        self.assertEqual(c.payment_terms_days, 14)
        # export contains the new columns
        body = self.client.get("/export/customers.csv").content.decode()
        self.assertIn("customer_type", body)
        self.assertIn("Imported Ltd", body)


class CustomerProfileTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, Payment, PaymentAllocation, CreditNote, CreditNoteLine
        from core.services.gl import post_customer_invoice, post_payment, post_credit_note
        self.tenant = Tenant.objects.create(name="Prof Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Profile Cust", email="p@example.com",
                                                customer_type="TRADE", credit_limit=Decimal("1000.00"))
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-P1")
        CustomerInvoiceLine.objects.create(invoice=inv, description="Item", qty=Decimal("2"), unit_price=Decimal("100.00"), tax_code=self.std)
        post_customer_invoice(inv)
        p = Payment.objects.create(tenant=self.tenant, direction=Payment.Direction.RECEIPT, customer=self.customer,
                                   amount=Decimal("100.00"), method="BANK")
        PaymentAllocation.objects.create(payment=p, customer_invoice=inv, amount=Decimal("100.00"))
        post_payment(p)
        cn = CreditNote.objects.create(tenant=self.tenant, kind=CreditNote.Kind.SALES, credit_note_number="CN-P1", customer=self.customer)
        CreditNoteLine.objects.create(credit_note=cn, description="Adj", qty=Decimal("1"), unit_amount=Decimal("10.00"), tax_code=self.std)
        post_credit_note(cn)
        self.user = User.objects.create_user("profu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="profu", password="pw")

    def test_profile_page_shows_everything(self):
        resp = self.client.get(f"/customers/{self.customer.id}/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "INV-P1")        # invoice
        self.assertContains(resp, "CN-P1")         # credit note
        self.assertContains(resp, "Receipt")       # payment
        self.assertContains(resp, "Activity timeline")
        # outstanding = invoice 240 - 100 paid = 140 (standalone credit note not allocated to it)
        self.assertEqual(resp.context["c"].outstanding_balance, Decimal("140.00"))

    def test_timeline_is_populated_and_sorted(self):
        resp = self.client.get(f"/customers/{self.customer.id}/")
        timeline = resp.context["timeline"]
        self.assertGreaterEqual(len(timeline), 3)  # invoice + payment + credit note
        dates = [e["date"] for e in timeline]
        self.assertEqual(dates, sorted(dates, reverse=True))


class CustomerSearchDedupTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Search Co")
        self.user = User.objects.create_user("srchu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="srchu", password="pw")
        Customer.objects.create(tenant=self.tenant, name="Alpha Retail", email="a@alpha.example",
                                phone="111", vat_number="GB111", customer_type="TRADE", status="ACTIVE", tags="VIP")
        Customer.objects.create(tenant=self.tenant, name="Beta Wholesale", email="b@beta.example",
                                phone="222", customer_type="WHOLESALE", status="INACTIVE", tags="Bulk")

    def test_search_by_text(self):
        resp = self.client.get("/customers/?q=alpha")
        names = [c.name for c in resp.context["customers"]]
        self.assertEqual(names, ["Alpha Retail"])
        resp = self.client.get("/customers/?q=222")  # phone
        self.assertEqual([c.name for c in resp.context["customers"]], ["Beta Wholesale"])

    def test_filter_by_type_status_tag(self):
        self.assertEqual([c.name for c in self.client.get("/customers/?type=TRADE").context["customers"]], ["Alpha Retail"])
        self.assertEqual([c.name for c in self.client.get("/customers/?status=INACTIVE").context["customers"]], ["Beta Wholesale"])
        self.assertEqual([c.name for c in self.client.get("/customers/?tag=VIP").context["customers"]], ["Alpha Retail"])

    def test_duplicate_detection_blocks_then_confirms(self):
        # New customer reusing Alpha's email -> flagged as duplicate, not saved.
        data = {"name": "Alpha Retail North", "customer_type": "COMPANY", "status": "ACTIVE",
                "email": "a@alpha.example", "credit_limit": "0"}
        resp = self.client.post("/customers/new/", data)
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.context["duplicates"])
        self.assertFalse(Customer.objects.filter(tenant=self.tenant, name="Alpha Retail North").exists())
        # Confirm "save anyway".
        data["confirm_duplicate"] = "1"
        resp = self.client.post("/customers/new/", data)
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(Customer.objects.filter(tenant=self.tenant, name="Alpha Retail North").exists())

    def test_duplicate_helper_matches_multiple_fields(self):
        from core.views import _find_customer_duplicates
        from core.models import Customer as C
        probe = C(tenant=self.tenant, name="X", email="a@alpha.example", phone="222")
        dups = _find_customer_duplicates(self.tenant, probe)
        matched = {(d["name"], d["match"]) for d in dups}
        self.assertIn(("Alpha Retail", "email"), matched)
        self.assertIn(("Beta Wholesale", "phone"), matched)


class SupplierRecordTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Sup Co")
        self.user = User.objects.create_user("supu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="supu", password="pw")

    def test_create_with_all_fields(self):
        resp = self.client.post("/suppliers/new/", {
            "name": "Globex", "status": "ACTIVE", "currency_code": "GBP", "contact_person": "Pat",
            "email": "p@globex.example", "phone": "111", "vat_number": "GB1", "company_number": "1122",
            "address": "Globex House", "payment_terms_days": "30", "bank_name": "Barclays",
            "bank_account_name": "Globex Ltd", "bank_sort_code": "20-00-00", "bank_account_number": "12345678",
            "categories": "Raw materials, Logistics", "notes": "Primary",
        })
        self.assertEqual(resp.status_code, 302)
        s = Supplier.objects.get(tenant=self.tenant, name="Globex")
        self.assertEqual(s.payment_terms_days, 30)
        self.assertEqual(s.bank_account_number, "12345678")
        self.assertEqual(s.category_list, ["Raw materials", "Logistics"])

    def test_outstanding_payables(self):
        from core.models import PurchaseOrder, GoodsReceipt, Location, SupplierInvoice, SupplierInvoiceLine, Product
        from core.services.gl import post_supplier_invoice
        s = Supplier.objects.create(tenant=self.tenant, name="Bills Co")
        std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        prod = Product.objects.create(tenant=self.tenant, sku="SKU-S1", name="Part")
        loc = Location.objects.create(tenant=self.tenant, name="WH")
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-S1", supplier=s)
        grn = GoodsReceipt.objects.create(tenant=self.tenant, po=po, grn_number="GRN-S1", received_to=loc, status=GoodsReceipt.Status.POSTED)
        inv = SupplierInvoice.objects.create(tenant=self.tenant, supplier=s, po=po, receipt=grn, invoice_number="BILL-1")
        SupplierInvoiceLine.objects.create(invoice=inv, product=prod, qty=Decimal("10"), unit_cost=Decimal("5.00"), tax_code=std)
        post_supplier_invoice(inv)  # 60 owed
        self.assertEqual(s.outstanding_payables, Decimal("60.00"))

    def test_import_export_round_trip(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv = (b"name,contact_person,email,phone,vat_number,company_number,address,currency_code,payment_terms_days,categories\n"
               b"Imported Supplies,Sam,sam@imp.example,123,GB9,9988,Addr,GBP,14,Raw materials\n")
        resp = self.client.post("/suppliers/import/", {"file": SimpleUploadedFile("s.csv", csv, content_type="text/csv")})
        self.assertEqual(resp.status_code, 200)
        s = Supplier.objects.get(tenant=self.tenant, name="Imported Supplies")
        self.assertEqual(s.payment_terms_days, 14)
        self.assertEqual(s.contact_person, "Sam")
        body = self.client.get("/export/suppliers.csv").content.decode()
        self.assertIn("contact_person", body)
        self.assertIn("Imported Supplies", body)


class SupplierProfileTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, PurchaseOrder, PurchaseOrderLine, GoodsReceipt, Location, SupplierInvoice, SupplierInvoiceLine
        from core.services.gl import post_supplier_invoice
        self.tenant = Tenant.objects.create(name="SupProf Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Globex", contact_person="Pat", categories="Raw materials")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH")
        self.prod = Product.objects.create(tenant=self.tenant, sku="SKU-PR1", name="Widget")
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-PR1", supplier=self.supplier)
        PurchaseOrderLine.objects.create(po=po, product=self.prod, ordered_qty=Decimal("10"), unit_cost=Decimal("4.00"))
        grn = GoodsReceipt.objects.create(tenant=self.tenant, po=po, grn_number="GRN-PR1", received_to=self.loc, status=GoodsReceipt.Status.POSTED)
        inv = SupplierInvoice.objects.create(tenant=self.tenant, supplier=self.supplier, po=po, receipt=grn, invoice_number="BILL-PR1")
        SupplierInvoiceLine.objects.create(invoice=inv, product=self.prod, qty=Decimal("10"), unit_cost=Decimal("4.00"), tax_code=self.std)
        post_supplier_invoice(inv)
        self.user = User.objects.create_user("spu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="spu", password="pw")

    def test_profile_shows_pos_bills_products_price_history(self):
        resp = self.client.get(f"/suppliers/{self.supplier.id}/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "PO-PR1")
        self.assertContains(resp, "BILL-PR1")
        self.assertContains(resp, "Products supplied")
        self.assertContains(resp, "Price history")
        self.assertContains(resp, "SKU-PR1")
        # products supplied + price history populated
        self.assertEqual(len(resp.context["products_supplied"]), 1)
        self.assertEqual(resp.context["products_supplied"][0]["last_cost"], Decimal("4.00"))
        self.assertGreaterEqual(len(resp.context["price_history"]), 1)
        # outstanding payables = 48 (40 net + 20% VAT)
        self.assertEqual(resp.context["s"].outstanding_payables, Decimal("48.00"))

    def test_timeline_sorted(self):
        resp = self.client.get(f"/suppliers/{self.supplier.id}/")
        dates = [e["date"] for e in resp.context["timeline"]]
        self.assertEqual(dates, sorted(dates, reverse=True))
        self.assertGreaterEqual(len(dates), 2)  # PO + bill


class SupplierDedupPreferredTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="SD Co")
        self.user = User.objects.create_user("sdu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="sdu", password="pw")
        Supplier.objects.create(tenant=self.tenant, name="Globex", email="g@globex.example", phone="111",
                                vat_number="GB1", status="ACTIVE", categories="Raw materials")
        Supplier.objects.create(tenant=self.tenant, name="Acme", email="a@acme.example", phone="222",
                                status="INACTIVE", categories="Logistics")

    def test_search_and_filter(self):
        self.assertEqual([s.name for s in self.client.get("/suppliers/?q=globex").context["suppliers"]], ["Globex"])
        self.assertEqual([s.name for s in self.client.get("/suppliers/?q=222").context["suppliers"]], ["Acme"])
        self.assertEqual([s.name for s in self.client.get("/suppliers/?status=INACTIVE").context["suppliers"]], ["Acme"])
        self.assertEqual([s.name for s in self.client.get("/suppliers/?category=Raw").context["suppliers"]], ["Globex"])

    def test_duplicate_detection(self):
        data = {"name": "Globex North", "status": "ACTIVE", "currency_code": "GBP", "email": "g@globex.example"}
        resp = self.client.post("/suppliers/new/", data)
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.context["duplicates"])
        self.assertFalse(Supplier.objects.filter(tenant=self.tenant, name="Globex North").exists())
        data["confirm_duplicate"] = "1"
        resp = self.client.post("/suppliers/new/", data)
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(Supplier.objects.filter(tenant=self.tenant, name="Globex North").exists())

    def test_preferred_supplier_link_on_product(self):
        from core.models import OrgMembership
        s = Supplier.objects.get(tenant=self.tenant, name="Globex")
        prod = Product.objects.create(tenant=self.tenant, sku="SKU-PS1", name="Widget")
        # set preferred supplier via the product edit form
        resp = self.client.post(f"/products/{prod.id}/edit/", {
            "sku": "SKU-PS1", "name": "Widget", "product_type": "STOCK", "uom": "each",
            "cost_method": "AVERAGE", "standard_cost": "5.00", "sales_price": "0",
            "reorder_level": "0", "preferred_supplier": s.id,
        })
        self.assertEqual(resp.status_code, 302)
        prod.refresh_from_db()
        self.assertEqual(prod.preferred_supplier_id, s.id)
        # appears under products supplied on the supplier profile
        resp = self.client.get(f"/suppliers/{s.id}/")
        skus = [p["product"].sku for p in resp.context["products_supplied"]]
        self.assertIn("SKU-PS1", skus)


class DetailPageRenderRegressionTests(TestCase):
    """Guards against |default: chains resolving attributes on None FKs."""
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Render Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.user = User.objects.create_user("rru", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="rru", password="pw")

    def test_credit_note_detail_renders_when_linked_to_invoice(self):
        from core.models import CreditNote, CreditNoteLine
        from core.services.gl import post_customer_invoice, post_credit_note
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer, invoice_number="INV-R1")
        CustomerInvoiceLine.objects.create(invoice=inv, description="X", qty=Decimal("1"), unit_price=Decimal("100"), tax_code=self.std)
        post_customer_invoice(inv)
        cn = CreditNote.objects.create(tenant=self.tenant, kind=CreditNote.Kind.SALES, credit_note_number="CN-R1",
                                       customer=self.customer, customer_invoice=inv)
        CreditNoteLine.objects.create(credit_note=cn, description="Adj", qty=Decimal("1"), unit_amount=Decimal("10"), tax_code=self.std)
        post_credit_note(cn)
        self.assertEqual(self.client.get(f"/credit-notes/{cn.id}/").status_code, 200)
        self.assertTrue(self.client.get(f"/credit-notes/{cn.id}/pdf/").content[:5] == b"%PDF-")

    def test_recurring_detail_renders_with_productless_line(self):
        from core.models import RecurringInvoice, RecurringInvoiceLine
        import datetime
        t = RecurringInvoice.objects.create(tenant=self.tenant, customer=self.customer, name="Retainer",
                                            frequency="MONTHLY", interval=1, start_date=datetime.date(2026, 1, 1),
                                            next_run_date=datetime.date(2026, 1, 1))
        RecurringInvoiceLine.objects.create(template=t, product=None, description="Support retainer",
                                            qty=Decimal("1"), unit_price=Decimal("300"), tax_code=self.std)
        self.assertEqual(self.client.get(f"/recurring-invoices/{t.id}/").status_code, 200)


class ProductRecordTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, Location
        self.tenant = Tenant.objects.create(name="Prod Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.loc = Location.objects.create(tenant=self.tenant, name="Main WH", type=Location.Type.WAREHOUSE)
        self.user = User.objects.create_user("produ", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="produ", password="pw")

    def _post(self, **over):
        data = {
            "sku": "SKU-A1", "name": "Widget", "product_type": "STOCK", "brand": "Acme",
            "description": "A widget", "is_active": "on", "uom": "each", "sales_price": "19.99",
            "tax_code": self.std.id, "cost_method": "AVERAGE", "standard_cost": "8.00",
            "reorder_level": "10", "barcode": "5012345678900",
            "opening_stock": "50", "opening_location": self.loc.id,
        }
        data.update(over)
        return self.client.post("/products/new/", data)

    def test_create_with_fields_and_opening_stock(self):
        from core.models import Product, InventoryBalance, ProductBarcode
        resp = self._post()
        self.assertEqual(resp.status_code, 302)
        p = Product.objects.get(tenant=self.tenant, sku="SKU-A1")
        self.assertEqual(p.product_type, "STOCK")
        self.assertEqual(p.sales_price, Decimal("19.99"))
        self.assertEqual(p.tax_code_id, self.std.id)
        self.assertTrue(ProductBarcode.objects.filter(tenant=self.tenant, code="5012345678900").exists())
        bal = InventoryBalance.objects.get(tenant=self.tenant, product=p, location=self.loc)
        self.assertEqual(bal.on_hand, Decimal("50.00"))
        # margin = 19.99 - 8.00 cost
        self.assertEqual(p.margin, Decimal("11.99"))

    def test_duplicate_sku_friendly_error(self):
        self._post()
        resp = self._post(barcode="")  # same SKU
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "SKU already exists")

    def test_duplicate_barcode_friendly_error(self):
        self._post()
        resp = self._post(sku="SKU-A2")  # same barcode
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "already assigned to another product")

    def test_import_creates_category_and_subcategory(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from core.models import Product, ProductCategory
        csv = (b"sku,name,product_type,category,brand,description,uom,cost_method,standard_cost,sales_price,is_active,barcode\n"
               b"SKU-IMP1,Imported Cam,STOCK,Electronics / Webcams,Acme,HD cam,each,AVERAGE,9.99,19.99,yes,5099999999999\n")
        resp = self.client.post("/products/import/", {"file": SimpleUploadedFile("p.csv", csv, content_type="text/csv")})
        self.assertEqual(resp.status_code, 200)
        p = Product.objects.get(tenant=self.tenant, sku="SKU-IMP1")
        self.assertEqual(p.product_type, "STOCK")
        self.assertEqual(p.sales_price, Decimal("19.99"))
        self.assertEqual(str(p.category), "Electronics / Webcams")
        self.assertTrue(ProductCategory.objects.filter(tenant=self.tenant, name="Webcams", parent__name="Electronics").exists())
        body = self.client.get("/export/products.csv").content.decode()
        self.assertIn("product_type", body)
        self.assertIn("SKU-IMP1", body)


class ProductDetailTests(TestCase):
    def setUp(self):
        from core.models import (OrgMembership, Location, InventoryMovement, PurchaseOrder,
                                 PurchaseOrderLine, Supplier)
        from core.services.inventory import apply_movement
        from core.services.gl import post_customer_invoice
        self.tenant = Tenant.objects.create(name="PD Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.loc = Location.objects.create(tenant=self.tenant, name="Main WH", type=Location.Type.WAREHOUSE)
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Globex")
        self.product = Product.objects.create(tenant=self.tenant, sku="SKU-PD1", name="Widget",
                                              sales_price=Decimal("20.00"), standard_cost=Decimal("8.00"),
                                              preferred_supplier=self.supplier, reorder_level=Decimal("10"))
        # opening stock + cost
        apply_movement(tenant=self.tenant, product=self.product, location=self.loc,
                       movement_type=InventoryMovement.MovementType.RECEIVE, qty_delta=Decimal("100"),
                       ref_type="SEED", ref_id="X", unit_cost=Decimal("8.00"))
        # a PO line (purchase + price history)
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-PD1", supplier=self.supplier)
        PurchaseOrderLine.objects.create(po=po, product=self.product, ordered_qty=Decimal("50"), unit_cost=Decimal("7.50"))
        from core.services.purchasing import record_po_prices
        record_po_prices(po)  # populates supplier price history
        # a sale
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=Customer.objects.create(tenant=self.tenant, name="C"), invoice_number="INV-PD1")
        CustomerInvoiceLine.objects.create(invoice=inv, product=self.product, qty=Decimal("5"), unit_price=Decimal("20.00"), tax_code=self.std)
        post_customer_invoice(inv)  # deducts 5 from stock + records sale
        self.user = User.objects.create_user("pdu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="pdu", password="pw")

    def test_detail_page_renders_all_sections(self):
        resp = self.client.get(f"/products/{self.product.id}/")
        self.assertEqual(resp.status_code, 200)
        for needle in ["Stock by location", "Sales history", "Purchase history",
                       "Price history", "Stock movements", "Suppliers",
                       "INV-PD1", "PO-PD1", "Globex", "Main WH"]:
            self.assertContains(resp, needle)

    def test_detail_aggregates(self):
        resp = self.client.get(f"/products/{self.product.id}/")
        self.assertEqual(resp.context["qty_sold"], Decimal("5.00"))
        self.assertEqual(resp.context["revenue"], Decimal("100.00"))
        self.assertEqual(resp.context["qty_purchased"], Decimal("50.00"))
        self.assertEqual(len(resp.context["price_history"]), 1)
        # margin: 20 sales - 8 avg cost = 12
        self.assertEqual(resp.context["p"].margin, Decimal("12.00"))
        # on hand: 100 received - 5 sold = 95
        self.assertEqual(resp.context["p"].on_hand_total, Decimal("95.00"))


class ProductCategorySearchTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, ProductCategory
        self.tenant = Tenant.objects.create(name="PC Co")
        self.user = User.objects.create_user("pcu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="pcu", password="pw")
        self.elec = ProductCategory.objects.create(tenant=self.tenant, name="Electronics")
        Product.objects.create(tenant=self.tenant, sku="SKU-EL1", name="Webcam", brand="Acme",
                               product_type="STOCK", category=self.elec, is_active=True)
        Product.objects.create(tenant=self.tenant, sku="SKU-SV1", name="Setup service", brand="SwifPro",
                               product_type="SERVICE", is_active=False)

    def test_search_and_filters(self):
        self.assertEqual([p.sku for p in self.client.get("/products/?q=webcam").context["products"]], ["SKU-EL1"])
        self.assertEqual([p.sku for p in self.client.get("/products/?q=acme").context["products"]], ["SKU-EL1"])
        self.assertEqual([p.sku for p in self.client.get("/products/?type=SERVICE").context["products"]], ["SKU-SV1"])
        self.assertEqual([p.sku for p in self.client.get(f"/products/?category={self.elec.id}").context["products"]], ["SKU-EL1"])
        self.assertEqual([p.sku for p in self.client.get("/products/?status=inactive").context["products"]], ["SKU-SV1"])

    def test_category_create_and_subcategory_and_delete(self):
        from core.models import ProductCategory
        # create a top-level
        self.client.post("/product-categories/", {"name": "Office"})
        office = ProductCategory.objects.get(tenant=self.tenant, name="Office")
        # create a subcategory under it
        self.client.post("/product-categories/", {"name": "Stationery", "parent": office.id})
        sub = ProductCategory.objects.get(tenant=self.tenant, name="Stationery")
        self.assertEqual(sub.parent_id, office.id)
        # page renders
        self.assertEqual(self.client.get("/product-categories/").status_code, 200)
        # delete: products keep data (FK SET_NULL)
        self.client.post(f"/product-categories/{self.elec.id}/delete/")
        self.assertFalse(ProductCategory.objects.filter(id=self.elec.id).exists())
        self.assertTrue(Product.objects.filter(sku="SKU-EL1").exists())

    def test_category_parent_choices_exclude_subcategories(self):
        from core.forms import ProductCategoryForm
        from core.models import ProductCategory
        from core.current import set_current_tenant
        set_current_tenant(self.tenant)
        sub = ProductCategory.objects.create(tenant=self.tenant, name="Cameras", parent=self.elec)
        form = ProductCategoryForm()
        parents = list(form.fields["parent"].queryset)
        self.assertIn(self.elec, parents)
        self.assertNotIn(sub, parents)  # subcategories can't be parents
        set_current_tenant(None)


class InventoryLedgerFoundationTests(TestCase):
    def test_movement_records_user_and_new_types_exist(self):
        from core.models import InventoryMovement, Location, Product
        from core.services.inventory import apply_movement
        t = Tenant.objects.create(name="Ledger Co")
        u = User.objects.create_user("ledu", password="pw")
        loc = Location.objects.create(tenant=t, name="Van A", type=Location.Type.VAN)
        p = Product.objects.create(tenant=t, sku="SKU-L1", name="Item")
        m = apply_movement(tenant=t, product=p, location=loc,
                           movement_type=InventoryMovement.MovementType.DAMAGE,
                           qty_delta=Decimal("-2"), ref_type="ADJ", ref_id="1",
                           notes="broken", user=u)
        self.assertEqual(m.user_id, u.id)
        self.assertEqual(m.movement_type, "DAMAGE")
        # new movement types + location types are registered
        mt = dict(InventoryMovement.MovementType.choices)
        for k in ("DAMAGE", "WRITE_OFF", "RETURN_SUPPLIER"):
            self.assertIn(k, mt)
        lt = dict(Location.Type.choices)
        for k in ("OFFICE", "VAN", "POPUP"):
            self.assertIn(k, lt)


class StockAdjustmentTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, Location, InventoryMovement
        from core.services.inventory import apply_movement
        self.tenant = Tenant.objects.create(name="Adj Co", stock_adjustment_approval_threshold=Decimal("100.00"))
        self.loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        self.product = Product.objects.create(tenant=self.tenant, sku="SKU-AJ1", name="Widget", standard_cost=Decimal("10.00"))
        apply_movement(tenant=self.tenant, product=self.product, location=self.loc,
                       movement_type=InventoryMovement.MovementType.RECEIVE, qty_delta=Decimal("100"),
                       ref_type="SEED", ref_id="X", unit_cost=Decimal("10.00"))
        self.user = User.objects.create_user("aju", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="aju", password="pw")

    def _bal(self):
        from core.models import InventoryBalance
        return InventoryBalance.objects.get(tenant=self.tenant, product=self.product, location=self.loc).on_hand

    def test_small_adjustment_auto_posts(self):
        from core.models import StockAdjustment, InventoryMovement
        # 3 units @ 10 = 30 < 100 threshold -> posts immediately
        resp = self.client.post("/inventory/adjustments/new/", {
            "product": self.product.id, "location": self.loc.id, "reason": "DAMAGE",
            "qty_delta": "-3", "notes": "broken"})
        self.assertEqual(resp.status_code, 302)
        adj = StockAdjustment.objects.get(tenant=self.tenant)
        self.assertEqual(adj.status, "POSTED")
        self.assertEqual(self._bal(), Decimal("97.00"))
        m = InventoryMovement.objects.filter(tenant=self.tenant, movement_type="DAMAGE").first()
        self.assertEqual(m.user_id, self.user.id)

    def test_large_adjustment_needs_approval(self):
        from core.models import StockAdjustment
        # 20 @ 10 = 200 >= 100 -> pending, not posted
        self.client.post("/inventory/adjustments/new/", {
            "product": self.product.id, "location": self.loc.id, "reason": "WRITE_OFF",
            "qty_delta": "-20", "notes": "lost"})
        adj = StockAdjustment.objects.get(tenant=self.tenant)
        self.assertEqual(adj.status, "PENDING")
        self.assertEqual(self._bal(), Decimal("100.00"))  # unchanged
        # approve -> posts
        self.client.post(f"/inventory/adjustments/{adj.id}/approve/")
        adj.refresh_from_db()
        self.assertEqual(adj.status, "POSTED")
        self.assertEqual(adj.approved_by_id, self.user.id)
        self.assertEqual(self._bal(), Decimal("80.00"))

    def test_reject_does_not_post(self):
        from core.models import StockAdjustment
        self.client.post("/inventory/adjustments/new/", {
            "product": self.product.id, "location": self.loc.id, "reason": "WRITE_OFF",
            "qty_delta": "-50", "notes": "x"})
        adj = StockAdjustment.objects.get(tenant=self.tenant)
        self.client.post(f"/inventory/adjustments/{adj.id}/reject/")
        adj.refresh_from_db()
        self.assertEqual(adj.status, "REJECTED")
        self.assertEqual(self._bal(), Decimal("100.00"))

    def test_zero_qty_rejected(self):
        resp = self.client.post("/inventory/adjustments/new/", {
            "product": self.product.id, "location": self.loc.id, "reason": "ADJUSTMENT", "qty_delta": "0"})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "cannot be zero")


class LowStockAndLedgerTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, Location, InventoryMovement
        from core.services.inventory import apply_movement
        self.tenant = Tenant.objects.create(name="LS Co")
        Location.objects.filter(tenant=self.tenant).delete()  # WH is the only (auto-selected) site
        self.loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        self.user = User.objects.create_user("lsu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="lsu", password="pw")
        # low: reorder 50, on hand 10
        self.low = Product.objects.create(tenant=self.tenant, sku="SKU-LOW", name="Low", reorder_level=Decimal("50"))
        apply_movement(tenant=self.tenant, product=self.low, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="SEED", ref_id="1", unit_cost=Decimal("1"), user=self.user)
        # ok: reorder 5, on hand 100
        self.ok = Product.objects.create(tenant=self.tenant, sku="SKU-OK", name="Ok", reorder_level=Decimal("5"))
        apply_movement(tenant=self.tenant, product=self.ok, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("100"), ref_type="SEED", ref_id="2", unit_cost=Decimal("1"))

    def test_low_stock_lists_only_below_reorder(self):
        resp = self.client.get("/inventory/low-stock/")
        self.assertEqual(resp.status_code, 200)
        skus = [r["product"].sku for r in resp.context["rows"]]
        self.assertEqual(skus, ["SKU-LOW"])
        self.assertEqual(resp.context["rows"][0]["shortfall"], Decimal("40.00"))

    def test_stock_movements_ledger_and_filter(self):
        resp = self.client.get("/inventory/movements/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context["movements"]), 2)
        # filter by product
        resp = self.client.get(f"/inventory/movements/?product={self.low.id}")
        self.assertEqual([m.product.sku for m in resp.context["movements"]], ["SKU-LOW"])
        # filter by type
        resp = self.client.get("/inventory/movements/?type=RECEIVE")
        self.assertEqual(len(resp.context["movements"]), 2)
        # the ledger shows the user who made the movement
        self.assertContains(self.client.get("/inventory/movements/"), "lsu")


class POCompletenessTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, Supplier, Location, GoodsReceipt, SupplierInvoice, SupplierInvoiceLine
        self.tenant = Tenant.objects.create(name="POC Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Globex", email="g@globex.example", address="Globex House")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        self.prod = Product.objects.create(tenant=self.tenant, sku="SKU-PO1", name="Part")
        self.user = User.objects.create_user("pocu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="pocu", password="pw")

    def test_po_line_vat_and_totals(self):
        from core.models import PurchaseOrder, PurchaseOrderLine
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-1", supplier=self.supplier, delivery_address="Dock 4")
        PurchaseOrderLine.objects.create(po=po, product=self.prod, ordered_qty=Decimal("10"), unit_cost=Decimal("5.00"), tax_code=self.std)
        self.assertEqual(po.subtotal, Decimal("50.00"))
        self.assertEqual(po.tax_total, Decimal("10.00"))
        self.assertEqual(po.total, Decimal("60.00"))

    def test_po_pdf_renders(self):
        from core.models import PurchaseOrder, PurchaseOrderLine
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-2", supplier=self.supplier)
        PurchaseOrderLine.objects.create(po=po, product=self.prod, ordered_qty=Decimal("3"), unit_cost=Decimal("5.00"), tax_code=self.std)
        resp = self.client.get(f"/po/{po.id}/pdf/")
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.content[:5] == b"%PDF-")

    def test_bill_posting_sets_po_billed(self):
        from core.models import PurchaseOrder, GoodsReceipt, SupplierInvoice, SupplierInvoiceLine
        from core.services.gl import post_supplier_invoice
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-3", supplier=self.supplier, status=PurchaseOrder.Status.RECEIVED)
        grn = GoodsReceipt.objects.create(tenant=self.tenant, po=po, grn_number="GRN-3", received_to=self.loc, status=GoodsReceipt.Status.POSTED)
        inv = SupplierInvoice.objects.create(tenant=self.tenant, supplier=self.supplier, po=po, receipt=grn, invoice_number="BILL-3")
        SupplierInvoiceLine.objects.create(invoice=inv, product=self.prod, qty=Decimal("10"), unit_cost=Decimal("5.00"), tax_code=self.std)
        post_supplier_invoice(inv)
        po.refresh_from_db()
        self.assertEqual(po.status, "BILLED")

    def test_po_email_attaches_pdf(self):
        from django.core import mail
        from core.models import PurchaseOrder, PurchaseOrderLine
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-4", supplier=self.supplier, status=PurchaseOrder.Status.APPROVED)
        PurchaseOrderLine.objects.create(po=po, product=self.prod, ordered_qty=Decimal("2"), unit_cost=Decimal("5.00"), tax_code=self.std)
        with self.settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"):
            self.client.post(f"/po/{po.id}/send/")
            self.assertEqual(len(mail.outbox), 1)
            self.assertTrue(mail.outbox[0].attachments)
            self.assertEqual(mail.outbox[0].attachments[0][2], "application/pdf")

    def test_backorders_view(self):
        from core.models import PurchaseOrder, PurchaseOrderLine
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-5", supplier=self.supplier, status=PurchaseOrder.Status.PARTIALLY_RECEIVED)
        PurchaseOrderLine.objects.create(po=po, product=self.prod, ordered_qty=Decimal("10"), received_qty=Decimal("4"), unit_cost=Decimal("5.00"))
        resp = self.client.get("/po/backorders/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context["rows"]), 1)
        self.assertEqual(resp.context["rows"][0]["open_qty"], Decimal("6.00"))


class PurchaseRequisitionTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Req Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Globex", email="g@globex.example")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        self.prod = Product.objects.create(tenant=self.tenant, sku="SKU-R1", name="Part",
                                           standard_cost=Decimal("3.00"), preferred_supplier=self.supplier)
        from core.models import Department
        self.dept = Department.objects.create(tenant=self.tenant, name="Ops")
        self.user = User.objects.create_user("requ", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="requ", password="pw")

    def _make_req(self, status=None):
        from core.models import PurchaseRequisition, PurchaseRequisitionLine
        req = PurchaseRequisition.objects.create(
            tenant=self.tenant, req_number="PR-T1", preferred_supplier=self.supplier,
            status=status or PurchaseRequisition.Status.DRAFT,
        )
        PurchaseRequisitionLine.objects.create(
            requisition=req, product=self.prod, quantity=Decimal("10"), estimated_unit_cost=Decimal("4.00"),
        )
        return req

    def test_create_via_view(self):
        from core.models import PurchaseRequisition
        resp = self.client.post("/requisitions/new/", {
            "department": self.dept.id, "preferred_supplier": self.supplier.id, "needed_by": "",
            "justification": "Restock", "action": "submit",
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "0",
            "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-product": self.prod.id, "lines-0-quantity": "5",
            "lines-0-estimated_unit_cost": "4.00", "lines-0-notes": "",
        })
        self.assertEqual(resp.status_code, 302)
        req = PurchaseRequisition.objects.get(tenant=self.tenant)
        self.assertEqual(req.status, PurchaseRequisition.Status.SUBMITTED)
        self.assertEqual(req.department, self.dept)
        self.assertEqual(req.requested_by, self.user)
        self.assertEqual(req.lines.count(), 1)
        self.assertEqual(req.estimated_total, Decimal("20.00"))

    def test_submit_and_approve(self):
        from core.models import PurchaseRequisition
        req = self._make_req()
        self.assertEqual(self.client.post(f"/requisitions/{req.id}/submit/").status_code, 302)
        req.refresh_from_db()
        self.assertEqual(req.status, PurchaseRequisition.Status.SUBMITTED)
        self.assertEqual(self.client.post(f"/requisitions/{req.id}/approve/").status_code, 302)
        req.refresh_from_db()
        self.assertEqual(req.status, PurchaseRequisition.Status.APPROVED)
        self.assertEqual(req.approved_by, self.user)

    def test_reject(self):
        from core.models import PurchaseRequisition
        req = self._make_req(status=PurchaseRequisition.Status.SUBMITTED)
        self.client.post(f"/requisitions/{req.id}/reject/", {"reason": "Over budget"})
        req.refresh_from_db()
        self.assertEqual(req.status, PurchaseRequisition.Status.REJECTED)
        self.assertEqual(req.rejected_reason, "Over budget")

    def test_convert_to_po_carries_lines(self):
        from core.models import PurchaseRequisition, PurchaseOrder
        req = self._make_req(status=PurchaseRequisition.Status.APPROVED)
        resp = self.client.post(f"/requisitions/{req.id}/convert/")
        self.assertEqual(resp.status_code, 302)
        req.refresh_from_db()
        self.assertEqual(req.status, PurchaseRequisition.Status.CONVERTED)
        self.assertIsNotNone(req.converted_po)
        po = req.converted_po
        self.assertEqual(po.status, PurchaseOrder.Status.DRAFT)
        self.assertEqual(po.supplier, self.supplier)
        line = po.lines.get()
        self.assertEqual(line.product, self.prod)
        self.assertEqual(line.ordered_qty, Decimal("10.00"))
        self.assertEqual(line.unit_cost, Decimal("4.00"))

    def test_convert_requires_approved(self):
        from core.models import PurchaseRequisition
        req = self._make_req(status=PurchaseRequisition.Status.DRAFT)
        self.client.post(f"/requisitions/{req.id}/convert/")
        req.refresh_from_db()
        self.assertEqual(req.status, PurchaseRequisition.Status.DRAFT)
        self.assertIsNone(req.converted_po)

    def test_detail_and_list_render(self):
        from core.models import PurchaseRequisition
        req = self._make_req(status=PurchaseRequisition.Status.APPROVED)
        self.assertEqual(self.client.get("/requisitions/").status_code, 200)
        self.assertEqual(self.client.get(f"/requisitions/{req.id}/").status_code, 200)


class InventoryAnalyticsTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        from core.services.inventory import apply_movement
        from django.utils import timezone
        self.tenant = Tenant.objects.create(name="IA Co")
        Location.objects.filter(tenant=self.tenant).delete()  # drop auto-seeded Main Location; this class manages its own
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.loc1 = Location.objects.create(tenant=self.tenant, name="WH1", type=Location.Type.WAREHOUSE)
        self.loc2 = Location.objects.create(tenant=self.tenant, name="WH2", type=Location.Type.WAREHOUSE)
        self.prod = Product.objects.create(tenant=self.tenant, sku="IA1", name="P")
        self.today = timezone.localdate()
        apply_movement(tenant=self.tenant, product=self.prod, location=self.loc1,
                       movement_type="RECEIVE", qty_delta=Decimal("100"), ref_type="S", ref_id="1", unit_cost=Decimal("5.00"))
        apply_movement(tenant=self.tenant, product=self.prod, location=self.loc2,
                       movement_type="RECEIVE", qty_delta=Decimal("20"), ref_type="S", ref_id="2", unit_cost=Decimal("5.00"))
        self.user = User.objects.create_user("iau", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="iau", password="pw")

    def test_valuation_by_location_and_turnover(self):
        from core.services import reports
        from core.services.gl import post_customer_invoice
        from datetime import timedelta
        cust = Customer.objects.create(tenant=self.tenant, name="C")
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=cust, invoice_number="INV-IA")
        CustomerInvoiceLine.objects.create(invoice=inv, product=self.prod, qty=Decimal("10"),
                                           unit_price=Decimal("9"), tax_code=self.std)
        post_customer_invoice(inv)  # COGS 50, stock now 110 @ 5
        data = reports.inventory_analytics(self.tenant, self.today - timedelta(days=1),
                                           self.today + timedelta(days=1))
        self.assertEqual(data["current_value"], Decimal("550.00"))
        self.assertEqual(len(data["by_location"]), 2)
        self.assertEqual(data["cogs"], Decimal("50.00"))
        self.assertIsNotNone(data["turnover"])
        self.assertIsNotNone(data["days_inventory"])

    def test_page_renders(self):
        resp = self.client.get("/reports/inventory-analytics/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Inventory Analytics")


class SupplierScorecardTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        from django.utils import timezone
        self.tenant = Tenant.objects.create(name="Score Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Globex")
        self.prod = Product.objects.create(tenant=self.tenant, sku="SC1", name="P")
        self.today = timezone.localdate()
        self.user = User.objects.create_user("scu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="scu", password="pw")

    def test_spend_otd_and_variance(self):
        from core.models import (PurchaseOrder, PurchaseOrderLine, GoodsReceipt,
                                 SupplierInvoice, SupplierInvoiceLine)
        from core.services.gl import post_supplier_invoice
        from core.services import purchasing
        from datetime import timedelta
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-SC", supplier=self.supplier,
                                          status=PurchaseOrder.Status.RECEIVED, expected_date=self.today)
        pol = PurchaseOrderLine.objects.create(po=po, product=self.prod, ordered_qty=Decimal("10"),
                                               unit_cost=Decimal("10.00"), tax_code=self.std)
        grn = GoodsReceipt.objects.create(tenant=self.tenant, po=po, grn_number="GRN-SC", received_to=self.loc,
                                          status=GoodsReceipt.Status.POSTED)
        inv = SupplierInvoice.objects.create(tenant=self.tenant, supplier=self.supplier, po=po,
                                             receipt=grn, invoice_number="BILL-SC", invoice_date=self.today)
        SupplierInvoiceLine.objects.create(invoice=inv, product=self.prod, po_line=pol,
                                           qty=Decimal("10"), unit_cost=Decimal("11.00"), tax_code=self.std)
        post_supplier_invoice(inv)

        data = purchasing.supplier_scorecard(self.tenant, self.today - timedelta(days=1),
                                             self.today + timedelta(days=1))
        r = data["rows"][0]
        self.assertEqual(r["bills"], 1)
        self.assertEqual(r["receipts"], 1)
        self.assertEqual(r["on_time"], 1)
        self.assertEqual(r["otd_pct"], Decimal("100.0"))
        self.assertEqual(r["price_variance"], Decimal("10.00"))  # 10 units x (11-10)
        self.assertEqual(r["spend"], inv.total)

    def test_page_renders(self):
        resp = self.client.get("/reports/supplier-scorecard/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Supplier Scorecard")


class ProfitabilityReportTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        from core.services.inventory import apply_movement
        self.tenant = Tenant.objects.create(name="Profit Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        self.prod = Product.objects.create(tenant=self.tenant, sku="PF1", name="Widget")
        apply_movement(tenant=self.tenant, product=self.prod, location=self.loc,
                       movement_type="RECEIVE", qty_delta=Decimal("100"), ref_type="SEED", ref_id="1",
                       unit_cost=Decimal("6.00"))
        self.cust = Customer.objects.create(tenant=self.tenant, name="Acme")
        self.user = User.objects.create_user("pfu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="pfu", password="pw")

    def _sell(self, qty, price):
        from core.services.gl import post_customer_invoice
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.cust,
                                             invoice_number=f"INV-PF-{qty}-{price}")
        CustomerInvoiceLine.objects.create(invoice=inv, product=self.prod, qty=Decimal(qty),
                                           unit_price=Decimal(price), tax_code=self.std)
        post_customer_invoice(inv)  # deducts stock + posts COGS
        return inv

    def test_margin_by_product_and_customer(self):
        from core.services import sales_reports
        from django.utils import timezone
        from datetime import timedelta
        self._sell("10", "10.00")  # revenue 100, COGS 60, margin 40
        d_from = timezone.localdate() - timedelta(days=1)
        d_to = timezone.localdate() + timedelta(days=1)
        data = sales_reports.profitability(self.tenant, d_from, d_to)
        prow = data["by_product"][0]
        self.assertEqual(prow["revenue"], Decimal("100.00"))
        self.assertEqual(prow["cogs"], Decimal("60.00"))
        self.assertEqual(prow["margin"], Decimal("40.00"))
        self.assertEqual(prow["margin_pct"], Decimal("40.0"))
        crow = data["by_customer"][0]
        self.assertEqual(crow["margin"], Decimal("40.00"))
        self.assertEqual(data["totals"]["margin"], Decimal("40.00"))

    def test_report_page_renders(self):
        self._sell("5", "12.00")
        resp = self.client.get("/sales/reports/profitability/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Profitability")
        self.assertContains(resp, "Widget")


class DunningReminderTests(TestCase):
    def setUp(self):
        from datetime import timedelta
        from django.utils import timezone
        self.tenant = Tenant.objects.create(name="Dun Co", dunning_enabled=True, dunning_interval_days=7)
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.today = timezone.localdate()
        self.cust = Customer.objects.create(tenant=self.tenant, name="C", email="c@example.com")
        self.cust_noemail = Customer.objects.create(tenant=self.tenant, name="NoMail")
        self.due = self.today - timedelta(days=10)

    def _invoice(self, customer, number, due=None):
        from core.services.gl import post_customer_invoice
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=customer, invoice_number=number,
                                             due_date=(due or self.due))
        CustomerInvoiceLine.objects.create(invoice=inv, description="X", qty=Decimal("1"),
                                           unit_price=Decimal("100"), tax_code=self.std)
        post_customer_invoice(inv)  # -> ISSUED
        return inv

    def test_overdue_reminder_sent_once_per_interval(self):
        from django.core import mail
        from core.services import housekeeping
        inv = self._invoice(self.cust, "INV-D1")
        with self.settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"):
            sent = housekeeping.send_overdue_reminders(self.tenant, today=self.today)
            self.assertEqual(sent, 1)
            self.assertEqual(len(mail.outbox), 1)
            self.assertIn("INV-D1", mail.outbox[0].subject)
            inv.refresh_from_db()
            self.assertEqual(inv.reminder_count, 1)
            self.assertEqual(inv.last_reminder_at, self.today)
            # Same day re-run: throttled, nothing more.
            self.assertEqual(housekeeping.send_overdue_reminders(self.tenant, today=self.today), 0)

    def test_no_email_no_reminder(self):
        from core.services import housekeeping
        self._invoice(self.cust_noemail, "INV-D2")
        with self.settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"):
            self.assertEqual(housekeeping.send_overdue_reminders(self.tenant, today=self.today), 0)

    def test_not_overdue_skipped(self):
        from datetime import timedelta
        from core.services import housekeeping
        self._invoice(self.cust, "INV-D3", due=self.today + timedelta(days=5))
        with self.settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"):
            self.assertEqual(housekeeping.send_overdue_reminders(self.tenant, today=self.today), 0)

    def test_disabled_tenant_sends_nothing(self):
        from core.services import housekeeping
        self.tenant.dunning_enabled = False
        self.tenant.save()
        self._invoice(self.cust, "INV-D4")
        with self.settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"):
            self.assertEqual(housekeeping.send_overdue_reminders(self.tenant, today=self.today), 0)

    def test_paid_invoice_not_reminded(self):
        from core.services import housekeeping
        from core.services.gl import post_payment
        from core.models import Payment, PaymentAllocation
        inv = self._invoice(self.cust, "INV-D5")
        pay = Payment.objects.create(tenant=self.tenant, customer=self.cust,
                                     direction=Payment.Direction.RECEIPT, amount=Decimal("120"),
                                     payment_date=self.today)
        PaymentAllocation.objects.create(payment=pay, customer_invoice=inv, amount=Decimal("120"))
        post_payment(pay)
        with self.settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"):
            self.assertEqual(housekeeping.send_overdue_reminders(self.tenant, today=self.today), 0)


class ReturnToSupplierTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, GLAccount
        from core.services.inventory import apply_movement
        self.tenant = Tenant.objects.create(name="Ret Co")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Globex")
        self.prod = Product.objects.create(tenant=self.tenant, sku="RT1", name="P")
        apply_movement(tenant=self.tenant, product=self.prod, location=self.loc,
                       movement_type="RECEIVE", qty_delta=Decimal("100"), ref_type="SEED", ref_id="1",
                       unit_cost=Decimal("4.00"))
        self.ap = GLAccount.objects.get(tenant=self.tenant, code="2000")
        self.inv = GLAccount.objects.get(tenant=self.tenant, code="1000")
        self.adj5200 = GLAccount.objects.get(tenant=self.tenant, code="5200")
        self.user = User.objects.create_user("rtu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="rtu", password="pw")

    def _bal(self, acc):
        from core.models import JournalLine
        from django.db.models import Sum
        agg = JournalLine.objects.filter(account=acc).aggregate(d=Sum("debit"), c=Sum("credit"))
        return (agg["d"] or Decimal("0.00")) - (agg["c"] or Decimal("0.00"))

    def test_return_creates_purchase_credit_note(self):
        from core.models import StockAdjustment, CreditNote
        resp = self.client.post("/inventory/adjustments/new/", {
            "product": self.prod.id, "location": self.loc.id, "reason": "RETURN_SUPPLIER",
            "supplier": self.supplier.id, "qty_delta": "-5", "notes": "faulty batch",
        })
        self.assertEqual(resp.status_code, 302)
        adj = StockAdjustment.objects.get(tenant=self.tenant)
        self.assertEqual(adj.status, StockAdjustment.Status.POSTED)
        self.assertIsNotNone(adj.credit_note)
        cn = adj.credit_note
        self.assertEqual(cn.kind, CreditNote.Kind.PURCHASE)
        self.assertEqual(cn.status, CreditNote.Status.POSTED)
        # 5 units @ GBP4 = 20: DR AP (reduces payable) / CR Inventory; no shrinkage.
        self.assertEqual(self._bal(self.ap), Decimal("20.00"))   # debit reduces the credit-normal AP
        self.assertEqual(self._bal(self.inv), Decimal("-20.00"))  # credit reduces inventory
        self.assertEqual(self._bal(self.adj5200), Decimal("0.00"))

    def test_return_requires_supplier(self):
        from core.models import StockAdjustment
        resp = self.client.post("/inventory/adjustments/new/", {
            "product": self.prod.id, "location": self.loc.id, "reason": "RETURN_SUPPLIER",
            "qty_delta": "-5",
        })
        self.assertEqual(resp.status_code, 200)  # re-render with error
        self.assertContains(resp, "Choose the supplier")
        self.assertEqual(StockAdjustment.objects.count(), 0)

    def test_return_must_be_negative(self):
        resp = self.client.post("/inventory/adjustments/new/", {
            "product": self.prod.id, "location": self.loc.id, "reason": "RETURN_SUPPLIER",
            "supplier": self.supplier.id, "qty_delta": "5",
        })
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "must remove stock")


class CustomerOrderReservationTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, InventoryBalance, CustomerOrder, CustomerOrderLine, Customer
        self.tenant = Tenant.objects.create(name="Resv Co")
        Location.objects.filter(tenant=self.tenant).delete()  # drop auto-seeded Main Location; this class manages its own
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        self.prod = Product.objects.create(tenant=self.tenant, sku="RV1", name="P")
        InventoryBalance.objects.create(tenant=self.tenant, product=self.prod, location=self.loc,
                                        on_hand=Decimal("50"), reserved=Decimal("0"))
        self.cust = Customer.objects.create(tenant=self.tenant, name="C")
        self.order = CustomerOrder.objects.create(tenant=self.tenant, customer=self.cust, order_number="SO-RV1",
                                                  status=CustomerOrder.Status.DRAFT)
        CustomerOrderLine.objects.create(order=self.order, product=self.prod, qty=Decimal("8"),
                                         unit_price=Decimal("10"), tax_code=self.std)
        self.user = User.objects.create_user("rvu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="rvu", password="pw")

    def _reserved(self):
        from core.models import InventoryBalance
        return InventoryBalance.objects.get(tenant=self.tenant, product=self.prod, location=self.loc).reserved

    def test_confirm_reserves_stock(self):
        from core.models import InventoryReservation
        self.client.post(f"/customer-orders/{self.order.id}/status/confirm/")
        self.assertEqual(self._reserved(), Decimal("8.00"))
        self.assertEqual(InventoryReservation.objects.filter(
            tenant=self.tenant, ref_type="CUSTOMER_ORDER", ref_id=str(self.order.id),
            status=InventoryReservation.Status.ACTIVE).count(), 1)

    def test_cancel_releases_stock(self):
        self.client.post(f"/customer-orders/{self.order.id}/status/confirm/")
        self.assertEqual(self._reserved(), Decimal("8.00"))
        self.client.post(f"/customer-orders/{self.order.id}/status/cancel/")
        self.assertEqual(self._reserved(), Decimal("0.00"))

    def test_invoice_releases_reservation(self):
        self.client.post(f"/customer-orders/{self.order.id}/status/confirm/")
        self.assertEqual(self._reserved(), Decimal("8.00"))
        self.client.post(f"/customer-orders/{self.order.id}/to-invoice/")
        self.assertEqual(self._reserved(), Decimal("0.00"))

    def test_edit_resyncs_reservation(self):
        self.client.post(f"/customer-orders/{self.order.id}/status/confirm/")
        line = self.order.lines.get()
        resp = self.client.post(f"/customer-orders/{self.order.id}/edit/", {
            "customer": self.cust.id, "order_number": "SO-RV1", "order_date": "2026-01-01",
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "1",
            "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-id": line.id, "lines-0-product": self.prod.id, "lines-0-qty": "20",
            "lines-0-unit_price": "10", "lines-0-tax_code": self.std.id,
        })
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self._reserved(), Decimal("20.00"))


class SupplierPriceHistoryTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Price Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Globex")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        self.prod = Product.objects.create(tenant=self.tenant, sku="PH1", name="Part")
        self.user = User.objects.create_user("phu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="phu", password="pw")

    def test_record_is_idempotent(self):
        from core.services.purchasing import record_supplier_price
        from core.models import SupplierPriceHistory
        for _ in range(2):
            record_supplier_price(tenant=self.tenant, supplier=self.supplier, product=self.prod,
                                  unit_cost=Decimal("5.00"), source="PO", reference="PO-X")
        self.assertEqual(SupplierPriceHistory.objects.filter(tenant=self.tenant).count(), 1)

    def test_zero_cost_skipped(self):
        from core.services.purchasing import record_supplier_price
        from core.models import SupplierPriceHistory
        record_supplier_price(tenant=self.tenant, supplier=self.supplier, product=self.prod,
                              unit_cost=Decimal("0.00"), source="PO", reference="PO-Z")
        self.assertEqual(SupplierPriceHistory.objects.count(), 0)

    def test_po_submit_records_price(self):
        from core.models import PurchaseOrder, PurchaseOrderLine, SupplierPriceHistory
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-PH", supplier=self.supplier,
                                          status=PurchaseOrder.Status.DRAFT)
        PurchaseOrderLine.objects.create(po=po, product=self.prod, ordered_qty=Decimal("10"),
                                         unit_cost=Decimal("7.50"), tax_code=self.std)
        self.client.post(f"/po/{po.id}/submit/")
        rec = SupplierPriceHistory.objects.get(tenant=self.tenant, supplier=self.supplier, product=self.prod, source="PO")
        self.assertEqual(rec.unit_cost, Decimal("7.50"))

    def test_bill_posting_records_actual_price(self):
        from core.models import (PurchaseOrder, GoodsReceipt, SupplierInvoice, SupplierInvoiceLine,
                                 SupplierPriceHistory)
        from core.services.gl import post_supplier_invoice
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-B", supplier=self.supplier,
                                          status=PurchaseOrder.Status.RECEIVED)
        grn = GoodsReceipt.objects.create(tenant=self.tenant, po=po, grn_number="GRN-B", received_to=self.loc,
                                          status=GoodsReceipt.Status.POSTED)
        inv = SupplierInvoice.objects.create(tenant=self.tenant, supplier=self.supplier, po=po, receipt=grn,
                                             invoice_number="BILL-B")
        SupplierInvoiceLine.objects.create(invoice=inv, product=self.prod, qty=Decimal("10"),
                                           unit_cost=Decimal("8.25"), tax_code=self.std)
        post_supplier_invoice(inv)
        rec = SupplierPriceHistory.objects.get(tenant=self.tenant, supplier=self.supplier, product=self.prod, source="BILL")
        self.assertEqual(rec.unit_cost, Decimal("8.25"))

    def test_last_prices_and_json_endpoint(self):
        from core.services.purchasing import record_supplier_price, last_prices_for_supplier
        from django.utils import timezone
        from datetime import timedelta
        record_supplier_price(tenant=self.tenant, supplier=self.supplier, product=self.prod,
                              unit_cost=Decimal("5.00"), source="PO", reference="PO-1",
                              recorded_at=timezone.localdate() - timedelta(days=10))
        record_supplier_price(tenant=self.tenant, supplier=self.supplier, product=self.prod,
                              unit_cost=Decimal("6.00"), source="BILL", reference="BILL-1",
                              recorded_at=timezone.localdate())
        last = last_prices_for_supplier(self.tenant, self.supplier)
        self.assertEqual(last[self.prod.id], Decimal("6.00"))  # most recent
        resp = self.client.get(f"/po/supplier/{self.supplier.id}/prices/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["prices"][str(self.prod.id)], "6.00")


class StockAdjustmentGLTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, GLAccount
        from core.services.inventory import apply_movement
        self.tenant = Tenant.objects.create(name="Adj Co")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        self.prod = Product.objects.create(tenant=self.tenant, sku="ADJ1", name="P")
        # Seed 100 units @ £2 so the product has an average cost.
        apply_movement(tenant=self.tenant, product=self.prod, location=self.loc,
                       movement_type="RECEIVE", qty_delta=Decimal("100"), ref_type="SEED", ref_id="1",
                       unit_cost=Decimal("2.00"))
        self.inv_acc = GLAccount.objects.get(tenant=self.tenant, code="1000")
        self.adj_acc = GLAccount.objects.get(tenant=self.tenant, code="5200")
        self.user = User.objects.create_user("adju", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="adju", password="pw")

    def _acc_balance(self, acc):
        from core.models import JournalLine
        from django.db.models import Sum
        agg = JournalLine.objects.filter(account=acc).aggregate(d=Sum("debit"), c=Sum("credit"))
        return (agg["d"] or Decimal("0.00")) - (agg["c"] or Decimal("0.00"))

    def test_write_off_posts_loss_to_gl(self):
        from core.models import StockAdjustment, JournalEntry
        resp = self.client.post("/inventory/adjustments/new/", {
            "product": self.prod.id, "location": self.loc.id, "reason": "WRITE_OFF",
            "qty_delta": "-10", "notes": "broken",
        })
        self.assertEqual(resp.status_code, 302)
        adj = StockAdjustment.objects.get(tenant=self.tenant)
        self.assertEqual(adj.status, StockAdjustment.Status.POSTED)
        JournalEntry.objects.get(tenant=self.tenant, ref_type="STOCK_ADJ", ref_id=str(adj.id))
        # 10 units @ £2 = £20 loss: DR 5200 / CR 1000 (only GL postings hit acct 1000)
        self.assertEqual(self._acc_balance(self.adj_acc), Decimal("20.00"))
        self.assertEqual(self._acc_balance(self.inv_acc), Decimal("-20.00"))

    def test_found_stock_posts_gain_to_gl(self):
        from core.models import StockAdjustment
        self.client.post("/inventory/adjustments/new/", {
            "product": self.prod.id, "location": self.loc.id, "reason": "ADJUSTMENT",
            "qty_delta": "5", "notes": "found",
        })
        adj = StockAdjustment.objects.get(tenant=self.tenant)
        self.assertEqual(adj.status, StockAdjustment.Status.POSTED)
        # 5 units @ £2 = £10 gain: DR 1000 / CR 5200
        self.assertEqual(self._acc_balance(self.adj_acc), Decimal("-10.00"))
        self.assertEqual(self._acc_balance(self.inv_acc), Decimal("10.00"))

    def test_je_balances(self):
        from core.models import StockAdjustment, JournalEntry
        from django.db.models import Sum
        self.client.post("/inventory/adjustments/new/", {
            "product": self.prod.id, "location": self.loc.id, "reason": "DAMAGE",
            "qty_delta": "-3", "notes": "",
        })
        adj = StockAdjustment.objects.get(tenant=self.tenant)
        je = JournalEntry.objects.get(tenant=self.tenant, ref_type="STOCK_ADJ", ref_id=str(adj.id))
        agg = je.lines.aggregate(d=Sum("debit"), c=Sum("credit"))
        self.assertEqual(agg["d"], agg["c"])


class CreditLimitEnforcementTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Credit Co")
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.user = User.objects.create_user("credu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="credu", password="pw")

    def _invoice(self, customer, net):
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=customer,
                                             invoice_number=f"INV-{customer.id}-{net}")
        CustomerInvoiceLine.objects.create(invoice=inv, description="X", qty=Decimal("1"),
                                           unit_price=Decimal(net), tax_code=self.std)
        return inv

    def test_credit_status_method(self):
        c = Customer.objects.create(tenant=self.tenant, name="A", credit_limit=Decimal("500.00"))
        ok, _ = c.credit_status(Decimal("100.00"))
        self.assertTrue(ok)
        ok, reason = c.credit_status(Decimal("600.00"))
        self.assertFalse(ok)
        self.assertIn("Credit limit exceeded", reason)

    def test_no_limit_allows_anything(self):
        c = Customer.objects.create(tenant=self.tenant, name="B", credit_limit=Decimal("0.00"))
        ok, _ = c.credit_status(Decimal("999999.00"))
        self.assertTrue(ok)

    def test_on_hold_blocks(self):
        c = Customer.objects.create(tenant=self.tenant, name="C", status=Customer.Status.ON_HOLD)
        ok, reason = c.credit_status(Decimal("1.00"))
        self.assertFalse(ok)
        self.assertIn("on hold", reason)

    def test_issue_blocked_when_over_limit(self):
        c = Customer.objects.create(tenant=self.tenant, name="D", credit_limit=Decimal("100.00"))
        inv = self._invoice(c, "200")  # 240 gross > 100 limit
        resp = self.client.post(f"/ar/invoices/{inv.id}/issue/")
        self.assertEqual(resp.status_code, 302)
        inv.refresh_from_db()
        self.assertEqual(inv.status, CustomerInvoice.Status.DRAFT)  # not issued

    def test_issue_succeeds_when_within_limit(self):
        c = Customer.objects.create(tenant=self.tenant, name="E", credit_limit=Decimal("1000.00"))
        inv = self._invoice(c, "200")  # 240 gross < 1000
        resp = self.client.post(f"/ar/invoices/{inv.id}/issue/")
        self.assertEqual(resp.status_code, 302)
        inv.refresh_from_db()
        self.assertEqual(inv.status, CustomerInvoice.Status.ISSUED)


class LowStockReorderTests(TestCase):
    def setUp(self):
        from core.models import OrgMembership, InventoryBalance
        self.tenant = Tenant.objects.create(name="Reorder Co")
        self.loc = Location.objects.create(tenant=self.tenant, name="WH", type=Location.Type.WAREHOUSE)
        self.supA = Supplier.objects.create(tenant=self.tenant, name="Supplier A")
        self.supB = Supplier.objects.create(tenant=self.tenant, name="Supplier B")
        # Two products below reorder for supplier A, one for B, one unassigned.
        self.p1 = Product.objects.create(tenant=self.tenant, sku="R1", name="P1", reorder_level=Decimal("20"),
                                         preferred_supplier=self.supA, standard_cost=Decimal("2.00"))
        self.p2 = Product.objects.create(tenant=self.tenant, sku="R2", name="P2", reorder_level=Decimal("30"),
                                         preferred_supplier=self.supA, standard_cost=Decimal("5.00"))
        self.p3 = Product.objects.create(tenant=self.tenant, sku="R3", name="P3", reorder_level=Decimal("10"),
                                         preferred_supplier=self.supB)
        self.p4 = Product.objects.create(tenant=self.tenant, sku="R4", name="P4", reorder_level=Decimal("10"))
        for p, oh in [(self.p1, "5"), (self.p2, "0"), (self.p3, "2"), (self.p4, "1")]:
            InventoryBalance.objects.create(tenant=self.tenant, product=p, location=self.loc, on_hand=Decimal(oh))
        self.user = User.objects.create_user("ro", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="ro", password="pw")

    def test_low_stock_lists_shortfalls(self):
        resp = self.client.get("/inventory/low-stock/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context["rows"]), 4)

    def test_reorder_groups_by_supplier(self):
        from core.models import PurchaseRequisition
        resp = self.client.post("/inventory/low-stock/reorder/", {
            "select": [str(self.p1.id), str(self.p2.id), str(self.p3.id), str(self.p4.id)],
            f"qty_{self.p1.id}": "15", f"qty_{self.p2.id}": "30",
            f"qty_{self.p3.id}": "8", f"qty_{self.p4.id}": "9",
        })
        self.assertEqual(resp.status_code, 302)
        # 3 groups: supA (2 lines), supB (1), unassigned (1)
        reqs = PurchaseRequisition.objects.filter(tenant=self.tenant)
        self.assertEqual(reqs.count(), 3)
        a = reqs.get(preferred_supplier=self.supA)
        self.assertEqual(a.lines.count(), 2)
        self.assertEqual(a.status, PurchaseRequisition.Status.DRAFT)
        line1 = a.lines.get(product=self.p1)
        self.assertEqual(line1.quantity, Decimal("15.00"))
        self.assertEqual(line1.estimated_unit_cost, Decimal("2.00"))
        self.assertEqual(reqs.get(preferred_supplier=self.supB).lines.count(), 1)
        self.assertEqual(reqs.filter(preferred_supplier__isnull=True).count(), 1)

    def test_reorder_skips_zero_qty(self):
        from core.models import PurchaseRequisition
        self.client.post("/inventory/low-stock/reorder/", {
            "select": [str(self.p1.id)], f"qty_{self.p1.id}": "0",
        })
        self.assertEqual(PurchaseRequisition.objects.filter(tenant=self.tenant).count(), 0)

    def test_reorder_requires_selection(self):
        from core.models import PurchaseRequisition
        self.client.post("/inventory/low-stock/reorder/", {})
        self.assertEqual(PurchaseRequisition.objects.filter(tenant=self.tenant).count(), 0)


class DefaultLocationTests(TestCase):
    """A fresh organisation is provisioned with a single 'Main Location'."""

    def test_new_tenant_gets_main_location(self):
        from core.models import Location
        t = Tenant.objects.create(name="Fresh Co")
        locs = Location.objects.filter(tenant=t)
        self.assertEqual(locs.count(), 1)
        loc = locs.get()
        self.assertEqual(loc.name, "Main Location")
        self.assertEqual(loc.type, Location.Type.WAREHOUSE)
        self.assertTrue(loc.holds_stock)
        self.assertTrue(loc.is_active)

    def test_seed_is_idempotent_on_resave(self):
        from core.models import Location
        t = Tenant.objects.create(name="Resave Co")
        t.name = "Resave Co Ltd"
        t.save()
        self.assertEqual(Location.objects.filter(tenant=t).count(), 1)

    def test_does_not_seed_when_locations_exist(self):
        # Simulate a tenant that already had a location before the save signal
        # re-fires: no duplicate Main Location is added.
        from core.models import Location
        t = Tenant.objects.create(name="Has Loc Co")
        Location.objects.filter(tenant=t).delete()
        Location.objects.create(tenant=t, name="Depot", type=Location.Type.WAREHOUSE)
        t.save()
        names = set(Location.objects.filter(tenant=t).values_list("name", flat=True))
        self.assertEqual(names, {"Depot"})

    def test_new_organisation_view_provisions_location(self):
        from core.models import Location, OrgMembership
        u = User.objects.create_user("orgcreator", password="pw")
        self.client.login(username="orgcreator", password="pw")
        resp = self.client.post("/onboarding/new-organisation/", {
            "name": "View Co", "currency_code": "GBP", "country": "United Kingdom",
        })
        self.assertEqual(resp.status_code, 302)
        t = Tenant.objects.get(name="View Co")
        self.assertEqual(Location.objects.filter(tenant=t, name="Main Location").count(), 1)


class LocationAccessEnforcementTests(TestCase):
    """A location-restricted user only sees/acts on their granted locations in
    stock adjustments, transfers and goods receipts."""

    def setUp(self):
        from core.models import (OrgMembership, UserLocationAccess, StockAdjustment,
                                 InventoryTransfer, InventoryBalance, Site)
        self.tenant = Tenant.objects.create(name="Acc Co")
        Location.objects.filter(tenant=self.tenant).delete()
        # locA under the default Main Site; locB under a second site.
        self.siteB = Site.objects.create(tenant=self.tenant, name="Site B", site_type=Site.Type.CITY_BRANCH)
        self.locA = Location.objects.create(tenant=self.tenant, name="WH-A", type=Location.Type.WAREHOUSE)
        self.locB = Location.objects.create(tenant=self.tenant, site=self.siteB, name="WH-B", type=Location.Type.WAREHOUSE)
        self.prod = Product.objects.create(tenant=self.tenant, sku="ACC1", name="P")
        # Warehouse user restricted to WH-A only.
        self.user = User.objects.create_user("whA", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="WAREHOUSE", is_default=True)
        UserLocationAccess.objects.create(tenant=self.tenant, user=self.user, location=self.locA)
        self.client.login(username="whA", password="pw")
        # One adjustment at each location.
        self.adjA = StockAdjustment.objects.create(tenant=self.tenant, product=self.prod, location=self.locA,
                                                   qty_delta=Decimal("1"), requested_by=self.user)
        self.adjB = StockAdjustment.objects.create(tenant=self.tenant, product=self.prod, location=self.locB,
                                                   qty_delta=Decimal("1"), requested_by=self.user)
        # A transfer entirely within WH-B (not accessible to whA).
        self.trB = InventoryTransfer.objects.create(tenant=self.tenant, transfer_number="TR-B",
                                                    from_location=self.locB, to_location=self.locB)

    def test_adjustment_list_filtered_to_accessible(self):
        resp = self.client.get("/inventory/adjustments/")
        ids = [a.id for a in resp.context["adjustments"]]
        self.assertIn(self.adjA.id, ids)
        self.assertNotIn(self.adjB.id, ids)

    def test_transfer_list_filtered_to_accessible(self):
        from core.models import InventoryTransfer
        trA = InventoryTransfer.objects.create(tenant=self.tenant, transfer_number="TR-A",
                                               from_location=self.locA, to_location=self.locB)
        resp = self.client.get("/transfers/")
        ids = [t.id for t in resp.context["transfers"]]
        self.assertIn(trA.id, ids)        # touches WH-A
        self.assertNotIn(self.trB.id, ids)  # entirely in WH-B

    def test_transfer_post_blocked_for_inaccessible_locations(self):
        resp = self.client.post(f"/transfers/{self.trB.id}/post/")
        self.assertEqual(resp.status_code, 403)

    def test_admin_also_scoped_to_selected_site(self):
        # Even an admin sees only the selected site - there is no combined view.
        from core.models import OrgMembership
        admin = User.objects.create_user("accadmin", password="pw")
        OrgMembership.objects.create(user=admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="accadmin", password="pw")  # auto-selects Main Site (first)
        ids = [a.id for a in self.client.get("/inventory/adjustments/").context["adjustments"]]
        self.assertIn(self.adjA.id, ids)
        self.assertNotIn(self.adjB.id, ids)  # adjB is at a location under another site
        # Switching site reveals the other site's data (and hides the first).
        self.client.post("/switch-site/", {"site": self.siteB.id})
        ids = [a.id for a in self.client.get("/inventory/adjustments/").context["adjustments"]]
        self.assertIn(self.adjB.id, ids)
        self.assertNotIn(self.adjA.id, ids)

    def test_can_access_location_helper(self):
        from core.views import _can_access_location

        class _Req:
            user = self.user
        self.assertTrue(_can_access_location(_Req(), self.tenant, self.locA.id))
        self.assertFalse(_can_access_location(_Req(), self.tenant, self.locB.id))
        self.assertFalse(_can_access_location(_Req(), self.tenant, self.locA.id, self.locB.id))
        self.assertTrue(_can_access_location(_Req(), self.tenant, None))  # None ignored


class DepartmentTests(TestCase):
    """Department/Team tier: model, CRUD pages, membership link, manager scoping."""

    def setUp(self):
        from core.models import OrgMembership
        self.tenant = Tenant.objects.create(name="Dept Co")
        self.other = Tenant.objects.create(name="Other Co")
        self.admin = User.objects.create_user("deptadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.member = User.objects.create_user("deptmember", password="pw")
        OrgMembership.objects.create(user=self.member, tenant=self.tenant, role="WAREHOUSE")
        # A user in another org only - must not appear as a manager option here.
        self.outsider = User.objects.create_user("outsider", password="pw")
        OrgMembership.objects.create(user=self.outsider, tenant=self.other, role="ADMIN")
        self.client.login(username="deptadmin", password="pw")

    def test_create_department_via_view(self):
        from core.models import Department
        resp = self.client.post("/departments/new/", {
            "name": "Warehouse Ops", "code": "WHO", "manager": self.member.id, "is_active": "on",
        })
        self.assertEqual(resp.status_code, 302)
        d = Department.objects.get(tenant=self.tenant, name="Warehouse Ops")
        self.assertEqual(d.code, "WHO")
        self.assertEqual(d.manager, self.member)
        self.assertTrue(d.is_active)

    def test_manager_choices_limited_to_org_members(self):
        from core.forms import DepartmentForm
        from core.current import set_current_tenant, clear_current_tenant
        set_current_tenant(self.tenant)
        try:
            ids = set(DepartmentForm().fields["manager"].queryset.values_list("id", flat=True))
        finally:
            clear_current_tenant()
        self.assertIn(self.admin.id, ids)
        self.assertIn(self.member.id, ids)
        self.assertNotIn(self.outsider.id, ids)

    def test_membership_department_link(self):
        from core.models import Department, OrgMembership
        d = Department.objects.create(tenant=self.tenant, name="Sales")
        m = OrgMembership.objects.get(user=self.member, tenant=self.tenant)
        m.department = d
        m.save()
        self.assertEqual(d.members.count(), 1)
        self.assertEqual(d.members.get(), m)

    def test_delete_department_unassigns_members(self):
        from core.models import Department, OrgMembership
        d = Department.objects.create(tenant=self.tenant, name="Temp")
        m = OrgMembership.objects.get(user=self.member, tenant=self.tenant)
        m.department = d
        m.save()
        resp = self.client.post(f"/departments/{d.id}/delete/")
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Department.objects.filter(id=d.id).exists())
        m.refresh_from_db()
        self.assertIsNone(m.department_id)  # SET_NULL

    def test_duplicate_name_rejected(self):
        from core.models import Department
        Department.objects.create(tenant=self.tenant, name="Finance")
        resp = self.client.post("/departments/new/", {"name": "Finance", "is_active": "on"})
        self.assertEqual(resp.status_code, 200)  # re-rendered with error
        self.assertEqual(Department.objects.filter(tenant=self.tenant, name="Finance").count(), 1)

    def test_list_page_renders(self):
        from core.models import Department
        Department.objects.create(tenant=self.tenant, name="Ops")
        resp = self.client.get("/departments/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Ops")

    def test_non_admin_cannot_create(self):
        self.client.login(username="deptmember", password="pw")
        resp = self.client.post("/departments/new/", {"name": "Sneaky", "is_active": "on"})
        self.assertEqual(resp.status_code, 403)


class SalesLocationTests(TestCase):
    """Sales documents carry a location; fulfilment and reports honour it."""

    def setUp(self):
        from core.models import OrgMembership, InventoryMovement
        from core.services.inventory import apply_movement
        self.tenant = Tenant.objects.create(name="SalesLoc Co")
        Location.objects.filter(tenant=self.tenant).delete()
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.locA = Location.objects.create(tenant=self.tenant, name="Shop A", type=Location.Type.SHOP)
        self.locB = Location.objects.create(tenant=self.tenant, name="Shop B", type=Location.Type.SHOP)
        self.prod = Product.objects.create(tenant=self.tenant, sku="SL1", name="Widget",
                                           cost_method=Product.CostMethod.AVERAGE)
        for loc in (self.locA, self.locB):
            apply_movement(tenant=self.tenant, product=self.prod, location=loc,
                           movement_type=InventoryMovement.MovementType.RECEIVE, qty_delta=Decimal("100"),
                           ref_type="SEED", ref_id=f"OPEN-{loc.id}", unit_cost=Decimal("4.00"))
        self.customer = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.user = User.objects.create_user("sluser", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="sluser", password="pw")

    def test_invoice_fulfils_from_its_location(self):
        from core.models import InventoryBalance
        from core.services.gl import post_customer_invoice
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer,
                                             invoice_number="INV-SL1", location=self.locB)
        CustomerInvoiceLine.objects.create(invoice=inv, product=self.prod, qty=Decimal("10"),
                                           unit_price=Decimal("25.00"), tax_code=self.std)
        post_customer_invoice(inv)
        balA = InventoryBalance.objects.get(tenant=self.tenant, product=self.prod, location=self.locA)
        balB = InventoryBalance.objects.get(tenant=self.tenant, product=self.prod, location=self.locB)
        self.assertEqual(balA.on_hand, Decimal("100.00"))  # untouched
        self.assertEqual(balB.on_hand, Decimal("90.00"))   # deducted here

    def test_invoice_without_location_falls_back(self):
        from core.models import InventoryBalance
        from core.services.gl import post_customer_invoice
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer,
                                             invoice_number="INV-SL2")  # no location
        CustomerInvoiceLine.objects.create(invoice=inv, product=self.prod, qty=Decimal("5"),
                                           unit_price=Decimal("25.00"), tax_code=self.std)
        post_customer_invoice(inv)
        # Falls back to first stock location (locA by id) - still deducts somewhere.
        total = sum(b.on_hand for b in InventoryBalance.objects.filter(tenant=self.tenant, product=self.prod))
        self.assertEqual(total, Decimal("195.00"))  # 200 - 5

    def test_create_view_defaults_location(self):
        resp = self.client.post("/ar/invoices/new/", {
            "customer": self.customer.id, "invoice_date": "2026-01-01", "action": "save",
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "0",
            "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-product": self.prod.id, "lines-0-qty": "1", "lines-0-unit_price": "10",
            "lines-0-tax_code": self.std.id,
        })
        self.assertEqual(resp.status_code, 302)
        inv = CustomerInvoice.objects.get(tenant=self.tenant, customer=self.customer)
        self.assertIsNotNone(inv.location_id)  # defaulted to a stock location

    def test_order_to_invoice_carries_location(self):
        from core.models import CustomerOrder, CustomerOrderLine
        o = CustomerOrder.objects.create(tenant=self.tenant, customer=self.customer,
                                         order_number="SO-SL1", location=self.locB,
                                         status=CustomerOrder.Status.CONFIRMED)
        CustomerOrderLine.objects.create(order=o, product=self.prod, qty=Decimal("3"),
                                         unit_price=Decimal("10"), tax_code=self.std)
        resp = self.client.post(f"/customer-orders/{o.id}/to-invoice/")
        self.assertEqual(resp.status_code, 302)
        inv = CustomerInvoice.objects.get(source_order=o)
        self.assertEqual(inv.location_id, self.locB.id)

    def test_sales_history_location_filter(self):
        from core.services import sales_reports
        from datetime import date
        for n, loc in (("INV-A", self.locA), ("INV-B", self.locB)):
            inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.customer,
                                                 invoice_number=n, location=loc,
                                                 status=CustomerInvoice.Status.ISSUED,
                                                 invoice_date=date(2026, 1, 15))
            CustomerInvoiceLine.objects.create(invoice=inv, product=self.prod, qty=Decimal("1"),
                                               unit_price=Decimal("50"), tax_code=self.std)
        df, dt = date(2026, 1, 1), date(2026, 12, 31)
        all_rows = sales_reports.sales_history(self.tenant, df, dt)["rows"]
        b_rows = sales_reports.sales_history(self.tenant, df, dt, location_ids=[self.locB.id])["rows"]
        self.assertEqual(len(all_rows), 2)
        self.assertEqual(len(b_rows), 1)
        self.assertEqual(b_rows[0]["invoice"].invoice_number, "INV-B")


class POReceivingLocationTests(TestCase):
    """Purchase orders carry a structured receiving location; shipments and
    goods receipts land stock there."""

    def setUp(self):
        from core.models import OrgMembership, Supplier
        self.tenant = Tenant.objects.create(name="POLoc Co")
        Location.objects.filter(tenant=self.tenant).delete()
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.locA = Location.objects.create(tenant=self.tenant, name="WH A", type=Location.Type.WAREHOUSE)
        self.locB = Location.objects.create(tenant=self.tenant, name="WH B", type=Location.Type.WAREHOUSE)
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Supp")
        self.prod = Product.objects.create(tenant=self.tenant, sku="POL1", name="Widget")
        self.user = User.objects.create_user("poluser", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="poluser", password="pw")

    def _post_po(self, receiving_location=None, action="save"):
        data = {
            "supplier": self.supplier.id, "expected_date": "2026-02-01", "action": action,
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "0",
            "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-product": self.prod.id, "lines-0-ordered_qty": "10",
            "lines-0-unit_cost": "4.00", "lines-0-tax_code": self.std.id,
        }
        if receiving_location is not None:
            data["receiving_location"] = receiving_location.id
        return self.client.post("/po/new/", data)

    def test_po_defaults_receiving_location(self):
        from core.models import PurchaseOrder
        resp = self._post_po()
        self.assertEqual(resp.status_code, 302)
        po = PurchaseOrder.objects.get(tenant=self.tenant)
        self.assertIsNotNone(po.receiving_location_id)  # defaulted to a stock location

    def test_explicit_receiving_location_drives_shipment_destination(self):
        from core.models import PurchaseOrder, Shipment
        resp = self._post_po(receiving_location=self.locB, action="submit")
        self.assertEqual(resp.status_code, 302)
        po = PurchaseOrder.objects.get(tenant=self.tenant)
        self.assertEqual(po.receiving_location_id, self.locB.id)
        shipment = Shipment.objects.get(po=po)
        self.assertEqual(shipment.destination_id, self.locB.id)

    def test_po_destination_helper(self):
        from core.models import PurchaseOrder
        from core.views import _po_destination
        po = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-H1",
                                          supplier=self.supplier, receiving_location=self.locB)
        self.assertEqual(_po_destination(po), self.locB)
        po2 = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-H2", supplier=self.supplier)
        self.assertEqual(_po_destination(po2), self.locA)  # fallback to first location


class RequisitionDepartmentFKTests(TestCase):
    """PurchaseRequisition.department is a structured Department FK; access-request
    approval can assign a department to the new member."""

    def setUp(self):
        from core.models import OrgMembership, Department
        self.tenant = Tenant.objects.create(name="ReqDept Co")
        self.other = Tenant.objects.create(name="ReqDept Other")
        self.dept = Department.objects.create(tenant=self.tenant, name="Procurement")
        self.foreign_dept = Department.objects.create(tenant=self.other, name="Foreign")
        self.admin = User.objects.create_user("rdadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="rdadmin", password="pw")

    def test_requisition_form_scopes_department_to_tenant(self):
        from core.forms import PurchaseRequisitionForm
        from core.current import set_current_tenant, clear_current_tenant
        set_current_tenant(self.tenant)
        try:
            ids = set(PurchaseRequisitionForm().fields["department"].queryset.values_list("id", flat=True))
        finally:
            clear_current_tenant()
        self.assertIn(self.dept.id, ids)
        self.assertNotIn(self.foreign_dept.id, ids)  # another org's dept excluded

    def test_requisition_links_department(self):
        from core.models import PurchaseRequisition
        req = PurchaseRequisition.objects.create(tenant=self.tenant, req_number="PR-D1", department=self.dept)
        self.assertEqual(req.department, self.dept)
        self.assertEqual(self.dept.requisitions.get(), req)

    def test_deleting_department_nulls_requisition(self):
        from core.models import PurchaseRequisition, Department
        d = Department.objects.create(tenant=self.tenant, name="Temp")
        req = PurchaseRequisition.objects.create(tenant=self.tenant, req_number="PR-D2", department=d)
        d.delete()
        req.refresh_from_db()
        self.assertIsNone(req.department_id)  # SET_NULL

    def test_access_request_approval_assigns_department(self):
        from core.models import AccessRequest, OrgMembership
        ar = AccessRequest.objects.create(name="Jane Doe", email="jane@x.example", team="Procurement")
        resp = self.client.post(f"/access-requests/{ar.id}/action/", {
            "action": "approve", "role": "PURCHASING", "department": self.dept.id,
        })
        self.assertEqual(resp.status_code, 302)
        ar.refresh_from_db()
        m = OrgMembership.objects.get(user=ar.created_user, tenant=self.tenant)
        self.assertEqual(m.department, self.dept)

    def test_access_request_approval_without_department(self):
        from core.models import AccessRequest, OrgMembership
        ar = AccessRequest.objects.create(name="No Dept", email="nd@x.example", team="")
        resp = self.client.post(f"/access-requests/{ar.id}/action/", {
            "action": "approve", "role": "SALES",
        })
        self.assertEqual(resp.status_code, 302)
        ar.refresh_from_db()
        m = OrgMembership.objects.get(user=ar.created_user, tenant=self.tenant)
        self.assertIsNone(m.department_id)


class SiteContextTests(TestCase):
    """Mandatory company + site context: auto-select, gate, switching, 403, audit.
    There is never an 'all sites' option."""
    client_class = Client  # exercise the real gate, not the auto-context client

    def setUp(self):
        from core.models import OrgMembership, Site
        self.tenant = Tenant.objects.create(name="Ctx Co")
        self.main_site = Site.objects.get(tenant=self.tenant, is_default=True)
        self.user = User.objects.create_user("ctxu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)

    def _login(self):
        self.client.login(username="ctxu", password="pw")

    def test_single_company_single_site_autoselects(self):
        self._login()
        resp = self.client.get("/inventory/")
        self.assertEqual(resp.status_code, 200)  # no redirect - auto-selected
        self.assertEqual(self.client.session["active_tenant_id"], self.tenant.id)
        self.assertEqual(self.client.session["active_site_id"], self.main_site.id)

    def test_multiple_sites_force_selection(self):
        from core.models import AuditLog, Site
        siteB = Site.objects.create(tenant=self.tenant, name="Leicester", site_type=Site.Type.CITY_BRANCH)
        self._login()
        resp = self.client.get("/inventory/")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/select-site/", resp["Location"])
        # Pick a specific site -> proceeds, context set, audited.
        resp2 = self.client.post("/select-site/", {"site": siteB.id, "next": "/inventory/"})
        self.assertRedirects(resp2, "/inventory/")
        self.assertEqual(self.client.session["active_site_id"], siteB.id)
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="SITE_SELECTED").exists())

    def test_site_picker_has_no_all_sites_option(self):
        from core.models import Site
        Site.objects.create(tenant=self.tenant, name="Leicester", site_type=Site.Type.CITY_BRANCH)
        self._login()
        resp = self.client.get("/select-site/")
        self.assertEqual(resp.status_code, 200)
        self.assertNotContains(resp, "All sites")
        self.assertNotContains(resp, "All Sites")

    def test_multiple_companies_force_company_then_site(self):
        from core.models import OrgMembership, Tenant as T
        t2 = T.objects.create(name="Ctx Co 2")  # gets its own default Site via signal
        OrgMembership.objects.create(user=self.user, tenant=t2, role="ADMIN")
        self._login()
        resp = self.client.get("/inventory/")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/select-org/", resp["Location"])
        # Choose company -> then site auto-resolves (each has one) -> dashboard reachable.
        self.client.post("/select-org/", {"tenant": self.tenant.id})
        self.assertEqual(self.client.session["active_tenant_id"], self.tenant.id)

    def test_switch_company_clears_site(self):
        from core.models import OrgMembership, Tenant as T, AuditLog
        t2 = T.objects.create(name="Ctx Co 2")
        OrgMembership.objects.create(user=self.user, tenant=t2, role="ADMIN")
        self._login()
        s = self.client.session
        s["active_tenant_id"] = self.tenant.id
        s["active_site_id"] = self.main_site.id
        s.save()
        resp = self.client.post("/switch-company/", {"tenant": t2.id})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.client.session["active_tenant_id"], t2.id)
        self.assertNotIn("active_site_id", self.client.session)  # site cleared
        self.assertTrue(AuditLog.objects.filter(action="COMPANY_SWITCHED").exists())

    def test_switch_site_stays_on_page(self):
        from core.models import AuditLog, Site
        siteB = Site.objects.create(tenant=self.tenant, name="Leicester", site_type=Site.Type.CITY_BRANCH)
        self._login()
        self.client.get("/inventory/")  # auto-selects... multiple sites now -> may redirect; set context
        s = self.client.session
        s["active_site_id"] = self.main_site.id
        s.save()
        resp = self.client.post("/switch-site/", {"site": siteB.id, "next": "/inventory/"})
        self.assertRedirects(resp, "/inventory/")
        self.assertEqual(self.client.session["active_site_id"], siteB.id)
        self.assertTrue(AuditLog.objects.filter(action="SITE_SWITCHED").exists())

    def test_unauthorised_site_returns_403_and_audits(self):
        from core.models import Tenant as T, Site, AuditLog
        other = T.objects.create(name="Other Ctx")
        other_site = Site.objects.filter(tenant=other).first()
        self._login()
        s = self.client.session
        s["active_tenant_id"] = self.tenant.id
        s["active_site_id"] = self.main_site.id
        s.save()
        resp = self.client.post("/switch-site/", {"site": other_site.id})
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(AuditLog.objects.filter(action="UNAUTHORISED_SITE_ACCESS").exists())

    def test_unauthorised_company_returns_403(self):
        from core.models import Tenant as T
        other = T.objects.create(name="Other Ctx 2")  # user is not a member
        self._login()
        resp = self.client.post("/switch-company/", {"tenant": other.id})
        self.assertEqual(resp.status_code, 403)

    def test_restricted_user_no_site_shows_no_site_page(self):
        from core.models import OrgMembership, UserSiteAccess
        # A warehouse user granted only an inactive site -> no selectable site.
        self.main_site.is_active = False
        self.main_site.save()
        u2 = User.objects.create_user("ctxnos", password="pw")
        OrgMembership.objects.create(user=u2, tenant=self.tenant, role="WAREHOUSE", is_default=True)
        UserSiteAccess.objects.create(tenant=self.tenant, user=u2, site=self.main_site)
        self.client.login(username="ctxnos", password="pw")
        resp = self.client.get("/inventory/", follow=True)
        self.assertContains(resp, "No site available")


class WorkspaceSwitcherTests(TestCase):
    """The two-pane workspace picker switches company + site atomically via
    /switch-workspace/, validating access to both before applying either."""
    client_class = Client

    def setUp(self):
        from core.models import OrgMembership, Tenant as T, Site
        self.t1 = Tenant.objects.create(name="WS One")
        self.t1_main = Site.objects.get(tenant=self.t1, is_default=True)
        self.t1_leic = Site.objects.create(tenant=self.t1, name="Leicester", site_type=Site.Type.CITY_BRANCH)
        self.t2 = T.objects.create(name="WS Two")
        self.t2_main = Site.objects.get(tenant=self.t2, is_default=True)
        self.user = User.objects.create_user("wsu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t1, role="ADMIN", is_default=True)
        OrgMembership.objects.create(user=self.user, tenant=self.t2, role="ADMIN")
        self.client.login(username="wsu", password="pw")
        self._set_ctx(self.t1.id, self.t1_main.id)

    def _set_ctx(self, tid, sid):
        s = self.client.session
        s["active_tenant_id"] = tid
        s["active_site_id"] = sid
        s.save()

    def test_switch_both_company_and_site(self):
        from core.models import AuditLog
        resp = self.client.post("/switch-workspace/",
                                {"tenant": self.t2.id, "site": self.t2_main.id, "next": "/"})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.client.session["active_tenant_id"], self.t2.id)
        self.assertEqual(self.client.session["active_site_id"], self.t2_main.id)
        self.assertTrue(AuditLog.objects.filter(action="COMPANY_SWITCHED").exists())
        self.assertTrue(AuditLog.objects.filter(action="SITE_SWITCHED").exists())

    def test_switch_site_within_same_company(self):
        from core.models import AuditLog
        resp = self.client.post("/switch-workspace/",
                                {"tenant": self.t1.id, "site": self.t1_leic.id, "next": "/inventory/"})
        self.assertRedirects(resp, "/inventory/", fetch_redirect_response=False)
        self.assertEqual(self.client.session["active_tenant_id"], self.t1.id)
        self.assertEqual(self.client.session["active_site_id"], self.t1_leic.id)
        # Company did not change -> no COMPANY_SWITCHED audit event.
        self.assertFalse(AuditLog.objects.filter(action="COMPANY_SWITCHED").exists())

    def test_unauthorised_company_403_and_no_change(self):
        from core.models import Tenant as T, Site
        other = T.objects.create(name="WS Stranger")
        other_site = Site.objects.filter(tenant=other, is_default=True).first()
        resp = self.client.post("/switch-workspace/", {"tenant": other.id, "site": other_site.id})
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(self.client.session["active_tenant_id"], self.t1.id)  # unchanged

    def test_site_outside_company_403_and_no_change(self):
        # A site the user can access, but not within the posted company.
        resp = self.client.post("/switch-workspace/", {"tenant": self.t1.id, "site": self.t2_main.id})
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(self.client.session["active_site_id"], self.t1_main.id)  # unchanged

    def test_context_lists_each_company_with_its_sites(self):
        resp = self.client.get("/", follow=True)
        wc = resp.context["workspace_companies"]
        sites_by_company = {w["tenant"].id: {s.id for s in w["sites"]} for w in wc}
        self.assertEqual(sites_by_company.get(self.t1.id), {self.t1_main.id, self.t1_leic.id})
        self.assertEqual(sites_by_company.get(self.t2.id), {self.t2_main.id})

    def test_dashboard_renders_switcher_modal(self):
        # The modal must render on the dashboard (in the body-level `modals`
        # block, so Bootstrap's fixed positioning isn't trapped by .skn-page).
        resp = self.client.get("/", follow=True)
        self.assertContains(resp, 'id="workspaceSwitcher"')
        self.assertContains(resp, 'action="/switch-workspace/"')
        # No leaked Django template comments (e.g. a multi-line {# #} that the
        # lexer fails to parse and renders as literal text).
        self.assertNotContains(resp, "{#")


class SiteDataScopingTests(TestCase):
    """Module lists show only the selected site's data; switching site swaps it.
    There is no combined view."""

    def setUp(self):
        from core.models import (OrgMembership, InventoryBalance, CustomerOrder,
                                 CustomerOrderLine, Supplier, Site)
        self.tenant = Tenant.objects.create(name="Scope Co")
        Location.objects.filter(tenant=self.tenant).delete()
        Site.objects.filter(tenant=self.tenant).delete()
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        # Two sites, each with one inventory location.
        self.siteA = Site.objects.create(tenant=self.tenant, name="A Site", site_type=Site.Type.CITY_BRANCH)
        self.siteB = Site.objects.create(tenant=self.tenant, name="B Site", site_type=Site.Type.CITY_BRANCH)
        self.locA = Location.objects.create(tenant=self.tenant, site=self.siteA, name="A Warehouse", type=Location.Type.WAREHOUSE)
        self.locB = Location.objects.create(tenant=self.tenant, site=self.siteB, name="B Warehouse", type=Location.Type.WAREHOUSE)
        self.prod = Product.objects.create(tenant=self.tenant, sku="SC1", name="Widget")
        self.cust = Customer.objects.create(tenant=self.tenant, name="Cust")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="Supp")
        # Inventory at both sites' locations.
        InventoryBalance.objects.create(tenant=self.tenant, product=self.prod, location=self.locA, on_hand=Decimal("5"))
        InventoryBalance.objects.create(tenant=self.tenant, product=self.prod, location=self.locB, on_hand=Decimal("9"))
        # An invoice, order and PO whose location sits under each site.
        self.invA = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.cust, invoice_number="INV-A", location=self.locA)
        self.invB = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.cust, invoice_number="INV-B", location=self.locB)
        self.ordA = CustomerOrder.objects.create(tenant=self.tenant, customer=self.cust, order_number="SO-A", location=self.locA)
        self.ordB = CustomerOrder.objects.create(tenant=self.tenant, customer=self.cust, order_number="SO-B", location=self.locB)
        self.poA = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-A", supplier=self.supplier, receiving_location=self.locA)
        self.poB = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-B", supplier=self.supplier, receiving_location=self.locB)
        self.user = User.objects.create_user("scopeu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="scopeu", password="pw")  # auto-selects A Site (first)

    def _ctx_ids(self, url, key):
        return [o.id for o in self.client.get(url).context[key]]

    def test_lists_scoped_to_selected_site(self):
        self.assertEqual(self.client.session["active_site_id"], self.siteA.id)
        # Inventory (locations under the selected site)
        bals = self.client.get("/inventory/").context["balances"]
        self.assertEqual({b.location_id for b in bals}, {self.locA.id})
        # Invoices / orders / POs
        self.assertEqual(self._ctx_ids("/ar/invoices/", "invoices"), [self.invA.id])
        self.assertEqual(self._ctx_ids("/customer-orders/", "orders"), [self.ordA.id])
        self.assertEqual(self._ctx_ids("/po/", "pos"), [self.poA.id])

    def test_switching_site_swaps_data(self):
        self.client.post("/switch-site/", {"site": self.siteB.id})
        self.assertEqual(self.client.session["active_site_id"], self.siteB.id)
        bals = self.client.get("/inventory/").context["balances"]
        self.assertEqual({b.location_id for b in bals}, {self.locB.id})
        self.assertEqual(self._ctx_ids("/ar/invoices/", "invoices"), [self.invB.id])
        self.assertEqual(self._ctx_ids("/customer-orders/", "orders"), [self.ordB.id])
        self.assertEqual(self._ctx_ids("/po/", "pos"), [self.poB.id])


class DefaultSiteTests(TestCase):
    """A fresh company gets a default Site, and the Main Location sits under it."""

    def test_new_tenant_gets_default_site_with_location(self):
        from core.models import Site, Location
        t = Tenant.objects.create(name="Site Boot Co")
        sites = Site.objects.filter(tenant=t)
        self.assertEqual(sites.count(), 1)
        site = sites.get()
        self.assertEqual(site.name, "Main Site")
        self.assertTrue(site.is_default)
        self.assertEqual(site.site_type, Site.Type.OPERATING_SITE)
        loc = Location.objects.get(tenant=t, name="Main Location")
        self.assertEqual(loc.site_id, site.id)  # location belongs to the site

    def test_existing_location_without_site_is_attached(self):
        from core.models import Site, Location
        t = Tenant.objects.create(name="Orphan Co")
        # Simulate a legacy orphan location (no site).
        orphan = Location.objects.create(tenant=t, name="Legacy WH", type=Location.Type.WAREHOUSE)
        orphan.site = None
        orphan.save(update_fields=["site"])
        # The backfill helper (same logic as migration 0066) re-attaches it.
        site = Site.objects.filter(tenant=t, is_default=True).first()
        Location.objects.filter(tenant=t, site__isnull=True).update(site=site)
        orphan.refresh_from_db()
        self.assertEqual(orphan.site_id, site.id)

    def test_site_type_choices_present(self):
        from core.models import Site
        keys = set(dict(Site.Type.choices))
        self.assertEqual(keys, {"region", "country_division", "city_branch", "business_unit", "operating_site"})


class SiteAccessTests(TestCase):
    """UserSiteAccess gates which sites a user may work in (distinct from
    location access)."""

    def setUp(self):
        from core.models import OrgMembership, Site
        self.tenant = Tenant.objects.create(name="SA Co")
        self.main = Site.objects.get(tenant=self.tenant, is_default=True)
        self.leic = Site.objects.create(tenant=self.tenant, name="Leicester", site_type=Site.Type.CITY_BRANCH)
        self.lon = Site.objects.create(tenant=self.tenant, name="London", site_type=Site.Type.CITY_BRANCH)
        self.admin = User.objects.create_user("saadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.staff = User.objects.create_user("sastaff", password="pw")
        OrgMembership.objects.create(user=self.staff, tenant=self.tenant, role="WAREHOUSE")

    def test_no_grants_unrestricted(self):
        from core.access import accessible_site_ids
        self.assertIsNone(accessible_site_ids(self.staff, self.tenant))

    def test_admin_unrestricted_even_with_grants(self):
        from core.access import accessible_site_ids
        from core.models import UserSiteAccess
        UserSiteAccess.objects.create(tenant=self.tenant, user=self.admin, site=self.leic)
        self.assertIsNone(accessible_site_ids(self.admin, self.tenant))

    def test_grants_restrict_staff(self):
        from core.access import accessible_site_ids, selectable_sites
        from core.models import UserSiteAccess
        UserSiteAccess.objects.create(tenant=self.tenant, user=self.staff, site=self.leic)
        self.assertEqual(accessible_site_ids(self.staff, self.tenant), {self.leic.id})
        names = set(selectable_sites(self.staff, self.tenant).values_list("name", flat=True))
        self.assertEqual(names, {"Leicester"})

    def test_selectable_sites_excludes_inactive(self):
        from core.access import selectable_sites
        self.lon.is_active = False
        self.lon.save()
        names = set(selectable_sites(self.admin, self.tenant).values_list("name", flat=True))
        self.assertEqual(names, {"Main Site", "Leicester"})

    def test_site_access_scoped_to_tenant(self):
        from core.access import accessible_site_ids
        from core.models import UserSiteAccess, Tenant as T
        other = T.objects.create(name="SA Other")
        # A grant in another tenant must not leak.
        UserSiteAccess.objects.create(tenant=self.tenant, user=self.staff, site=self.leic)
        self.assertIsNone(accessible_site_ids(self.staff, other))  # no grants there -> unrestricted in 'other'


class ActiveSiteResolutionTests(TestCase):
    """get_active_site resolves from the session, validates access, and falls
    back to the selected location's site during the transition."""

    def setUp(self):
        from core.models import OrgMembership, Site
        self.tenant = Tenant.objects.create(name="AS Co")
        self.main_site = Site.objects.get(tenant=self.tenant, is_default=True)
        self.main_loc = Location.objects.get(tenant=self.tenant, name="Main Location")
        self.user = User.objects.create_user("asu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)

    def _request(self, **session):
        from django.test import RequestFactory
        from django.contrib.sessions.backends.db import SessionStore
        req = RequestFactory().get("/")
        req.user = self.user
        s = SessionStore()
        for k, v in session.items():
            s[k] = v
        s.save()
        req.session = s
        return req

    def test_explicit_site_selected(self):
        from core.access import get_active_site, SESSION_TENANT_KEY, SESSION_SITE_KEY
        req = self._request(**{SESSION_TENANT_KEY: self.tenant.id, SESSION_SITE_KEY: self.main_site.id})
        self.assertEqual(get_active_site(req), self.main_site)

    def test_fallback_from_selected_location(self):
        from core.access import get_active_site, SESSION_TENANT_KEY, SESSION_LOCATION_KEY
        # No site key, but a location is selected -> derive its site.
        req = self._request(**{SESSION_TENANT_KEY: self.tenant.id, SESSION_LOCATION_KEY: self.main_loc.id})
        self.assertEqual(get_active_site(req), self.main_site)

    def test_invalid_site_not_returned(self):
        from core.access import get_active_site, SESSION_TENANT_KEY, SESSION_SITE_KEY
        from core.models import Site, Tenant as T
        other = T.objects.create(name="AS Other")
        foreign = Site.objects.filter(tenant=other).first()
        req = self._request(**{SESSION_TENANT_KEY: self.tenant.id, SESSION_SITE_KEY: foreign.id})
        self.assertIsNone(get_active_site(req))  # not in this company -> not returned


class StockSiteAndWorkflowTests(TestCase):
    """Stock balances/movements carry site_id; inventory workflow location
    pickers are scoped to the selected site (cross-site transfer destinations
    stay open)."""

    def setUp(self):
        from core.models import OrgMembership, Site
        self.tenant = Tenant.objects.create(name="S6 Co")
        Location.objects.filter(tenant=self.tenant).delete()
        Site.objects.filter(tenant=self.tenant).delete()
        self.siteA = Site.objects.create(tenant=self.tenant, name="A Site", site_type=Site.Type.CITY_BRANCH, is_default=True)
        self.siteB = Site.objects.create(tenant=self.tenant, name="B Site", site_type=Site.Type.CITY_BRANCH)
        self.locA = Location.objects.create(tenant=self.tenant, site=self.siteA, name="A WH", type=Location.Type.WAREHOUSE)
        self.locB = Location.objects.create(tenant=self.tenant, site=self.siteB, name="B WH", type=Location.Type.WAREHOUSE)
        self.prod = Product.objects.create(tenant=self.tenant, sku="S6-1", name="Widget")
        self.user = User.objects.create_user("s6u", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="s6u", password="pw")  # auto-selects A Site

    def test_movement_and_balance_record_site(self):
        from core.services.inventory import apply_movement
        from core.models import InventoryBalance, InventoryMovement
        apply_movement(tenant=self.tenant, product=self.prod, location=self.locB,
                       movement_type="RECEIVE", qty_delta=Decimal("10"), ref_type="SEED", ref_id="1",
                       unit_cost=Decimal("2"))
        bal = InventoryBalance.objects.get(tenant=self.tenant, product=self.prod, location=self.locB)
        self.assertEqual(bal.site_id, self.siteB.id)  # derived from location.site
        mv = InventoryMovement.objects.filter(tenant=self.tenant, location=self.locB).first()
        self.assertEqual(mv.site_id, self.siteB.id)

    def test_adjustment_location_picker_scoped_to_active_site(self):
        from core.forms import StockAdjustmentForm
        # Active site is A Site -> adjustment location options limited to A's locations.
        resp = self.client.get("/inventory/adjustments/new/")
        form = resp.context["form"]
        ids = set(form.fields["location"].queryset.values_list("id", flat=True))
        self.assertEqual(ids, {self.locA.id})

    def test_transfer_from_is_site_scoped_to_is_open(self):
        from core.forms import InventoryTransferForm
        resp = self.client.get("/transfers/new/")
        form = resp.context["form"]
        from_ids = set(form.fields["from_location"].queryset.values_list("id", flat=True))
        to_ids = set(form.fields["to_location"].queryset.values_list("id", flat=True))
        self.assertEqual(from_ids, {self.locA.id})            # within the working site
        self.assertEqual(to_ids, {self.locA.id, self.locB.id})  # cross-site allowed


class DocumentSiteScopingTests(TestCase):
    """Sales orders/invoices and POs carry site_id (stamped from the active site
    on create, or derived), and lists scope by it."""

    def setUp(self):
        from core.models import OrgMembership, Site, Supplier
        self.tenant = Tenant.objects.create(name="Doc Co")
        Location.objects.filter(tenant=self.tenant).delete()
        Site.objects.filter(tenant=self.tenant).delete()
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.siteA = Site.objects.create(tenant=self.tenant, name="A Site", site_type=Site.Type.CITY_BRANCH, is_default=True)
        self.siteB = Site.objects.create(tenant=self.tenant, name="B Site", site_type=Site.Type.CITY_BRANCH)
        self.locA = Location.objects.create(tenant=self.tenant, site=self.siteA, name="A WH", type=Location.Type.WAREHOUSE)
        self.locB = Location.objects.create(tenant=self.tenant, site=self.siteB, name="B WH", type=Location.Type.WAREHOUSE)
        self.cust = Customer.objects.create(tenant=self.tenant, name="C")
        self.supplier = Supplier.objects.create(tenant=self.tenant, name="S")
        self.prod = Product.objects.create(tenant=self.tenant, sku="DOC1", name="W")
        self.user = User.objects.create_user("docu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="docu", password="pw")  # auto-selects A Site

    def test_invoice_site_derived_from_location(self):
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.cust,
                                             invoice_number="INV-1", location=self.locB)
        self.assertEqual(inv.site_id, self.siteB.id)

    def test_document_without_location_falls_back_to_default_site(self):
        # No location -> default Site (A Site is default).
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.cust, invoice_number="INV-2")
        self.assertEqual(inv.site_id, self.siteA.id)

    def test_create_view_stamps_active_site(self):
        # Switch to B Site, then create an invoice via the view -> stamped to B.
        self.client.post("/switch-site/", {"site": self.siteB.id})
        resp = self.client.post("/ar/invoices/new/", {
            "customer": self.cust.id, "invoice_date": "2026-01-01", "action": "save",
            "lines-TOTAL_FORMS": "1", "lines-INITIAL_FORMS": "0",
            "lines-MIN_NUM_FORMS": "0", "lines-MAX_NUM_FORMS": "1000",
            "lines-0-product": self.prod.id, "lines-0-qty": "1", "lines-0-unit_price": "10",
            "lines-0-tax_code": self.std.id,
        })
        self.assertEqual(resp.status_code, 302)
        inv = CustomerInvoice.objects.get(tenant=self.tenant, customer=self.cust)
        self.assertEqual(inv.site_id, self.siteB.id)

    def test_po_list_scoped_by_site(self):
        poA = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-A", supplier=self.supplier, receiving_location=self.locA)
        poB = PurchaseOrder.objects.create(tenant=self.tenant, po_number="PO-B", supplier=self.supplier, receiving_location=self.locB)
        ids = [p.id for p in self.client.get("/po/").context["pos"]]  # A Site active
        self.assertEqual(ids, [poA.id])


class UKSeedTests(TestCase):
    """The seed_uk_demo command builds Company UK -> city/region Sites -> named
    inventory locations, idempotently."""

    def test_seed_creates_sites_and_locations(self):
        from django.core.management import call_command
        from core.models import Tenant, Site, Location
        call_command("seed_uk_demo")
        call_command("seed_uk_demo")  # idempotent
        uk = Tenant.objects.get(name="UK")
        names = set(Site.objects.filter(tenant=uk).values_list("name", flat=True))
        for s in ["London", "Leicester", "Manchester", "Birmingham",
                  "England", "Wales", "Scotland", "Northern Ireland"]:
            self.assertIn(s, names)
        leic = Site.objects.get(tenant=uk, name="Leicester")
        leic_locs = set(Location.objects.filter(tenant=uk, site=leic).values_list("name", flat=True))
        self.assertEqual(leic_locs, {
            "Leicester Main Warehouse", "Leicester Shop Floor", "Leicester Back Room",
            "Leicester Returns Area", "Leicester Delivery Van"})
        # Region sites carry the region type.
        self.assertEqual(Site.objects.get(tenant=uk, name="Scotland").site_type, Site.Type.REGION)
        self.assertEqual(leic.site_type, Site.Type.CITY_BRANCH)


class SiteAccessMatrixTests(TestCase):
    """Admin UI to grant per-site access (UserSiteAccess), with audit."""

    def setUp(self):
        from core.models import OrgMembership, Site
        self.tenant = Tenant.objects.create(name="SAM Co")
        self.main = Site.objects.get(tenant=self.tenant, is_default=True)
        self.leic = Site.objects.create(tenant=self.tenant, name="Leicester", site_type=Site.Type.CITY_BRANCH)
        self.admin = User.objects.create_user("samadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)
        self.staff = User.objects.create_user("samstaff", password="pw")
        OrgMembership.objects.create(user=self.staff, tenant=self.tenant, role="WAREHOUSE")

    def test_admin_saves_site_grants_and_audits(self):
        from core.models import UserSiteAccess, AuditLog
        self.client.login(username="samadmin", password="pw")
        self.assertEqual(self.client.get("/sites/access/").status_code, 200)
        resp = self.client.post("/sites/access/", {f"grant_{self.staff.id}_{self.leic.id}": "on"})
        self.assertEqual(resp.status_code, 302)
        grants = set(UserSiteAccess.objects.filter(tenant=self.tenant, user=self.staff).values_list("site_id", flat=True))
        self.assertEqual(grants, {self.leic.id})
        self.assertTrue(AuditLog.objects.filter(tenant=self.tenant, action="SITE_ACCESS_CHANGED").exists())

    def test_non_admin_forbidden(self):
        self.client.login(username="samstaff", password="pw")
        self.assertEqual(self.client.get("/sites/access/").status_code, 403)

    def test_grant_restricts_site_picker(self):
        from core.models import UserSiteAccess
        from core.access import selectable_sites
        UserSiteAccess.objects.create(tenant=self.tenant, user=self.staff, site=self.leic)
        names = set(selectable_sites(self.staff, self.tenant).values_list("name", flat=True))
        self.assertEqual(names, {"Leicester"})  # restricted to the granted site


class GLSiteDimensionTests(TestCase):
    """Journal entries carry a Site (from their source document) so P&L /
    balance sheet can be filtered by site; company-wide is the default."""

    def setUp(self):
        from core.models import OrgMembership, Site, InventoryMovement, GLAccount
        from core.services.inventory import apply_movement
        self.tenant = Tenant.objects.create(name="GLS Co")
        Location.objects.filter(tenant=self.tenant).delete()
        Site.objects.filter(tenant=self.tenant).delete()
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.siteA = Site.objects.create(tenant=self.tenant, name="A Site", site_type=Site.Type.CITY_BRANCH, is_default=True)
        self.siteB = Site.objects.create(tenant=self.tenant, name="B Site", site_type=Site.Type.CITY_BRANCH)
        self.locA = Location.objects.create(tenant=self.tenant, site=self.siteA, name="A WH", type=Location.Type.WAREHOUSE)
        self.locB = Location.objects.create(tenant=self.tenant, site=self.siteB, name="B WH", type=Location.Type.WAREHOUSE)
        self.cust = Customer.objects.create(tenant=self.tenant, name="C")
        self.prod = Product.objects.create(tenant=self.tenant, sku="GLS1", name="W",
                                            cost_method=Product.CostMethod.AVERAGE)
        for loc in (self.locA, self.locB):
            apply_movement(tenant=self.tenant, product=self.prod, location=loc,
                           movement_type=InventoryMovement.MovementType.RECEIVE, qty_delta=Decimal("100"),
                           ref_type="SEED", ref_id=f"O{loc.id}", unit_cost=Decimal("4.00"))

    def _invoice(self, number, location, qty, price):
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.cust,
                                             invoice_number=number, location=location)
        CustomerInvoiceLine.objects.create(invoice=inv, product=self.prod, qty=Decimal(qty),
                                           unit_price=Decimal(price), tax_code=self.std)
        post_customer_invoice(inv)
        return inv

    def test_invoice_je_carries_site(self):
        from core.models import JournalEntry
        self._invoice("INV-A", self.locA, "10", "25")
        for ref in ("AR_INVOICE", "COGS"):
            je = JournalEntry.objects.get(tenant=self.tenant, ref_type=ref, ref_id="INV-A")
            self.assertEqual(je.site_id, self.siteA.id)

    def test_pnl_filtered_by_site(self):
        from core.services import reports
        self._invoice("INV-A", self.locA, "10", "25")  # site A: revenue 250, COGS 40
        self._invoice("INV-B", self.locB, "4", "25")   # site B: revenue 100, COGS 16
        company = reports.profit_and_loss(self.tenant)
        self.assertEqual(company["income_total"], Decimal("350.00"))  # combined default
        a = reports.profit_and_loss(self.tenant, site_ids=[self.siteA.id])
        self.assertEqual(a["income_total"], Decimal("250.00"))
        self.assertEqual(a["cogs_total"], Decimal("40.00"))
        b = reports.profit_and_loss(self.tenant, site_ids=[self.siteB.id])
        self.assertEqual(b["income_total"], Decimal("100.00"))

    def test_expense_je_carries_site(self):
        from core.models import Expense, GLAccount, JournalEntry
        from core.services.gl import post_expense
        acc = GLAccount.objects.filter(tenant=self.tenant, code="6100").first()
        e = Expense.objects.create(tenant=self.tenant, site=self.siteB, payee="Rent Co",
                                   category=acc, net_amount=Decimal("100.00"))
        post_expense(e)
        je = JournalEntry.objects.get(tenant=self.tenant, ref_type="EXPENSE", ref_id=str(e.id))
        self.assertEqual(je.site_id, self.siteB.id)
        from core.services import reports
        b = reports.profit_and_loss(self.tenant, site_ids=[self.siteB.id])
        self.assertEqual(b["expense_total"], Decimal("100.00"))
        a = reports.profit_and_loss(self.tenant, site_ids=[self.siteA.id])
        self.assertEqual(a["expense_total"], Decimal("0.00"))

    def test_pnl_page_site_filter(self):
        from core.models import OrgMembership
        u = User.objects.create_user("glsu", password="pw")
        OrgMembership.objects.create(user=u, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="glsu", password="pw")
        self._invoice("INV-A", self.locA, "10", "25")
        self._invoice("INV-B", self.locB, "4", "25")
        # Company-wide (default).
        resp = self.client.get("/reports/profit-and-loss/?from=2026-01-01&to=2026-12-31")
        self.assertEqual(resp.context["data"]["income_total"], Decimal("350.00"))
        # Filtered to site A.
        resp = self.client.get(f"/reports/profit-and-loss/?from=2026-01-01&to=2026-12-31&site={self.siteA.id}")
        self.assertEqual(resp.context["data"]["income_total"], Decimal("250.00"))


class AuditSiteTests(TestCase):
    """Audit records capture the working Site, and the log can be filtered by it."""

    def setUp(self):
        from core.models import OrgMembership, Site
        self.tenant = Tenant.objects.create(name="AudSite Co")
        self.main = Site.objects.get(tenant=self.tenant, is_default=True)
        self.leic = Site.objects.create(tenant=self.tenant, name="Leicester", site_type=Site.Type.CITY_BRANCH)
        self.admin = User.objects.create_user("audadmin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.tenant, role="ADMIN", is_default=True)

    def test_log_audit_stamps_current_site(self):
        from core.models import AuditLog
        from core.current import set_current_site, clear_current_site
        from core.audit import log_audit
        set_current_site(self.leic)
        try:
            log_audit(action="DATA_EXPORTED", tenant=self.tenant, username="x", detail="test")
        finally:
            clear_current_site()
        log = AuditLog.objects.filter(tenant=self.tenant, action="DATA_EXPORTED").first()
        self.assertEqual(log.site_id, self.leic.id)

    def test_explicit_site_overrides_threadlocal(self):
        from core.models import AuditLog
        from core.audit import log_audit
        log_audit(action="RECORD_DELETED", tenant=self.tenant, site=self.main, username="x")
        log = AuditLog.objects.filter(tenant=self.tenant, action="RECORD_DELETED").first()
        self.assertEqual(log.site_id, self.main.id)

    def test_audit_log_view_filters_by_site(self):
        from core.models import AuditLog
        AuditLog.objects.create(tenant=self.tenant, site=self.leic, action="LOGIN", username="a")
        AuditLog.objects.create(tenant=self.tenant, site=self.main, action="LOGIN", username="b")
        self.client.login(username="audadmin", password="pw")
        resp = self.client.get(f"/audit/?site={self.leic.id}")
        self.assertEqual(resp.status_code, 200)
        usernames = {l.username for l in resp.context["logs"]}
        self.assertIn("a", usernames)
        self.assertNotIn("b", usernames)


class DashboardSiteScopingTests(TestCase):
    """Dashboard KPIs are scoped to the selected site (site-dimensioned data)."""

    def setUp(self):
        from core.models import OrgMembership, Site, InventoryMovement
        from core.services.inventory import apply_movement
        self.tenant = Tenant.objects.create(name="DashSite Co")
        Location.objects.filter(tenant=self.tenant).delete()
        Site.objects.filter(tenant=self.tenant).delete()
        self.std = TaxCode.objects.get(tenant=self.tenant, code="STD")
        self.siteA = Site.objects.create(tenant=self.tenant, name="A Site", site_type=Site.Type.CITY_BRANCH, is_default=True)
        self.siteB = Site.objects.create(tenant=self.tenant, name="B Site", site_type=Site.Type.CITY_BRANCH)
        self.locA = Location.objects.create(tenant=self.tenant, site=self.siteA, name="A WH", type=Location.Type.WAREHOUSE)
        self.locB = Location.objects.create(tenant=self.tenant, site=self.siteB, name="B WH", type=Location.Type.WAREHOUSE)
        self.cust = Customer.objects.create(tenant=self.tenant, name="C")
        self.prod = Product.objects.create(tenant=self.tenant, sku="DS1", name="W", cost_method=Product.CostMethod.AVERAGE)
        for loc in (self.locA, self.locB):
            apply_movement(tenant=self.tenant, product=self.prod, location=loc,
                           movement_type=InventoryMovement.MovementType.RECEIVE, qty_delta=Decimal("100"),
                           ref_type="SEED", ref_id=f"O{loc.id}", unit_cost=Decimal("4.00"))

    def _invoice(self, number, loc, qty, price):
        inv = CustomerInvoice.objects.create(tenant=self.tenant, customer=self.cust, invoice_number=number, location=loc)
        CustomerInvoiceLine.objects.create(invoice=inv, product=self.prod, qty=Decimal(qty), unit_price=Decimal(price), tax_code=self.std)
        post_customer_invoice(inv)
        return inv

    def test_kpis_scoped_to_site(self):
        from core.views import _dashboard_kpis
        from core.access import accessible_locations
        self._invoice("INV-A", self.locA, "10", "25")  # site A this month
        self._invoice("INV-B", self.locB, "4", "25")   # site B this month
        a_locs = list(accessible_locations(None, self.tenant).filter(site=self.siteA).values_list("id", flat=True))
        kA = {c["label"]: c["value"] for c in _dashboard_kpis(self.tenant, "ADMIN", site_id=self.siteA.id, location_ids=a_locs)}
        # "Sales (this month)" should reflect only site A's revenue.
        self.assertEqual(kA["Sales (this month)"], Decimal("250.00"))
        company = {c["label"]: c["value"] for c in _dashboard_kpis(self.tenant, "ADMIN")}
        self.assertEqual(company["Sales (this month)"], Decimal("350.00"))

    def test_dashboard_page_renders_with_site(self):
        from core.models import OrgMembership
        u = User.objects.create_user("dsu", password="pw")
        OrgMembership.objects.create(user=u, tenant=self.tenant, role="ADMIN", is_default=True)
        self.client.login(username="dsu", password="pw")  # auto-selects A Site
        resp = self.client.get("/dashboard/admin")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["active_site"], self.siteA)


class UkRetailDemoScenarioTests(TestCase):
    """End-to-end proof that the ERP separates the three tiers using the
    `seed_uk_retail_demo` data:

        Company  = legal/business entity (UK Retail Group Ltd)
        Site     = operating/reporting branch (London/Leicester/Manchester/Birmingham)
        Inventory Location = physical stock storage (warehouse/shop/back-room/returns)

    Covers the 10 required scenarios plus structural validation. The suite-wide
    `_CtxClient` is the default client; raw-gate scenarios use `django.test.Client`.
    """

    def setUp(self):
        from io import StringIO
        from django.core.management import call_command
        from core.models import Site, OrgMembership
        call_command("seed_uk_retail_demo", stdout=StringIO())
        self.tenant = Tenant.objects.get(name="UK Retail Group Ltd")
        self.london = Site.objects.get(tenant=self.tenant, name="London")
        self.leicester = Site.objects.get(tenant=self.tenant, name="Leicester")
        self.manchester = Site.objects.get(tenant=self.tenant, name="Manchester")
        self.birmingham = Site.objects.get(tenant=self.tenant, name="Birmingham")
        self.p1 = Product.objects.get(tenant=self.tenant, sku="UKR-001")
        self.p2 = Product.objects.get(tenant=self.tenant, sku="UKR-002")
        self.lon_wh = Location.objects.get(tenant=self.tenant, name="London Main Warehouse")
        self.man_wh = Location.objects.get(tenant=self.tenant, name="Manchester Main Warehouse")
        self.users = {m.role: m.user for m in OrgMembership.objects.filter(
            tenant=self.tenant, user__username__endswith="@ukretail.demo")}

    # -- helpers -------------------------------------------------------------
    def _ctx(self, user, site, client=None):
        """Log a user in and pin the active company + site."""
        from core.access import SESSION_TENANT_KEY, SESSION_SITE_KEY
        c = client or self.client
        c.force_login(user)
        s = c.session
        s[SESSION_TENANT_KEY] = self.tenant.id
        s[SESSION_SITE_KEY] = site.id
        s.save()
        return c

    # === Structural validation =============================================
    def test_company_sites_and_locations_structure(self):
        from core.models import Site
        self.assertEqual(Tenant.objects.filter(name="UK Retail Group Ltd").count(), 1)
        self.assertTrue(self.tenant.vat_registered)
        self.assertEqual(self.tenant.currency_code, "GBP")
        active = Site.objects.filter(tenant=self.tenant, is_active=True)
        self.assertEqual(set(active.values_list("name", flat=True)),
                         {"London", "Leicester", "Manchester", "Birmingham"})
        self.assertEqual(Site.objects.get(tenant=self.tenant, is_default=True).name, "London")
        for site in active:
            locs = Location.objects.filter(tenant=self.tenant, site=site)
            self.assertEqual(locs.count(), 4)
            self.assertEqual(set(locs.values_list("type", flat=True)),
                             {Location.Type.WAREHOUSE, Location.Type.SHOP_FLOOR,
                              Location.Type.BACK_ROOM, Location.Type.RETURNS})
            for loc in locs:
                self.assertEqual(loc.site_id, site.id)

    def test_eight_role_users_with_explicit_site_access(self):
        from core.models import OrgMembership, UserSiteAccess
        self.assertEqual(OrgMembership.objects.filter(
            tenant=self.tenant, user__username__endswith="@ukretail.demo").count(), 8)

        def granted(role):
            return set(UserSiteAccess.objects.filter(
                tenant=self.tenant, user=self.users[role]).values_list("site__name", flat=True))

        self.assertEqual(granted("SALES"), {"London"})
        self.assertEqual(granted("WAREHOUSE"), {"Manchester"})
        self.assertEqual(granted("MANAGER"), {"London", "Leicester"})
        self.assertEqual(granted("PURCHASING"), {"Birmingham"})
        for role in self.users:
            self.assertTrue(UserSiteAccess.objects.filter(
                tenant=self.tenant, user=self.users[role]).exists())

    # === Scenario 1: log in and select Company + Site ======================
    def test_scenario_01_login_selects_company_and_site(self):
        from core.access import SESSION_TENANT_KEY, SESSION_SITE_KEY
        c = Client()  # raw client -> exercise the real post-login context gate
        c.force_login(self.users["SALES"])  # SALES has exactly one site (London)
        resp = c.get("/", follow=True)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(c.session.get(SESSION_TENANT_KEY), self.tenant.id)
        self.assertEqual(c.session.get(SESSION_SITE_KEY), self.london.id)

    # === Scenario 2: switch from Leicester to London =======================
    def test_scenario_02_switch_leicester_to_london(self):
        from core.access import SESSION_SITE_KEY
        c = self._ctx(self.users["MANAGER"], self.leicester, client=Client())
        self.assertEqual(c.session[SESSION_SITE_KEY], self.leicester.id)
        resp = c.post("/switch-workspace/",
                      {"tenant": self.tenant.id, "site": self.london.id, "next": "/"})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(c.session[SESSION_SITE_KEY], self.london.id)

    # === Scenario 3: products remain company-level =========================
    def test_scenario_03_products_are_company_level(self):
        from django.core.exceptions import FieldDoesNotExist
        self.assertEqual(Product.objects.filter(tenant=self.tenant).count(), 2)
        with self.assertRaises(FieldDoesNotExist):
            Product._meta.get_field("site")  # no per-site product dimension

    # === Scenario 4: stock changes only inside selected site + location ====
    def test_scenario_04_stock_scoped_to_site_and_location(self):
        lon = InventoryBalance.objects.get(tenant=self.tenant, product=self.p1, location=self.lon_wh)
        man = InventoryBalance.objects.get(tenant=self.tenant, product=self.p1, location=self.man_wh)
        self.assertEqual(lon.on_hand, Decimal("93.00"))   # 100 opening - 5 invoice - 2 adj
        self.assertEqual(man.on_hand, Decimal("100.00"))  # untouched
        self.assertEqual(lon.site_id, self.london.id)     # site auto-synced from location
        self.assertEqual(man.site_id, self.manchester.id)

    # === Scenario 5: sales filtered by selected site =======================
    def test_scenario_05_sales_filtered_by_site(self):
        self._ctx(self.users["ADMIN"], self.london)
        invoices = [i.invoice_number for i in self.client.get("/ar/invoices/").context["invoices"]]
        self.assertIn("INV-0001", invoices)
        self._ctx(self.users["ADMIN"], self.manchester)
        invoices = [i.invoice_number for i in self.client.get("/ar/invoices/").context["invoices"]]
        self.assertNotIn("INV-0001", invoices)

    # === Scenario 6: purchasing filtered by selected site ==================
    def test_scenario_06_purchasing_filtered_by_site(self):
        self._ctx(self.users["ADMIN"], self.london)
        pos = [p.po_number for p in self.client.get("/po/").context["pos"]]
        self.assertIn("PO-0001", pos)
        self._ctx(self.users["ADMIN"], self.manchester)
        pos = [p.po_number for p in self.client.get("/po/").context["pos"]]
        self.assertNotIn("PO-0001", pos)

    # === Scenario 7: reports show only selected company + site =============
    def test_scenario_07_reports_filtered_by_site(self):
        from core.services import reports
        self.assertGreater(reports.net_income(self.tenant, site_ids=[self.london.id]), 0)
        self.assertEqual(reports.net_income(self.tenant, site_ids=[self.manchester.id]), 0)

    # === Scenario 8: warehouse user cannot access another site =============
    def test_scenario_08_warehouse_user_cannot_access_other_site(self):
        from core.access import selectable_sites
        from core.models import AuditLog
        wh = self.users["WAREHOUSE"]
        self.assertEqual(set(selectable_sites(wh, self.tenant).values_list("name", flat=True)),
                         {"Manchester"})
        c = self._ctx(wh, self.manchester, client=Client())
        resp = c.post("/switch-workspace/", {"tenant": self.tenant.id, "site": self.london.id})
        self.assertEqual(resp.status_code, 403)  # London not granted to this user
        self.assertTrue(AuditLog.objects.filter(
            tenant=self.tenant, action="UNAUTHORISED_SITE_ACCESS").exists())

    # === Scenario 9: inventory location chosen only inside inventory workflows
    def test_scenario_09_inventory_location_only_in_inventory_workflows(self):
        c = self._ctx(self.users["ADMIN"], self.london)
        sess = c.session
        self.assertIn("active_tenant_id", sess)
        self.assertIn("active_site_id", sess)
        self.assertIsNone(sess.get("active_location_id"))  # no global location in the context
        resp = c.get("/inventory/")
        self.assertEqual(resp.status_code, 200)
        for bal in resp.context["balances"]:
            self.assertEqual(bal.location.site_id, self.london.id)

    # === Scenario 10: no "All Sites" option anywhere =======================
    def test_scenario_10_no_all_sites_option(self):
        from core.access import selectable_sites
        names = set(selectable_sites(self.users["ADMIN"], self.tenant).values_list("name", flat=True))
        self.assertEqual(names, {"London", "Leicester", "Manchester", "Birmingham"})
        self._ctx(self.users["ADMIN"], self.london)
        resp = self.client.get("/", follow=True)
        self.assertNotContains(resp, "All Sites")
        self.assertNotContains(resp, "All sites")

    # === Bonus: inventory list scoped to a restricted warehouse user =======
    def test_scenario_10b_inventory_scoped_for_warehouse_user(self):
        c = self._ctx(self.users["WAREHOUSE"], self.manchester, client=Client())
        resp = c.get("/inventory/")
        self.assertEqual(resp.status_code, 200)
        for bal in resp.context["balances"]:
            self.assertEqual(bal.location.site_id, self.manchester.id)


class GlobalSearchAndNavTests(TestCase):
    """Permission-aware global search + grouped/collapsible hamburger menu, all
    driven from the navigation registry in core.roles."""

    def setUp(self):
        from core.models import OrgMembership
        self.t = Tenant.objects.create(name="Search Co")
        self.admin = User.objects.create_user("srch_admin", password="pw")
        OrgMembership.objects.create(user=self.admin, tenant=self.t, role="ADMIN", is_default=True)
        self.sales = User.objects.create_user("srch_sales", password="pw")
        OrgMembership.objects.create(user=self.sales, tenant=self.t, role="SALES", is_default=True)

    def _urls(self, role, q):
        from core.roles import search_nav
        return [r["url"] for r in search_nav(role, q, limit=None)]

    # ---- registry-level search ----
    def test_search_returns_known_pages(self):
        from core.roles import ADMIN
        self.assertEqual(self._urls(ADMIN, "purchase orders")[0], "/po/")
        self.assertIn("/inventory/", self._urls(ADMIN, "inventory"))
        self.assertIn("/uoms/", self._urls(ADMIN, "units of measure"))
        self.assertIn("/stock-takes/", self._urls(ADMIN, "stock take"))

    def test_aliases_resolve(self):
        from core.roles import ADMIN
        self.assertEqual(self._urls(ADMIN, "po")[0], "/po/")
        self.assertIn("/requisitions/", self._urls(ADMIN, "pr"))
        self.assertIn("/shipments/", self._urls(ADMIN, "grn"))
        self.assertTrue(set(self._urls(ADMIN, "uom")) & {"/uoms/", "/uom-conversions/"})
        self.assertTrue(set(self._urls(ADMIN, "gl")) & {"/gl/journal/", "/gl/accounts/"})
        self.assertIn("/returns/", self._urls(ADMIN, "rma"))
        self.assertTrue(set(self._urls(ADMIN, "vat")) & {"/vat/", "/tax-codes/"})
        self.assertIn("/gl/accounts/", self._urls(ADMIN, "coa"))
        self.assertIn("/invoices/", self._urls(ADMIN, "ap"))

    def test_short_alias_does_not_false_match(self):
        # "po" must surface Purchase Orders, never "Reports" / "Products".
        from core.roles import ADMIN
        urls = self._urls(ADMIN, "po")
        self.assertEqual(urls[0], "/po/")
        self.assertNotIn("/reports/", urls)
        self.assertNotIn("/products/", urls)

    def test_search_is_permission_aware(self):
        from core.roles import SALES
        # Sales has no GL or procurement access -> those pages never appear.
        self.assertNotIn("/gl/journal/", self._urls(SALES, "journal"))
        self.assertNotIn("/po/", self._urls(SALES, "po"))
        # But Sales-accessible pages still resolve.
        self.assertIn("/customers/", self._urls(SALES, "customers"))

    # ---- view / endpoint ----
    def test_search_page_renders(self):
        self.client.login(username="srch_admin", password="pw")
        resp = self.client.get("/search/?q=inventory")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Inventory")

    def test_suggest_endpoint_json(self):
        import json
        self.client.login(username="srch_admin", password="pw")
        resp = self.client.get("/search/suggest/?q=uom")
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.content)
        urls = [r["url"] for r in data["results"]]
        self.assertTrue(set(urls) & {"/uoms/", "/uom-conversions/"})

    def test_suggest_endpoint_permission_aware(self):
        import json
        self.client.login(username="srch_sales", password="pw")
        resp = self.client.get("/search/suggest/?q=journal")
        data = json.loads(resp.content)
        self.assertNotIn("/gl/journal/", [r["url"] for r in data["results"]])

    def test_menu_renders_grouped_collapsible_with_filter(self):
        self.client.login(username="srch_admin", password="pw")
        resp = self.client.get("/", follow=True)
        self.assertContains(resp, 'id="menuFilter"')         # menu search box
        self.assertContains(resp, "skn-navgroup")            # collapsible groups
        self.assertContains(resp, "Procurement")             # a grouped section
        self.assertContains(resp, 'id="globalSearchInput"')  # header search bar

    def test_important_modules_reachable_for_admin(self):
        from core.roles import ADMIN
        for q in ["inventory", "purchase orders", "customers", "uom",
                  "stock takes", "chart of accounts", "vat returns"]:
            self.assertTrue(self._urls(ADMIN, q), f"no results for {q!r}")


class StockTakeTests(TestCase):
    """Full physical stock-take: snapshot, blind entry, valuation, staleness
    guard, approval, GL posting, idempotency, closed periods, tenant isolation."""

    def setUp(self):
        import datetime
        from core.models import OrgMembership, Site, Bin
        from core.services.inventory import apply_movement
        self.t = Tenant.objects.create(name="StockTake Co")   # signal seeds GL
        self.site = Site.objects.create(tenant=self.t, name="HQ")
        self.loc = Location.objects.create(tenant=self.t, name="WH", site=self.site)
        # Lot/serial-tracked FIFO product.
        self.lp = Product.objects.create(tenant=self.t, sku="SKU-L", name="Lotted",
                                         cost_method=Product.CostMethod.FIFO,
                                         track_lots=True, track_expiry=True)
        self.sp = Product.objects.create(tenant=self.t, sku="SKU-S", name="Serial",
                                         cost_method=Product.CostMethod.FIFO, track_serial=True)
        # Plain (location-level) product.
        self.pp = Product.objects.create(tenant=self.t, sku="SKU-P", name="Plain")
        self.bin = Bin.objects.create(tenant=self.t, location=self.loc, code="A1")
        self.eA = datetime.date(2026, 6, 1)

        apply_movement(tenant=self.t, product=self.lp, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="la", unit_cost=Decimal("2.00"),
                       lot_code="A", expiry_date=self.eA)
        apply_movement(tenant=self.t, product=self.sp, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("1"), ref_type="T", ref_id="s1", unit_cost=Decimal("5.00"),
                       serial_number="S1")
        apply_movement(tenant=self.t, product=self.pp, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("20"), ref_type="T", ref_id="pa", unit_cost=Decimal("3.00"))

        self.user = User.objects.create_user("stu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t, role="ADMIN", is_default=True)
        self.client.login(username="stu", password="pw")

    # ---- helpers -----------------------------------------------------------
    def _session(self, scope="LOCATION"):
        from core.models import StockTakeSession
        s = StockTakeSession(tenant=self.t, status=StockTakeSession.Status.DRAFT,
                             reference="ST-T1")
        if scope == "SITE":
            s.scope = StockTakeSession.Scope.SITE
            s.site = self.site
        else:
            s.scope = StockTakeSession.Scope.LOCATION
            s.location = self.loc
        s.save()
        return s

    def _count(self, line, qty):
        from core.services.stock_take import recompute_variance
        from core.models import StockTakeLine
        line.counted_qty = Decimal(qty)
        line.count_status = StockTakeLine.CountStatus.COUNTED
        recompute_variance(line)
        line.save()

    def _line_for(self, session, product, **kw):
        return session.lines.get(product=product, **kw)

    def _approve(self, session):
        from core.models import StockTakeSession
        session.status = StockTakeSession.Status.APPROVED
        session.approved_by = self.user
        session.save()

    # ---- tests -------------------------------------------------------------
    def test_create_session_view(self):
        from django.urls import reverse
        from core.models import StockTakeSession
        resp = self.client.post(reverse("stock_take_create"), {
            "scope": "LOCATION", "location": self.loc.id,
            "count_date": "2026-06-09", "blind": "on", "notes": "first"})
        self.assertEqual(resp.status_code, 302)
        s = StockTakeSession.objects.get(tenant=self.t)
        self.assertEqual(s.location_id, self.loc.id)
        self.assertTrue(s.reference)
        self.assertEqual(s.started_by, self.user)

    def test_snapshot_generates_lines_for_whole_location(self):
        from core.services.stock_take import generate_snapshot
        from core.models import StockTakeSession
        s = self._session()
        n = generate_snapshot(s, user=self.user)
        self.assertEqual(n, 3)  # one lot line, one serial line, one plain line
        s.refresh_from_db()
        self.assertEqual(s.status, StockTakeSession.Status.SNAPSHOTTED)
        self.assertIsNotNone(s.snapshot_at)
        lot_line = self._line_for(s, self.lp)
        self.assertEqual(lot_line.lot_code, "A")
        self.assertEqual(lot_line.expected_qty_snapshot, Decimal("10.00"))
        self.assertEqual(lot_line.expected_unit_cost, Decimal("2.0000"))  # lot layer cost
        self.assertEqual(lot_line.expected_value_snapshot, Decimal("20.00"))

    def test_snapshot_whole_site_scope(self):
        from core.services.stock_take import generate_snapshot
        s = self._session(scope="SITE")
        n = generate_snapshot(s, user=self.user)
        self.assertEqual(n, 3)
        self.assertEqual({l.location_id for l in s.lines.all()}, {self.loc.id})

    def test_blind_count_entry_hides_expected(self):
        from django.urls import reverse
        from core.services.stock_take import generate_snapshot
        s = self._session()
        s.blind = True
        s.save()
        generate_snapshot(s, user=self.user)
        resp = self.client.get(reverse("stock_take_count", args=[s.id]))
        self.assertEqual(resp.status_code, 200)
        self.assertNotContains(resp, "Expected")  # blind hides the expected column

    def test_count_entry_view_and_zero_count(self):
        from django.urls import reverse
        from core.services.stock_take import generate_snapshot
        from core.models import StockTakeLine
        s = self._session()
        generate_snapshot(s, user=self.user)
        plain = self._line_for(s, self.pp)
        # Record an explicit zero count for the plain product (full shrinkage).
        resp = self.client.post(reverse("stock_take_count", args=[s.id]),
                                {f"counted_{plain.id}": "0"})
        self.assertEqual(resp.status_code, 302)
        plain.refresh_from_db()
        self.assertEqual(plain.counted_qty, Decimal("0.00"))
        self.assertEqual(plain.variance_qty, Decimal("-20.00"))
        self.assertEqual(plain.count_status, StockTakeLine.CountStatus.COUNTED)

    def test_positive_and_negative_variance_post(self):
        from core.services.stock_take import generate_snapshot, post_session
        s = self._session()
        generate_snapshot(s, user=self.user)
        # Plain product: found 5 extra (positive); lot A: lost 1 (negative).
        self._count(self._line_for(s, self.pp), "25")
        self._count(self._line_for(s, self.lp), "9")
        self._approve(s)
        ok, reason = post_session(s, user=self.user)
        self.assertTrue(ok)
        bal_p = InventoryBalance.objects.get(tenant=self.t, product=self.pp, location=self.loc)
        self.assertEqual(bal_p.on_hand, Decimal("25.00"))   # 20 + 5
        bal_l = InventoryBalance.objects.get(tenant=self.t, product=self.lp, location=self.loc)
        self.assertEqual(bal_l.on_hand, Decimal("9.00"))    # 10 - 1

    def test_lot_specific_valuation(self):
        from core.services.stock_take import generate_snapshot, post_session
        s = self._session()
        generate_snapshot(s, user=self.user)
        self._count(self._line_for(s, self.lp), "9")   # -1 of lot A @ cost 2.00
        self._approve(s)
        post_session(s, user=self.user)
        mv = InventoryMovement.objects.get(tenant=self.t, ref_type="STOCK_TAKE", product=self.lp)
        self.assertEqual(mv.value, Decimal("-2.00"))   # lot cost, not product average

    def test_serial_specific_valuation(self):
        from core.services.stock_take import generate_snapshot, post_session
        s = self._session()
        generate_snapshot(s, user=self.user)
        self._count(self._line_for(s, self.sp), "0")   # serial S1 missing
        self._approve(s)
        post_session(s, user=self.user)
        mv = InventoryMovement.objects.get(tenant=self.t, ref_type="STOCK_TAKE", product=self.sp)
        self.assertEqual(mv.value, Decimal("-5.00"))   # the serial's own cost

    def test_bin_level_line(self):
        from core.services.inventory import apply_movement
        from core.services.stock_take import generate_snapshot, post_session
        from core.models import InventoryBinBalance
        # A bin-tracked product (no lot tracking) so a bin line is generated.
        bp = Product.objects.create(tenant=self.t, sku="SKU-B", name="Binned")
        apply_movement(tenant=self.t, product=bp, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("12"), ref_type="T", ref_id="bb", unit_cost=Decimal("1.00"),
                       bin=self.bin)
        s = self._session()
        generate_snapshot(s, user=self.user)
        bin_line = self._line_for(s, bp)
        self.assertEqual(bin_line.bin_id, self.bin.id)
        self.assertEqual(bin_line.expected_qty_snapshot, Decimal("12.00"))
        self._count(bin_line, "10")
        self._approve(s)
        ok, _ = post_session(s, user=self.user)
        self.assertTrue(ok)
        bb = InventoryBinBalance.objects.get(tenant=self.t, product=bp, location=self.loc, bin=self.bin)
        self.assertEqual(bb.on_hand, Decimal("10.00"))

    def test_stale_line_detected_and_cannot_post_silently(self):
        from core.services.inventory import apply_movement
        from core.services.stock_take import generate_snapshot, post_session
        from core.models import StockTakeSession, StockTakeLine, JournalEntry
        s = self._session()
        generate_snapshot(s, user=self.user)
        self._count(self._line_for(s, self.pp), "20")  # no variance at snapshot
        self._approve(s)
        # Stock moves AFTER approval (book now 25, snapshot said 20).
        apply_movement(tenant=self.t, product=self.pp, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("5"), ref_type="T", ref_id="late", unit_cost=Decimal("3.00"))
        ok, reason = post_session(s, user=self.user)
        self.assertFalse(ok)
        self.assertEqual(reason, "stale")
        s.refresh_from_db()
        self.assertEqual(s.status, StockTakeSession.Status.REVIEW)   # bounced back
        self.assertIsNone(s.approved_by_id)
        line = self._line_for(s, self.pp)
        self.assertEqual(line.count_status, StockTakeLine.CountStatus.STALE)
        self.assertEqual(line.expected_qty_snapshot, Decimal("25.00"))  # refreshed to live
        # Nothing posted: no stock-take movement or journal.
        self.assertFalse(InventoryMovement.objects.filter(tenant=self.t, ref_type="STOCK_TAKE").exists())
        self.assertFalse(JournalEntry.objects.filter(tenant=self.t, ref_type="STOCK_TAKE").exists())

    def test_approval_required_before_posting(self):
        from core.services.stock_take import generate_snapshot, post_session
        from core.models import StockTakeSession
        s = self._session()
        generate_snapshot(s, user=self.user)
        self._count(self._line_for(s, self.lp), "9")
        s.status = StockTakeSession.Status.REVIEW
        s.save()
        ok, reason = post_session(s, user=self.user)
        self.assertFalse(ok)
        self.assertEqual(reason, "not_approved")
        self.assertFalse(InventoryMovement.objects.filter(tenant=self.t, ref_type="STOCK_TAKE").exists())

    def test_posting_creates_movement_and_gl(self):
        from core.services.stock_take import generate_snapshot, post_session
        from core.models import JournalEntry
        s = self._session()
        generate_snapshot(s, user=self.user)
        self._count(self._line_for(s, self.lp), "9")   # -1 @ 2.00 = -2.00
        self._approve(s)
        post_session(s, user=self.user)
        self.assertTrue(InventoryMovement.objects.filter(tenant=self.t, ref_type="STOCK_TAKE").exists())
        je = JournalEntry.objects.get(tenant=self.t, ref_type="STOCK_TAKE", ref_id=str(s.id))
        self.assertEqual(_account_balance(self.t, "1000"), Decimal("-2.00"))  # inventory credited
        self.assertTrue(je.lines.exists())

    def test_posting_is_idempotent(self):
        from core.services.stock_take import generate_snapshot, post_session
        from core.models import JournalEntry
        s = self._session()
        generate_snapshot(s, user=self.user)
        self._count(self._line_for(s, self.lp), "9")
        self._approve(s)
        post_session(s, user=self.user)
        mv1 = InventoryMovement.objects.filter(tenant=self.t, ref_type="STOCK_TAKE").count()
        ok, reason = post_session(s, user=self.user)   # second call
        self.assertTrue(ok)
        self.assertEqual(reason, "already_posted")
        self.assertEqual(InventoryMovement.objects.filter(tenant=self.t, ref_type="STOCK_TAKE").count(), mv1)
        self.assertEqual(JournalEntry.objects.filter(tenant=self.t, ref_type="STOCK_TAKE").count(), 1)

    def test_closed_period_posts_to_current_period(self):
        import datetime
        from django.utils import timezone
        from core.services.stock_take import generate_snapshot, post_session
        from core.models import JournalEntry
        s = self._session()
        s.count_date = datetime.date(2026, 1, 1)
        s.save()
        generate_snapshot(s, user=self.user)
        self._count(self._line_for(s, self.lp), "9")
        self._approve(s)
        post_session(s, user=self.user, lock_date=datetime.date(2026, 3, 31))
        je = JournalEntry.objects.get(tenant=self.t, ref_type="STOCK_TAKE", ref_id=str(s.id))
        self.assertEqual(je.entry_date, timezone.localdate())  # shifted out of closed period

    def test_tenant_isolation(self):
        from django.urls import reverse
        from core.models import OrgMembership, StockTakeSession
        from core.services.stock_take import generate_snapshot
        s = self._session()
        generate_snapshot(s, user=self.user)
        # A second tenant + user must not see tenant A's session.
        t2 = Tenant.objects.create(name="Other Co")
        u2 = User.objects.create_user("other", password="pw")
        OrgMembership.objects.create(user=u2, tenant=t2, role="ADMIN", is_default=True)
        c2 = Client()
        c2.login(username="other", password="pw")
        resp = c2.get(reverse("stock_take_detail", args=[s.id]))
        self.assertEqual(resp.status_code, 404)
        self.assertEqual(StockTakeSession.objects.filter(tenant=t2).count(), 0)


class BackNavigationTests(TestCase):
    """Hybrid global Back: pages declare a logical back_url via the base.html
    block; the header button renders it. Local Back/Cancel links stay intact."""

    def setUp(self):
        from core.models import OrgMembership, StockTakeSession
        self.t = Tenant.objects.create(name="Back Co")
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.user = User.objects.create_user("backu", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t, role="ADMIN", is_default=True)
        self.client.login(username="backu", password="pw")
        self.session = StockTakeSession.objects.create(
            tenant=self.t, scope=StockTakeSession.Scope.LOCATION, location=self.loc,
            status=StockTakeSession.Status.SNAPSHOTTED, reference="ST-BK")

    def test_search_page_declares_dashboard_back(self):
        resp = self.client.get("/search/?q=po")
        self.assertContains(resp, 'data-back-url="/dashboard/"')

    def test_report_page_declares_reports_index_back(self):
        resp = self.client.get("/reports/stock-take/")
        self.assertContains(resp, 'data-back-url="/reports/"')

    def test_stock_take_detail_back_to_list(self):
        resp = self.client.get(f"/stock-takes/{self.session.id}/")
        self.assertContains(resp, 'data-back-url="/stock-takes/"')

    def test_stock_take_count_back_to_detail_and_keeps_local_link(self):
        resp = self.client.get(f"/stock-takes/{self.session.id}/count/")
        detail_url = f"/stock-takes/{self.session.id}/"
        # Global declared back -> detail.
        self.assertContains(resp, f'data-back-url="{detail_url}"')
        # Local "Back" button (to the same detail page) is NOT removed.
        self.assertContains(resp, f'href="{detail_url}">Back</a>')

    def test_create_form_declares_list_back_and_keeps_cancel(self):
        resp = self.client.get("/stock-takes/new/")
        self.assertContains(resp, 'data-back-url="/stock-takes/"')
        self.assertContains(resp, 'href="/stock-takes/">Cancel</a>')   # local cancel intact

    def test_po_pages_declare_list_back(self):
        self.assertContains(self.client.get("/po/new/"), 'data-back-url="/po/"')

    def test_dashboard_has_no_declared_back(self):
        # Root/landing pages declare no logical parent (JS also hides the button).
        resp = self.client.get("/", follow=True)
        self.assertContains(resp, 'data-back-url=""')


class BackNavCoverageTests(TestCase):
    """Representative back_url coverage across Sales, Procurement, Finance,
    Inventory and Administration (list/form/report pages)."""

    def setUp(self):
        from core.models import OrgMembership
        self.t = Tenant.objects.create(name="Cover Co")
        self.user = User.objects.create_user("coveru", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t, role="ADMIN", is_default=True)
        self.client.login(username="coveru", password="pw")

    # (url, expected data-back-url) — one or more per ERP area.
    CASES = [
        # Sales
        ("/customers/", "/dashboard/"), ("/customers/new/", "/customers/"),
        ("/quotes/", "/dashboard/"), ("/customer-orders/", "/dashboard/"),
        ("/ar/invoices/", "/dashboard/"), ("/ar/invoices/new/", "/ar/invoices/"),
        ("/returns/", "/dashboard/"), ("/recurring-invoices/", "/dashboard/"),
        ("/sales-orders/", "/dashboard/"),
        # Procurement
        ("/suppliers/", "/dashboard/"), ("/suppliers/new/", "/suppliers/"),
        ("/requisitions/", "/dashboard/"), ("/shipments/", "/dashboard/"),
        ("/invoices/", "/dashboard/"), ("/po/backorders/", "/po/"),
        # Finance
        ("/payments/", "/dashboard/"), ("/expenses/", "/dashboard/"),
        ("/expenses/new/", "/expenses/"), ("/credit-notes/", "/dashboard/"),
        ("/bank/transactions/", "/dashboard/"), ("/tax-codes/", "/dashboard/"),
        ("/gl/journal/", "/dashboard/"), ("/gl/accounts/", "/dashboard/"),
        # Inventory / master data
        ("/po/", "/dashboard/"), ("/inventory/", "/dashboard/"), ("/cycle-counts/", "/dashboard/"),
        ("/inventory/adjustments/", "/dashboard/"), ("/inventory/movements/", "/dashboard/"),
        ("/inventory/low-stock/", "/inventory/"), ("/transfers/", "/dashboard/"),
        ("/sites/", "/dashboard/"), ("/locations/", "/dashboard/"), ("/bins/", "/dashboard/"),
        ("/products/", "/dashboard/"), ("/products/new/", "/products/"),
        ("/product-categories/", "/dashboard/"), ("/boms/", "/dashboard/"),
        ("/uoms/", "/dashboard/"), ("/uom-conversions/", "/dashboard/"),
        # Reports
        ("/reports/profit-and-loss/", "/reports/"), ("/reports/balance-sheet/", "/reports/"),
        ("/reports/trial-balance/", "/reports/"), ("/reports/stock-valuation/", "/reports/"),
        # Administration
        ("/departments/", "/dashboard/"), ("/departments/new/", "/departments/"),
        ("/audit/", "/dashboard/"), ("/email-log/", "/dashboard/"),
        ("/settings/tenant/", "/dashboard/"), ("/users/", "/dashboard/"),
    ]

    def test_back_url_coverage(self):
        for url, expected in self.CASES:
            resp = self.client.get(url)
            self.assertEqual(resp.status_code, 200, f"{url} did not return 200")
            self.assertContains(resp, f'data-back-url="{expected}"',
                                msg_prefix=f"{url} missing back_url {expected}")


class SerialCorrectnessTests(TestCase):
    """Serial-number correctness package: ledger cardinality guard, AR-path serial
    carry-through + COGS, and serial-aware stock adjustments."""

    def setUp(self):
        from core.models import OrgMembership, Customer
        self.t = Tenant.objects.create(name="Serial Co")          # signal seeds GL
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.loc2 = Location.objects.create(tenant=self.t, name="WH2")
        self.sp = Product.objects.create(tenant=self.t, sku="SN-PROD", name="Serial Widget",
                                         cost_method=Product.CostMethod.FIFO, track_serial=True)
        self.plain = Product.objects.create(tenant=self.t, sku="PLAIN", name="Plain")
        self.cust = Customer.objects.create(tenant=self.t, name="Acme")
        self.user = User.objects.create_user("seru", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t, role="ADMIN", is_default=True)

    # ---- helpers -----------------------------------------------------------
    def _recv(self, serial, cost="7.00", product=None, location=None, tenant=None):
        from core.services.inventory import apply_movement
        return apply_movement(
            tenant=tenant or self.t, product=product or self.sp, location=location or self.loc,
            movement_type="RECEIVE", qty_delta=Decimal("1"), ref_type="T", ref_id="r" + serial,
            unit_cost=Decimal(cost), serial_number=serial)

    def _lot(self, serial, location=None, tenant=None):
        from core.models import InventoryLotBalance
        return InventoryLotBalance.objects.filter(
            tenant=tenant or self.t, product=self.sp, location=location or self.loc,
            serial_number=serial).first()

    # ---- Part 1: cardinality guard ----------------------------------------
    def test_valid_receipt_0_to_1(self):
        self._recv("SN1")
        self.assertEqual(self._lot("SN1").on_hand, Decimal("1.00"))

    def test_duplicate_receipt_blocked(self):
        from django.core.exceptions import ValidationError
        self._recv("SN1")
        with self.assertRaises(ValidationError):
            self._recv("SN1")
        self.assertEqual(self._lot("SN1").on_hand, Decimal("1.00"))  # unchanged

    def test_serial_on_hand_cannot_exceed_one(self):
        from django.core.exceptions import ValidationError
        from core.services.inventory import apply_movement
        with self.assertRaises(ValidationError):
            apply_movement(tenant=self.t, product=self.sp, location=self.loc,
                           movement_type="RECEIVE", qty_delta=Decimal("2"),
                           ref_type="T", ref_id="x", unit_cost=Decimal("7.00"), serial_number="SN1")

    def test_issue_below_zero_blocked_even_without_flag(self):
        from django.core.exceptions import ValidationError
        from core.services.inventory import apply_movement
        self.assertFalse(self.t.block_negative_stock)   # default off
        with self.assertRaises(ValidationError):        # serial not in stock
            apply_movement(tenant=self.t, product=self.sp, location=self.loc,
                           movement_type="SALE", qty_delta=Decimal("-1"),
                           ref_type="T", ref_id="y", serial_number="GHOST")

    def test_serial_required_on_movement(self):
        from django.core.exceptions import ValidationError
        from core.services.inventory import apply_movement
        with self.assertRaises(ValidationError):
            apply_movement(tenant=self.t, product=self.sp, location=self.loc,
                           movement_type="SALE", qty_delta=Decimal("-1"), ref_type="T", ref_id="z")

    def test_valid_sale_1_to_0(self):
        from core.services.inventory import apply_movement
        self._recv("SN1")
        apply_movement(tenant=self.t, product=self.sp, location=self.loc, movement_type="SALE",
                       qty_delta=Decimal("-1"), ref_type="T", ref_id="s", serial_number="SN1")
        self.assertEqual(self._lot("SN1").on_hand, Decimal("0.00"))

    def test_valid_return_0_to_1(self):
        from core.services.inventory import apply_movement
        self._recv("SN1")
        apply_movement(tenant=self.t, product=self.sp, location=self.loc, movement_type="SALE",
                       qty_delta=Decimal("-1"), ref_type="T", ref_id="s", serial_number="SN1")
        apply_movement(tenant=self.t, product=self.sp, location=self.loc, movement_type="RETURN",
                       qty_delta=Decimal("1"), ref_type="RMA", ref_id="r", serial_number="SN1")
        self.assertEqual(self._lot("SN1").on_hand, Decimal("1.00"))

    # ---- Part 2: AR path ---------------------------------------------------
    def _invoice(self, serial=None, product=None):
        from core.models import CustomerInvoice, CustomerInvoiceLine
        inv = CustomerInvoice.objects.create(
            tenant=self.t, customer=self.cust, invoice_number="INV-" + (serial or "x"),
            location=self.loc, status=CustomerInvoice.Status.DRAFT)
        CustomerInvoiceLine.objects.create(
            invoice=inv, product=product or self.sp, qty=Decimal("1"),
            unit_price=Decimal("10.00"), serial_number=serial)
        return inv

    def test_invoice_line_carries_serial(self):
        inv = self._invoice(serial="SN1")
        self.assertEqual(inv.lines.first().serial_number, "SN1")

    def test_invoice_cogs_posts_with_serial_and_issue_cost(self):
        from core.services.gl import post_customer_invoice
        from core.models import InventoryMovement, InventoryIssueCost, JournalEntry
        self._recv("SN1", cost="7.00")
        inv = self._invoice(serial="SN1")
        post_customer_invoice(inv, user=self.user)
        mv = InventoryMovement.objects.get(tenant=self.t, ref_type="AR_INVOICE", product=self.sp)
        self.assertEqual(mv.serial_number, "SN1")
        self.assertEqual(mv.value, Decimal("-7.00"))             # the serial's own layer cost
        self.assertEqual(self._lot("SN1").on_hand, Decimal("0.00"))
        ic = InventoryIssueCost.objects.get(movement=mv)
        self.assertEqual(ic.serial_number, "SN1")
        self.assertIsNotNone(ic.cost_layer)                      # linked to the serial's layer
        self.assertTrue(JournalEntry.objects.filter(tenant=self.t, ref_type="COGS", ref_id=inv.invoice_number).exists())

    def test_issue_serial_product_without_serial_blocked(self):
        from django.core.exceptions import ValidationError
        from core.services.gl import post_customer_invoice
        from core.models import CustomerInvoice
        self._recv("SN1")
        inv = self._invoice(serial=None)                         # serial-tracked, no serial
        with self.assertRaises(ValidationError):
            post_customer_invoice(inv, user=self.user)
        inv.refresh_from_db()
        self.assertEqual(inv.status, CustomerInvoice.Status.DRAFT)   # atomic rollback
        self.assertEqual(self._lot("SN1").on_hand, Decimal("1.00"))  # stock untouched

    def test_order_to_invoice_preserves_serial(self):
        from core.models import CustomerOrder, CustomerOrderLine
        from core.views import _invoice_from_lines
        order = CustomerOrder.objects.create(tenant=self.t, customer=self.cust, order_number="SO-1")
        CustomerOrderLine.objects.create(order=order, product=self.sp, qty=Decimal("1"),
                                         unit_price=Decimal("10"), serial_number="SN1")
        inv = _invoice_from_lines(self.t, self.cust, "GBP", "", "", order, "source_order", self.user)
        self.assertEqual(inv.lines.first().serial_number, "SN1")

    def test_quote_to_order_preserves_serial(self):
        from core.models import SalesQuote, SalesQuoteLine, CustomerOrder, CustomerOrderLine
        from core.views import _copy_sales_lines
        q = SalesQuote.objects.create(tenant=self.t, customer=self.cust, quote_number="Q-1")
        SalesQuoteLine.objects.create(quote=q, product=self.sp, qty=Decimal("1"),
                                      unit_price=Decimal("10"), serial_number="SN1")
        order = CustomerOrder.objects.create(tenant=self.t, customer=self.cust, order_number="SO-2")
        _copy_sales_lines(q, order, CustomerOrderLine, "order")
        self.assertEqual(order.lines.first().serial_number, "SN1")

    # ---- Part 3: stock adjustment -----------------------------------------
    def test_stock_adjustment_writeoff_with_serial(self):
        from core.models import StockAdjustment, InventoryMovement
        from core.views import _post_stock_adjustment
        self._recv("SN1")
        adj = StockAdjustment.objects.create(
            tenant=self.t, product=self.sp, location=self.loc,
            reason=StockAdjustment.Reason.WRITE_OFF, qty_delta=Decimal("-1"), serial_number="SN1")
        _post_stock_adjustment(adj, self.user)
        adj.refresh_from_db()
        self.assertEqual(adj.status, StockAdjustment.Status.POSTED)
        mv = InventoryMovement.objects.get(tenant=self.t, ref_type="STOCK_ADJ", ref_id=str(adj.id))
        self.assertEqual(mv.serial_number, "SN1")
        self.assertEqual(self._lot("SN1").on_hand, Decimal("0.00"))

    def test_stock_adjustment_serial_product_without_serial_blocked(self):
        from core.forms import StockAdjustmentForm
        form = StockAdjustmentForm(data={
            "product": self.sp.id, "location": self.loc.id,
            "reason": "WRITE_OFF", "qty_delta": "-1"})
        self.assertFalse(form.is_valid())
        self.assertIn("serial_number", form.errors)

    # ---- Regression: existing paths still work -----------------------------
    def test_channel_sales_path_still_works(self):
        from core.models import SalesOrder, SalesOrderLine, InventoryMovement
        from core.views import _post_sales_order
        self._recv("SN1")
        order = SalesOrder.objects.create(tenant=self.t, order_number="CH-1",
                                          ship_from_location=self.loc)
        SalesOrderLine.objects.create(order=order, product=self.sp, qty=Decimal("1"),
                                      serial_number="SN1", unit_price=Decimal("10"))
        _post_sales_order(order)
        mv = InventoryMovement.objects.get(tenant=self.t, ref_type="SALES_ORDER", product=self.sp)
        self.assertEqual(mv.serial_number, "SN1")
        self.assertEqual(self._lot("SN1").on_hand, Decimal("0.00"))

    def test_transfer_with_serial_still_works(self):
        from core.models import InventoryTransfer, InventoryTransferLine
        from core.views import _post_transfer
        self._recv("SN1")
        tr = InventoryTransfer.objects.create(
            tenant=self.t, transfer_number="TR-1", from_location=self.loc, to_location=self.loc2)
        InventoryTransferLine.objects.create(transfer=tr, product=self.sp, qty=Decimal("1"),
                                             serial_number="SN1")
        _post_transfer(tr, None)
        self.assertEqual(self._lot("SN1", location=self.loc).on_hand, Decimal("0.00"))
        self.assertEqual(self._lot("SN1", location=self.loc2).on_hand, Decimal("1.00"))

    def test_tenant_isolation(self):
        from core.models import OrgMembership
        other = Tenant.objects.create(name="Other Serial Co")
        oloc = Location.objects.create(tenant=other, name="OWH")
        op = Product.objects.create(tenant=other, sku="SN-PROD", name="Serial Widget",
                                    cost_method=Product.CostMethod.FIFO, track_serial=True)
        self._recv("SN1")                                        # tenant A
        # Same serial in another tenant is independent (no clash).
        self._recv("SN1", product=op, location=oloc, tenant=other)
        self.assertEqual(self._lot("SN1").on_hand, Decimal("1.00"))         # A unaffected
        from core.models import InventoryLotBalance
        ob = InventoryLotBalance.objects.get(tenant=other, product=op, location=oloc, serial_number="SN1")
        self.assertEqual(ob.on_hand, Decimal("1.00"))


class SerialPickerUiTests(TestCase):
    """Serial availability service + picker endpoint + availability page + serial
    visibility in tables (package 2)."""

    def setUp(self):
        from core.models import OrgMembership, Customer
        self.t = Tenant.objects.create(name="Picker Co")
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.sp = Product.objects.create(tenant=self.t, sku="SN-P", name="Serial Widget",
                                         cost_method=Product.CostMethod.FIFO, track_serial=True)
        self.plain = Product.objects.create(tenant=self.t, sku="PLAIN", name="Plain")
        self.cust = Customer.objects.create(tenant=self.t, name="Acme")
        self.user = User.objects.create_user("picku", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t, role="ADMIN", is_default=True)
        self.client.login(username="picku", password="pw")

    def _recv(self, serial, cost="7.00"):
        from core.services.inventory import apply_movement
        apply_movement(tenant=self.t, product=self.sp, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("1"), ref_type="T", ref_id="r" + serial,
                       unit_cost=Decimal(cost), serial_number=serial)

    def _issue(self, serial, mtype="SALE"):
        from core.services.inventory import apply_movement
        apply_movement(tenant=self.t, product=self.sp, location=self.loc, movement_type=mtype,
                       qty_delta=Decimal("-1"), ref_type="T", ref_id="i" + serial, serial_number=serial)

    def _avail(self, **kw):
        from core.services.inventory import available_serials
        return [r["serial_number"] for r in available_serials(self.t, **kw)]

    # ---- service ----
    def test_available_returns_only_available(self):
        self._recv("SN1"); self._recv("SN2")
        self.assertEqual(set(self._avail()), {"SN1", "SN2"})

    def test_issued_serial_excluded(self):
        self._recv("SN1"); self._recv("SN2")
        self._issue("SN1")                       # sold
        self.assertEqual(set(self._avail()), {"SN2"})

    def test_writeoff_and_rts_excluded(self):
        self._recv("SN1"); self._recv("SN2")
        self._issue("SN1", mtype="WRITE_OFF")
        self._issue("SN2", mtype="RETURN_SUPPLIER")
        self.assertEqual(self._avail(), [])

    def test_reserved_serial_excluded(self):
        from core.models import InventoryLotBalance
        self._recv("SN1")
        # Mark the unit as held by an active posting (reserved == on_hand).
        InventoryLotBalance.objects.filter(tenant=self.t, serial_number="SN1").update(reserved=Decimal("1.00"))
        self.assertEqual(self._avail(), [])

    def test_service_tenant_isolation(self):
        self._recv("SN1")
        other = Tenant.objects.create(name="Other Picker Co")
        from core.services.inventory import available_serials
        self.assertEqual(available_serials(other), [])

    # ---- endpoint ----
    def test_options_endpoint_returns_available(self):
        import json
        self._recv("SN1")
        resp = self.client.get(f"/inventory/serials/options/?product={self.sp.id}")
        data = json.loads(resp.content)
        self.assertTrue(data["track_serial"])
        self.assertIn("SN1", [s["serial"] for s in data["serials"]])

    def test_options_endpoint_non_serial_product_empty(self):
        import json
        resp = self.client.get(f"/inventory/serials/options/?product={self.plain.id}")
        data = json.loads(resp.content)
        self.assertFalse(data["track_serial"])
        self.assertEqual(data["serials"], [])

    def test_options_endpoint_tenant_scoped(self):
        import json
        from core.models import OrgMembership
        other = Tenant.objects.create(name="Outsider Co")
        op = Product.objects.create(tenant=other, sku="X", name="X", track_serial=True)
        resp = self.client.get(f"/inventory/serials/options/?product={op.id}")  # not my tenant
        data = json.loads(resp.content)
        self.assertFalse(data["track_serial"])

    # ---- duplicate / qty>1 backend enforcement ----
    def _invoice_two_lines(self, serial_a, serial_b, qty=Decimal("1")):
        from core.models import CustomerInvoice, CustomerInvoiceLine
        inv = CustomerInvoice.objects.create(tenant=self.t, customer=self.cust,
                                             invoice_number="INV-D", location=self.loc,
                                             status=CustomerInvoice.Status.DRAFT)
        CustomerInvoiceLine.objects.create(invoice=inv, product=self.sp, qty=qty,
                                           unit_price=Decimal("10"), serial_number=serial_a, description="a")
        if serial_b is not None:
            CustomerInvoiceLine.objects.create(invoice=inv, product=self.sp, qty=qty,
                                               unit_price=Decimal("10"), serial_number=serial_b, description="b")
        return inv

    def test_duplicate_serial_same_document_blocked(self):
        from django.core.exceptions import ValidationError
        from core.services.gl import post_customer_invoice
        self._recv("SN1")
        inv = self._invoice_two_lines("SN1", "SN1")
        with self.assertRaises(ValidationError):
            post_customer_invoice(inv, user=self.user)

    def test_qty_gt_one_with_one_serial_rejected(self):
        from django.core.exceptions import ValidationError
        from core.services.gl import post_customer_invoice
        self._recv("SN1")
        inv = self._invoice_two_lines("SN1", None, qty=Decimal("2"))
        with self.assertRaises(ValidationError):
            post_customer_invoice(inv, user=self.user)

    # ---- availability page ----
    def test_availability_page_renders_and_filters(self):
        self._recv("SN1"); self._recv("SN2")
        resp = self.client.get("/inventory/serials/")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "SN1")
        self.assertContains(resp, "SN2")
        resp = self.client.get("/inventory/serials/?q=SN1")
        self.assertContains(resp, "SN1")
        self.assertNotContains(resp, "SN2")

    # ---- serial visibility in tables ----
    def test_stock_movements_shows_and_filters_serial(self):
        self._recv("SN1")
        resp = self.client.get("/inventory/movements/?serial=SN1")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "SN1")

    def test_lot_trace_shows_serial(self):
        self._recv("SN1")
        resp = self.client.get(f"/reports/lot-trace/?product={self.sp.id}&serial=SN1")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "SN1")

    # ---- global search ----
    def test_global_search_finds_serial_availability(self):
        from core.roles import search_nav, ADMIN
        urls = [r["url"] for r in search_nav(ADMIN, "serial availability", limit=None)]
        self.assertIn("/inventory/serials/", urls)


class RmaDispositionTests(TestCase):
    """RMA receipt routes each return line by disposition: restock (sellable),
    quarantine/repair/RTS (owned hold, not sellable), scrap (write-off + GL)."""

    def setUp(self):
        from core.models import OrgMembership
        self.t = Tenant.objects.create(name="RMA Disp Co")          # signal seeds GL
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.sp = Product.objects.create(tenant=self.t, sku="SN-R", name="Serial Widget",
                                         cost_method=Product.CostMethod.FIFO, track_serial=True,
                                         average_cost=Decimal("5.0000"))
        self.plain = Product.objects.create(tenant=self.t, sku="PLAIN-R", name="Plain",
                                            average_cost=Decimal("3.0000"))
        self.user = User.objects.create_user("rmau", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t, role="ADMIN", is_default=True)

    def _rma(self, disposition, serial="SN1", product=None, qty=Decimal("1"), num=None):
        from core.models import ReturnAuthorization, ReturnLine
        rma = ReturnAuthorization.objects.create(
            tenant=self.t, channel="SHOPIFY",
            rma_number=num or f"RMA-{disposition}-{serial}", receive_location=self.loc)
        ReturnLine.objects.create(rma=rma, product=product or self.sp, qty=qty,
                                  serial_number=serial if (product or self.sp).track_serial else None,
                                  disposition=disposition)
        return rma

    def _avail(self):
        from core.services.inventory import available_serials
        return [r["serial_number"] for r in available_serials(self.t)]

    def _lot(self, serial):
        from core.models import InventoryLotBalance
        return InventoryLotBalance.objects.filter(tenant=self.t, product=self.sp, serial_number=serial).first()

    # ---- dispositions ----
    def test_restock_makes_serial_available(self):
        from core.services.returns import receive_return
        from core.models import ReturnAuthorization
        rma = self._rma("RESTOCK", "SN1")
        receive_return(rma, user=self.user)
        rma.refresh_from_db()
        self.assertEqual(rma.status, ReturnAuthorization.Status.RECEIVED)
        self.assertIn("SN1", self._avail())
        lb = self._lot("SN1")
        self.assertEqual(lb.on_hand, Decimal("1.00"))
        self.assertEqual(lb.location_id, self.loc.id)

    def test_quarantine_owned_but_not_available(self):
        from core.services.returns import receive_return
        rma = self._rma("QUARANTINE", "SN2")
        receive_return(rma, user=self.user)
        self.assertNotIn("SN2", self._avail())            # excluded from availability
        lb = self._lot("SN2")
        self.assertEqual(lb.on_hand, Decimal("1.00"))     # still owned
        self.assertEqual(lb.location.type, "QUARANTINE")  # routed to hold location

    def test_scrap_not_available_and_posts_writeoff_gl(self):
        from core.services.returns import receive_return
        from core.models import InventoryMovement, JournalEntry
        rma = self._rma("SCRAP", "SN3")
        receive_return(rma, user=self.user)
        self.assertNotIn("SN3", self._avail())
        self.assertEqual(self._lot("SN3").on_hand, Decimal("0.00"))   # received then written off
        wo = InventoryMovement.objects.get(tenant=self.t, ref_type="RMA_SCRAP", product=self.sp)
        self.assertEqual(wo.movement_type, "WRITE_OFF")
        self.assertEqual(wo.value, Decimal("-5.00"))                  # the unit's cost
        self.assertTrue(JournalEntry.objects.filter(tenant=self.t, ref_type="RMA_SCRAP").exists())
        self.assertEqual(_account_balance(self.t, "5200"), Decimal("5.00"))   # loss booked

    def test_default_disposition_is_quarantine(self):
        from core.models import ReturnAuthorization, ReturnLine
        rma = ReturnAuthorization.objects.create(tenant=self.t, channel="SHOPIFY",
                                                 rma_number="RMA-DEF", receive_location=self.loc)
        line = ReturnLine.objects.create(rma=rma, product=self.sp, qty=Decimal("1"), serial_number="SNx")
        self.assertEqual(line.disposition, ReturnLine.Disposition.QUARANTINE)

    def test_repair_and_rts_held_not_available(self):
        from core.services.returns import receive_return
        receive_return(self._rma("REPAIR", "SN4"), user=self.user)
        receive_return(self._rma("RETURN_TO_SUPPLIER", "SN5"), user=self.user)
        avail = self._avail()
        self.assertNotIn("SN4", avail)
        self.assertNotIn("SN5", avail)
        self.assertEqual(self._lot("SN4").on_hand, Decimal("1.00"))   # owned hold

    def test_traceability_shows_disposition(self):
        from core.services.returns import receive_return
        from core.models import InventoryMovement
        receive_return(self._rma("QUARANTINE", "SN6"), user=self.user)
        mv = InventoryMovement.objects.get(tenant=self.t, ref_type="RMA", serial_number="SN6")
        self.assertIn("Quarantine", mv.notes)

    def test_non_serial_restock_and_scrap(self):
        from core.services.returns import receive_return
        from core.models import InventoryBalance, InventoryMovement
        receive_return(self._rma("RESTOCK", serial=None, product=self.plain, qty=Decimal("2"), num="RMA-NS1"),
                       user=self.user)
        bal = InventoryBalance.objects.get(tenant=self.t, product=self.plain, location=self.loc)
        self.assertEqual(bal.on_hand, Decimal("2.00"))
        receive_return(self._rma("SCRAP", serial=None, product=self.plain, qty=Decimal("1"), num="RMA-NS2"),
                       user=self.user)
        self.assertTrue(InventoryMovement.objects.filter(tenant=self.t, ref_type="RMA_SCRAP",
                                                         product=self.plain).exists())

    def test_idempotent_receive(self):
        from core.services.returns import receive_return
        from core.models import InventoryMovement, JournalEntry
        rma = self._rma("SCRAP", "SN7")
        receive_return(rma, user=self.user)
        receive_return(rma, user=self.user)   # second call no-op (status RECEIVED)
        self.assertEqual(InventoryMovement.objects.filter(tenant=self.t, ref_type="RMA",
                                                          serial_number="SN7").count(), 1)
        self.assertEqual(InventoryMovement.objects.filter(tenant=self.t, ref_type="RMA_SCRAP").count(), 1)
        self.assertEqual(JournalEntry.objects.filter(tenant=self.t, ref_type="RMA_SCRAP").count(), 1)

    def test_tenant_isolation(self):
        from core.services.returns import receive_return
        from core.services.inventory import available_serials
        receive_return(self._rma("RESTOCK", "SN8"), user=self.user)
        other = Tenant.objects.create(name="Other RMA Co")
        self.assertEqual(available_serials(other), [])
        self.assertIn("SN8", self._avail())


class RmaAccountingAndHoldTests(TestCase):
    """RMA return GL accounting (inventory <-> GL stays reconciled) and the
    post-receipt hold resolution workflow (release to sellable / scrap)."""

    def setUp(self):
        from core.models import OrgMembership
        self.t = Tenant.objects.create(name="RMA Acct Co")          # signal seeds GL
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.sp = Product.objects.create(tenant=self.t, sku="SN-A", name="Serial Widget",
                                         cost_method=Product.CostMethod.FIFO, track_serial=True,
                                         average_cost=Decimal("5.0000"))
        self.user = User.objects.create_user("rmaacct", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t, role="ADMIN", is_default=True)

    def _rma_receive(self, disposition, serial):
        from core.models import ReturnAuthorization, ReturnLine
        from core.services.returns import receive_return
        rma = ReturnAuthorization.objects.create(
            tenant=self.t, channel="SHOPIFY", rma_number=f"RMA-{disposition}-{serial}",
            receive_location=self.loc)
        line = ReturnLine.objects.create(rma=rma, product=self.sp, qty=Decimal("1"),
                                         serial_number=serial, disposition=disposition)
        receive_return(rma, user=self.user)
        return rma, ReturnLine.objects.get(pk=line.pk)

    def _variance(self):
        from core.services.reports import inventory_gl_reconciliation
        return inventory_gl_reconciliation(self.t)["variance"]

    def _avail(self):
        from core.services.inventory import available_serials
        return [r["serial_number"] for r in available_serials(self.t)]

    # ---- accounting ----
    def test_restock_posts_inventory_debit_cogs_credit_and_reconciles(self):
        self._rma_receive("RESTOCK", "SN1")
        self.assertEqual(_account_balance(self.t, "1000"), Decimal("5.00"))   # DR Inventory
        self.assertEqual(_account_balance(self.t, "5000"), Decimal("-5.00"))  # CR COGS (reversal)
        self.assertEqual(self._variance(), Decimal("0.00"))                   # no drift

    def test_quarantine_reconciles_without_making_available(self):
        self._rma_receive("QUARANTINE", "SN2")
        self.assertEqual(_account_balance(self.t, "1000"), Decimal("5.00"))   # owned inventory
        self.assertNotIn("SN2", self._avail())                               # but not sellable
        self.assertEqual(self._variance(), Decimal("0.00"))

    def test_scrap_nets_correctly_and_reconciles(self):
        self._rma_receive("SCRAP", "SN3")
        self.assertEqual(_account_balance(self.t, "1000"), Decimal("0.00"))   # in then out -> 0
        self.assertEqual(_account_balance(self.t, "5200"), Decimal("5.00"))   # write-off loss
        self.assertEqual(_account_balance(self.t, "5000"), Decimal("-5.00"))  # COGS reversed
        self.assertEqual(self._variance(), Decimal("0.00"))

    # ---- hold resolution ----
    def test_quarantine_release_to_sellable(self):
        from core.services.returns import resolve_hold
        from core.models import ReturnLine
        rma, line = self._rma_receive("QUARANTINE", "SN4")
        self.assertNotIn("SN4", self._avail())
        resolve_hold(line, ReturnLine.Disposition.RESTOCK, user=self.user)
        line.refresh_from_db()
        self.assertEqual(line.final_disposition, ReturnLine.Disposition.RESTOCK)
        self.assertIn("SN4", self._avail())                  # now sellable
        self.assertEqual(self._variance(), Decimal("0.00"))  # value-neutral move

    def test_quarantine_scrap_from_hold(self):
        from core.services.returns import resolve_hold
        from core.models import ReturnLine, InventoryMovement
        rma, line = self._rma_receive("QUARANTINE", "SN5")
        resolve_hold(line, ReturnLine.Disposition.SCRAP, user=self.user)
        line.refresh_from_db()
        self.assertEqual(line.final_disposition, ReturnLine.Disposition.SCRAP)
        self.assertNotIn("SN5", self._avail())
        self.assertTrue(InventoryMovement.objects.filter(tenant=self.t, ref_type="RMA_HOLD_SCRAP",
                                                         serial_number="SN5").exists())
        self.assertEqual(_account_balance(self.t, "5200"), Decimal("5.00"))   # loss booked
        self.assertEqual(self._variance(), Decimal("0.00"))

    def test_serial_preserved_through_release(self):
        from core.services.returns import resolve_hold
        from core.models import ReturnLine, InventoryMovement
        rma, line = self._rma_receive("REPAIR", "SN6")
        resolve_hold(line, ReturnLine.Disposition.RESTOCK, user=self.user)
        mv = InventoryMovement.objects.filter(tenant=self.t, ref_type="RMA_RELEASE",
                                              movement_type="TRANSFER_IN").first()
        self.assertEqual(mv.serial_number, "SN6")
        self.assertEqual(mv.location_id, self.loc.id)        # back at sellable location

    def test_invalid_duplicate_transition_blocked(self):
        from django.core.exceptions import ValidationError
        from core.services.returns import resolve_hold
        from core.models import ReturnLine, InventoryMovement
        rma, line = self._rma_receive("QUARANTINE", "SN7")
        resolve_hold(line, ReturnLine.Disposition.RESTOCK, user=self.user)
        line.refresh_from_db()
        with self.assertRaises(ValidationError):             # already resolved
            resolve_hold(line, ReturnLine.Disposition.SCRAP, user=self.user)
        # No extra movements from the blocked second transition.
        self.assertFalse(InventoryMovement.objects.filter(tenant=self.t, ref_type="RMA_HOLD_SCRAP",
                                                          serial_number="SN7").exists())

    def test_cannot_resolve_non_hold_or_unreceived(self):
        from django.core.exceptions import ValidationError
        from core.services.returns import resolve_hold
        from core.models import ReturnAuthorization, ReturnLine
        # A RESTOCK line is not a hold -> cannot be "resolved".
        rma, line = self._rma_receive("RESTOCK", "SN8")
        with self.assertRaises(ValidationError):
            resolve_hold(line, ReturnLine.Disposition.SCRAP, user=self.user)

    def test_tenant_isolation(self):
        from core.services.inventory import available_serials
        self._rma_receive("RESTOCK", "SN9")
        other = Tenant.objects.create(name="Other Acct Co")
        self.assertEqual(available_serials(other), [])
        self.assertEqual(self._variance(), Decimal("0.00"))


class ReplenishmentTests(TestCase):
    """Replenishment planning: projected availability, suggestions (MOQ/pack/max),
    status, ABC, transfer suggestion, and requisition creation."""

    def setUp(self):
        from core.models import OrgMembership, Supplier
        self.t = Tenant.objects.create(name="Replen Co")
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.loc2 = Location.objects.create(tenant=self.t, name="WH2")
        self.sup = Supplier.objects.create(tenant=self.t, name="Acme Supply")
        self.p = Product.objects.create(tenant=self.t, sku="P1", name="Widget",
                                        average_cost=Decimal("10.0000"))
        self.user = User.objects.create_user("repl", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t, role="ADMIN", is_default=True)

    def _bal(self, on_hand, reserved="0", product=None, location=None):
        from core.models import InventoryBalance
        return InventoryBalance.objects.create(
            tenant=self.t, product=product or self.p, location=location or self.loc,
            on_hand=Decimal(on_hand), reserved=Decimal(reserved))

    def _policy(self, product=None, location=None, **kw):
        from core.models import ReplenishmentPolicy
        return ReplenishmentPolicy.objects.create(
            tenant=self.t, product=product or self.p, location=location, **kw)

    def _row(self, product=None, location=None, **plankw):
        from core.services import replenishment as rsvc
        product = product or self.p
        location = location or self.loc
        for r in rsvc.plan(self.t, **plankw):
            if r["product"].id == product.id and r["location"].id == location.id:
                return r
        return None

    # ---- formula / suggestions ----
    def test_below_reorder_suggests_purchase_refill_to_max(self):
        self._bal("5")
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"))
        r = self._row()
        self.assertEqual(r["status"], "below_reorder")
        self.assertEqual(r["projected_available"], Decimal("5.00"))
        self.assertEqual(r["suggested_purchase_qty"], Decimal("45.00"))   # 50 - 5

    def test_projected_includes_inbound_po(self):
        from core.models import PurchaseOrder, PurchaseOrderLine
        self._bal("5")
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"))
        po = PurchaseOrder.objects.create(tenant=self.t, po_number="PO1", supplier=self.sup,
                                          receiving_location=self.loc, status=PurchaseOrder.Status.APPROVED)
        PurchaseOrderLine.objects.create(po=po, product=self.p, ordered_qty=Decimal("20"),
                                         unit_cost=Decimal("10"))
        r = self._row()
        self.assertEqual(r["inbound_po"], Decimal("20"))
        self.assertEqual(r["projected_available"], Decimal("25.00"))       # 5 + 20
        self.assertEqual(r["status"], "okay")                              # covered

    def test_projected_includes_in_transit(self):
        from core.models import InventoryTransfer, InventoryTransferLine
        self._bal("5")
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"))
        tr = InventoryTransfer.objects.create(tenant=self.t, transfer_number="TR1",
                                              from_location=self.loc2, to_location=self.loc,
                                              status=InventoryTransfer.Status.DISPATCHED)
        InventoryTransferLine.objects.create(transfer=tr, product=self.p, qty=Decimal("15"),
                                             dispatched_qty=Decimal("15"), received_qty=Decimal("0"))
        r = self._row()
        self.assertEqual(r["in_transit"], Decimal("15"))
        self.assertEqual(r["projected_available"], Decimal("20.00"))       # 5 + 15

    def test_reserved_reduces_projected(self):
        self._bal("20", reserved="8")
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"))
        r = self._row()
        self.assertEqual(r["projected_available"], Decimal("12.00"))       # 20 - 8

    def test_moq_respected(self):
        self._bal("0")
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"), moq=Decimal("100"))
        r = self._row()
        self.assertEqual(r["suggested_purchase_qty"], Decimal("100.00"))   # need 50, MOQ 100

    def test_pack_size_rounding(self):
        self._bal("5")
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"), pack_size=Decimal("12"))
        r = self._row()
        self.assertEqual(r["suggested_purchase_qty"], Decimal("48.00"))    # 45 -> 4 packs of 12

    def test_overstock_status(self):
        self._bal("80")
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"))
        r = self._row()
        self.assertEqual(r["status"], "overstock")
        self.assertEqual(r["suggested_purchase_qty"], Decimal("0.00"))

    def test_preferred_supplier_from_policy(self):
        from core.models import Supplier
        other_sup = Supplier.objects.create(tenant=self.t, name="Override Supply")
        self.p.preferred_supplier = self.sup
        self.p.save()
        self._bal("5")
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"), preferred_supplier=other_sup)
        r = self._row()
        self.assertEqual(r["preferred_supplier"].id, other_sup.id)

    def test_transfer_suggestion_from_excess_location(self):
        self._bal("5", location=self.loc)
        self._bal("100", location=self.loc2)
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"))  # product-level
        r = self._row(location=self.loc)
        self.assertGreater(r["suggested_transfer_qty"], Decimal("0"))
        self.assertEqual(r["transfer_from"].id, self.loc2.id)

    # ---- ABC ----
    def test_abc_classification(self):
        from core.services.replenishment import abc_classes
        pa = Product.objects.create(tenant=self.t, sku="A1", name="A", average_cost=Decimal("10"))
        pb = Product.objects.create(tenant=self.t, sku="B1", name="B", average_cost=Decimal("1"))
        pc = Product.objects.create(tenant=self.t, sku="C1", name="C", average_cost=Decimal("1"))
        self._bal("8", product=pa)    # value 80
        self._bal("15", product=pb)   # value 15
        self._bal("5", product=pc)    # value 5
        cls = abc_classes(self.t)
        self.assertEqual(cls[pa.id], "A")
        self.assertEqual(cls[pb.id], "B")
        self.assertEqual(cls[pc.id], "C")

    # ---- requisition integration ----
    def test_requisition_created_from_suggestion(self):
        from core.models import PurchaseRequisition, PurchaseRequisitionLine
        self.p.preferred_supplier = self.sup
        self.p.save()
        self._bal("5")
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"))
        self.client.login(username="repl", password="pw")
        token = f"{self.p.id}:{self.loc.id}"
        resp = self.client.post("/inventory/replenishment/reorder/",
                                {"select": [token], f"qty_{token}": "45"})
        self.assertEqual(resp.status_code, 302)
        req = PurchaseRequisition.objects.get(tenant=self.t)
        self.assertEqual(req.preferred_supplier_id, self.sup.id)
        line = req.lines.get(product=self.p)
        self.assertEqual(line.quantity, Decimal("45"))

    def test_duplicate_open_requisition_skipped(self):
        from core.models import PurchaseRequisition, PurchaseRequisitionLine
        self._bal("5")
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"))
        existing = PurchaseRequisition.objects.create(
            tenant=self.t, req_number="PR-OLD", status=PurchaseRequisition.Status.DRAFT,
            requested_by=self.user)
        PurchaseRequisitionLine.objects.create(requisition=existing, product=self.p, quantity=Decimal("10"))
        self.client.login(username="repl", password="pw")
        token = f"{self.p.id}:{self.loc.id}"
        self.client.post("/inventory/replenishment/reorder/", {"select": [token], f"qty_{token}": "45"})
        # No new requisition created for the already-open product.
        self.assertEqual(PurchaseRequisition.objects.filter(tenant=self.t).count(), 1)

    # ---- isolation / nav ----
    def test_tenant_isolation(self):
        from core.services import replenishment as rsvc
        self._bal("5")
        self._policy(reorder_point=Decimal("10"), max_stock=Decimal("50"))
        other = Tenant.objects.create(name="Other Replen Co")
        self.assertEqual(rsvc.plan(other), [])

    def test_search_finds_replenishment(self):
        from core.roles import search_nav, ADMIN
        for q in ("reorder planning", "purchase suggestions", "safety stock", "abc analysis"):
            urls = [r["url"] for r in search_nav(ADMIN, q, limit=None)]
            self.assertIn("/inventory/replenishment/", urls, f"{q!r} did not find replenishment")


class LandedCostAccrualTests(TestCase):
    """Landed-cost accrual (2150) is credited at receipt and cleared (DR 2150 /
    into AP) when the supplier invoice posts — idempotent and bounded."""

    def setUp(self):
        self.t = Tenant.objects.create(name="Landed Co")   # signal seeds GL
        self.sup = Supplier.objects.create(tenant=self.t, name="S")
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.p = Product.objects.create(tenant=self.t, sku="LC", name="P")

    def _receipt(self, num, goods_qty="10", goods_cost="5.00", landed="0.00",
                 product=None, inventory_value=None):
        from core.models import PurchaseOrder, PurchaseOrderLine, GoodsReceipt, GoodsReceiptLine, LandedCostCharge
        from core.services.gl import post_inventory_receipt
        p = product or self.p
        goods_qty = Decimal(goods_qty); goods_cost = Decimal(goods_cost); landed = Decimal(landed)
        po = PurchaseOrder.objects.create(tenant=self.t, po_number="PO-" + num, supplier=self.sup)
        pol = PurchaseOrderLine.objects.create(po=po, product=p, ordered_qty=goods_qty, unit_cost=goods_cost)
        grn = GoodsReceipt.objects.create(tenant=self.t, po=po, grn_number="GRN-" + num,
                                          received_to=self.loc, status=GoodsReceipt.Status.POSTED)
        GoodsReceiptLine.objects.create(receipt=grn, po_line=pol, product=p,
                                        qty_received=goods_qty, unit_cost=goods_cost)
        if landed > 0:
            LandedCostCharge.objects.create(tenant=self.t, receipt=grn, name="Freight", amount=landed)
        goods_value = goods_qty * goods_cost
        inv_val = inventory_value if inventory_value is not None else (goods_value + landed)
        post_inventory_receipt(self.t, goods_value, grn.grn_number, landed_value=landed, inventory_value=inv_val)
        return po, grn

    def _invoice(self, po, grn, number, qty="10", cost="5.00", product=None):
        from core.models import SupplierInvoice, SupplierInvoiceLine
        inv = SupplierInvoice.objects.create(tenant=self.t, supplier=self.sup, po=po, receipt=grn,
                                             invoice_number=number)
        SupplierInvoiceLine.objects.create(invoice=inv, product=product or self.p,
                                            qty=Decimal(qty), unit_cost=Decimal(cost))  # tax_code None -> no VAT
        return inv

    # ---- receipt ----
    def test_receipt_with_landed_credits_accruals(self):
        self._receipt("R1", landed="20.00")
        self.assertEqual(_account_balance(self.t, "2150"), Decimal("-20.00"))  # accrual credit
        self.assertEqual(_account_balance(self.t, "2100"), Decimal("-50.00"))  # GRNI credit

    def test_receipt_without_landed_does_not_touch_2150(self):
        from core.models import JournalLine
        self._receipt("R0", landed="0.00")
        self.assertEqual(_account_balance(self.t, "2150"), Decimal("0.00"))
        self.assertFalse(JournalLine.objects.filter(entry__tenant=self.t, account__code="2150").exists())

    # ---- invoice clears ----
    def test_supplier_invoice_clears_accrual_and_grni(self):
        from core.services.gl import post_supplier_invoice
        po, grn = self._receipt("R2", landed="20.00")
        inv = self._invoice(po, grn, "SINV-R2")
        je = post_supplier_invoice(inv)
        inv.refresh_from_db()
        self.assertEqual(je.total_debit, je.total_credit)                      # balanced
        self.assertEqual(_account_balance(self.t, "2100"), Decimal("0.00"))    # GRNI cleared
        self.assertEqual(_account_balance(self.t, "2150"), Decimal("0.00"))    # accrual cleared
        self.assertEqual(inv.landed_cleared, Decimal("20.00"))

    def test_ap_credit_equals_invoice_payable(self):
        from core.services.gl import post_supplier_invoice
        po, grn = self._receipt("R3", landed="20.00")
        inv = self._invoice(po, grn, "SINV-R3")
        post_supplier_invoice(inv)
        inv.refresh_from_db()
        self.assertEqual(inv.total, Decimal("70.00"))                          # 50 goods + 0 VAT + 20 landed
        self.assertEqual(_account_balance(self.t, "2000"), -inv.total)         # AP credit == payable

    def test_repost_does_not_double_clear(self):
        from core.services.gl import post_supplier_invoice
        from core.models import JournalEntry
        po, grn = self._receipt("R4", landed="20.00")
        inv = self._invoice(po, grn, "SINV-R4")
        post_supplier_invoice(inv)
        post_supplier_invoice(inv)   # idempotent re-post
        self.assertEqual(_account_balance(self.t, "2150"), Decimal("0.00"))    # not +20 again
        self.assertEqual(_account_balance(self.t, "2000"), Decimal("-70.00"))  # AP not doubled
        self.assertEqual(JournalEntry.objects.filter(tenant=self.t, ref_type="AP_INVOICE",
                                                     ref_id="SINV-R4").count(), 1)

    def test_partial_multiple_invoices_clear_accrual_once(self):
        from core.services.gl import post_supplier_invoice
        po, grn = self._receipt("R5", landed="20.00")
        inv1 = self._invoice(po, grn, "SINV-R5a", qty="6", cost="5.00")
        inv2 = self._invoice(po, grn, "SINV-R5b", qty="4", cost="5.00")
        post_supplier_invoice(inv1)
        post_supplier_invoice(inv2)
        inv1.refresh_from_db(); inv2.refresh_from_db()
        self.assertEqual(inv1.landed_cleared, Decimal("20.00"))   # first invoice clears full accrual
        self.assertEqual(inv2.landed_cleared, Decimal("0.00"))    # second clears nothing (bounded)
        self.assertEqual(_account_balance(self.t, "2150"), Decimal("0.00"))   # never over-cleared

    def test_trial_balance_no_uncleared_landed_after_invoice(self):
        from core.services.gl import post_supplier_invoice
        po, grn = self._receipt("R6", landed="15.00")
        post_supplier_invoice(self._invoice(po, grn, "SINV-R6"))
        self.assertEqual(_account_balance(self.t, "2150"), Decimal("0.00"))

    # ---- standard cost documentation ----
    def test_standard_cost_landed_to_ppv_but_accrual_still_clears(self):
        from core.services.gl import post_supplier_invoice
        sp = Product.objects.create(tenant=self.t, sku="LC-STD", name="Std",
                                    cost_method=Product.CostMethod.STANDARD, standard_cost=Decimal("9.00"))
        # goods 10@10=100 actual, landed 20, inventory carried at standard 10*9=90.
        po, grn = self._receipt("R7", goods_qty="10", goods_cost="10.00", landed="20.00",
                                product=sp, inventory_value=Decimal("90.00"))
        self.assertEqual(_account_balance(self.t, "1000"), Decimal("90.00"))   # inventory at standard
        self.assertEqual(_account_balance(self.t, "2150"), Decimal("-20.00"))  # landed still accrued
        self.assertEqual(_account_balance(self.t, "5100"), Decimal("30.00"))   # PPV absorbs (120-90), incl landed
        # Invoice still clears the accrual.
        post_supplier_invoice(self._invoice(po, grn, "SINV-R7", qty="10", cost="10.00", product=sp))
        self.assertEqual(_account_balance(self.t, "2150"), Decimal("0.00"))

    def test_tenant_isolation(self):
        from core.services.gl import post_supplier_invoice
        po, grn = self._receipt("R8", landed="20.00")
        post_supplier_invoice(self._invoice(po, grn, "SINV-R8"))
        other = Tenant.objects.create(name="Other Landed Co")
        self.assertEqual(_account_balance(other, "2150"), Decimal("0.00"))


class PerformanceHardeningTests(TestCase):
    """N+1 elimination (available_serials, near-expiry, dashboard low-stock),
    pagination, and index presence — behaviour must be unchanged."""

    def setUp(self):
        from core.models import OrgMembership
        self.t = Tenant.objects.create(name="Perf Co")
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.sp = Product.objects.create(tenant=self.t, sku="SN-PERF", name="Serial",
                                         cost_method=Product.CostMethod.FIFO, track_serial=True)
        self.user = User.objects.create_user("perf", password="pw")
        OrgMembership.objects.create(user=self.user, tenant=self.t, role="ADMIN", is_default=True)

    def _recv_serial(self, serial, cost="7.00"):
        from core.services.inventory import apply_movement
        apply_movement(tenant=self.t, product=self.sp, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("1"), ref_type="GRN", ref_id="r" + serial,
                       unit_cost=Decimal(cost), serial_number=serial)

    # ---- available_serials ----
    def test_available_serials_values_correct_after_batching(self):
        from core.services.inventory import available_serials
        self._recv_serial("SN1", "7.00")
        self._recv_serial("SN2", "9.00")
        by_serial = {r["serial_number"]: r for r in available_serials(self.t)}
        self.assertEqual(set(by_serial), {"SN1", "SN2"})
        self.assertEqual(by_serial["SN1"]["unit_cost"], Decimal("7.0000"))
        self.assertEqual(by_serial["SN2"]["unit_cost"], Decimal("9.0000"))
        self.assertEqual(by_serial["SN1"]["source"], "GRN rSN1")
        self.assertIsNotNone(by_serial["SN1"]["received_at"])

    def test_available_serials_query_count_is_constant(self):
        from core.services.inventory import available_serials
        for i in range(6):
            self._recv_serial(f"S{i}", "5.00")
        with self.assertNumQueries(2):          # 1 balances + 1 cost layers, regardless of count
            rows = available_serials(self.t)
        self.assertEqual(len(rows), 6)

    def test_available_serials_excludes_unavailable(self):
        from core.services.inventory import available_serials, apply_movement
        from core.models import InventoryLotBalance
        self._recv_serial("AVAIL", "5.00")
        self._recv_serial("RESV", "5.00")
        self._recv_serial("SOLD", "5.00")
        InventoryLotBalance.objects.filter(tenant=self.t, serial_number="RESV").update(reserved=Decimal("1.00"))
        apply_movement(tenant=self.t, product=self.sp, location=self.loc, movement_type="SALE",
                       qty_delta=Decimal("-1"), ref_type="T", ref_id="s", serial_number="SOLD")
        avail = {r["serial_number"] for r in available_serials(self.t)}
        self.assertEqual(avail, {"AVAIL"})

    # ---- near-expiry ----
    def test_near_expiry_values_correct_and_batched(self):
        import datetime
        from core.services.inventory import apply_movement
        from core.services.reports import near_expiry_lots
        from core.models import InventoryLotBalance
        today = datetime.date(2026, 6, 10)
        soon = today + datetime.timedelta(days=5)
        p = Product.objects.create(tenant=self.t, sku="EXP", name="E",
                                   cost_method=Product.CostMethod.FIFO, track_lots=True, track_expiry=True)
        apply_movement(tenant=self.t, product=p, location=self.loc, movement_type="RECEIVE",
                       qty_delta=Decimal("10"), ref_type="T", ref_id="e", unit_cost=Decimal("3.00"),
                       lot_code="L1", expiry_date=soon)
        # A lot with no cost layer falls back to product cost.
        p2 = Product.objects.create(tenant=self.t, sku="EXP2", name="E2",
                                    average_cost=Decimal("4.00"), track_lots=True, track_expiry=True)
        InventoryLotBalance.objects.create(tenant=self.t, product=p2, location=self.loc,
                                           lot_code="L2", expiry_date=soon, on_hand=Decimal("2.00"))
        with self.assertNumQueries(2):          # 1 lot-balance scan + 1 cost-layer batch
            rows = near_expiry_lots(self.t, today=today)
        by_sku = {r["product"].sku: r for r in rows}
        self.assertEqual(by_sku["EXP"]["value"], Decimal("30.00"))          # 10 @ 3 (layer)
        self.assertEqual(by_sku["EXP"]["valuation_source"], "lot_layer")
        self.assertEqual(by_sku["EXP2"]["value"], Decimal("8.00"))          # 2 @ 4 (product cost)
        self.assertEqual(by_sku["EXP2"]["valuation_source"], "product_cost")

    # ---- dashboard low-stock ----
    def test_dashboard_low_stock_count_correct(self):
        from core.views import _dashboard_kpis
        from core.models import InventoryBalance
        below = Product.objects.create(tenant=self.t, sku="LOW", name="Low", reorder_level=Decimal("10"))
        InventoryBalance.objects.create(tenant=self.t, product=below, location=self.loc, on_hand=Decimal("3"))
        ok = Product.objects.create(tenant=self.t, sku="OK", name="Ok", reorder_level=Decimal("5"))
        InventoryBalance.objects.create(tenant=self.t, product=ok, location=self.loc, on_hand=Decimal("20"))
        nostock = Product.objects.create(tenant=self.t, sku="NONE", name="None", reorder_level=Decimal("2"))
        cards = {c["label"]: c["value"] for c in _dashboard_kpis(self.t, "ADMIN")}
        self.assertEqual(cards["Low-stock items"], 2)   # LOW (3<10) + NONE (0<2); OK excluded

    # ---- pagination ----
    def test_paginated_pages_render_and_keep_filters(self):
        for i in range(60):
            self._recv_serial(f"P{i:03d}", "5.00")
        self.client.login(username="perf", password="pw")
        resp = self.client.get("/inventory/serials/?q=P0")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["page_obj"].number, 1)
        self.assertIn("q=P0", resp.context["base_qs"])           # filter persists across pages
        # Second page also renders.
        resp2 = self.client.get("/inventory/serials/?q=P0&page=2")
        self.assertEqual(resp2.status_code, 200)

    def test_movements_page_paginates(self):
        for i in range(60):
            self._recv_serial(f"M{i:03d}", "5.00")
        self.client.login(username="perf", password="pw")
        resp = self.client.get("/inventory/movements/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context["page_obj"].object_list), 50)   # default page size
        self.assertTrue(resp.context["page_obj"].has_next())

    # ---- indexes ----
    def test_movement_indexes_present(self):
        from core.models import InventoryMovement
        idx_fields = {tuple(i.fields) for i in InventoryMovement._meta.indexes}
        self.assertIn(("tenant", "created_at"), idx_fields)
        self.assertIn(("tenant", "ref_type", "ref_id"), idx_fields)

    def test_tenant_isolation_preserved(self):
        from core.services.inventory import available_serials
        self._recv_serial("SN1", "5.00")
        other = Tenant.objects.create(name="Other Perf Co")
        self.assertEqual(available_serials(other), [])


class SerialReadinessAuditTests(TestCase):
    """Audit-first legacy serial readiness: detect-only, never mutates stock/GL."""

    def setUp(self):
        self.t = Tenant.objects.create(name="Audit Co")
        self.loc = Location.objects.create(tenant=self.t, name="WH")
        self.loc2 = Location.objects.create(tenant=self.t, name="WH2")

    def _sp(self, sku, fifo=False):
        return Product.objects.create(
            tenant=self.t, sku=sku, name=sku, track_serial=True,
            cost_method=Product.CostMethod.FIFO if fifo else Product.CostMethod.AVERAGE)

    def _lot(self, product, serial, on_hand, location=None):
        from core.models import InventoryLotBalance
        return InventoryLotBalance.objects.create(
            tenant=self.t, product=product, location=location or self.loc,
            serial_number=serial, on_hand=Decimal(on_hand))

    def _types(self, tenant=None):
        from core.services.serial_audit import audit_serial_readiness
        return {i["issue_type"] for i in audit_serial_readiness(tenant=tenant or self.t)}

    # ---- detections ----
    def test_clean_data_reports_no_issues(self):
        from core.models import InventoryBalance, InventoryCostLayer
        from core.services.serial_audit import audit_serial_readiness
        p = self._sp("CLEAN", fifo=True)
        self._lot(p, "SN1", "1")
        InventoryBalance.objects.create(tenant=self.t, product=p, location=self.loc, on_hand=Decimal("1"))
        InventoryCostLayer.objects.create(tenant=self.t, product=p, location=self.loc,
                                          serial_number="SN1", qty_received=Decimal("1"),
                                          qty_remaining=Decimal("1"), unit_cost=Decimal("5.00"))
        self.assertEqual(audit_serial_readiness(tenant=self.t), [])

    def test_onhand_gt_1_detected(self):
        self._lot(self._sp("GT1"), "SN1", "2")
        self.assertIn("ONHAND_GT_1", self._types())

    def test_blank_serial_balance_detected(self):
        self._lot(self._sp("BLANK"), "", "3")
        self.assertIn("BLANK_SERIAL_BALANCE", self._types())

    def test_blank_serial_movement_detected(self):
        from core.models import InventoryMovement
        p = self._sp("MOVE")
        InventoryMovement.objects.create(tenant=self.t, product=p, location=self.loc,
                                         movement_type="RECEIVE", qty_delta=Decimal("1"),
                                         ref_type="LEGACY", ref_id="x", serial_number="")
        self.assertIn("SERIALLESS_MOVEMENT", self._types())

    def test_duplicate_serial_across_locations_detected(self):
        p = self._sp("DUP")
        self._lot(p, "SN1", "1", location=self.loc)
        self._lot(p, "SN1", "1", location=self.loc2)
        self.assertIn("DUPLICATE_SERIAL", self._types())

    def test_negative_serial_balance_detected(self):
        self._lot(self._sp("NEG"), "SN1", "-1")
        self.assertIn("NEGATIVE_SERIAL_BALANCE", self._types())

    def test_cost_layer_missing_serial_detected(self):
        from core.models import InventoryCostLayer
        p = self._sp("LAYER", fifo=True)
        InventoryCostLayer.objects.create(tenant=self.t, product=p, location=self.loc,
                                          serial_number="", qty_received=Decimal("5"),
                                          qty_remaining=Decimal("5"), unit_cost=Decimal("2.00"))
        self.assertIn("COST_LAYER_MISSING_SERIAL", self._types())

    def test_serial_missing_cost_layer_detected(self):
        # FIFO serial on hand with no remaining cost layer.
        self._lot(self._sp("NOLAYER", fifo=True), "SN9", "1")
        self.assertIn("SERIAL_MISSING_COST_LAYER", self._types())

    def test_untracked_onhand_detected(self):
        from core.models import InventoryBalance
        p = self._sp("COVER")
        self._lot(p, "SN1", "1")                                  # 1 unit on a serial
        InventoryBalance.objects.create(tenant=self.t, product=p, location=self.loc, on_hand=Decimal("3"))
        self.assertIn("UNTRACKED_ONHAND", self._types())          # 2 units uncovered

    # ---- safety / isolation ----
    def test_dry_run_does_not_mutate(self):
        from io import StringIO
        from django.core.management import call_command
        from core.models import AuditLog, InventoryLotBalance
        lb = self._lot(self._sp("DR"), "SN1", "2")
        call_command("audit_serial_readiness", "--tenant", self.t.name, stdout=StringIO())
        self.assertEqual(AuditLog.objects.filter(tenant=self.t).count(), 0)      # nothing written
        self.assertEqual(InventoryLotBalance.objects.get(pk=lb.pk).on_hand, Decimal("2"))  # unchanged

    def test_apply_writes_audit_flags_only(self):
        from io import StringIO
        from django.core.management import call_command
        from core.models import AuditLog, InventoryLotBalance
        lb = self._lot(self._sp("AP"), "SN1", "2")
        call_command("audit_serial_readiness", "--tenant", self.t.name, "--apply", stdout=StringIO())
        self.assertTrue(AuditLog.objects.filter(tenant=self.t, action="SERIAL_AUDIT").exists())
        self.assertEqual(InventoryLotBalance.objects.get(pk=lb.pk).on_hand, Decimal("2"))  # stock untouched

    def test_csv_output(self):
        import os, tempfile
        from io import StringIO
        from django.core.management import call_command
        self._lot(self._sp("CSVP"), "SN1", "2")
        path = os.path.join(tempfile.mkdtemp(), "out.csv")
        call_command("audit_serial_readiness", "--tenant", self.t.name, "--csv", path, stdout=StringIO())
        with open(path, encoding="utf-8") as fh:
            content = fh.read()
        self.assertIn("issue_type", content)           # header
        self.assertIn("ONHAND_GT_1", content)          # at least one row

    def test_tenant_isolation(self):
        self._lot(self._sp("ISO"), "SN1", "2")          # issue in tenant A
        other = Tenant.objects.create(name="Other Audit Co")
        from core.services.serial_audit import audit_serial_readiness
        self.assertEqual(audit_serial_readiness(tenant=other), [])
